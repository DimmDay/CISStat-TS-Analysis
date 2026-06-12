import time
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io, os, re, json
from datetime import datetime, date
from pathlib import Path
from sqlalchemy import create_engine
import clickhouse_connect
import numpy as np
import re
import seaborn as sns
from scipy import stats
from scipy.stats import boxcox
from statsmodels.tsa.stattools import adfuller
from statsmodels.stats.diagnostic import acorr_ljungbox
from plotly.subplots import make_subplots
from statsmodels.tsa.seasonal import STL, seasonal_decompose
from typing import Dict, List, Tuple, Optional
from scipy.fft import fft, fftfreq
from scipy.signal import find_peaks, periodogram, welch
from statsmodels.graphics.tsaplots import plot_acf, plot_pacf
import pywt
import matplotlib.pyplot as plt
from statsmodels.tsa.stattools import acf, adfuller
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import hashlib

# ─────────────────────────────────────────────────────────────
# 🔧 ИМПОРТЫ МОДУЛЕЙ ВАЛИДАЦИИ
# ─────────────────────────────────────────────────────────────
from validation.engine import (
    load_rules,
    validate_dataframe,
    validate_formats,
    validate_consistency,
    validate_ranges,
    validate_referential,
    validate_text_quality,
    validate_regular_step,
    validate_sufficiency,
    auto_generate_rules
)
from validation.missing import analyze_missing, get_expert_list_df
from validation.outliers import detect_outliers, get_outliers_df
from validation.reporter import save_validated_dataset, generate_correction_report
from validation.audit import log_expert_action

from src.catalog.recommender import CISStatRecommender

# Инициализация рекомендателя
if "recommender" not in st.session_state:
    st.session_state.recommender = CISStatRecommender()

# ────────────────────────────────────────────────────────────
#  ⬤ ◉ ◎ ◌ ◍ ● ○ ◐ ◑ ◒ ◓ • ‣ ⁃ ∙ ∘ ∙ ∘ ∙ ⁝ ⁞ ⋮ ⋯ … ... ⋯ ⚫ ⚪ ⬛ ⬜ ◼️ ◻️ ◾ ◽ ▪️ ▫️ 🔴 🟠 🟡 🟢 🔵 🟣 🟤 ⚫ ⚪ ⭕

# ────────────────────────────────────────────────────────────
#  НАСТРОЙКА ШРИФТА
# ─────────────────────────────────────────────────────────────
# Выберите шрифт здесь.
# Можно использовать: 'Arial', 'Verdana', 'Helvetica', 'Times New Roman', 'Georgia'
FONT_FAMILY = "'Helvetica'" #, 'Arial', sans-serif"

st.markdown(f"""
<style>
    /* Основной шрифт для всего приложения */
    html, body, [class*="css"]  {{
        font-family: {FONT_FAMILY} !important;
    }}

    /* Заголовки */
    h1, h2, h3, h4, h5, h6, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {{
        font-family: {FONT_FAMILY} !important;
        font-weight: 700 !important;
    }}

    /* Текст в метриках */
    [data-testid="stMetricValue"], [data-testid="stMetricLabel"] {{
        font-family: {FONT_FAMILY} !important;
    }}

    /* Текст в таблицах (dataframe) */
    [data-testid="stDataFrame"] {{
        font-family: {FONT_FAMILY} !important;
    }}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# 🔧 УНИВЕРСАЛЬНАЯ ФУНКЦИЯ РАСЧЁТА ПАСПОРТА СВОЙСТВ РЯДА
# ─────────────────────────────────────────────────────────────
def calculate_ts_passport(analysis_series: pd.Series,
                          df_filtered: pd.DataFrame = None,
                          ct_f: dict = None,
                          target_col: str = None) -> dict:
    """
    Рассчитывает полный паспорт свойств временного ряда (13 метрик).
    Структура идентична паспорту во вкладке "Загрузка".
    """
    from scipy.stats import linregress, jarque_bera
    from scipy.signal import periodogram, find_peaks
    from scipy.fft import fft, fftfreq
    from statsmodels.tsa.stattools import adfuller, acf as acf_func
    from statsmodels.tsa.seasonal import STL
    from statsmodels.stats.diagnostic import acorr_ljungbox

    props = {}

    if len(analysis_series) < 30:
        return {"error": "Недостаточно данных (нужно > 30 точек)"}

    try:
        # 0. ЧАСТОТА РЯДА
        try:
            inferred_freq = pd.infer_freq(analysis_series.index.drop_duplicates().sort_values())
        except:
            inferred_freq = None
        props['freq'] = {
            'value': inferred_freq if inferred_freq else 'Нерегулярная',
            'is_ok': inferred_freq is not None
        }

        # 1. СТАЦИОНАРНОСТЬ (ADF)
        adf_res = adfuller(analysis_series.dropna(), autolag='AIC')
        adf_p = adf_res[1]
        is_stationary = adf_p < 0.05
        props['stationarity'] = {
            'value': float(adf_p),
            'is_stationary': bool(is_stationary),
            'is_ok': bool(is_stationary)
        }

        # 2. ДЕТЕРМИНИРОВАННОСТЬ (R² тренда)
        slope, intercept, r_value, p_value, std_err = linregress(
            range(len(analysis_series)), analysis_series.values
        )
        r_squared = float(r_value**2)
        is_deterministic = r_squared >= 0.7
        props['determinism'] = {
            'value': r_squared,
            'slope': float(slope),
            'is_deterministic': bool(is_deterministic)
        }

        # 3. АВТОКОРРЕЛЯЦИЯ (Ljung-Box)
        lb_res = acorr_ljungbox(analysis_series, lags=[10])
        if isinstance(lb_res, pd.DataFrame):
            lb_p = float(lb_res['lb_pvalue'].iloc[0])
        else:
            lb_p = float(lb_res[1][0])
        is_white_noise = lb_p > 0.05
        props['autocorrelation'] = {
            'value': lb_p,
            'is_white_noise': bool(is_white_noise),
            'is_ok': bool(is_white_noise)
        }

        # 4. НОРМАЛЬНОСТЬ (Jarque-Bera)
        jb_res = jarque_bera(analysis_series.dropna())
        jb_p = float(jb_res.pvalue) if hasattr(jb_res, 'pvalue') else float(jb_res[1])
        is_normal = jb_p > 0.05
        props['normality'] = {
            'value': jb_p,
            'is_normal': bool(is_normal),
            'is_ok': bool(is_normal)
        }

        # 5. НАПРАВЛЕНИЕ ТРЕНДА
        if slope > 0:
            trend_dir = 'up'
        elif slope < 0:
            trend_dir = 'down'
        else:
            trend_dir = 'flat'
        props['trend'] = {
            'slope': float(slope),
            'direction': trend_dir
        }

        # 6. КОРРЕЛЯЦИЯ ПРИЗНАКОВ
        props['correlations'] = {}
        if df_filtered is not None and ct_f is not None and target_col:
            try:
                num_cols = ct_f.get("num", [])
                if len(num_cols) >= 2 and target_col in num_cols:
                    corr_df = df_filtered[num_cols].corr()
                    target_corr = corr_df[target_col].drop(target_col).sort_values(key=abs, ascending=False)
                    top_corrs = {col: float(val) for col, val in target_corr.head(3).items()}
                    props['correlations'] = {
                        'top3': top_corrs,
                        'max_abs_corr': float(target_corr.iloc[0]) if len(target_corr) > 0 else 0.0
                    }
            except:
                pass

        # 7. СЕЗОННОСТЬ (STL strength)
        period = 7 if (inferred_freq and 'D' in str(inferred_freq)) else 12
        try:
            stl_res = STL(analysis_series, period=period, robust=True).fit()
            var_total = float(analysis_series.var())
            var_resid = float(stl_res.resid.var())
            var_detrended = var_total - float(stl_res.trend.var())
            strength_seasonality = max(0, 1 - var_resid / var_detrended) if var_detrended > 0 else 0
            is_seasonal = strength_seasonality > 0.6
        except:
            strength_seasonality = 0.0
            is_seasonal = False
        props['seasonality'] = {
            'strength': float(strength_seasonality),
            'is_seasonal': bool(is_seasonal)
        }

        # 8. СЕЗОННЫЕ ПЕРИОДЫ (ACF)
        try:
            max_lag = min(60, len(analysis_series) // 4)
            acf_values = acf_func(analysis_series, nlags=max_lag)
            confidence = 1.96 / np.sqrt(len(analysis_series))
            significant_lags = np.where(np.abs(acf_values) > confidence)[0][1:]

            seasonal_periods_acf = []
            for i, lag in enumerate(significant_lags):
                if i > 0 and lag - significant_lags[i-1] < 3:
                    continue
                if lag > 2:
                    seasonal_periods_acf.append(int(lag))

            props['seasonal_periods'] = {
                'periods': seasonal_periods_acf[:3],
                'count': len(seasonal_periods_acf[:3])
            }
        except:
            props['seasonal_periods'] = {'periods': [], 'count': 0}

        # 9. ДОЛГАЯ ПАМЯТЬ (Hurst)
        def hurst_exponent(series, max_lag=20):
            lags = range(2, max_lag)
            tau = [max(np.std(np.subtract(series[lag:], series[:-lag])), 1e-8) for lag in lags]
            try:
                return float(np.polyfit(np.log(lags), np.log(tau), 1)[0])
            except:
                return 0.5

        hurst_val = hurst_exponent(analysis_series.values)
        if hurst_val < 0.45:
            memory_type = 'anti_persistent'
        elif hurst_val > 0.55:
            memory_type = 'persistent'
        else:
            memory_type = 'random_walk'
        props['hurst'] = {
            'value': float(hurst_val),
            'type': memory_type
        }

        # 10. ДОМИНИРУЮЩИЕ ЧАСТОТЫ (FFT)
        try:
            n = len(analysis_series)
            y = analysis_series.values - analysis_series.mean()
            yf = fft(y)
            xf = fftfreq(n, 1)[:n//2]
            amplitude = 2.0/n * np.abs(yf[0:n//2])

            peaks, _ = find_peaks(amplitude, height=np.mean(amplitude) + np.std(amplitude))
            fft_periods = [1/xf[p] for p in peaks if xf[p] > 0 and xf[p] < 0.5]
            fft_dominant = sorted(fft_periods)[:3] if fft_periods else []
            props['fft'] = {
                'dominant_periods': fft_dominant,
                'count': len(fft_dominant)
            }
        except:
            props['fft'] = {'dominant_periods': [], 'count': 0}

        # 11. ПЕРИОДОГРАММА
        try:
            freq_per, pxx_per = periodogram(analysis_series.values, fs=1.0, window='hann')
            peaks_per, _ = find_peaks(pxx_per, height=np.median(pxx_per)*2)
            periodogram_periods = sorted([1/freq_per[p] for p in peaks_per if freq_per[p] > 0])[:3]
            props['periodogram'] = {
                'periods': periodogram_periods,
                'count': len(periodogram_periods)
            }
        except:
            props['periodogram'] = {'periods': [], 'count': 0}

        # 12. ВЕЙВЛЕТ-МАСШТАБЫ
        try:
            import pywt
            widths = np.arange(1, min(128, len(analysis_series)//4))
            cwtmatr, _ = pywt.cwt(
                analysis_series.values - analysis_series.mean(),
                widths, 'morl', sampling_period=1
            )
            mean_power = np.mean(np.abs(cwtmatr), axis=1)
            wavelet_peaks, _ = find_peaks(mean_power, height=np.mean(mean_power))
            wavelet_scales = widths[wavelet_peaks][:3].tolist() if len(wavelet_peaks) > 0 else []
            props['wavelet'] = {
                'scales': wavelet_scales,
                'count': len(wavelet_scales)
            }
        except:
            props['wavelet'] = {'scales': [], 'count': 0}

        # 13. БАЗОВЫЕ СТАТИСТИКИ
        props['basic_stats'] = {
            'n': int(len(analysis_series)),
            'mean': float(analysis_series.mean()),
            'std': float(analysis_series.std()),
            'min': float(analysis_series.min()),
            'max': float(analysis_series.max())
        }

        props['timestamp'] = pd.Timestamp.now().isoformat()

    except Exception as e:
        props['error'] = str(e)

    return props


def _compare_ts_props(props_old: dict, props_new: dict) -> dict:
    """
    Сравнивает ВСЕ 13 свойств паспорта v1.0 и v1.1.
    Поддерживает: числа, строки, списки, булевы значения, вложенные dict.
    """
    comparison = {
        'metrics': {},          # Числовые метрики с delta
        'qualitative_changes': [],  # Качественные изменения
        'categorical_changes': {},  # Категориальные/строковые изменения
        'list_changes': {},     # Изменения списков (FFT, wavelet, periods)
        'boolean_changes': {},  # Изменения булевых флагов
        'summary': ''
    }

    # ═══════════════════════════════════════════════════════
    # 1. ЧИСЛОВЫЕ МЕТРИКИ (10 свойств)
    # ═══════════════════════════════════════════════════════
    numeric_comparisons = [
        ('n', 'basic_stats', 'Число наблюдений'),
        ('mean', 'basic_stats', 'Среднее'),
        ('std', 'basic_stats', 'Стандартное отклонение'),
        ('value', 'stationarity', 'ADF p-value (стационарность)'),
        ('value', 'determinism', 'R² тренда (детерминированность)'),
        ('value', 'autocorrelation', 'Ljung-Box p-value (автокорреляция)'),
        ('value', 'normality', 'Jarque-Bera p-value (нормальность)'),
        ('slope', 'trend', 'Наклон тренда'),
        ('strength', 'seasonality', 'Сила сезонности'),
        ('value', 'hurst', 'Показатель Хёрста'),
    ]

    for key, section, label in numeric_comparisons:
        try:
            old_val = props_old.get(section, {}).get(key)
            new_val = props_new.get(section, {}).get(key)

            if old_val is not None and new_val is not None:
                if pd.notna(old_val) and pd.notna(new_val):
                    delta = new_val - old_val
                    # Защита от деления на ноль
                    if abs(old_val) > 1e-10:
                        delta_pct = (delta / abs(old_val)) * 100
                    else:
                        delta_pct = 0.0 if abs(delta) < 1e-10 else 100.0

                    comparison['metrics'][label] = {
                        'v_old': float(old_val),
                        'v_new': float(new_val),
                        'delta': float(delta),
                        'delta_pct': float(delta_pct),
                        'type': 'numeric'
                    }
        except Exception as e:
            continue

    # ═══════════════════════════════════════════════════════
    # 2. КАТЕГОРИАЛЬНЫЕ СВОЙСТВА (3 свойства)
    # ═══════════════════════════════════════════════════════
    categorical_comparisons = [
        ('value', 'freq', 'Частота ряда'),
        ('direction', 'trend', 'Направление тренда'),
    ]

    for key, section, label in categorical_comparisons:
        try:
            old_val = props_old.get(section, {}).get(key)
            new_val = props_new.get(section, {}).get(key)

            if old_val is not None and new_val is not None:
                old_str = str(old_val)
                new_str = str(new_val)

                if old_str != new_str:
                    comparison['categorical_changes'][label] = {
                        'v_old': old_str,
                        'v_new': new_str,
                        'changed': True,
                        'type': 'categorical'
                    }
                else:
                    comparison['categorical_changes'][label] = {
                        'v_old': old_str,
                        'v_new': new_str,
                        'changed': False,
                        'type': 'categorical'
                    }
        except Exception as e:
            continue

    # ══════════════════════════════════════════════════════
    # 3. СПИСКИ (3 свойства: сезонные периоды, FFT, wavelet)
    # ═══════════════════════════════════════════════════════
    list_comparisons = [
        ('periods', 'seasonal_periods', 'Сезонные периоды (ACF)'),
        ('dominant_periods', 'fft', 'Доминирующие частоты (FFT)'),
        ('scales', 'wavelet', 'Доминирующие масштабы (Wavelet)'),
    ]

    for key, section, label in list_comparisons:
        try:
            old_list = props_old.get(section, {}).get(key, [])
            new_list = props_new.get(section, {}).get(key, [])

            old_list = list(old_list) if old_list else []
            new_list = list(new_list) if new_list else []

            # Сравниваем содержимое
            old_set = set(old_list)
            new_set = set(new_list)

            added = new_set - old_set
            removed = old_set - new_set

            comparison['list_changes'][label] = {
                'v_old': old_list,
                'v_new': new_list,
                'added': list(added),
                'removed': list(removed),
                'changed': old_list != new_list,
                'type': 'list'
            }
        except Exception as e:
            continue

    # ═══════════════════════════════════════════════════════
    # 4. БУЛЕВЫ ФЛАГИ (5 свойств)
    # ═══════════════════════════════════════════════════════
    boolean_comparisons = [
        ('is_stationary', 'stationarity', 'Стационарность'),
        ('is_deterministic', 'determinism', 'Детерминированность'),
        ('is_white_noise', 'autocorrelation', 'Белый шум (нет автокорреляции)'),
        ('is_normal', 'normality', 'Нормальность распределения'),
        ('is_seasonal', 'seasonality', 'Наличие сезонности'),
    ]

    for key, section, label in boolean_comparisons:
        try:
            old_val = props_old.get(section, {}).get(key)
            new_val = props_new.get(section, {}).get(key)

            if old_val is not None and new_val is not None:
                comparison['boolean_changes'][label] = {
                    'v_old': bool(old_val),
                    'v_new': bool(new_val),
                    'changed': old_val != new_val,
                    'type': 'boolean'
                }

                # Добавляем в качественные изменения, если изменилось
                if old_val != new_val:
                    status_new = "✅ Да" if new_val else "❌ Нет"
                    comparison['qualitative_changes'].append(
                        f"{label}: стало {status_new} (было {'✅ Да' if old_val else '❌ Нет'})"
                    )
        except Exception as e:
            continue

    # ═══════════════════════════════════════════════════════
    # 5. ИТОГОВОЕ РЕЗЮМЕ
    # ═══════════════════════════════════════════════════════
    n_numeric_changes = len([m for m in comparison['metrics'].values() if abs(m.get('delta_pct', 0)) > 5])
    n_categorical_changes = len([c for c in comparison['categorical_changes'].values() if c.get('changed')])
    n_list_changes = len([l for l in comparison['list_changes'].values() if l.get('changed')])
    n_boolean_changes = len([b for b in comparison['boolean_changes'].values() if b.get('changed')])

    total_changes = n_numeric_changes + n_categorical_changes + n_list_changes + n_boolean_changes

    if total_changes > 0:
        parts = []
        if n_numeric_changes > 0:
            parts.append(f"{n_numeric_changes} числовых метрик")
        if n_categorical_changes > 0:
            parts.append(f"{n_categorical_changes} категориальных")
        if n_list_changes > 0:
            parts.append(f"{n_list_changes} списков")
        if n_boolean_changes > 0:
            parts.append(f"{n_boolean_changes} булевых флагов")

        comparison['summary'] = (
            f"⚠️ Валидация повлияла на свойства ряда: "
            f"изменено {total_changes} свойств ({', '.join(parts)}). "
            f"Рекомендуется изучить различия перед выбором модели."
        )
    else:
        comparison['summary'] = (
            "✅ Валидация незначительно повлияла на свойства ряда. "
            "Все 13 свойств стабильны."
        )

    return comparison

# ─────────────────────────────────────────────────────────────
# 📁 ФУНКЦИЯ ЧТЕНИЯ ФАЙЛА (С КЭШИРОВАНИЕМ)
# ─────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=" Чтение и парсинг файла...")
def read_uploaded_file(uploaded_file):
    """
    Читает загруженный файл с автодетектом формата и заголовков.
    Кэшируется для предотвращения повторного чтения.
    Возвращает: pd.DataFrame, extension
    """
    if uploaded_file is None:
        raise ValueError("Файл не загружен")

    # 🔑 ОПРЕДЕЛЕНИЕ РАСШИРЕНИЯ (до использования!)
    file_name = uploaded_file.name or "unknown.file"
    ext = file_name.split('.')[-1].lower()

    # Чтение файла
    if ext == "csv":
        df = pd.read_csv(
            uploaded_file,
            sep=None,
            engine='python',
            encoding='utf-8-sig',
            on_bad_lines='skip',
            header=None
        )
        # ... (Ваш код автодетекта заголовков без изменений) ...
        first_col_sample = df[0].head(10).astype(str)
        is_date_like = first_col_sample.str.contains(r'\d{4}[-/]\d{2}[-/]\d{2}', regex=True).mean() > 0.8
        if is_date_like:
            new_headers = ['date'] + [f'col_{i}' for i in range(1, len(df.columns))]
            df.columns = new_headers
        else:
            df.columns = [f'col_{i}' for i in range(len(df.columns))]

    elif ext in ["xlsx", "xls"]:
        df = pd.read_excel(uploaded_file)
        if isinstance(df.columns[0], (int, float)):
            df.columns = [f'col_{i}' for i in range(len(df.columns))]

    elif ext == "json":
        # ... (Ваш код JSON без изменений) ...
        uploaded_file.seek(0)
        content = uploaded_file.read().decode('utf-8-sig')
        data = json.loads(content)
        df = pd.json_normalize(data) if isinstance(data, list) else pd.DataFrame([data])
    else:
        raise ValueError(f"Формат .{ext} не поддерживается.")

    if df.empty:
        raise ValueError("Файл пуст или не содержит табличных данных.")

    return df, ext


# ─────────────────────────────────────────────────────────────
# 📅 ФУНКЦИЯ ROBUST DATETIME DETECTOR (С КЭШИРОВАНИЕМ)
# ─────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=" Анализ дат и конвертация...")
def robust_datetime_detector(df: pd.DataFrame, min_confidence: float = 0.7) -> Tuple[pd.DataFrame, List[str], bool, Optional[str]]:
    """
    Ищет и конвертирует даты.
    Кэшируется: результат сохраняется, пока не изменится сам DataFrame.
    """
    df_work = df.copy()
    original_columns = df_work.columns.tolist()
    # Нормализация имён
    df_work.columns = [str(c).strip().lower().replace(' ', '_').replace('-', '_').replace('.', '_') for c in df_work.columns]

    detected_cols = []
    potential_date_col = None
    max_confidence = 0

    # ── РАСШИРЕННЫЕ ПАТТЕРНЫ ДАТ ──────────────────────────────
    DATE_PATTERNS = [
        # ISO и стандартные
        (r'^\d{4}-\d{2}-\d{2}[T\s]\d{2}:\d{2}:\d{2}', 'iso_datetime'),
        (r'^\d{4}-\d{2}-\d{2}$', 'iso_date'),
        (r'^\d{2}\.\d{2}\.\d{4}$', 'dd.mm.yyyy'),
        (r'^\d{2}/\d{2}/\d{4}$', 'dd/mm/yyyy'),
        (r'^\d{4}/\d{2}/\d{2}$', 'yyyy/mm/dd'),
        (r'^\d{4}-\d{2}$', 'yyyy-mm'),
        (r'^\d{4}\.\d{2}$', 'yyyy.mm'),
        (r'^\d{2}\.\d{4}$', 'mm.yyyy'),
        (r'^\d{1,2}/\d{4}$', 'm/yyyy'),
        (r'^\d{1,2}-\d{4}$', 'm-yyyy'),
        # Гибкие форматы (без ведущих нулей)
        (r'^\d{1,2}/\d{1,2}/\d{4}$', 'us_slash_flexible'),
        (r'^\d{1,2}-\d{1,2}-\d{4}$', 'us_dash_flexible'),
        (r'^\d{1,2}\.\d{1,2}\.\d{4}$', 'eu_dot_flexible'),
        # Только год
        (r'^\d{4}$', 'year_only'),
        # Unix timestamp
        (r'^\d{10}$', 'unix_s'),
        (r'^\d{13}$', 'unix_ms'),
    ]

    # ── РАСШИРЕННЫЕ КЛЮЧЕВЫЕ СЛОВА ───────────────────────────
    TIME_KEYWORDS = [
        # English
        'date', 'time', 'datetime', 'timestamp', 'year', 'month', 'day', 'period',
        'quarter', 'week', 'hour', 'minute', 'second', 'start', 'end', 'begin', 'finish',
        'report_date', 'reporting', 'fiscal', 'calendar', 'observation', 'record_date',
        # Russian
        'дата', 'время', 'год', 'месяц', 'день', 'период', 'квартал', 'неделя',
        'час', 'минута', 'секунда', 'отчетный', 'отчётный', 'начало', 'конец',
        # Other languages / variations
        'jahr', 'année', 'ano', 'anno', 'fecha', 'data', 'datum', 'dat', 'date_',
        'year_', 'yr', 'y_', 'mon', 'm_', 'd_', 'period_', 'time_',
        # FAO / CIS specific
        'reference_year', 'ref_year', 'report_year', 'data_year', 'observation_year'
    ]

    for idx, col in enumerate(df_work.columns):
        # 1. Пропускаем уже datetime
        if pd.api.types.is_datetime64_any_dtype(df_work[col]):
            if col not in detected_cols:
                detected_cols.append(col)
                if potential_date_col is None:
                    potential_date_col = col
            continue

        col_str = str(col).lower()

        # Проверка по ключевым словам
        check_col = any(kw in col_str for kw in TIME_KEYWORDS)

        # Первая колонка — приоритетный кандидат
        if idx == 0 and not check_col:
            check_col = True

        # Дополнительная эвристика для числовых колонок
        if not check_col and df_work[col].dtype in ['int64', 'float64', 'object']:
            sample_check = df_work[col].dropna().head(100)
            if len(sample_check) > 0:
                if sample_check.astype(str).str.match(r'^\d{4}$').mean() > 0.8:
                    check_col = True
                elif sample_check.astype(str).str.contains(r'\d{1,2}[./-]\d{1,2}[./-]\d{2,4}').mean() > 0.8:
                    check_col = True

        if not check_col:
            continue

        # 2. Сэмплирование для анализа
        sample = df_work[col].dropna()
        if len(sample) == 0:
            continue

        sample_vals = sample.head(min(500, len(sample)))
        sample_str = sample_vals.astype(str).str.strip()
        is_numeric = pd.api.types.is_numeric_dtype(sample_vals)

        best_fmt = None
        best_match_ratio = 0

        # 3. Определение формата
        if is_numeric:
            # Проверка на годы (1800-2100) — ГИБКАЯ: 80% вместо 100%
            year_like = sample_vals.between(1800, 2100) & (sample_vals % 1 == 0)
            if year_like.mean() >= 0.8 and len(sample_vals[year_like]) >= 2:
                best_fmt = 'year_only'
                best_match_ratio = year_like.mean()
            # Проверка на Unix timestamp
            elif sample_vals.min() > 1e9:
                best_fmt = 'unix_s' if sample_vals.max() < 1e12 else 'unix_ms'
                best_match_ratio = 1.0
        else:
            # Строковые паттерны (Regex)
            for pattern, fmt_name in DATE_PATTERNS:
                match_ratio = sample_str.str.match(pattern, case=False).mean()
                if match_ratio > best_match_ratio:
                    best_match_ratio = match_ratio
                    best_fmt = fmt_name

            # Fallback: авто-парсинг pandas
            if best_match_ratio < min_confidence:
                try:
                    test_parse = pd.to_datetime(sample_vals, infer_datetime_format=True, errors='coerce')
                    success = test_parse.notna().mean()
                    if success >= min_confidence and success > best_match_ratio:
                        best_fmt = 'auto_infer'
                        best_match_ratio = success
                except Exception:
                    pass

        # 4. Конвертация
        if best_fmt and best_match_ratio >= min_confidence:
            try:
                converted = None

                if best_fmt == 'year_only':
                    converted = pd.to_datetime(sample_vals.astype(int).astype(str), format='%Y', errors='coerce')
                elif best_fmt == 'unix_s':
                    converted = pd.to_datetime(sample_vals, unit='s', errors='coerce')
                elif best_fmt == 'unix_ms':
                    converted = pd.to_datetime(sample_vals, unit='ms', errors='coerce')
                elif best_fmt == 'auto_infer':
                    converted = pd.to_datetime(sample_vals, infer_datetime_format=True, errors='coerce')
                else:
                    converted = pd.to_datetime(sample_vals, format='mixed', errors='coerce')

                success_rate = converted.notna().mean()

                if success_rate >= min_confidence:
                    # Применяем ко всему столбцу
                    if best_fmt == 'year_only':
                        df_work[col] = pd.to_datetime(df_work[col].astype(float).astype(int).astype(str), format='%Y', errors='coerce')
                    elif best_fmt in ['unix_s', 'unix_ms']:
                        unit = 's' if best_fmt == 'unix_s' else 'ms'
                        df_work[col] = pd.to_datetime(df_work[col], unit=unit, errors='coerce')
                    elif best_fmt == 'auto_infer':
                        df_work[col] = pd.to_datetime(df_work[col], infer_datetime_format=True, errors='coerce')
                    else:
                        df_work[col] = pd.to_datetime(df_work[col], format='mixed', errors='coerce')

                    detected_cols.append(col)

                    # Расчет уверенности
                    fill_rate = df_work[col].notna().sum() / max(len(df_work[col]), 1)
                    confidence = success_rate * fill_rate

                    if confidence > max_confidence:
                        max_confidence = confidence
                        potential_date_col = col

            except Exception as e:
                # Логирование для отладки
                # print(f"⚠️ Ошибка конвертации {col}: {e}")
                pass

    # 5. Восстановление оригинальных имён колонок
    for i, orig_col in enumerate(original_columns):
        current_col = df_work.columns[i]
        if current_col in detected_cols:
            df_work.rename(columns={current_col: orig_col}, inplace=True)
            detected_cols = [orig_col if c == current_col else c for c in detected_cols]
            if potential_date_col == current_col:
                potential_date_col = orig_col

    # 6. Активация TS и сортировка
    ts_active = len(detected_cols) > 0
    if ts_active and potential_date_col:
        df_work = df_work.sort_values(potential_date_col).reset_index(drop=True)

    return df_work, detected_cols, ts_active, potential_date_col

# CSS для уменьшения заголовков
st.markdown("""
            <style>
            .st-emotion-cache-3o718f h3 {
                font-size: 1.25rem !important;
                font-weight: 600;
                padding: 0.5rem 0px 0.75rem;
            }
            </style>
            """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# ФУНКЦИЯ РАСЧЁТА СВОЙСТВ ВРЕМЕННОГО РЯДА
# ─────────────────────────────────────────────────────────────
def _calc_ts_props(series: pd.Series) -> dict:
    """
    Рассчитывает ключевые свойства временного ряда.

    Args:
        series: pd.Series с временным рядом (index=datetime, values=numeric)

    Returns:
        dict с ключами:
        - n: число наблюдений
        - mean, std, min, max: базовые статистики
        - adf_pvalue: p-value теста Дики-Фуллера (стационарность)
        - is_stationary: True если p < 0.05
        - has_trend: наличие тренда (через линейную регрессию)
        - has_seasonality: наличие сезонности (через FFT)
        - trend_strength: сила тренда (R²)
        - seasonal_strength: сила сезонности
    """
    props = {
        'n': len(series),
        'mean': series.mean(),
        'std': series.std(),
        'min': series.min(),
        'max': series.max(),
        'adf_pvalue': None,
        'is_stationary': None,
        'has_trend': False,
        'has_seasonality': False,
        'trend_strength': 0.0,
        'seasonal_strength': 0.0
    }

    if len(series) < 10:
        return props

    # 1. Тест Дики-Фуллера (стационарность)
    try:
        from statsmodels.tsa.stattools import adfuller
        adf_result = adfuller(series.dropna(), autolag='AIC')
        props['adf_pvalue'] = adf_result[1]
        props['is_stationary'] = adf_result[1] < 0.05
    except Exception:
        pass

    # 2. Проверка тренда (линейная регрессия)
    try:
        from scipy import stats
        x = np.arange(len(series))
        slope, intercept, r_value, p_value, std_err = stats.linregress(x, series.values)
        props['has_trend'] = p_value < 0.05
        props['trend_strength'] = r_value**2
    except Exception:
        pass

    # 3. Проверка сезонности (FFT)
    try:
        from scipy.fft import fft
        fft_vals = np.abs(fft(series.values - series.mean()))
        fft_vals = fft_vals[1:len(fft_vals)//2]  # Убираем постоянную составляющую
        if len(fft_vals) > 0:
            # Ищем доминирующую частоту
            dominant_freq_idx = np.argmax(fft_vals)
            dominant_amplitude = fft_vals[dominant_freq_idx]
            mean_amplitude = np.mean(fft_vals)
            # Если амплитуда доминирующей частоты в 3+ раза выше средней — есть сезонность
            props['has_seasonality'] = dominant_amplitude > 3 * mean_amplitude
            props['seasonal_strength'] = dominant_amplitude / mean_amplitude if mean_amplitude > 0 else 0
    except Exception:
        pass

    return props


# ─────────────────────────────────────────────────────────────
# 🎨 GOOGLE MATERIAL ICONS
# ─────────────────────────────────────────────────────────────
st.markdown("""
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200&icon_names=check_small,warning,error,info" />
<style>
.material-symbols-outlined {
  font-family: 'Material Symbols Outlined';
  font-variation-settings: 'FILL' 0, 'wght' 400, 'GRAD' 0, 'opsz' 24;
  font-size: 24px;
  line-height: 1;
  display: inline-block;
  white-space: nowrap;
  -webkit-font-smoothing: antialiased;
}
.icon-green { color: #16a34a !important; }
.icon-red { color: #dc2626 !important; }
.icon-blue { color: #2563eb !important; }
.icon-yellow { color: #ca8a04 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# 📊 ГЕНЕРАЦИЯ ОТЧЁТА ВАЛИДАЦИИ
# ─────────────────────────────────────────────────────────────
def generate_validation_report(df, val_results):
    import pandas as pd
    import datetime

    miss_summary = val_results.get('miss', {}).get('summary', {})
    outl_summary = val_results.get('outl', {}).get('summary', {})
    ts_data = val_results.get('ts', {})
    total_rows = len(df)

    missing_count = miss_summary.get('total_missing', 0)
    missing_pct = miss_summary.get('missing_rate_pct', (missing_count / total_rows * 100) if total_rows > 0 else 0.0)
    outlier_count = outl_summary.get('total_outliers', 0)
    outlier_pct = outl_summary.get('outlier_rate_pct', (outlier_count / total_rows * 100) if total_rows > 0 else 0.0)

    summary_data = {
        "Параметр": [
            "Название файла", "Дата анализа", "Всего записей", "Всего колонок",
            "Найдено пропусков (шт)", "Найдено пропусков (%)",
            "Найдено выбросов (шт)", "Найдено выбросов (%)",
            "Стационарность ряда (ADF)", "Частота ряда (Inferred)"
        ],
        "Значение": [
            st.session_state.get("original_filename", "Unknown"),
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            total_rows, len(df.columns),
            missing_count, f"{missing_pct:.2f}%",
            outlier_count, f"{outlier_pct:.2f}%",
            "Да" if ts_data.get('is_stationary') else "Нет",
            ts_data.get('frequency', 'N/A')
        ]
    }
    df_summary = pd.DataFrame(summary_data)

    all_issues = []
    # ... (Сбор проблем из val_results: Пропуски, Выбросы, Диапазоны, Уникальность)
    for col, stats in val_results.get('miss', {}).get('columns', {}).items():
        if isinstance(stats, dict) and stats.get('count', 0) > 0:
            all_issues.append({"Тип проверки": "Пропуски", "Колонка": col, "Проблема": f"{stats['count']} шт ({stats.get('percent', 0):.1f}%)", "Рекомендация": "Заполнить"})
    for col, stats in val_results.get('outl', {}).get('columns', {}).items():
        if isinstance(stats, dict) and stats.get('count', 0) > 0:
            all_issues.append({"Тип проверки": "Выбросы", "Колонка": col, "Проблема": f"{stats['count']} шт ({stats.get('percent', 0):.1f}%)", "Рекомендация": "Кэпировать"})
    for issue in val_results.get('range_results', []):
        if isinstance(issue, dict):
            all_issues.append({"Тип проверки": "Диапазоны", "Колонка": issue.get('Колонка'), "Проблема": f"{issue.get('Нарушений')} нарушений", "Рекомендация": "Кэпировать"})

    df_issues = pd.DataFrame(all_issues) if all_issues else pd.DataFrame(columns=["Тип проверки", "Колонка", "Проблема", "Рекомендация"])

    ts_summary = {
        "Метрика": ["Стационарность (ADF p-value)", "Частота (Frequency)", "Макс. разрыв (Max Gap)", "Статус TS"],
        "Значение": [ts_data.get('adf_pvalue', 'N/A'), ts_data.get('frequency', 'N/A'), str(ts_data.get('max_gap', 'N/A')), ts_data.get('error', 'Готово к анализу')]
    }
    df_ts = pd.DataFrame(ts_summary)

    filename = f"Statcom_DQ_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='1_Сводка', index=False)
        df_issues.to_excel(writer, sheet_name='2_Проблемы', index=False)
        df_ts.to_excel(writer, sheet_name='3_TS_Props', index=False)

    return filename

def add_log(level: str, message: str):
    entry = {"⏱️ Время": datetime.now().strftime("%H:%M:%S"), "📊 Уровень": level, "📝 Сообщение": message}
    st.session_state.error_log.append(entry)
    if len(st.session_state.error_log) > 50: st.session_state.error_log = st.session_state.error_log[-50:]

# ─────────────────────────────────────────────────────────────
# 🎨 ТЕМА И ШРИФТ (CISStat Branding)
# ─────────────────────────────────────────────────────────────
ASSETS_DIR = Path(__file__).parent / "assets"
logo_path = ASSETS_DIR / "Logo_CISSTAT.png"
favicon_path = ASSETS_DIR / "favicon.ico"

st.set_page_config(
    page_title="CISStat deep research",
    page_icon=str(favicon_path) if favicon_path.exists() else "📊",
    layout="wide"
)

if logo_path.exists():
    st.sidebar.image(str(logo_path), width=80, use_container_width=False)
else:
    st.sidebar.markdown("### 🏢 **CISStat TS Analysis**")

st.markdown("""
<style>
/* Адаптивный заголовок для мобильных */
@media (max-width: 768px) {
    div[data-testid="stMarkdown"] h1 {
        font-size: 28px !important;
    }
    div[data-testid="stMarkdown"] p {
        font-size: 18px !important;
    }
}
@media (max-width: 480px) {
    div[data-testid="stMarkdown"] h1 {
        font-size: 24px !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# 🎨 НАСТРОЙКИ ЗАГОЛОВКА (хэдер)
# ─────────────────────────────────────────────────────────────
import base64
from pathlib import Path

HEADER_FONT_SIZE = "36px"
SUBHEADER_FONT_SIZE = "21px"
HEADER_WEIGHT = "700"

# Загрузка иконки и кодирование в base64
icon_path = Path("assets/logo_platform1.png")  # или "logo.png" в корне
if icon_path.exists():
    with open(icon_path, "rb") as f:
        icon_base64 = base64.b64encode(f.read()).decode()
    icon_html = f'<img src="data:image/png;base64,{icon_base64}" style="height: 42px; width: auto; vertical-align: middle; margin-right: 12px;"/>'
else:
    icon_html = '📈'  # Fallback на эмодзи

st.markdown(f"""
    <div style='text-align: center; margin: -20px 0 15px 0; padding: 10px 25px;
              background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
              border-radius: 12px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);'>
        <h1 style='color: white; font-size: {HEADER_FONT_SIZE};
                   font-weight: {HEADER_WEIGHT}; margin: 0; line-height: 1.2;
                   display: flex; align-items: center; justify-content: center;'>
            {icon_html} CISStat TS Analysis
        </h1>
        <div style='display: flex; align-items: center; justify-content: center; margin: 12px 0;'>
            <div style='flex: 1; height: 2px; background: rgba(255, 255, 255, 0.7); margin-right: -2px;'></div>
            <svg width='60' height='40' style='margin: 0;' viewBox='0 0 60 40'>
                <path d='M 0 20 L 5 20 L 10 6 L 15 34 L 20 6 L 25 34 L 30 6 L 35 34 L 40 6 L 45 34 L 50 20 L 60 20'
                      stroke='rgba(255, 255, 255, 0.9)'
                      stroke-width='2'
                      fill='none'
                      stroke-linecap='round'
                      stroke-linejoin='round'/>
            </svg>
        </div>
        <p style='color: rgba(255,255,255,0.95); font-size: {SUBHEADER_FONT_SIZE};
                margin: 8px 0 0 0; font-weight: 400;'>
            профессиональная платформа анализа временных рядов
        </p>
    </div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# 🔐 АВТОРИЗАЦИЯ (Безопасная версия)
# ─────────────────────────────────────────────────────────────

# 🔑 Эталонный хэш токена (задаётся через переменную окружения CISSTAT_TOKEN_HASH)
# Если переменная не задана, используется хэш от "123" для локальных тестов
SECURE_TOKEN_HASH = os.environ.get(
    "CISSTAT_TOKEN_HASH",
    "a665a45920422f9d417e4867efdc4fb8a04a1f3fff1fa07e998e86f7f7a27ae3"
)

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.markdown(
    "<h2 style='text-align:center; margin-top: 40px; font-size: 26px;'>Вход в систему</h2>",
    unsafe_allow_html=True
    )
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        token_input = st.text_input("Токен доступа", type="password", placeholder="Введите пароль")
        if st.button("Войти", type="primary", use_container_width=True):
            # 🔒 Хешируем ввод пользователя и сверяем с эталоном
            input_hash = hashlib.sha256(token_input.encode('utf-8')).hexdigest()
            if input_hash == SECURE_TOKEN_HASH:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ Неверный токен доступа")
    st.stop()

# ─────────────────────────────────────────────────────────────
# 🧠 ИНИЦИАЛИЗАЦИЯ SESSION STATE
# ─────────────────────────────────────────────────────────────
if "loading_progress" not in st.session_state: st.session_state.loading_progress = 0.0
if "error_log" not in st.session_state: st.session_state.error_log = []
if "validation_ready" not in st.session_state: st.session_state.validation_ready = False
if "val_results" not in st.session_state: st.session_state.val_results = {"val": None, "miss": None, "outl": None, "ts": {}}
if "df" not in st.session_state: st.session_state.df = pd.DataFrame()
if "col_types" not in st.session_state: st.session_state.col_types = {"num": [], "cat": [], "date": []}
if "ts_mode_active" not in st.session_state: st.session_state.ts_mode_active = False
if "primary_date_col" not in st.session_state: st.session_state.primary_date_col = None
if "passport_csv" not in st.session_state:
    st.session_state.passport_csv = None
if "passport_ready" not in st.session_state:
    st.session_state.passport_ready = False

# ── ГЛОБАЛЬНЫЕ КОНСТАНТЫ ДЛЯ РЕЖИМОВ ВИЗУАЛИЗАЦИИ ───────────
MODE_TS = "⏱️ Временные ряды"
MODE_GEN = "🔍 Общий (категории)"


# ────────────────────────────────────────────────────────────
# 🔧 ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ (должна быть определена ДО использования)
# ────────────────────────────────────────────────────────────
def _safe_nunique(series: pd.Series, min_val: int = 1, max_val: int = 100) -> bool:
    """
    Безопасный подсчёт уникальных значений для колонок с возможными нехэшируемыми типами.
    Возвращает True, если количество уникальных значений в диапазоне [min_val, max_val).
    """
    try:
        sample = series.dropna().head(100)
        if len(sample) == 0:
            return False
        first_val = sample.iloc[0]
        if isinstance(first_val, (dict, list, set, pd.Series, pd.DataFrame)):
            return False
        uniq = series.nunique()
        return min_val < uniq < max_val
    except TypeError:
        return False
    except Exception:
        return False


# ────────────────────────────────────────────────────────────
# 🗄️ ФУНКЦИЯ ПОДКЛЮЧЕНИЯ К БД (С КЭШИРОВАНИЕМ РЕСУРСА)
# ────────────────────────────────────────────────────────────
@st.cache_resource(ttl=3600)
def init_db_connection(db_type: str, host: str, port: int, user: str, password: str, db_name: str):
    """
    Создаёт и кэширует подключение к базе данных.
    Повторные вызовы с теми же параметрами вернут тот же объект.
    """
    try:
        if db_type == "PostgreSQL":
            try:
                import psycopg2
            except ImportError:
                raise ImportError("psycopg2-binary")

            url = f"postgresql://{user}:{password}@{host}:{port}/{db_name}"
            engine = create_engine(
                url,
                connect_args={"connect_timeout": 10, "options": "-c statement_timeout=60000"},
                pool_pre_ping=True,
                pool_recycle=300
            )
            with engine.connect() as conn:
                conn.execute("SELECT 1")
            return engine

        elif db_type == "ClickHouse":
            try:
                import clickhouse_connect
            except ImportError:
                raise ImportError("clickhouse-connect")

            client = clickhouse_connect.get_client(
                host=host, port=port,
                username=user, password=password,
                database=db_name,
                secure=False,
                verify=False,
                connect_timeout=10,
                send_receive_timeout=60
            )
            client.ping()
            return client

    except ImportError as e:
        raise e
    except Exception as e:
        add_log("ERROR", f"DB Connection failed: {db_type}@{host}:{port}/{db_name} - {e}")
        raise ConnectionError(f"Не удалось подключиться к {db_type}")
    # 🔧 Функция ЗАВЕРШЕНА здесь (return выше)


# ────────────────────────────────────────────────────────────
#  БОКОВАЯ ПАНЕЛЬ
# ────────────────────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.markdown("### Источник данных")

data_source = st.sidebar.radio(
    "Выберите источник:",
    ["◉ Файл .xlsx, .xls, .csv, .json", "◉ База Данных (SQL)"],
    key="sidebar_data_source"
)

# ── ЗАГРУЗКА ФАЙЛОВ ──────────────────────────────────────────
if data_source == "◉ Файл .xlsx, .xls, .csv, .json":
    uploaded = st.sidebar.file_uploader(
        "Загрузите файл",
        type=["xlsx", "xls", "csv", "json"],
        key="file_uploader_main",
        help="Поддерживаются: CSV (автодетект разделителя), Excel (.xlsx/.xls), JSON (плоский или вложенный)"
    )

    if uploaded is not None:
        st.sidebar.success(f"Файл выбран: `{uploaded.name}`")

        if st.sidebar.button("Загрузить файл", type="primary", use_container_width=True, key="btn_load_main"):
            with st.spinner("⏳ Обработка данных..."):
                try:
                    file_name = uploaded.name or "unknown.file"
                    ext = file_name.split('.')[-1].lower()

                    if ext == "csv":
                        df = pd.read_csv(
                            uploaded,
                            sep=None,
                            engine='python',
                            encoding='utf-8-sig',
                            on_bad_lines='skip',
                            parse_dates=False
                        )
                    elif ext in ["xlsx", "xls"]:
                        df = pd.read_excel(uploaded, parse_dates=False)

                    elif ext == "json":
                        try:
                            uploaded.seek(0)
                            content = uploaded.read().decode('utf-8-sig')
                            import json as json_lib
                            from pandas import json_normalize

                            data = json_lib.loads(content)

                            # JSON-stat 2.0 обработка
                            if isinstance(data, dict) and data.get("version") == "2.0" and "value" in data and "dimension" in data:
                                dimensions = data.get("dimension", {})
                                dimension_ids = data.get("id", [])
                                sizes = data.get("size", [])

                                strides = [1] * len(sizes)
                                for j in range(len(sizes) - 2, -1, -1):
                                    strides[j] = strides[j + 1] * sizes[j + 1]

                                category_maps = {}
                                for dim_id in dimension_ids:
                                    dim_info = dimensions.get(dim_id, {})
                                    category_info = dim_info.get("category", {})
                                    index_map = category_info.get("index", {})
                                    label_map = category_info.get("label", {})
                                    reverse_index = {v: k for k, v in index_map.items()}
                                    category_maps[dim_id] = {
                                        "reverse_index": reverse_index,
                                        "label": label_map
                                    }

                                rows = []
                                for key_str, value in data["value"].items():
                                    try:
                                        linear_idx = int(key_str)
                                        indices = []
                                        remaining = linear_idx
                                        for j, size in enumerate(sizes):
                                            if j == len(sizes) - 1:
                                                indices.append(remaining)
                                            else:
                                                idx = remaining // strides[j]
                                                indices.append(idx)
                                                remaining = remaining % strides[j]

                                        row = {}
                                        for j, dim_id in enumerate(dimension_ids):
                                            cat_info = category_maps.get(dim_id, {})
                                            reverse_index = cat_info.get("reverse_index", {})
                                            label_map = cat_info.get("label", {})
                                            cat_code = reverse_index.get(indices[j])
                                            row[dim_id] = label_map.get(cat_code, cat_code) if cat_code else None
                                        row["value"] = value
                                        rows.append(row)
                                    except (ValueError, KeyError):
                                        continue

                                if rows:
                                    df = pd.DataFrame(rows)
                                    st.sidebar.success(f"✅ JSON-stat 2.0 распарсен: {len(df)} записей")
                                else:
                                    raise ValueError("JSON-stat 2.0 не содержит валидных данных")

                            # Обычный JSON
                            elif isinstance(data, list):
                                if len(data) > 0 and isinstance(data[0], dict):
                                    df = json_normalize(data)
                                elif len(data) > 0:
                                    df = pd.DataFrame({uploaded.name.rsplit('.', 1)[0]: data})
                                else:
                                    df = pd.DataFrame()
                            elif isinstance(data, dict):
                                df = pd.DataFrame([data])
                            else:
                                df = pd.DataFrame([{"value": str(data)}])

                            if df.empty:
                                raise ValueError("JSON пуст")

                            # Переименование колонок (не для JSON-stat)
                            if "version" not in data or data.get("version") != "2.0":
                                if len(df.columns) > 0 and isinstance(df.columns[0], (int, float)):
                                    df.columns = [f'col_{i}' for i in range(len(df.columns))]

                        except json.JSONDecodeError as je:
                            raise ValueError(f"❌ Ошибка парсинга JSON: {je}")
                        except Exception as e:
                            raise ValueError(f"❌ Ошибка обработки JSON: {e}")

                    else:
                        raise ValueError(f"Формат .{ext} не поддерживается")

                    if df.empty:
                        raise ValueError("Файл пуст")

                    # Очистка имён колонок
                    df.columns = df.columns.astype(str).str.strip().str.replace(r'\s+', '_', regex=True)

                    # Автодетект дат
                    df, detected_dates, ts_active, primary_date = robust_datetime_detector(df)

                    # ОЧИСТКА ОТ СИСТЕМНЫХ КОЛОНОК (ВСТАВИТЬ СЮДА, ДО СОХРАНЕНИЯ В СЕССИЮ)
                    service_cols = [
                        c for c in df.columns
                        if c.lower() in ['row_id', 'index', 'level_0', 'level_1', 'unnamed', 'unnamed: 0']
                    ]
                    if service_cols:
                        df = df.drop(columns=service_cols)
                        st.sidebar.toast(f"Удалены системные колонки: {service_cols}")

                    # Обновление session_state (теперь уже чистого df)
                    st.session_state.df = df
                    st.session_state.df_raw = df.copy()
                    st.session_state.col_types = {
                        "num": df.select_dtypes(include='number').columns.tolist(),
                        "date": detected_dates,
                        "cat": [
                            c for c in df.select_dtypes(include=['object', 'string']).columns
                            if _safe_nunique(df[c], min_val=1, max_val=100)
                        ]
                    }
                    st.session_state.ts_mode_active = ts_active
                    st.session_state.primary_date_col = primary_date if primary_date != "wide_format_detected" else None
                    st.session_state.original_filename = uploaded.name

                    add_log("INFO", f"✅ Загружен: {uploaded.name} ({len(df)} строк)")

                    if ts_active:
                        if primary_date == "wide_format_detected":
                            st.sidebar.warning("📊 Обнаружен широкий формат данных")
                        else:
                            st.sidebar.success(f"✅ TimeSeries активирован! `{primary_date}`")
                    else:
                        st.sidebar.info("ℹ️ Режим общего анализа данных")

                    st.success(f"✅ Загружено: {len(df)} строк × {len(df.columns)} колонок")
                    st.rerun()

                except Exception as e:
                    add_log("ERROR", f"❌ Ошибка загрузки: {e}")
                    st.sidebar.error(f"❌ Ошибка: {e}")
                    import traceback
                    st.sidebar.code(traceback.format_exc(), language="python")
                finally:
                    if uploaded:
                        uploaded.seek(0)

# ─── БАЗА ДАННЫХ (SQL) ──────────────────────────────────────
else:  # 🔧 Этот else на одном уровне с if выше!
    with st.sidebar.expander("⚙️ Настройки подключения", expanded=True):
        db_type = st.selectbox("● Тип БД", ["PostgreSQL", "ClickHouse"], key="db_type_sel")

        default_port = 5432 if db_type == "PostgreSQL" else 8123
        port_help = "Стандартный порт: 5432" if db_type == "PostgreSQL" else "Стандартный порт: 8123 (HTTP) / 9000 (Native)"

        host = st.text_input("● Host", "localhost", key="db_host", help="IP-адрес или домен сервера БД")
        port = st.number_input("● Port", value=default_port, key="db_port", help=port_help)
        db_name = st.text_input("● Database", "postgres", key="db_name", help="Имя базы данных")
        user = st.text_input("● User", "postgres", key="db_user", help="Пользователь БД")
        pwd = st.text_input("● Password", type="password", key="db_pwd", help="Пароль пользователя")
        query = st.text_area("● SQL Query", "SELECT * FROM your_table LIMIT 1000", key="db_query",
                        help="SQL-запрос для выборки данных. Рекомендуется использовать LIMIT для тестов.")

        c1, c2 = st.sidebar.columns(2)

        # 🔹 КНОПКА ТЕСТА ПОДКЛЮЧЕНИЯ
        with c1:
            if st.button("🔌 Тест", type="secondary", key="btn_db_test", help="Проверить доступность БД"):
                with st.spinner("Проверка подключения..."):
                    try:
                        conn_obj = init_db_connection(db_type, host, port, user, pwd, db_name)

                        if db_type == "PostgreSQL":
                            with conn_obj.connect() as conn:
                                result = conn.execute("SELECT version()").fetchone()
                                st.sidebar.success(f"✅ PostgreSQL: подключено! Версия: {result[0][:50]}...")
                        elif db_type == "ClickHouse":
                            result = conn_obj.query("SELECT version()").result_rows
                            st.sidebar.success(f"✅ ClickHouse: подключено! Версия: {result[0][0]}")

                        if db_type == "ClickHouse":
                            conn_obj.close()

                    except ImportError as ie:
                        driver = "psycopg2-binary" if "psycopg2" in str(ie) else "clickhouse-connect"
                        st.sidebar.error(f"❌ Требуется драйвер: `pip install {driver}`")
                        with st.sidebar.expander("🔧 Инструкция по установке"):
                            st.markdown(f"""
                            1. Откройте терминал в папке проекта
                            2. Активируйте виртуальное окружение:
                            ```bash
                            .venv\\Scripts\\activate  # Windows
                            source .venv/bin/activate  # Linux/Mac
                            ```
                            3. Установите драйвер:
                            ```bash
                            pip install {driver}
                            ```
                            4. Перезапустите Streamlit
                            """)
                    except ConnectionError as ce:
                        st.sidebar.error(f"🔴 {ce}")
                        with st.sidebar.expander("🔍 Возможные причины"):
                            st.markdown("""
                            - Сервер БД не запущен
                            - Порт заблокирован брандмауэром
                            - Неправильные учётные данные
                            - Сетевые проблемы
                            """)
                    except Exception as e:
                        safe_msg = str(e).replace(pwd, "****") if pwd else str(e)
                        st.sidebar.error(f"❌ Ошибка: {safe_msg[:200]}...")
                        add_log("ERROR", f"DB Test failed: {db_type}@{host}:{port} - {e}")

        # 🔹 КНОПКА ЗАГРУЗКИ ДАННЫХ
        with c2:
            if st.button("📥 Загрузить", type="primary", key="btn_db_load", help="Выполнить запрос и загрузить данные"):
                if not query.strip():
                    st.sidebar.warning("⚠️ Введите SQL-запрос")
                else:
                    with st.spinner("Загрузка данных из БД..."):
                        progress = st.progress(0)
                        try:
                            progress.progress(0.3, text="Подключение к БД...")
                            conn_obj = init_db_connection(db_type, host, port, user, pwd, db_name)

                            progress.progress(0.6, text="Выполнение запроса...")
                            if db_type == "PostgreSQL":
                                df_db = pd.read_sql(query, conn_obj)
                                conn_obj.dispose()
                            elif db_type == "ClickHouse":
                                df_db = conn_obj.query_df(query)
                                conn_obj.close()

                            progress.progress(0.9, text="Обработка данных...")

                            # Пост-обработка
                            df_db.columns = df_db.columns.astype(str).str.strip()
                            df_db, detected_dates, ts_active, primary_date = robust_datetime_detector(df_db)

                            # ОЧИСТКА ОТ СИСТЕМНЫХ КОЛОНОК (ВСТАВИТЬ СЮДА, ДО СОХРАНЕНИЯ В СЕССИЮ)
                            service_cols = [
                                c for c in df.columns
                                if c.lower() in ['row_id', 'index', 'level_0', 'level_1', 'unnamed', 'unnamed: 0']
                            ]
                            if service_cols:
                                df = df.drop(columns=service_cols)
                                st.sidebar.toast(f"Удалены системные колонки: {service_cols}")

                            # Обновление session_state
                            st.session_state.df = df_db
                            st.session_state.col_types = {
                                "num": df_db.select_dtypes(include='number').columns.tolist(),
                                "date": detected_dates,
                                "cat": [c for c in df_db.select_dtypes(include=['object', 'string']).columns
                                    if 1 < df_db[c].nunique() < 100]
                            }
                            st.session_state.ts_mode_active = ts_active
                            st.session_state.primary_date_col = primary_date
                            st.session_state.original_filename = f"db_{db_name}_{datetime.now().strftime('%Y%m%d')}.csv"

                            progress.progress(1.0, text="✅ Готово!")
                            add_log("INFO", f"✅ Загружено из БД: {db_name} ({len(df_db)} строк)")
                            st.sidebar.success(f"✅ Загружено: {len(df_db):,} строк × {len(df_db.columns)} колонок")
                            st.rerun()

                        except ImportError as ie:
                            driver = "psycopg2-binary" if "psycopg2" in str(ie) else "clickhouse-connect"
                            add_log("ERROR", f"❌ Ошибка импорта БД: {ie}")
                            st.sidebar.error(f"❌ Требуется драйвер: `pip install {driver}`")
                        except pd.errors.DatabaseError as de:
                            add_log("ERROR", f"❌ Ошибка SQL-запроса: {de}")
                            st.sidebar.error(f"❌ Ошибка выполнения запроса. Проверьте синтаксис SQL.")
                            with st.sidebar.expander("🔍 Детали ошибки"):
                                st.code(str(de).replace(pwd, "****") if pwd else str(de), language="sql")
                        except Exception as e:
                            add_log("ERROR", f"❌ Ошибка БД: {e}")
                            safe_msg = str(e).replace(pwd, "****") if pwd else str(e)
                            st.sidebar.error(f"❌ Ошибка: {safe_msg[:200]}...")
                            with st.sidebar.expander("🔍 Stack trace"):
                                st.code(traceback.format_exc(), language="python")
                        finally:
                            progress.empty()

# ─────────────────────────────────────────────────────────────
# 🔧 ИНИЦИАЛИЗАЦИЯ SESSION_STATE (если ещё не инициализировано)
# ─────────────────────────────────────────────────────────────
if "error_log" not in st.session_state:
    st.session_state.error_log = []
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame()
if "col_types" not in st.session_state:
    st.session_state.col_types = {"num": [], "cat": [], "date": []}
if "rules" not in st.session_state:
    st.session_state.rules = load_rules()


# ─────────────────────────────────────────────────────────────
# 📋 УПРАВЛЕНИЕ ПРАВИЛАМИ YAML
# ─────────────────────────────────────────────────────────────
# st.sidebar.divider()
with st.sidebar.expander("Управление правилами"):
    st.markdown("**Текущий файл правил:** `fao_prices.yaml`")

    # Инициализация состояния шаблона (Custom по умолчанию)
    if "rule_template" not in st.session_state:
        st.session_state.rule_template = "Custom (автогенерация)"  # ✅ По умолчанию автогенерация

    # Выбор шаблона
    template = st.selectbox(
        "Выберите шаблон правил:",
        ["Custom (автогенерация)", "FAO Prices (CIS)", "Macro indicators"],  # 🔀 Custom первым
        index=0,  # ✅ По умолчанию выбран первый элемент (Custom)
        key="rule_template_selector"
    )

    # Загрузка правил с обработкой ошибок
    try:
        if template == "Custom (автогенерация)":
            current_df = st.session_state.get("df", pd.DataFrame())
            if not current_df.empty:
                try:
                    rules = auto_generate_rules(current_df)
                    st.success(f"✅ Сгенерировано правил: {len(rules.get('ranges', []))} диапазонов")
                    st.info("💡 Автогенерация проанализировала ваши данные и создала персональные правила")
                except Exception as e:
                    st.error(f"❌ Ошибка автогенерации: {e}")
                    rules = {}
            else:
                st.info("ℹ️ Загрузите данные для автогенерации правил")
                rules = {}
        elif template == "FAO Prices (CIS)":
            rules = load_rules("rules/fao_prices.yaml")
            if not rules:
                st.warning("⚠️ Файл `fao_prices.yaml` не найден. Используется автогенерация.")
                current_df = st.session_state.get("df", pd.DataFrame())
                if not current_df.empty:
                    rules = auto_generate_rules(current_df)
                else:
                    rules = {}
        elif template == "Macro indicators":
            rules = load_rules("rules/macro.yaml")
            if not rules:
                st.warning("⚠️ Файл `macro.yaml` не найден. Используется автогенерация.")
                current_df = st.session_state.get("df", pd.DataFrame())
                if not current_df.empty:
                    rules = auto_generate_rules(current_df)
                else:
                    rules = {}
        else:
            rules = {}
    except Exception as e:
        st.error(f"❌ Ошибка загрузки правил: {e}")
        rules = {}

    # 🔧 Редактор правил (улучшенный)
    if rules:
        st.markdown("### Редактор диапазонов")

        # Сохраняем правила в session_state для редактирования
        if "editable_rules" not in st.session_state:
            st.session_state.editable_rules = rules.copy()

        for i, rule in enumerate(st.session_state.editable_rules.get("ranges", [])):
            with st.container(border=True):
                st.markdown(f"**Правило {i+1}:** {rule.get('name', 'Без имени')}")
                col_name = rule.get('keywords', ['?'])[0] if rule.get('keywords') else '?'
                st.caption(f"Колонка: `{col_name}`")

                c1, c2 = st.columns(2)
                with c1:
                    min_val = st.number_input(
                        "Минимум",
                        value=float(rule.get('min', 0)) if rule.get('min') is not None else 0.0,
                        key=f"min_{i}",
                        format="%.2f"
                    )
                with c2:
                    max_val = st.number_input(
                        "Максимум",
                        value=float(rule.get('max', 1000)) if rule.get('max') is not None else 1000.0,
                        key=f"max_{i}",
                        format="%.2f"
                    )

                # Обновляем правило в session_state
                st.session_state.editable_rules["ranges"][i]["min"] = min_val
                st.session_state.editable_rules["ranges"][i]["max"] = max_val

        # 🔧 Кнопка применения
        if st.button("Применить правила", use_container_width=True, key="btn_apply_rules"):
            st.session_state.rules = st.session_state.editable_rules.copy()
            st.session_state.validation_ready = False  # Сброс валидации
            st.success("✅ Правила обновлены! Перезапустите валидацию.")
            st.info("💡 Нажмите **🚀 Запустить валидацию** для проверки с новыми правилами.")

        # 🔧 Кнопка сброса к исходным
        if st.button("🔄 Сбросить к исходным", use_container_width=True, key="btn_reset_rules"):
            if "editable_rules" in st.session_state:
                del st.session_state.editable_rules
            st.rerun()
    else:
        st.info("ℹ️ Нет доступных правил для редактирования")


# ─────────────────────────────────────────────────────────────
# 📋 ЛОГ СОБЫТИЙ
# ─────────────────────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.markdown("### Лог событий")
if st.session_state.error_log:
    # 🔧 Исправлено: убран hide_index, добавлена обработка
    log_df = pd.DataFrame(st.session_state.error_log[-20:])  # Последние 20 записей
    st.sidebar.dataframe(
        log_df,
        use_container_width=True,
        height=150,
        column_config={
            "_index": st.column_config.Column("№", width="small")
        }
    )
    if st.sidebar.button("🗑️ Очистить лог", use_container_width=True, key="btn_clear_log"):
        st.session_state.error_log = []
        st.rerun()
else:
    st.sidebar.success("✅ Ошибок нет")

# ─────────────────────────────────────────────────────────────
# ℹ️ ИНФОРМАЦИЯ О ДАННЫХ
# ─────────────────────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.markdown("### Информация о данных")

# 🔧 Безопасная проверка существования данных
df_info = st.session_state.get("df", pd.DataFrame())
ct_info = st.session_state.get("col_types", {"num": [], "cat": [], "date": []})

if not df_info.empty:
    st.sidebar.info(f"""
    - **📊 Записей**: `{len(df_info):,}`
    - **📐 Колонок**: `{len(df_info.columns)}`
    - **🔢 Числовых**: `{len(ct_info.get('num', []))}`
    - **📋 Категорий**: `{len(ct_info.get('cat', []))}`
    - **📅 Даты**: `{len(ct_info.get('date', []))}`
    """)
else:
    st.sidebar.info("ℹ️ Данные ещё не загружены")


# ─────────────────────────────────────────────────────────────
# 📞 ПОДДЕРЖКА
# ─────────────────────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.markdown("### 📞 Поддержка")
st.sidebar.markdown("""
📧 support@cisstat.org
🌐 [new.cisstat.org](https://new.cisstat.org)
📚 Документация: `/docs`
""")

# ─────────────────────────────────────────────────────────────
# 💡 КАК РАБОТАТЬ С ПЛАТФОРМОЙ (Расширенная версия)
# ─────────────────────────────────────────────────────────────
with st.sidebar.expander("💡 Как работать с платформой", expanded=False):
    st.markdown("""
    ### 🚀 Быстрый старт

    #### 1️⃣ Загрузка данных
    | Источник | Действие | Примечание |
    |----------|----------|------------|
    | 📄 Файл | Выберите `.csv`, `.xlsx`, `.json` → Нажмите «Загрузить» | Автодетект кодировки, разделителей, дат |
    | 🗄️ БД | Укажите параметры → «Тест» → «Загрузить» | Поддержка: PostgreSQL, ClickHouse |

    #### 2️⃣ Валидация данных
    ```mermaid
    graph LR
    A[Загрузка] --> B[🚀 Запустить валидацию]
    B --> C{Проблемы?}
    C -->|✅ Нет| D[Переход к анализу]
    C -->|⚠️ Да| E[🔧 Исправить в пайплайне]
    E --> F[💾 Сохранить/Экспорт]
    ```

    **🔍 Что проверяется:**
    - ✅ Типы данных и форматы (Regex)
    - ✅ Диапазоны значений и логика
    - ✅ Пропуски (с анализом механизма MCAR/MAR)
    - ✅ Выбросы (IQR, Z-score, MAD)
    - ✅ Уникальность и ссылочная целостность
    - ✅ TS-метрики: стационарность, сезонность, частота

    #### 3️⃣ Предобработка и анализ
    - 🧪 **Песочница**: трансформации в реальном времени (Box-Cox, diff, smoothing)
    - 📊 **Визуализация**: интерактивные графики, корреляции, спектральный анализ
    - 🎯 **Рекомендации**: авто-подбор моделей на основе свойств ряда

    #### 4️⃣ Экспорт результатов
    | Формат | Содержимое | Когда использовать |
    |--------|------------|-------------------|
    | 📄 Excel | Паспорт качества + рекомендации | Отчёт заказчику |
    | 📄 CSV | Очищенные данные | Дальнейшая обработка |
    | 📊 PNG/PDF | Графики | Презентация |

    ---

    ### 🗄️ Подключение к базам данных

    #### PostgreSQL
    ```python
    # requirements.txt
    psycopg2-binary>=2.9.9
    sqlalchemy>=2.0.0
    ```
    **Настройка сервера (`postgresql.conf`):**
    ```conf
    listen_addresses = '*'          # Разрешить внешние подключения
    port = 5432                     # Стандартный порт
    ```
    **Настройка доступа (`pg_hba.conf`):**
    ```conf
    # Разрешить подключение с подсети 192.168.1.0/24
    host    all    all    192.168.1.0/24    md5
    ```

    #### ClickHouse
    ```python
    # requirements.txt
    clickhouse-connect>=0.7.0
    ```
    **Пример запроса:**
    ```sql
    SELECT
        toStartOfDay(timestamp) as date,
        avg(value) as metric
    FROM events
    WHERE date >= '2024-01-01'
    GROUP BY date
    ORDER BY date
    LIMIT 10000
    ```

    ---

    ### 🔧 Устранение частых проблем

    | Ошибка | Решение |
    |--------|---------|
    | `No module named 'psycopg2'` | `pip install psycopg2-binary` |
    | `Connection timed out` | Проверить фаервол, `listen_addresses`, `pg_hba.conf` |
    | `authentication failed` | Проверить пароль, метод аутентификации в `pg_hba.conf` |
    | `Файл пуст` | Проверить кодировку (должна быть UTF-8), наличие заголовков |
    | `Не найдены даты` | Убедиться, что колонка с датами имеет корректный формат (см. `robust_datetime_detector`) |

    ---

    ### 📚 Дополнительные ресурсы
    - 📘 [Документация по валидации](/docs/validation)
    - 📊 [Примеры дашбордов](/examples)
    - 🐛 [Сообщить об ошибке](https://github.com/cisstat/platform/issues)

    > 💡 **Совет**: Для воспроизводимости результатов сохраняйте `requirements.txt` и конфигурационные файлы в системе контроля версий.
    """)


# ────────────────────────────────────────────────────────────
# ГЛАВНАЯ СТРАНИЦА (до загрузки данных)
# ────────────────────────────────────────────────────────────
df = st.session_state.get("df", pd.DataFrame())

if df.empty:

    # 🔹 ПОДСКАЗКА ПО РАБОТЕ С ПЛАТФОРМОЙ
    with st.expander("ℹ️ Как начать работу?", expanded=False):
        st.markdown("""
        ###### Пошаговый алгоритм работы:

        **Загрузка данных**
        - В боковой панели выберите источник: **Файл** (CSV, Excel, JSON) или **База данных** (PostgreSQL, ClickHouse)
        - Загрузите ваш датасет

        **Настройка правил валидации**
        - После загрузки файла перейдите в раздел **Управление правилами** в боковой панели
        - **По умолчанию выбран шаблон "Custom (автогенерация)"** — платформа автоматически создаст правила на основе ваших метаданных
        - Или выберите готовый шаблон: **FAO Prices (CIS)**, **Macro indicators**
        - При необходимости отредактируйте диапазоны и нажмите **Применить правила**

        ---
        **Совет:** Для первого знакомства с платформой используйте автогенерацию правил — это самый быстрый способ начать!
        """)

    # 🔹 СООБЩЕНИЯ НА СТАРТОВОМ ЭКРАНЕ
    st.info("👈 **Загрузите данные через боковую панель для начала анализа.**")

    st.markdown("""
    <div style='background-color: #e3f2fd; padding: 15px; border-radius: 8px; border-left: 4px solid #2196f3; margin: 10px 0;'>
    <strong>Выберите шаблон для валидации датасета</strong><br>
    <small>После загрузки файла перейдите в <b>Управление правилами</b> в боковой панели и выберите подходящий шаблон (по умолчанию — Custom автогенерация)</small>
    </div>
    """, unsafe_allow_html=True)

    # 🔹 СКРЫВАЕМ ВКЛАДКИ ДО ЗАГРУЗКИ
    st.stop()

df = st.session_state.df
ct = st.session_state.col_types

# ────────────────────────────────────────────────────────────
#  БАЛАНСИРОВКА ВКЛАДОК (равномерное распределение, ограничено шириной страницы)
# ────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* 🔧 ГЛАВНОЕ: Ограничиваем контейнер вкладок шириной страницы */
div[data-testid="stTabs"] {
display: flex;
justify-content: center;
width: 100%;
max-width: 100vw;  /* 🔧 Не шире видимой области */
box-sizing: border-box;  /* 🔧 Учитывать padding в ширине */
margin: 0 auto;  /* 🔧 Центрирование внутри страницы */
padding: 0 1rem;  /* 🔧 Отступы по краям для мобильных */
gap: 0;
}
/* 🔽 ДОБАВЛЕНО: СТИЛИ ДЛЯ МЕТРИК */
[data-testid="stMetricValue"] { font-size: 18px !important; font-weight: 600 !important; color: #0F172A !important; }
[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 500 !important; text-transform: uppercase; letter-spacing: 0.5px; }
/* Кнопки вкладок — равная ширина с отступами */
div[data-testid="stTabs"] button[data-testid="stTab"] {
flex: 1 1 0;
min-width: 80px;  /* 🔧 Уменьшено для адаптивности */
max-width: 180px;  /* 🔧 Чуть меньше для запаса */
padding: 10px 6px;
margin: 0 2px;
text-align: center;
white-space: nowrap;
overflow: hidden;
text-overflow: ellipsis;
box-sizing: border-box;  /* 🔧 Важно для расчёта ширины */
}
/* Убираем стандартные отступы Streamlit */
div[data-testid="stTabs"] > div:first-child {
gap: 2px !important;
justify-content: center !important;
width: 100%;
max-width: 100%;  /* 🔧 Ограничение ширины */
}
/* Центрирование текста и иконок */
button[data-testid="stTab"] p,
button[data-testid="stTab"] span {
text-align: center;
width: 100%;
display: block;
}
/* 🔧 НОВЫЙ БЛОК: Адаптив для узких экранов */
@media (max-width: 1200px) {
div[data-testid="stTabs"] button[data-testid="stTab"] {
min-width: 70px;
max-width: 140px;
padding: 8px 4px;
font-size: 14px;
}
}
@media (max-width: 768px) {
div[data-testid="stTabs"] {
padding: 0 0.5rem;  /* 🔧 Меньшие отступы на мобильных */
}
div[data-testid="stTabs"] button[data-testid="stTab"] {
min-width: 60px;
max-width: 110px;
padding: 6px 3px;
font-size: 12px;
margin: 0 1px;
}
}
/* 🔧 НОВЫЙ БЛОК: Горизонтальный скролл если не вмещается */
@media (max-width: 480px) {
div[data-testid="stTabs"] {
overflow-x: auto;
justify-content: flex-start;
padding-bottom: 4px;  /* Место для скроллбара */
}
div[data-testid="stTabs"] button[data-testid="stTab"] {
flex: 0 0 auto;  /* 🔧 Фиксированная ширина на очень узких */
min-width: 90px;
max-width: none;
}
}
/* 🔧 FIX: Убираем лишние отступы у главного контейнера Streamlit */
.main .block-container {
padding-left: 1rem;
padding-right: 1rem;
max-width: 100% !important;  /* 🔧 Разрешаем использовать всю ширину */
}
</style>
""", unsafe_allow_html=True)

# ────────────────────────────────────────────────────────────
#  СОЗДАНИЕ ВКЛАДОК
# ────────────────────────────────────────────────────────────
tab_download, tab_validation, tab_preprocessing, tab_exploratory, tab_modeling, tab_forecasting, tab_taskset = st.tabs([
    "Загрузка",
    "Валидация",
    "Предобработка",
    "Разведочный EDA",
    "Моделирование",
    "Прогнозирование",
    "Задачи"
])


# ────────────────────────────────────────────────────────────
#  ВКЛАДКА 1: ЗАГРУЗКА
# ────────────────────────────────────────────────────────────
with tab_download:
    st.markdown("""
    <div style="padding-left: 20px; margin: 20px 0; text-align: right;">
        <p style="margin: 0 0 10px 0; color: #1e293b; line-height: 1.6; font-size: 18px; font-weight: 400;">
            "Следуя по дороге к научным открытиям, вы не найдете коротких путей, но каждый шаг приближает нас к цели".
        </p>
        <p style="margin: 0; color: #64748B; font-style: italic; font-size: 16px; line-height: 1.5;">
            — Георгий Николаевич Флёров, советский физик-ядерщик, академик АН СССР, один из основателей<br> Объединённого института ядерных исследований (ОИЯИ) в Дубне. 
        </p>
    </div>
    """, unsafe_allow_html=True)

    # Сообщение об активации TS-режима (из session_state)
    if st.session_state.get("ts_mode_active"):
        st.success(
            f"**Режим TimeSeries активирован!** Главная дата: `{st.session_state.primary_date_col}`"
            # 🔧 Убран параметр icon="📈"
        )
        st.info(f"Найдено временных столбцов: {', '.join(st.session_state.col_types.get('date', []))}")

    # ── ПРЕВЬЮ ДАННЫХ ────────────────────────────────────────
    st.markdown("### Превью структуры датасета")

    preview_df = df if len(df) > 0 else pd.DataFrame()
    if not preview_df.empty:
        if len(preview_df) <= 20:
            st.info(f"📊 Всего записей: {len(preview_df)}")
            st.dataframe(preview_df, use_container_width=True, height=300)
        else:
            st.info(f"Показано: первые 10 и последние 10 из {len(preview_df)} записей")
            st.markdown("**● Начало датасета:**")
            st.dataframe(preview_df.head(10), use_container_width=True, height=200)
            st.markdown(f"""
            <div style='text-align: center; padding: 12px; background: #f8fafc; border-radius: 6px; color: #64748B; font-size: 14px;'>
                ⋮   Пропущено {len(preview_df) - 20} строк ⋮  ⋮
            </div>
            """, unsafe_allow_html=True)
            st.markdown("**● Конец датасета:**")
            st.dataframe(preview_df.tail(10), use_container_width=True, height=200)
    else:
        st.info("ℹ️ Датасет пуст.")

    # ── ТЕХНИЧЕСКАЯ ИНФОРМАЦИЯ ─────────────────────────────
    st.divider()
    with st.expander(" Техническая информация о датасете", expanded=False):
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Всего строк", f"{len(df):,}".replace(",", " "))
        c2.metric("Всего колонок", len(df.columns))
        c3.metric("Память", f"{df.memory_usage(deep=True).sum() / 1024**2:.2f} MB")
        c4.metric("Числовых / Текстовых", f"{len(ct['num'])} / {len(ct['cat'])}")

        info_data = []
        for col in df.columns:
            dtype = str(df[col].dtype)
            non_null = df[col].count()
            nulls = df[col].isnull().sum()
            unique = df[col].nunique()
            type_icon = "🔢" if dtype in ['int64', 'float64', 'int32', 'float32'] else "📅" if dtype == 'datetime64[ns]' else "📋" if unique < 50 else "📝"
            info_data.append({
                "КОЛОНКА": col,
                "ТИП": f"{type_icon} {dtype}",
                "НЕ ПУСТО": f"{non_null:,}".replace(",", " "),
                "ПРОПУСКИ": f"{nulls:,}".replace(",", " ") if nulls > 0 else "—",
                "УНИКАЛЬНЫХ": f"{unique:,}".replace(",", " ")
            })

        info_df = pd.DataFrame(info_data)
        st.dataframe(
            info_df, use_container_width=True, hide_index=True,
            column_config={"Пропуски": st.column_config.ProgressColumn("Пропуски", help="Количество пропущенных значений", min_value=0, max_value=len(df), format="%d")},
            height=250
        )
        st.caption(f"Датасет загружен: {st.session_state.get('original_filename', 'неизвестно')}")

    # ────────────────────────────────────────────────────────────
    #  📊 ВИЗУАЛИЗАЦИЯ РАСПРЕДЕЛЕНИЯ ДАННЫХ
    # ────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### Визуализация распределения данных")
    st.caption("Интерактивный анализ распределения числовых признаков: точечный график, гистограмма и статистические метрики для оценки формы распределения, асимметрии и выбросов.")

    # ──  СПРАВКА ПО МЕТОДУ (скрыта по умолчанию) ────────────
    with st.expander(" Справка по описательным статистикам", expanded=False):
        st.markdown("""
        **Назначение:** Визуальная и статистическая оценка распределения данных для выбора подходящих методов анализа и моделей.

        **Используемые методы:**
        - **🔵 Точечный график** — показывает все наблюдения, выявляет выбросы и кластеры
        - **📊 Гистограмма** — демонстрирует распределение данных: отображает частоту их наблюдений через высоту столбцов
        - **📊 KDE (Kernel Density Estimation)** — сглаженная оценка плотности распределения

        **Статистики распределения:**
        | Метрика | Описание | Интерпретация |
        |---------|----------|---------------|
        | **Mean (среднее)** | Среднее арифметическое | Чувствительно к выбросам |
        | **Median (медиана)** | 50-й перцентиль | Устойчива к выбросам |
        | **Std (стандартное отклонение)** | Мера разброса данных | Чем больше, тем шире разброс |
        | **Skewness (асимметрия)** | Мера симметричности | >0: правосторонняя, <0: левосторонняя |
        | **Kurtosis (эксцесс)** | Мера остроты пика | >0: островершинное, <0: плосковершинное |
        | **IQR (межквартильный размах)** | Q3 - Q1 | Разброс средних 50% наблюдений |

        **Авто-детекция типа распределения:**
        Система автоматически определяет тип распределения через:
        1. **KS-тест (Kolmogorov-Smirnov)** — сравнение с теоретическими распределениями
        2. **Анализ асимметрии и эксцесса** — эвристическая классификация
        3. **Проверка на дискретность** — для целочисленных данных

        **Применение:**
        - Нормальное распределение → параметрические тесты, линейные модели
        - Асимметричное → логарифмическая трансформация, Box-Cox
        - Мультимодальное → сегментация данных, mixture models

        **⚠️ Бесплатный курс "Основы статистики"от Анатолия Карпова:**
        - часть 1: https://stepik.org/course/76/syllabus
        - часть 2: https://stepik.org/course/524/syllabus
        - часть 3: https://stepik.org/course/2152/syllabus

        **⚠️ Почитать:**
        - про дисперсию и стандартное отклонение: https://thecode.media/dispersiya-v-data-science/
        """)

    # ── 🔽 CSS ДЛЯ ГОЛУБОГО ФОНА ЗАГОЛОВКА ─────
    st.markdown("""
    <style>
    /* Голубой фон ТОЛЬКО для заголовка expander'а "Показать анализ распределения" */
    div[data-testid="stExpander"] details:has(summary span:nth-child(2)) {
        background-color: #e3f2fd !important;
        border-radius: 8px !important;
        border: 1px solid #bbdefb !important;
        margin-bottom: 10px !important;
    }
    div[data-testid="stExpander"] details summary {
        background-color: #e3f2fd !important;
        border-radius: 8px !important;
        padding: 8px 12px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── 🔽 ОСНОВНОЙ КОНТЕНТ СЕКЦИИ ─────
    with st.expander(" Показать анализ распределения", expanded=False):
        if st.session_state.col_types.get("num"):
            selected_col = st.selectbox(
                "Выберите числовую колонку для анализа распределения:",
                st.session_state.col_types["num"],
                key="dist_col_select"
            )

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("##### 🔵 Точечный график")
                fig_scatter = px.scatter(
                    df, y=selected_col, x=df.index,
                    title=f"Распределение: {selected_col}",
                    labels={"index": "Индекс", selected_col: "Значение"},
                    height=400, opacity=0.6
                )
                fig_scatter.update_traces(marker=dict(size=6, color='#1f77b4'))
                st.plotly_chart(fig_scatter, use_container_width=True)

            with col2:
                st.markdown("##### 📊 Гистограмма с метриками")

                mean_val = df[selected_col].mean()
                median_val = df[selected_col].median()
                q1_val = df[selected_col].quantile(0.25)
                q3_val = df[selected_col].quantile(0.75)
                std_val = df[selected_col].std()
                skew_val = df[selected_col].skew()
                kurt_val = df[selected_col].kurtosis()

                fig_hist = px.histogram(
                    df, x=selected_col, nbins=30,
                    title=f"Гистограмма: {selected_col}",
                    labels={selected_col: "Значение"},
                    height=400, opacity=0.7
                )

                fig_hist.add_vline(x=mean_val, line_dash="dash", line_color="red", line_width=2, annotation_text="Mean", annotation_position="top right")
                fig_hist.add_vline(x=median_val, line_dash="dash", line_color="green", line_width=2, annotation_text="Median", annotation_position="top right")
                fig_hist.add_vline(x=q1_val, line_dash="dot", line_color="orange", line_width=2, annotation_text="Q1", annotation_position="bottom right")
                fig_hist.add_vline(x=q3_val, line_dash="dot", line_color="purple", line_width=2, annotation_text="Q3", annotation_position="bottom right")

                fig_hist.update_layout(showlegend=True, legend=dict(itemsizing='constant', title="Статистики"))
                st.plotly_chart(fig_hist, use_container_width=True)

                st.markdown(
                    '<span style="font-weight: bold; font-size: 16px;"> Статистики распределения</span>',
                    unsafe_allow_html=True
                )

                def detect_distribution_type(series):
                    import numpy as np
                    from scipy import stats
                    data = series.dropna()
                    if len(data) < 30: return "Недостаточно данных для определения (<30 точек)"
                    if len(data) > 5000: data = data.sample(5000, random_state=42)
                    is_discrete = (data == data.astype(int)).all()
                    unique_count = data.nunique()
                    min_val = data.min()
                    mean_v = data.mean()
                    var_v = data.var()
                    skew = stats.skew(data)
                    kurt = stats.kurtosis(data)

                    if is_discrete and unique_count < 100:
                        if unique_count == 2 and min_val >= 0: return "Дискретное - Биномальное"
                        elif min_val >= 1 and var_v > mean_v**2: return "Дискретное - Геометрическое"
                        elif var_v > mean_v * 1.3: return "Дискретное - Отрицательное биномальное"
                        elif abs(var_v - mean_v) < mean_v * 0.25: return "Дискретное - Пуассона"
                        elif unique_count < len(data) * 0.4: return "Дискретное - Гипергеометрическое (оценка)"
                        return "Дискретное - Эмпирическое"

                    candidates = {
                        "Нормальное": stats.norm, "Логнормальное": stats.lognorm,
                        "Экспоненциальное": stats.expon, "Равномерное": stats.uniform,
                        "Стьюдента": stats.t, "Хи-квадрат": stats.chi2, "Гамма": stats.gamma
                    }
                    best_name, best_ks = None, np.inf
                    for name, dist in candidates.items():
                        try:
                            if name in ["Логнормальное", "Экспоненциальное", "Хи-квадрат"] and min_val <= 0: continue
                            params = dist.fit(data)
                            ks_stat, _ = stats.kstest(data, dist.name, args=params)
                            if ks_stat < best_ks: best_ks, best_name = ks_stat, name
                        except: continue

                    prefix = "Непрерывное - "
                    if best_name is None:
                        if abs(skew) < 0.5: return f"{prefix}Нормальное (по асимметрии)"
                        if skew > 0.5: return f"{prefix}Правосторонняя асимметрия"
                        if skew < -0.5: return f"{prefix}Левосторонняя асимметрия"
                        return f"{prefix}Неопределённое"
                    if best_ks < 0.06: return f"{prefix}{best_name}"
                    elif best_ks < 0.14: return f"{prefix}{best_name} (близко)"
                    else:
                        if skew > 0.6: return f"{prefix}Правосторонняя асимметрия"
                        if skew < -0.6: return f"{prefix}Левосторонняя асимметрия"
                        return f"{prefix}Эмпирическое (сложная форма)"

                dist_type = detect_distribution_type(df[selected_col])
                dist_emoji = "🔵" if "Нормальное" in dist_type else "🟠" if "асимметрия" in dist_type.lower() else "🟢" if "Равномерное" in dist_type else "🟣" if "Логнормальное" in dist_type else "⚪"

                st.markdown(f"""
                - Тип распределения: `{dist_type}`
                - Mean (среднее): `{mean_val:,.2f}`
                - Median (медиана): `{median_val:,.2f}`
                - Std (стандартное отклонение): `{std_val:,.2f}`
                - Skewness (асимметрия): `{skew_val:.3f}`
                - Kurtosis (эксцесс): `{kurt_val:.3f}`
                - Q1 (25-й перцентиль): `{q1_val:,.2f}`
                - Q3 (75-й перцентиль): `{q3_val:,.2f}`
                - IQR (межквартильный размах): `{q3_val - q1_val:,.2f}`
                """)
        else:
            st.warning("⚠️ В датасете нет числовых колонок для визуализации распределения.")

    # ────────────────────────────────────────────────────────────
    #  📊 КОРРЕЛЯЦИОННЫЙ АНАЛИЗ ЧИСЛОВЫХ ПРИЗНАКОВ
    # ────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### Корреляция числовых признаков")
    st.caption("Анализ линейных взаимосвязей между числовыми переменными через коэффициент корреляции Пирсона. Выявление сильных связей, мультиколлинеарности и независимых признаков для улучшения качества моделей.")

    # ── 📘 СПРАВКА ПО МЕТОДУ (скрыта по умолчанию) ────────────
    with st.expander(" Справка по методу", expanded=False):
        st.markdown("""
        **Назначение:** Обнаружение и оценка силы линейных взаимосвязей между парами числовых признаков.

        **Используемые методы:**
        - **Коэффициент корреляции Пирсона (r)** — мера линейной зависимости между двумя переменными
        - **Тепловая карта (heatmap)** — визуализация матрицы корреляций
        - **Пороговые значения** — фильтрация значимых связей

        **Интерпретация коэффициента Пирсона:**
        | Значение r | Сила связи | Интерпретация |
        |------------|------------|---------------|
        | **0.0 – 0.3** | Слабая/отсутствует | Практически нет линейной зависимости |
        | **0.3 – 0.5** | Умеренная | Есть заметная связь |
        | **0.5 – 0.7** | Заметная | Уверенная линейная зависимость |
        | **0.7 – 0.9** | Сильная | Тесная связь между признаками |
        | **0.9 – 1.0** | Очень сильная | Почти функциональная зависимость |

        **Критическая корреляция (|r| ≥ 0.85):**
        - Возможна **мультиколлинеарность** → нестабильность оценок в регрессии
        - Признаки могут быть **дубликатами** или сильно связаны
        - Рекомендуется: PCA, удаление одного признака, регуляризация

        **Применение:**
        - Отбор признаков для моделей машинного обучения
        - Обнаружение избыточных переменных
        - Понимание структуры данных
        - Подготовка к построению регрессионных моделей

        **Формула коэффициента Пирсона:**
        ```
        r = Σ[(xi - x̄)(yi - ȳ)] / √[Σ(xi - x̄)² × Σ(yi - ȳ)²]
        ```
        где:
        - xi, yi — значения переменных
        - x̄, ȳ — средние значения
        - r ∈ [-1; +1]

        **⚠️ Почитать:**
        - про корреляцию: https://thecode.media/chto-takoe-korrelyatsiya-obyasnyaem-prostymi-slovami-na-primerah/
        """)

    # Получаем список числовых колонок из типов данных
    num_cols = ct.get("num", [])

    # 🔹 ЭКСПАНДЕР: заголовок виден, контент скрыт по умолчанию
    with st.expander(" Показать корреляционный анализ", expanded=False):

        if len(num_cols) >= 2:
            col_corr1, col_corr2 = st.columns(2)

            # ── ЛЕВАЯ ЧАСТЬ: ГРАФИК ──────────────────────────────
            with col_corr1:
                st.markdown("### **Тепловая карта корреляции Пирсона**")
                # Рассчитываем матрицу корреляции Пирсона только для числовых признаков
                corr_matrix = df[num_cols].corr()

                fig, ax = plt.subplots(figsize=(8, 6))
                sns.heatmap(corr_matrix,
                            annot=True,
                            cmap='coolwarm',
                            fmt='.2f',
                            linewidths=0.5,
                            ax=ax,
                            cbar_kws={'label': 'Коэффициент корреляции (r)'})
                ax.set_title("Матрица корреляции числовых признаков", fontsize=12)
                st.pyplot(fig, use_container_width=True)

            # ── ПРАВАЯ ЧАСТЬ: ПОЯСНЕНИЯ ──────────────────────────
            with col_corr2:
                st.markdown("### **Пояснения к корреляциям**")

                # Анализ значимых связей (r >= 0.5 или r <= -0.5)
                significant_links = []
                for i in range(len(corr_matrix.columns)):
                    for j in range(i + 1, len(corr_matrix.columns)):
                        val = corr_matrix.iloc[i, j]
                        if abs(val) >= 0.5:  # Порог значимости
                            col1, col2 = corr_matrix.columns[i], corr_matrix.columns[j]
                            strength = "Сильная" if abs(val) >= 0.7 else "Умеренная"
                            direction = "прямая (+)" if val > 0 else "обратная (-)"
                            significant_links.append({
                                "pair": f"{col1} ↔ {col2}",
                                "val": val,
                                "desc": f"{strength} {direction} связь (`r = {val:.2f}`)"
                            })

                if significant_links:
                    st.info(f"Найдено {len(significant_links)} значимых связей (|r| ≥ 0.5):")
                    for item in significant_links:
                        st.markdown(f"- **{item['pair']}**: {item['desc']}")
                        if abs(item['val']) >= 0.85:
                            st.error("⚠️ **Критическая корреляция!** Возможна мультиколлинеарность или дублирование данных.")
                else:
                    st.success("✅ Значимых линейных связей между числовыми признаками не обнаружено.")
                    st.markdown("**Интерпретация:** Наблюдения независимы друг от друга, что является хорошим признаком для моделей машинного обучения.")

                    # Показываем слабые связи, если их нет значимых
                    if st.checkbox("Показать слабые связи (0.3 < |r| < 0.5)", key="show_weak_corr"):
                        weak_links = []
                        for i in range(len(corr_matrix.columns)):
                            for j in range(i + 1, len(corr_matrix.columns)):
                                val = corr_matrix.iloc[i, j]
                                if 0.3 <= abs(val) < 0.5:
                                    col1, col2 = corr_matrix.columns[i], corr_matrix.columns[j]
                                    weak_links.append(f"{col1} ↔ {col2} (`r={val:.2f}`)")
                        if weak_links:
                            st.caption(", ".join(weak_links))
        else:
            st.info("Для корреляционного анализа необходимо минимум 2 числовые колонки.")

    # ───────────────────────────────────────────────────────────
    #  ВАРИАТИВНАЯ ВИЗУАЛИЗАЦИЯ (df_filtered)
    # ───────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### Вариативная визуализация")
    st.caption("Интерактивная фильтрация данных по категориям и временным периодам. Динамическое переключение между общим анализом и визуализацией временных рядов для исследования паттернов в подвыборках.")

    # 🔹 ЭКСПАНДЕР: заголовок виден, контент скрыт по умолчанию
    with st.expander(" Показать панель фильтров и визуализации", expanded=False):

        # Категориальные колонки для фильтров
        cat_cols = df.select_dtypes(include=['object', 'string']).columns.tolist()
        if not cat_cols:
            cat_cols = [c for c in df.select_dtypes(include='number').columns if 1 < df[c].nunique() < 100]

        # Приоритетные колонки
        for p in ["Country", "Region", "Регион", "Страна", "Категория", "Product"]:
            if p in cat_cols:
                cat_cols.insert(0, cat_cols.pop(cat_cols.index(p)))

        category_col_1 = st.selectbox(
            "📂 Первая категория",
            options=cat_cols + ["(нет)"],
            index=0 if cat_cols else len(cat_cols),
            key="sel_cat1"
        )
        category_col_2 = st.selectbox(
            "📂 Вторая категория (опционально)",
            options=cat_cols + ["(нет)"],
            index=len(cat_cols),
            key="sel_cat2"
        )

        if category_col_1 != "(нет)" and category_col_1 == category_col_2:
            st.warning("⚠️ Вы выбрали одну и ту же колонку дважды. Вторая категория будет проигнорирована.")

        if category_col_1 != "(нет)":
            all_cat1 = sorted(df[category_col_1].dropna().astype(str).unique().tolist())
            selected_cat1 = st.multiselect(
                f"🔢 Значения: {category_col_1}",
                options=all_cat1,
                default=[],
                key="filt_cat1",
                placeholder="Выберите..."
            )
        else:
            selected_cat1 = []

        if category_col_2 != "(нет)" and category_col_2 != category_col_1:
            all_cat2 = sorted(df[category_col_2].dropna().astype(str).unique().tolist())
            selected_cat2 = st.multiselect(
                f"🔢 Значения: {category_col_2}",
                options=all_cat2,
                default=[],
                key="filt_cat2",
                placeholder="Выберите..."
            )
        else:
            selected_cat2 = []

        # Временная колонка
        datetime_cols = df.select_dtypes(include=['datetime64[ns]']).columns.tolist()
        num_year_cols = [c for c in df.select_dtypes(include='number').columns if 'year' in c.lower() or 'год' in c.lower()]
        time_cols = datetime_cols + [c for c in num_year_cols if c not in datetime_cols]

        for p in ["Year", "Date", "Год", "Дата"]:
            if p in time_cols:
                time_cols.insert(0, time_cols.pop(time_cols.index(p)))

        if time_cols:
            time_col = st.selectbox("📅 Временная колонка", options=time_cols, index=0, key="selector_time_col")
        else:
            time_col = None
            st.caption("ℹ️ В датасете не найдено колонок с датами или годами")

        # Фильтр по годам
        if time_col and time_cols:
            if time_col in datetime_cols:
                df['_tmp_year'] = df[time_col].dt.year.astype(str)
                all_years = sorted(df['_tmp_year'].dropna().unique().tolist(), reverse=True)
                time_col_filter = '_tmp_year'
            else:
                all_years = sorted(df[time_col].dropna().unique().astype(int).astype(str).tolist(), reverse=True)
                time_col_filter = time_col

            selected_years = st.multiselect(
                "📅 Годы",
                options=all_years,
                default=[],
                key="filt_years",
                placeholder="Выберите годы..."
            )
            if not selected_years:
                selected_years = all_years
        else:
            all_years, selected_years, time_col_filter = [], [], None

        # ── ПРИМЕНЕНИЕ ФИЛЬТРОВ ─────────────────────────────────
        df_filtered = df.copy()

        if selected_cat1 and category_col_1 != "(нет)":
            df_filtered = df_filtered[df_filtered[category_col_1].astype(str).isin(selected_cat1)]
        if selected_cat2 and category_col_2 != "(нет)":
            df_filtered = df_filtered[df_filtered[category_col_2].astype(str).isin(selected_cat2)]
        if selected_years and time_col_filter and time_col_filter in df_filtered.columns:
            df_filtered = df_filtered[df_filtered[time_col_filter].astype(str).isin(selected_years)]
        if time_col and time_col in df_filtered.columns:
            df_filtered = df_filtered.sort_values(time_col).reset_index(drop=True)
        if '_tmp_year' in df_filtered.columns:
            df_filtered = df_filtered.drop(columns=['_tmp_year'])

        st.caption(
            f"📊 Активно: {len(df_filtered)} записей | "
            f"{len(selected_cat1)} значений (кат.1) | "
            f"{len(selected_cat2)} значений (кат.2) | "
            f"{len(selected_years)} лет"
        )

        if df_filtered.empty:
            st.warning("⚠️ Нет данных для отображения. Измените фильтры.")
            st.stop()

        # ───────────────────────────────────────────────────────────
        #  2. ПОДГОТОВКА МЕТРИК И TS MODE (ct_f, ts_mode_active)
        # ───────────────────────────────────────────────────────────
        # Определяем константы режимов, если их нет глобально
        MODE_TS = "Временные ряды"
        MODE_GEN = "Общий (категории)"

        ct_f = {
            "num": df_filtered.select_dtypes(include='number').columns.tolist(),
            "cat": [c for c in df_filtered.select_dtypes(include=['object', 'string']).columns if 1 < df_filtered[c].nunique() < 100],
            "date": df_filtered.select_dtypes(include='datetime').columns.tolist()
        }

        # Инициализируем ts_mode_active
        ts_mode_active = False
        df_ts = pd.DataFrame()

        if ct_f["date"] and ct_f["num"]:
            date_col = ct_f["date"][0]
            df_ts = df_filtered.sort_values(by=date_col).set_index(date_col).copy()
            ts_mode_active = True

        # Автоматическая установка режима визуализации
        default_mode = MODE_TS if ts_mode_active else MODE_GEN

        if "viz_mode_toggle" not in st.session_state:
            st.session_state.viz_mode_toggle = default_mode
        elif ts_mode_active and st.session_state.viz_mode_toggle == MODE_GEN:
            # Если TS активирован, но выбран общий режим — переключаем на TS
            # (только при первой инициализации или смене данных)
            if "last_data_hash" not in st.session_state or st.session_state.get("last_data_hash") != hash(df_filtered.to_string()):
                st.session_state.viz_mode_toggle = MODE_TS
                st.session_state.last_data_hash = hash(df_filtered.to_string())

        # Сохраняем хэш данных для отслеживания изменений
        if "last_data_hash" not in st.session_state:
            st.session_state.last_data_hash = hash(df_filtered.to_string())

        if ts_mode_active:
            # Показываем метрики ТОЛЬКО если режим TS
            if st.session_state.viz_mode_toggle == MODE_TS:
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.metric("Период", f"{df_ts.index.min().date()} — {df_ts.index.max().date()}")
                with c2:
                    st.metric("Записей", f"{len(df_ts):,}".replace(",", " "))
                with c3:
                    # Безопасное определение частоты
                    try:
                        freq = pd.infer_freq(df_ts.index) if len(df_ts.index) >= 3 else None
                    except ValueError:
                        freq = None
                    freq_label = freq if freq else "Нерегулярная"
                    st.metric("Частота", freq_label)
        else:
            # Если TS не активен — принудительно общий режим
            st.session_state.viz_mode_toggle = MODE_GEN
            ts_mode_active = False
            st.info("ℹ️ Для анализа временных рядов необходимы колонки с датами и числовые данные. Включен общий режим.")

        # ───────────────────────────────────────────────────────────
        #  3. СОЗДАНИЕ ВКЛАДОК (tab_g, tab_k)
        # ───────────────────────────────────────────────────────────
        tab_g, tab_k = st.tabs(["📊 Графики", "📐 KPI и метрики"])

        # ── ВКЛАДКА: ГРАФИКИ ─────────────────────────────────────
        with tab_g:
            # Корректный radio с установкой индекса по умолчанию
            mode_options = [MODE_GEN, MODE_TS] if ts_mode_active else [MODE_GEN]
            default_index = 1 if (ts_mode_active and st.session_state.viz_mode_toggle == MODE_TS) else 0

            mode_toggle = st.radio(
                "**РЕЖИМ ВИЗУАЛИЗАЦИИ**",
                options=mode_options,
                index=default_index,
                horizontal=True,
                key="viz_mode_toggle",
                help="🔹 Временные ряды: анализ динамики, трендов, сезонности\n🔹 Общий: сравнение категорий, группировка"
            )

            if mode_toggle == "Общий (категории)":
                c1, c2, c3 = st.columns(3)
                with c1:
                    chart_type = st.selectbox(
                        "Тип графика",
                        ["📊 Столбчатая", "📈 Линейная", "📉 Площадь", "🔵 Точечная", "🔢 Box Plot", "🔢 Гистограмма", "🔢 Гистограмма+KDE", "🔢 Воронка"],
                        key="sel_chart"
                    )
                with c2:
                    x_opts = ct["cat"] + ct["date"] + ct["num"]
                    cat_col = st.selectbox("Ось X", options=x_opts if x_opts else [""], index=0 if x_opts else None, key="sel_x")
                    val_col = st.selectbox("Ось Y", options=ct["num"] if ct["num"] else [""], index=0 if ct["num"] else None, key="sel_y")
                with c3:
                    color_col = st.selectbox("Цвет", options=["Нет"] + (ct["cat"] if ct["cat"] else []), key="sel_color")
                    size_col = st.selectbox("Размер", options=["Нет"] + (ct["num"] if ct["num"] else []), key="sel_size")
                    agg_method = st.selectbox("Агрегация", ["mean", "sum", "min", "max", "count"], index=0, key="sel_agg")

                if st.button("📈 Построить", type="primary", use_container_width=True, key="btn_build_chart"):
                    try:
                        if not cat_col or not val_col or cat_col == "" or val_col == "":
                            st.error("❌ Пожалуйста, выберите колонки для осей X и Y.")
                        elif cat_col not in df.columns or val_col not in df.columns:
                            st.error(f"❌ Выбранные колонки не найдены в данных.")
                        else:
                            base_group_cols = [c for c in [cat_col, category_col_1, category_col_2] if c != "(нет)" and c != ""]
                            group_cols = list(dict.fromkeys(base_group_cols))
                            color_arg = category_col_1 if (category_col_1 != "(нет)" and len(selected_cat1) > 1) else (color_col if color_col != "Нет" else None)
                            is_multi = len(selected_cat1) > 1 or len(selected_cat2) > 1
                            has_dup = df.groupby(group_cols, observed=True)[val_col].transform('count').max() > 1 if len(df) > 0 and group_cols else False

                            if has_dup and group_cols:
                                agg_data = df.groupby(group_cols, observed=True)[val_col].agg(agg_method).reset_index()
                                title_note = f" ({agg_method})"
                            else:
                                cols_to_keep = [c for c in group_cols + [val_col] if c in df.columns]
                                agg_data = df[cols_to_keep].drop_duplicates().reset_index(drop=True)
                                title_note = ""

                            full_title = f"{val_col}{title_note} по {cat_col}"

                            if chart_type == "📊 Столбчатая":
                                fig = px.bar(agg_data, x=cat_col, y=val_col, color=color_arg, barmode="group", title=full_title)
                            elif chart_type == "📈 Линейная":
                                fig = px.line(agg_data, x=cat_col, y=val_col, color=color_arg, markers=True, title=full_title)
                            elif chart_type == "📉 Площадь":
                                fig = px.area(agg_data, x=cat_col, y=val_col, color=color_arg, title=full_title)
                            elif chart_type == "🔵 Точечная":
                                fig = px.scatter(df, x=cat_col, y=val_col, color=color_arg, size=(size_col if size_col!="Нет" else None), title=full_title)
                            elif chart_type == "🔢 Box Plot":
                                fig = px.box(df, x=cat_col, y=val_col, color=color_arg, points="all", title=full_title)
                            elif chart_type == "📊 Гистограмма":
                                fig = px.histogram(df, x=val_col, color=color_arg, nbins=30, title=f"Распределение {val_col}")
                            elif chart_type == "📊 Гистограмма+KDE":
                                fig = px.histogram(df, x=val_col, color=color_arg, nbins=30, title=f"Распределение {val_col}", histnorm='probability density')
                                kde_x = np.linspace(df[val_col].min(), df[val_col].max(), 100)
                                kde_y = stats.gaussian_kde(df[val_col].dropna())(kde_x)
                                fig.add_trace(go.Scatter(x=kde_x, y=kde_y, mode='lines', name='KDE', line=dict(color='red', width=2)))
                            elif chart_type == "🔻 Воронка":
                                f_data = agg_data.sort_values(val_col, ascending=True)
                                fig = px.funnel(f_data, x=val_col, y=cat_col, title=f"Воронка: {val_col}")
                            else:
                                raise ValueError("Неподдерживаемый тип графика")

                            fig.update_layout(template="plotly_white", height=500, showlegend=bool(is_multi or color_arg), hovermode='x unified')
                            st.session_state["current_chart"] = fig
                            st.session_state["chart_built"] = True
                            st.success("✅ График построен!")

                    except Exception as e:
                        st.error(f"❌ Ошибка построения: {str(e)}")
                        st.session_state["chart_built"] = False

                if st.session_state.get("chart_built") and "current_chart" in st.session_state:
                    st.plotly_chart(st.session_state["current_chart"], use_container_width=True)
                    with st.expander("🔢 Показать данные графика", expanded=False):
                        safe_data = locals().get("agg_data", df)
                        st.dataframe(safe_data, use_container_width=True, height=300)

            else:  # MODE_TS
                if not ts_mode_active:
                    st.warning("⚠️ Для режима временных рядов нужна колонка с датами. Переключитесь на общий режим или загрузите данные с датами.")
                else:
                    c1, c2 = st.columns(2)
                    with c1:
                        ts_metric = st.selectbox("Метрика (Y)", ct["num"], key="ts_metric")
                    with c2:
                        ts_chart_type = st.selectbox(
                            "Тип TS-анализа",
                            ["📈 Линейный тренд", "🔄 Скользящее среднее", "📊 Автокорреляция (ACF)", "🧩 Декомпозиция (STL)"],
                            key="ts_chart_type"
                        )

                    window, lags = None, None
                    if ts_chart_type == "🔄 Скользящее среднее":
                        window = st.slider("Окно сглаживания (периодов)", 3, 30, 7, key="ts_window")
                    elif ts_chart_type == "📊 Автокорреляция (ACF)":
                        lags = st.slider("Количество лагов", 10, 50, 20, key="ts_lags")

                    show_trend = show_seasonal = show_cyclical = show_residual = True

                    # ── ПОДСКАЗКИ ДЛЯ КОМПОНЕНТОВ ДЕКОМПОЗИЦИИ (ТОЛЬКО ДЛЯ STL) ──
                    if ts_chart_type == "🧩 Декомпозиция (STL)":
                        st.markdown("##### 🔄 Выберите компоненты для отображения:")
                        c_chk1, c_chk2, c_chk3, c_chk4 = st.columns(4)

                        with c_chk1:
                            show_trend = st.checkbox(
                                "📈 Тренд",
                                value=True,
                                key="stl_chk_trend",
                                help=(
                                    "**📊 Тренд — долгосрочная направленность ряда**\n\n"
                                    "**Описание:** Показывает основную тенденцию изменения данных "
                                    "(рост, падение или стабильность) после удаления сезонных и "
                                    "циклических колебаний.\n\n"
                                    "**Метод:** STL использует локальную регрессию LOESS "
                                    "(Locally Estimated Scatterplot Smoothing) с трикубическими весами.\n\n"
                                    "**Формула:** `w(x) = (1 - |x|³)³` для |x| < 1\n\n"
                                    "**Параметры:** Окно тренда: 21 период (нечётное число), "
                                    "робастность: устойчив к выбросам до 15%.\n\n"
                                    "**Интерпретация:** Положительный наклон → рост, "
                                    "отрицательный → падение, горизонтальный → стабильность."
                                )
                            )

                        with c_chk2:
                            show_seasonal = st.checkbox(
                                "🔄 Сезонность",
                                value=True,
                                key="stl_chk_seasonal",
                                help=(
                                    "**📊 Сезонность — регулярные циклы**\n\n"
                                    "**Описание:** Периодические колебания фиксированной частоты, "
                                    "повторяющиеся через равные интервалы времени "
                                    "(календарная зависимость).\n\n"
                                    "**Метод:** Извлечение сезонной компоненты через STL-декомпозицию "
                                    "(итеративная процедура с циклическим сглаживанием).\n\n"
                                    "**Формула:** `Y(t) = Trend(t) + Seasonal(t) + Residual(t)`\n\n"
                                    "**Параметры:** Период: автоматически определяется через ACF "
                                    "(автокорреляцию), аддитивная модель.\n\n"
                                    "**Примеры:** 12 месяцев (год), 4 квартала, 7 дней (неделя), "
                                    "30 дней (месяц)."
                                )
                            )

                        with c_chk3:
                            show_cyclical = st.checkbox(
                                "🔁 Цикличность",
                                value=False,
                                key="stl_chk_cyclical",
                                help=(
                                    "**📊 Цикличность — среднесрочные колебания**\n\n"
                                    "**Описание:** Нерегулярные колебания, связанные с "
                                    "экономическими или бизнес-циклами (не календарная сезонность).\n\n"
                                    "**Метод:** Разница между трендом и его долгосрочным сглаживанием "
                                    "через простое скользящее среднее.\n\n"
                                    "**Формула:** `Cyclical(t) = Trend(t) - SMA₃₀(Trend(t))`\n\n"
                                    "**Параметры:** SMA₃₀ — простое скользящее среднее за 30 периодов "
                                    "(настраивается).\n\n"
                                    "**Отличие от сезонности:** Циклы нерегулярны (3-10 лет), "
                                    "амплитуда меняется, связаны с экономикой."
                                )
                            )

                        with c_chk4:
                            show_residual = st.checkbox(
                                "🔢 Остаток",
                                value=False,
                                key="stl_chk_residual",
                                help=(
                                    "**📊 Остаток — случайная компонента**\n\n"
                                    "**Описание:** Непредсказуемая часть ряда после извлечения всех "
                                    "систематических компонент (тренд, сезонность, циклы).\n\n"
                                    "**Метод:** Вычитание всех систематических компонент из исходного ряда.\n\n"
                                    "**Формула:** `Residual(t) = Y(t) - Trend(t) - Seasonal(t)`\n\n"
                                    "**Параметры:** Содержит: случайный шум, аномалии, выбросы, "
                                    "ошибки измерения, непредсказуемые события.\n\n"
                                    "**Диагностика:** ✅ Хорошо = белый шум (нет паттернов), "
                                    "⚠️ Плохо = есть автокорреляция или тренд."
                                )
                            )

                    # 🔧 ИСПРАВЛЕНИЕ: Уникальный ключ для кнопки, чтобы избежать дублирования
                    if st.button("Построить TS-график", type="primary", use_container_width=True, key="btn_ts_chart_unique"):
                        try:
                            fig = go.Figure()
                            series = df_ts[ts_metric].resample('D').mean().dropna()

                            if len(series) < 20:
                                st.error(f"❌ Недостаточно данных: {len(series)} точек (минимум 20)")
                            else:
                                if ts_chart_type == "📈 Линейный тренд":
                                    fig.add_trace(go.Scatter(x=series.index, y=series, mode='lines', name=ts_metric, line=dict(color='#2563EB', width=2)))
                                    fig.update_layout(title=f"Динамика {ts_metric}", xaxis_title="Дата", yaxis_title="Значение", template="plotly_white", height=500)
                                elif ts_chart_type == "🔄 Скользящее среднее":
                                    rolling = series.rolling(window=window).mean()
                                    fig.add_trace(go.Scatter(x=series.index, y=series, mode='lines', name='Исходный', opacity=0.3, line=dict(color='gray')))
                                    fig.add_trace(go.Scatter(x=series.index, y=rolling, mode='lines', name=f'MA({window})', line=dict(color='red', width=3)))
                                    fig.update_layout(title=f"Скользящее среднее (Window={window})", xaxis_title="Дата", template="plotly_white", height=500)
                                elif ts_chart_type == "📊 Автокорреляция (ACF)":
                                    acf_vals = [series.autocorr(lag=i) for i in range(min(lags, len(series)//2))]
                                    fig = px.bar(x=list(range(len(acf_vals))), y=acf_vals, title=f"ACF для {ts_metric}", labels={'x': 'Лаг', 'y': 'Корреляция'})
                                    fig.add_hline(y=0, line_dash="dot", line_color="gray")
                                    fig.add_hline(y=1.96/np.sqrt(len(series)), line_dash="dash", line_color="green")
                                    fig.add_hline(y=-1.96/np.sqrt(len(series)), line_dash="dash", line_color="green")
                                    fig.update_layout(template="plotly_white", height=400, showlegend=False)
                                elif ts_chart_type == "🧩 Декомпозиция (STL)":
                                    from statsmodels.tsa.seasonal import STL
                                    from statsmodels.tsa.stattools import acf

                                    try:
                                        acf_vals = acf(series.dropna().nlargest(100), nlags=min(50, len(series)//2), fft=True)
                                        peaks = [i for i in range(2, len(acf_vals)) if acf_vals[i] > acf_vals[i-1] and acf_vals[i] > 0.3]
                                        period = peaks[0] if peaks else 7
                                        period = max(2, min(period, len(series)//4))
                                    except Exception:
                                        period = 7

                                    seasonal_window = min(13, period*2 + 1)
                                    trend_window = min(21, len(series)//4)
                                    if trend_window <= period: trend_window = period + 1
                                    if trend_window % 2 == 0: trend_window += 1

                                    stl = STL(series, period=period, seasonal=seasonal_window, trend=trend_window, robust=True)
                                    result = stl.fit()
                                    trend, seasonal, residual = result.trend, result.seasonal, result.resid
                                    cyclical = trend - trend.rolling(window=min(30, len(trend)//4), center=True, min_periods=1).mean()

                                    fig = make_subplots(rows=1, cols=2, subplot_titles=(" Исходный ряд", " Компоненты"), horizontal_spacing=0.1, column_widths=[0.4, 0.6])
                                    fig.add_trace(go.Scatter(x=series.index, y=series, mode='lines', name='Исходный', line=dict(color='#2563EB', width=2)), row=1, col=1)

                                    colors = {'trend': '#16a34a', 'seasonal': '#dc2626', 'cyclical': '#9333ea', 'residual': '#f59e0b'}
                                    if show_trend:
                                        fig.add_trace(go.Scatter(x=trend.index, y=trend, mode='lines', name='Тренд', line=dict(color=colors['trend'], width=2)), row=1, col=2)
                                    if show_seasonal:
                                        fig.add_trace(go.Scatter(x=seasonal.index, y=seasonal, mode='lines', name='Сезонность', line=dict(color=colors['seasonal'], width=1.5, dash='dash')), row=1, col=2)
                                    if show_cyclical:
                                        fig.add_trace(go.Scatter(x=cyclical.index, y=cyclical, mode='lines', name='Цикличность', line=dict(color=colors['cyclical'], width=1.5, dash='dot')), row=1, col=2)
                                    if show_residual:
                                        fig.add_trace(go.Scatter(x=residual.index, y=residual, mode='lines', name='Остаток', opacity=0.7, line=dict(color=colors['residual'], width=1)), row=1, col=2)

                                    fig.update_layout(height=500, showlegend=True, hovermode='x unified', template="plotly_white", title=f"Декомпозиция {ts_metric} (STL)")
                                    fig.update_xaxes(title_text="Дата", row=1, col=1)
                                    fig.update_xaxes(title_text="Дата", row=1, col=2)
                                    fig.update_yaxes(title_text=ts_metric, row=1, col=1)
                                    fig.update_yaxes(title_text="Значение", row=1, col=2)

                                st.session_state.ts_current_chart = fig
                                st.session_state.ts_chart_built = True
                                st.success(f"✅ График '{ts_chart_type}' построен!")
                                st.rerun()

                        except Exception as e:
                            st.error(f"❌ Ошибка TS-графика: {e}")
                            st.session_state.ts_chart_built = False

                    if st.session_state.get("ts_chart_built") and st.session_state.get("ts_current_chart") is not None:
                        st.plotly_chart(st.session_state.ts_current_chart, use_container_width=True)
                        if ts_chart_type == "🧩 Декомпозиция (STL)":
                            with st.expander("📊 Статистика компонент", expanded=False):
                                try:
                                    from statsmodels.tsa.seasonal import STL
                                    stl_stat = STL(df_ts[ts_metric].resample('D').mean().dropna(), period=7, robust=True).fit()
                                    c1, c2, c3, c4 = st.columns(4)
                                    def fmt_var(x): return f"{x:,.2f}".replace(',', ' ')
                                    c1.metric("📈 Дисперсия тренда", fmt_var(stl_stat.trend.var()))
                                    c2.metric("🔄 Дисперсия сезонности", fmt_var(stl_stat.seasonal.var()))
                                    c3.metric("🔁 Дисперсия цикличности", fmt_var((stl_stat.trend - stl_stat.trend.rolling(30, min_periods=1).mean()).var()))
                                    c4.metric("🔢 Дисперсия остатка", fmt_var(stl_stat.resid.var()))
                                except Exception:
                                    st.caption("ℹ️ Статистика недоступна")

        # ── ВКЛАДКА: KPI И МЕТРИКИ ──────────────────────────────
        with tab_k:
            st.markdown("### 📐 KPI и метрики")

            if not ct["num"]:
                st.info("ℹ️ Нет числовых колонок для расчёта метрик")
            else:
                c1, c2 = st.columns(2)
                with c1:
                    kpi_m = st.selectbox("Метрика", ct["num"], key="kpi_m_kpi")
                    kpi_f = st.selectbox("Функция", ["sum", "mean", "median", "min", "max"], key="kpi_f_kpi")
                with c2:
                    show_trend = st.checkbox("Тренд", value=True, key="show_trend_kpi")
                    trend_opts = ct["date"] + ct["num"]
                    trend_c = st.selectbox("Ось времени", trend_opts if trend_opts else [""], disabled=not show_trend, key="trend_c_kpi")

                if st.button("Рассчитать KPI", type="secondary", key="btn_calc_kpi"):
                    if kpi_m in df.columns:
                        val = getattr(df[kpi_m], kpi_f)()
                        st.metric(kpi_m, f"{val:,.2f}")
                        if show_trend and trend_c and trend_c in df.columns:
                            df_s = df.sort_values(trend_c)
                            if len(df_s) >= 2:
                                h1 = df_s.iloc[:len(df_s)//2][kpi_m].sum()
                                h2 = df_s.iloc[len(df_s)//2:][kpi_m].sum()
                                delta = ((h2-h1)/h1*100) if h1 != 0 else 0
                                st.metric("Тренд", f"{delta:+.1f}%", delta=f"{delta:+.1f}%")
                    else:
                        st.error(f"❌ Колонка '{kpi_m}' не найдена в данных")

    # ────────────────────────────────────────────────────────────
    #  СПЕКТРАЛЬНЫЙ АНАЛИЗ ВРЕМЕННОГО РЯДА (FFT, Wavelet, ACF/PACF)
    # ────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("###  Предварительный спектральный анализ временного ряда")
    st.caption("Выявление скрытых периодичностей и частотных составляющих для улучшения качества прогнозирования")

    # ── 📘 СПРАВКА ПО МЕТОДАМ (скрыта по умолчанию) ────────────
    with st.expander(" Справка по методам спектрального анализа", expanded=False):
        st.markdown("""
        **Назначение:** Спектральный анализ позволяет выявить скрытые периодичности, сезонность и циклические паттерны во временных рядах, которые не видны при визуальном осмотре.

        ######  1. Автокорреляционный анализ (ACF/PACF)

        **Что это:**
        - **ACF (Autocorrelation Function)** — корреляция ряда с самим собой при различных временных сдвигах (лагах)
        - **PACF (Partial Autocorrelation Function)** — корреляция между наблюдениями на определённом лаге при удалении влияния промежуточных лагов

        **Как работает:**
        - Вычисляет коэффициент корреляции Пирсона между `Y(t)` и `Y(t-k)` для разных `k` (лагов)
        - Формула: `ACF(k) = Corr(Y_t, Y_{t-k})`
        - Доверительный интервал: `±1.96/√n` (95% уровень значимости)

        **Как читать графики:**
        - Вертикальные линии за пределами синей зоны → статистически значимая автокорреляция
        - Пики на регулярных лагах (7, 14, 21 или 12, 24, 36) → сезонность
        - Медленное затухание → наличие тренда (нестационарность)

        **Применение:**
        - Определение параметра `p` (AR) для ARIMA — по значимым лагам PACF
        - Определение параметра `q` (MA) для ARIMA — по значимым лагам ACF
        - Обнаружение сезонности `m` — по периодическим пикам

        **⚠️ Почитать:**
        - коэффициент корреляции Пирсона: https://education.yandex.ru/handbook/ml/article/analitika-vremennyh-ryadov

        ---

        ######  2. Преобразование ФУРЬЕ (FFT — Fast Fourier Transform)

        **Что это:**
        Математическое преобразование, которое раскладывает временной ряд на сумму синусоид (гармоник) разных частот и амплитуд.

        **Как работает:**
        - Переводит сигнал из **временной области** (time domain) в **частотную область** (frequency domain)
        - Формула: `X(f) = Σ x(t) · e^{-2πift}`
        - Быстрое преобразование (FFT) — алгоритм сложности O(n log n) вместо O(n²)

        **Как читать график:**
        - **Ось X:** Частота (циклы на единицу времени)
        - **Ось Y:** Амплитуда (сила/мощность данной частоты)
        - **Пики:** Доминирующие периодичности в данных
        - **Период = 1 / Частота** (например, частота 0.0027 → период 365 дней)

        **Применение:**
        - Создание **Fourier features** для ML-моделей: `sin(2πt/P)`, `cos(2πt/P)`
        - Обнаружение скрытых циклов (недельных, месячных, годовых)
        - Фильтрация шума (удаление высокочастотных компонент)

        **⚠️ Почитать:**
        - что такое преобразование Фурье?: https://habr.com/ru/articles/967798/

        ---

        ######  3. Периодограмма (Spectral Density|Periodogram)

        **Что это:**
        Оценка спектральной плотности мощности (PSD — Power Spectral Density) сигнала.

        **Как работает:**
        - Показывает распределение дисперсии (мощности) сигнала по частотам
        - **Метод Уэлча (Welch):** Усреднение периодограмм по перекрывающимся сегментам для сглаживания
        - Окно Ханна (Hann window) уменьшает спектральную утечку (spectral leakage)

        **Как читать график:**
        - **Ось X:** Частота
        - **Ось Y:** Спектральная плотность мощности (PSD)
        - **Пики мощности:** Частоты, вносящие наибольший вклад в изменчивость ряда
        - Логарифмическая шкала Y для лучшей визуализации

        **Применение:**
        - Более устойчивая оценка спектра по сравнению с FFT (меньше шума)
        - Выявление доминирующих циклов для параметра `m` в SARIMA
        - Анализ вклада разных частот в общую дисперсию

        **⚠️ Почитать:**
        - немного про периодограммы временных рядов: https://habr.com/ru/articles/505738/

        ---

        ######  4. Вейвлет-преобразование (Wavelet Transform)

        **Что это:**
        Анализ сигнала одновременно во **времени** и **частоте** с помощью вейвлетов (маленьких волн).

        **Как работает:**
        - **CWT (Continuous Wavelet Transform):** Свёртка сигнала с масштабируемым вейвлетом Морле
        - В отличие от FFT (стационарный спектр), показывает **КОГДА** происходят изменения частот
        - Формула: `W(a,b) = ∫ x(t) · ψ*((t-b)/a) dt`, где `a` — масштаб (частота), `b` — время

        **Как читать скалограмму:**
        - **Ось X:** Время (наблюдения)
        - **Ось Y:** Период/масштаб (обратная частота)
        - **Цвет:** Мощность (яркие цвета = высокая мощность)
        - **Горизонтальные полосы:** Устойчивые периодичности
        - **Смещение полос:** Изменение частоты во времени

        **Применение:**
        - Анализ **нестационарных** рядов (где частоты меняются со временем)
        - Обнаружение структурных изменений, сдвигов режимов
        - Адаптивное прогнозирование с меняющимися параметрами
        - Выделение тренда, сезонности и остатка на разных масштабах

        **⚠️ Почитать:**
        - вейвлет–анализ. Основы: https://habr.com/ru/articles/449646/

        ---

        ###### Сравнение методов:

        | Метод | Стационарность | Временная локализация | Основное применение |
        |-------|----------------|----------------------|---------------------|
        | **ACF/PACF** | Требует стационарности | Нет | Параметры ARIMA, сезонность |
        | **FFT** | Требует стационарности | Нет | Fourier features, частоты |
        | **Periodogram** | Требует стационарности | Нет | Устойчивая оценка спектра |
        | **Wavelet** | Не требует | Есть ✅ | Нестационарные ряды, изменение циклов |

        **Рекомендация:** Начните с ACF для подбора ARIMA, используйте FFT/Periodogram для поиска сезонности, примените Wavelet для сложных нестационарных рядов.
        """)

    # ── ВЫБОР СТОЛБЦА ДЛЯ АНАЛИЗА ───────────────────────────────
    num_cols = ct_f.get("num", [])

    if num_cols:
        # 🔧 ДОБАВЛЕНО: Селектор столбца
        target_col = st.selectbox(
            " Выберите числовой признак для спектрального анализа:",
            options=num_cols,
            index=0,
            key="spectral_analysis_target_col",
            help="Выберите целевую переменную временного ряда для анализа частотных характеристик"
        )
        
        # Подготовка данных для выбранного столбца
        if ts_mode_active:
            df_ts_temp = df_ts.copy()
            if target_col in df_ts_temp.columns:
                analysis_series = df_ts_temp[target_col].resample('D').mean().dropna().astype(float)
            else:
                analysis_series = pd.Series()
        else:
            if target_col in df_filtered.columns:
                analysis_series = df_filtered[target_col].dropna().astype(float)
            else:
                analysis_series = pd.Series()
        
        # Проверка достаточности данных
        if len(analysis_series) >= 30:
            st.success(f"✅ **Анализируется:** `{target_col}` | **Длина ряда:** {len(analysis_series)} наблюдений")
            
            # ── 1. ACF И PACF (Автокорреляция) ──────────────────────
            with st.expander("● Автокорреляционный анализ (ACF/PACF)", expanded=False):
                st.markdown(f"**Анализируемый признак:** `{target_col}`")
                st.markdown("""
                **Назначение:** Обнаружение сезонности через анализ корреляции ряда с его лагами.                            
                **Алгоритм:** Расчет корреляции между наблюдениями с разными временными сдвигами.                           
                **Влияние на модель:** Определяет параметры `p`, `q` для ARIMA и сезонность `m`.
                """)

                from statsmodels.graphics.tsaplots import plot_acf, plot_pacf

                c1, c2 = st.columns(2)
                max_lag = min(60, len(analysis_series) // 4)

                with c1:
                    st.markdown("**ACF (Autocorrelation Function)**")
                    fig_acf, ax_acf = plt.subplots(figsize=(8, 3))
                    plot_acf(analysis_series, lags=max_lag, ax=ax_acf, alpha=0.05)
                    ax_acf.set_title(f"Автокорреляционная функция (ACF) — {target_col}", fontsize=10)
                    st.pyplot(fig_acf, use_container_width=True)

                with c2:
                    st.markdown("**PACF (Partial Autocorrelation Function)**")
                    fig_pacf, ax_pacf = plt.subplots(figsize=(8, 3))
                    plot_pacf(analysis_series, lags=max_lag, ax=ax_pacf, alpha=0.05)
                    ax_pacf.set_title(f"Частная автокорреляция (PACF) — {target_col}", fontsize=10)
                    st.pyplot(fig_pacf, use_container_width=True)

                # Автоматическое определение сезонности из ACF
                acf_values = acf(analysis_series, nlags=max_lag)
                confidence = 1.96 / np.sqrt(len(analysis_series))
                significant_lags = np.where(np.abs(acf_values) > confidence)[0][1:]  # Без lag=0

                if len(significant_lags) > 0:
                    # Ищем периодические пики
                    seasonal_candidates = []
                    for i, lag in enumerate(significant_lags):
                        if i > 0 and lag - significant_lags[i-1] < 3:
                            continue
                        if lag > 2:
                            seasonal_candidates.append(int(lag))

                    if seasonal_candidates:
                        st.success(f"✅ **Обнаружена сезонность** с периодами: {seasonal_candidates[:3]}")
                        st.info(f"ℹ️ **Рекомендация:** Используйте SARIMA с seasonal_order=(..., m={seasonal_candidates[0]})")
                else:
                    st.info("ℹ️ Явная сезонность не обнаружена")

            # ── 2. FFT (Быстрое преобразование Фурье) ────────────────
            with st.expander("● Преобразование ФУРЬЕ (FFT)", expanded=False):
                st.markdown(f"**Анализируемый признак:** `{target_col}`")
                st.markdown("""
                **Назначение:** Разложение временного ряда на гармонические составляющие.                                    
                **Алгоритм:** FFT преобразует сигнал из временной области в частотную.                                       
                **Влияние на модель:** Выявляет доминирующие частоты для создания Fourier features.
                """)

                # Подготовка данных
                n = len(analysis_series)
                y = analysis_series.values - analysis_series.mean()  # Центрируем

                # FFT
                yf = fft(y)
                xf = fftfreq(n, 1)[:n//2]
                amplitude = 2.0/n * np.abs(yf[0:n//2])

                # Поиск пиков
                peaks, properties = find_peaks(amplitude, height=np.mean(amplitude) + np.std(amplitude))
                dominant_periods = [1/xf[p] for p in peaks if xf[p] > 0]

                # Визуализация
                c1, c2 = st.columns([2, 1])

                with c1:
                    fig_fft, ax_fft = plt.subplots(figsize=(10, 3))
                    ax_fft.plot(xf[:n//4], amplitude[:n//4], 'b-', linewidth=1)
                    ax_fft.plot(xf[peaks], amplitude[peaks], 'ro', markersize=5, label='Пики')
                    ax_fft.set_xlabel('Частота (циклы/единица времени)')
                    ax_fft.set_ylabel('Амплитуда')
                    ax_fft.set_title(f'Амплитудный спектр (FFT) — {target_col}')
                    ax_fft.legend()
                    ax_fft.grid(True, alpha=0.3)
                    st.pyplot(fig_fft, use_container_width=True)

                with c2:
                    st.markdown("**Доминирующие периоды:**")
                    if dominant_periods:
                        for i, period in enumerate(sorted(dominant_periods)[:5], 1):
                            st.metric(f"Период {i}", f"{period:.1f}")

                        st.info(f"ℹ️ **Рекомендация:** Добавьте Fourier features с периодами {sorted(dominant_periods)[:3]}")
                    else:
                        st.info("Явные периодичности не обнаружены")

                # Прогноз с использованием значимых гармоник
                if len(peaks) > 0 and st.checkbox(" Показать прогноз по значимым гармоникам", key="fft_forecast"):
                    # Берем топ-5 гармоник
                    top_peaks = peaks[np.argsort(amplitude[peaks])[-5:]]

                    # Реконструкция сигнала
                    reconstructed = np.zeros(n)
                    for p in top_peaks:
                        if xf[p] > 0:
                            reconstructed += (amplitude[p] * np.cos(2 * np.pi * xf[p] * np.arange(n) + np.angle(yf[p])))

                    reconstructed += analysis_series.mean()

                    # ИЗМЕНЕНО: figsize=(10, 2.5) вместо (10, 3) для одинаковой высоты с FFT
                    fig_rec, ax_rec = plt.subplots(figsize=(10, 2.5))
                    ax_rec.plot(analysis_series.index[:100], analysis_series.values[:100], 'b-', alpha=0.5, label='Исходный ряд')
                    ax_rec.plot(analysis_series.index[:100], reconstructed[:100], 'r-', linewidth=2, label='Прогноз (FFT)')
                    ax_rec.set_title(f'Реконструкция ряда по значимым гармоникам FFT — {target_col}', fontsize=10)
                    ax_rec.legend(fontsize=8)
                    plt.tight_layout()  # Оптимизация отступов
                    st.pyplot(fig_rec, use_container_width=True)

            # ── 3. ПЕРИОДОГРАММА ─────────────────────────────────────
            with st.expander("● Периодограмма (Spectral Density)", expanded=False):
                st.markdown(f"**Анализируемый признак:** `{target_col}`")
                st.markdown("""
                **Назначение:** Оценка спектральной плотности мощности.                                                    
                **Алгоритм:** Метод Уэлча для сглаживания спектра.                                                             
                **Влияние на модель:** Показывает мощность различных частотных компонент.
                """)

                from scipy.signal import periodogram, welch

                c1, c2 = st.columns(2)

                # Периодограмма
                with c1:
                    st.markdown("**Периодограмма**")
                    freq_per, pxx_per = periodogram(analysis_series.values, fs=1.0, window='hann')

                    fig_per, ax_per = plt.subplots(figsize=(8, 3))
                    ax_per.semilogy(freq_per, pxx_per)
                    ax_per.set_xlabel('Частота')
                    ax_per.set_ylabel('Спектральная плотность мощности')
                    ax_per.set_title(f'Периодограмма — {target_col}')
                    ax_per.grid(True, alpha=0.3)
                    st.pyplot(fig_per, use_container_width=True)

                # Метод Уэлча
                with c2:
                    st.markdown("**Метод Уэлча (сглаженный)**")
                    freq_welch, pxx_welch = welch(analysis_series.values, fs=1.0, nperseg=min(256, len(analysis_series)//4))

                    fig_welch, ax_welch = plt.subplots(figsize=(8, 3))
                    ax_welch.semilogy(freq_welch, pxx_welch)
                    ax_welch.set_xlabel('Частота')
                    ax_welch.set_ylabel('PSD')
                    ax_welch.set_title(f'Welch PSD — {target_col}')
                    ax_welch.grid(True, alpha=0.3)
                    st.pyplot(fig_welch, use_container_width=True)

                # Значимые частоты
                peaks_welch, _ = find_peaks(pxx_welch, height=np.median(pxx_welch)*2)

                if len(peaks_welch) > 0:
                    significant_freqs = freq_welch[peaks_welch]
                    significant_periods = 1/significant_freqs[significant_freqs > 0]

                    st.success(f"✅ **Значимые периоды:** {', '.join([f'{p:.1f}' for p in sorted(significant_periods)[:5]])}")

            # ── 4. WAVELET-ПРЕОБРАЗОВАНИЕ ────────────────────────────
            with st.expander("● Вейвлет-преобразование (Wavelet Transform)", expanded=False):
                st.markdown(f"**Анализируемый признак:** `{target_col}`")
                st.markdown("""
                **Назначение:** Анализ частот во времени для нестационарных рядов.                                               
                **Алгоритм:** Continuous Wavelet Transform (CWT) с вейвлетом Морле.                                            
                **Влияние на модель:** Показывает, КОГДА происходят циклические изменения.
                """)

                try:
                    import pywt
                    from scipy import signal

                    # CWT
                    widths = np.arange(1, min(128, len(analysis_series)//4))
                    cwtmatr, freqs_cwt = pywt.cwt(analysis_series.values - analysis_series.mean(),
                                                widths, 'morl', sampling_period=1)

                    # Усредненный спектр по времени
                    mean_power = np.mean(np.abs(cwtmatr), axis=1)

                    # ── ГОРИЗОНТАЛЬНОЕ РАСПОЛОЖЕНИЕ С ЕДИНОЙ ОСЬЮ Y ──────
                    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5),
                                                    gridspec_kw={'width_ratios': [3, 1]},
                                                    sharey=True)  # 🔑 ОБЩАЯ ОСЬ Y!

                    # Левый график: Вейвлет-скалограмма
                    im = ax1.imshow(np.abs(cwtmatr), extent=[0, len(analysis_series), 1, len(widths)],
                                cmap='jet', aspect='auto', interpolation='bilinear')
                    ax1.set_xlabel('Время', fontsize=10)
                    ax1.set_ylabel('Период (масштаб)', fontsize=10)
                    ax1.set_title(f'Вейвлет-скалограмма — {target_col}', fontsize=10, fontweight='normal')
                    plt.colorbar(im, label='Мощность', ax=ax1)

                    # Правый график: Усредненный спектр (ГОРИЗОНТАЛЬНЫЙ!)
                    ax2.plot(mean_power, widths, 'b-', linewidth=2)
                    ax2.set_xlabel('Средняя мощность', fontsize=10)

                    # 🔧 ДОБАВЛЕНО: Подпись оси Y и включение цифр
                    ax2.set_ylabel('Период (масштаб)', fontsize=10)
                    ax2.tick_params(axis='y', labelleft=True)

                    ax2.set_title('Усредненный спектр', fontsize=10, fontweight='normal')
                    ax2.grid(True, alpha=0.3)

                    plt.tight_layout()
                    st.pyplot(fig, use_container_width=True)

                    # Пики в вейвлет-спектре
                    wavelet_peaks, _ = find_peaks(mean_power, height=np.mean(mean_power))
                    if len(wavelet_peaks) > 0:
                        dominant_scales = widths[wavelet_peaks]
                        st.info(f"**Доминирующие масштабы:** {dominant_scales[:5]}")

                except ImportError:
                    st.warning("⚠️ Установите PyWavelets: `pip install PyWavelets`")
                except Exception as e:
                    st.error(f"Ошибка вейвлет-анализа: {e}")
        else:
            st.warning(f"⚠️ Недостаточно данных для спектрального анализа: {len(analysis_series)} точек (минимум 30)")
    else:
        st.info("ℹ️ Для спектрального анализа необходимы числовые колонки.")


    # ────────────────────────────────────────────────────────────
    #  ИТОГОВАЯ ТАБЛИЦА СВОЙСТВ ВРЕМЕННОГО РЯДА (Расширенный анализ)
    # ────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("###  Паспорт свойств входных данных")
    st.caption("Фиксируем исходные свойства ряда. Сравнивая свойства до и после предобработки, получаем объективную метрику качества предобработки.")

    # ────────────────────────────────────────────────────────────
    # 📘 СПРАВКА ПО СВОЙСТВАМ ВРЕМЕННЫХ РЯДОВ
    # ────────────────────────────────────────────────────────────

    with st.expander(" Справка по свойствам временных рядов", expanded=False):
        st.markdown("""
        ###### Научное описание свойств временных рядов

        Платформа CISStat рассчитывает **13 ключевых метрик** для комплексной характеристики временного ряда.
        Ниже приведено подробное описание каждого свойства с точки зрения науки о данных.

        ---

        ###### **1. Частота ряда (Frequency)**
        **Что это:** Регулярность временных интервалов между наблюдениями.

        **Научное обоснование:**
        - Определяется через `pd.infer_freq()` + эвристический анализ
        - Критична для выбора моделей: ARIMA/SARIMA требуют регулярный шаг
        - Нерегулярная частота → необходимость ресемплинга или использования Gaussian Processes

        **Интерпретация:**
        - `D` (daily), `M` (monthly), `Q` (quarterly), `Y` (yearly) — стандартные частоты
        - `Нерегулярная` → пропуски, неравномерные интервалы

        **Влияние на моделирование:**
        - ✅ Регулярная → все классические TS-модели
        - ❌ Нерегулярная → только модели с временными метками (Prophet, GP)

        ---

        ###### **2. Стационарность (Stationarity, ADF Test)**
        **Что это:** Свойство ряда иметь постоянные статистические характеристики во времени.

        **Научное обоснование:**
        - **Тест Дики-Фуллера (ADF)**: проверяет наличие единичного корня (unit root)
        - H₀ (нулевая гипотеза): ряд **нестационарен** (есть тренд/единичный корень)
        - H₁ (альтернатива): ряд **стационарен**
        - p-value < 0.05 → отвергаем H₀ → ряд стационарен

        **Типы стационарности:**
        - **Строгая стационарность**: распределение не меняется во времени
        - **Слабая стационарность**: постоянны mean, variance, autocovariance

        **Влияние на моделирование:**
        - ✅ Стационарен → ARIMA(p,**0**,q), ARMA
        - ❌ Нестационарен → требуется дифференцирование (d≥1) → ARIMA(p,**d**,q)

        **⚠️ Почитать**
        - статья на Хабре_стационарность и ADF: https://habr.com/ru/articles/1043810/
        ---

        ###### **3. Детерминированность (Determinism, R² тренда)**
        **Что это:** Доля дисперсии, объяснённая детерминированным (предсказуемым) трендом.

        **Научное обоснование:**
        - R² тренда = коэффициент детерминации линейной регрессии Y(t) = α + β·t + ε
        - R² ≥ 0.7 → **сильный детерминированный тренд** (70% вариации объясняется временем)
        - R² < 0.3 → **стохастический ряд** (случайные колебания преобладают)

        **Интерпретация:**
        - **Детерминированный** (R² ≥ 0.7): ряд следует предсказуемой траектории
        - **Стохастический/Смешанный** (R² < 0.7): преобладают случайные флуктуации

        **Влияние на моделирование:**
        - ✅ Детерминированный → Linear Trend + ARMA, Polynomial Regression
        - ⚠️ Смешанный → STL-декомпозиция, Prophet
        - ❌ Стохастический → Random Walk, ARIMA без тренда

        ---

        ###### **4. Автокорреляция (Autocorrelation, Ljung-Box Test)**
        **Что это:** Зависимость текущих значений от предыдущих (лагов).

        **Научное обоснование:**
        - **Тест Льюнга-Бокса**: проверяет гипотезу о независимости наблюдений
        - H₀: автокорреляция = 0 (белый шум, нет зависимости)
        - H₁: есть автокорреляция (зависимость от лагов)
        - p-value > 0.05 → **белый шум** (независимые наблюдения)

        **Интерпретация:**
        - ✅ **Белый шум** (p > 0.05): значения независимы → нельзя прогнозировать по истории
        - ⚠️ **Есть автокорреляция** (p < 0.05): значения зависят от предыдущих → можно прогнозировать

        **Влияние на моделирование:**
        - ✅ Есть АК → ARIMA, Exponential Smoothing (используют историю)
        - ❌ Белый шум → External Regressors, Prophet (нужны внешние факторы)

        ---

        ###### **5. Нормальность (Normality, Jarque-Bera Test)**
        **Что это:** Соответствие распределения остатков нормальному закону.

        **Научное обоснование:**
        - **Тест Жарка-Бера**: проверяет асимметрию (skewness) и эксцесс (kurtosis)
        - H₀: распределение **нормальное** (симметричное, kurtosis = 3)
        - H₁: распределение **отклоняется от нормы**
        - p-value > 0.05 → нормальность не отвергается

        **Интерпретация:**
        - ✅ **Нормально**: симметричное распределение, редкие экстремумы
        - ⚠️ **Отклонение**: асимметрия (skewness ≠ 0) или тяжёлые хвосты (kurtosis > 3)

        **Влияние на моделирование:**
        - ✅ Нормально → параметрические тесты, ARIMA (MLE), Gaussian likelihood
        - ❌ Отклонение → робастные методы, Box-Cox трансформация, квантильная регрессия

        ---

        ###### **6. Направление тренда (Trend Direction)**
        **Что это:** Долгосрочная направленность изменения ряда.

        **Научное обоснование:**
        - Определяется через **OLS (Ordinary Least Squares)** регрессию
        - Slope (β) > 0 → **восходящий тренд** (рост)
        - Slope (β) < 0 → **нисходящий тренд** (падение)
        - Slope ≈ 0 → **горизонтальный** (стабильность)

        **Интерпретация:**
        - **Восходящий** (Slope > 0): систематический рост во времени
        - **Нисходящий** (Slope < 0): систематическое снижение
        - **Горизонтальный** (|Slope| < ε): отсутствие направленного движения

        **Влияние на моделирование:**
        - ✅ Восходящий/Нисходящий → Trend + Seasonality models (Prophet, Holt-Winters)
        - ⚠️ Горизонтальный → Simple Exponential Smoothing, Naive

        ---

        ###### **7. Корреляция признаков (Feature Correlation)**
        **Что это:** Линейная связь целевой метрики с другими числовыми признаками.

        **Научное обоснование:**
        - **Коэффициент Пирсона (r)**: мера линейной зависимости ∈ [-1, +1]
        - r > 0.7 → **сильная положительная** связь
        - r < -0.7 → **сильная отрицательная** связь
        - |r| < 0.3 → **слабая** связь

        **Интерпретация:**
        - 🟢 **r > 0.5**: сильный предиктор (можно использовать в регрессии)
        - 🔴 **r < -0.5**: сильный обратный предиктор
        - 🟡 **|r| < 0.5**: слабая линейная связь (возможно, нелинейная зависимость)

        **Влияние на моделирование:**
        - ✅ Сильные корреляции → Multiple Regression, XGBoost с feature selection
        - ⚠️ Мультиколлинеарность (|r| > 0.85 между признаками) → PCA, Ridge/Lasso

        ---

        ###### **8. Сезонность (сила, STL Strength)**
        **Что это:** Регулярные повторяющиеся колебания фиксированной частоты.

        **Научное обоснование:**
        - **STL-декомпозиция** (Seasonal-Trend-Loess): Y(t) = Trend(t) + Seasonal(t) + Residual(t)
        - Strength = 1 - Var(Residual) / Var(Detrended)
        - Strength > 0.6 → **сильная сезонность** (≥60% вариации объясняется сезонностью)

        **Интерпретация:**
        - ✅ **Сильная** (S > 0.6): явные сезонные паттерны (например, продажи зимой выше)
        - ⚠️ **Слабая/Нет** (S < 0.3): сезонность отсутствует или незначительна

        **Влияние на моделирование:**
        - ✅ Сильная → SARIMA (с сезонным параметром m), Prophet с seasonality, STL-ARIMA
        - ❌ Слабая → ARIMA без сезонности, Exponential Smoothing

        ---

        ###### **9. Сезонные периоды (ACF Periods)**
        **Что это:** Конкретные лаги (периоды), на которых наблюдается сезонность.

        **Научное обоснование:**
        - **ACF (Autocorrelation Function)**: корреляция ряда с самим собой на разных лагах
        - Значимые пики на лагах 7, 14, 21 → **недельная сезонность** (для дневных данных)
        - Значимые пики на лагах 12, 24, 36 → **годовая сезонность** (для месячных данных)
        - Порог значимости: ±1.96/√n (95% доверительный интервал)

        **Интерпретация:**
        - `7` → недельный цикл (повторяется каждые 7 дней)
        - `12` → годовой цикл (для месячных данных)
        - `4` → квартальный цикл

        **Влияние на моделирование:**
        - Определяет параметр **m** в SARIMA(p,d,q)(P,D,Q,**m**)
        - Пример: m=12 для месячных данных с годовой сезонностью

        ---

        ###### **10. Долгая память (Hurst Exponent)**
        **Что это:** Мера персистентности (устойчивости тренда) ряда.

        **Научное обоснование:**
        - **Показатель Хёрста (H)** ∈ [0, 1] через R/S-анализ (Rescaled Range)
        - H = 0.5 → **случайное блуждание** (Random Walk, нет памяти)
        - H > 0.5 → **персистентность** (тренд усиливается, положительная автокорреляция)
        - H < 0.5 → **антиперсистентность** (mean reversion, отрицательная автокорреляция)

        **Интерпретация:**
        - 🔴 **H > 0.55 (Устойчивый тренд)**: если ряд рос, продолжит расти; если падал — продолжит падать
        - ⚪ **H ≈ 0.5 (Случайное блуждание)**: прошлое не предсказывает будущее
        - 🔵 **H < 0.45 (Антиперсистентность)**: ряд стремится к среднему (mean reverting)

        **Влияние на моделирование:**
        - ✅ H > 0.5 → Trend-following модели, ARIMA с трендом
        - ⚠️ H < 0.5 → Mean reversion модели, Ornstein-Uhlenbeck process
        - ❌ H ≈ 0.5 → Random Walk, Naive forecast

        ---

        ###### **11. Доминирующие частоты (FFT)**
        **Что это:** Периоды с максимальной амплитудой в частотной области.

        **Научное обоснование:**
        - **FFT (Fast Fourier Transform)**: преобразование из временной области в частотную
        - Выявляет скрытые периодичности, невидимые во временной области
        - Пики амплитуды → доминирующие частоты (периоды)

        **Интерпретация:**
        - Частота f = 0.0027 → Период T = 1/f ≈ 365 дней (годовой цикл)
        - Частота f = 0.143 → Период T = 1/f ≈ 7 дней (недельный цикл)

        **Влияние на моделирование:**
        - Создание **Fourier features**: sin(2πt/P), cos(2πt/P) для ML-моделей
        - Улучшает точность Prophet, XGBoost, Neural Networks

        ---

        ###### **12. Значимые периоды (Периодограмма)**
        **Что это:** Частоты с максимальной мощностью (вкладом в дисперсию).

        **Научное обоснование:**
        - **Periodogram**: оценка спектральной плотности мощности (PSD)
        - Метод Уэлча (Welch) с окном Ханна для сглаживания
        - Показывает, какие частоты вносят наибольший вклад в изменчивость ряда

        **Интерпретация:**
        - Высокая мощность на частоте f → этот цикл сильно влияет на ряд
        - Несколько пиков → множественные периодичности (например, недельная + годовая)

        **Влияние на моделирование:**
        - Определение параметра **m** для SARIMA
        - Выбор периодов для сезонной декомпозиции

        ---

        ###### **13. Доминирующие масштабы (Wavelet)**
        **Что это:** Масштабы (периоды) с максимальной мощностью, меняющиеся во времени.

        **Научное обоснование:**
        - **CWT (Continuous Wavelet Transform)**: анализ одновременно во времени и частоте
        - Вейвлет Морле (Morlet) для детекции осцилляций
        - В отличие от FFT (стационарный спектр), показывает **КОГДА** происходят изменения

        **Интерпретация:**
        - Устойчивые горизонтальные полосы на скалограмме → стабильные периодичности
        - Смещение полос → изменение частоты во времени (нестационарность)

        **Влияние на моделирование:**
        - ✅ Выявление нестационарных циклов → Time-Varying Parameter (TVP) модели
        - ✅ Адаптивное прогнозирование с меняющимися параметрами
        - ✅ Обнаружение структурных изменений (change points)

        ---

        ###### Сводная таблица влияния свойств на выбор моделей

        | Свойство | Значение | Рекомендуемые модели |
        |----------|----------|---------------------|
        | **Стационарность** | ✅ Стационарен | ARIMA(p,**0**,q), ARMA |
        | | ❌ Нестационарен | ARIMA(p,**d≥1**,q), Prophet |
        | **Сезонность** | ✅ Сильная (S>0.6) | SARIMA, Prophet, Holt-Winters |
        | | ❌ Слабая | ARIMA, Exponential Smoothing |
        | **Детерминированность** | ✅ R²≥0.7 | Linear Trend + ARMA |
        | | ⚠️ R²<0.7 | STL, Random Walk |
        | **Долгая память** | ✅ H>0.55 | Trend-following модели |
        | | 🔵 H<0.45 | Mean reversion модели |
        | **Автокорреляция** | ✅ Есть (p<0.05) | ARIMA, ETS |
        | | ❌ Белый шум | External regressors |

        ---

        **Источники:**
        - Hyndman, R.J., & Athanasopoulos, G. (2021). *Forecasting: principles and practice*
        - Box, G.E.P., Jenkins, G.M., Reinsel, G.C. (2015). *Time Series Analysis: Forecasting and Control*
        - Hamilton, J.D. (1994). *Time Series Analysis*
        - CISStat Internal Standard: TS-ANALYSIS-2026-v1.0
        """)

    # Селектор для выбора анализируемой колонки
    # st.divider()
    if ct_f["num"]:
        target_col = st.selectbox(
            "Исследуемый признак:",
            options=ct_f["num"],
            index=0,
            key="target_analysis_col"
        )

        # Кнопка запуска анализа
        if st.button(" Рассчитать свойства ряда", type="primary", key="btn_calc_props"):
            with st.spinner("Выполняем статистические тесты..."):
                try:
                    # 1. Подготовка данных
                    if ts_mode_active and target_col in df_ts.columns:
                        analysis_series = df_ts[target_col].resample('D').mean().dropna().astype(float)
                    else:
                        analysis_series = df_filtered[target_col].dropna().astype(float)

                    if len(analysis_series) < 30:
                        st.warning("⚠️ Недостаточно данных (нужно > 30 точек) для полного анализа.")
                    else:
                        results_data = []

                        # ── 0. ЧАСТОТА РЯДА (новая метрика) ─────────────────────
                        inferred_freq = pd.infer_freq(analysis_series.index.drop_duplicates().sort_values())
                        freq_result = f"✅ {inferred_freq}" if inferred_freq else "⚠️ Нерегулярная"

                        results_data.append({
                            "Свойство": "Частота ряда",
                            "Метод": "pd.infer_freq() + эвристика",
                            "Описание": "Определяет регулярность временного интервала между наблюдениями.",
                            "Результат": freq_result
                        })

                        # ── 1. Стационарность (ADF Test) ──────────────────────────
                        adf_res = adfuller(analysis_series, autolag='AIC')
                        adf_p = adf_res[1]
                        is_stationary = adf_p < 0.05

                        results_data.append({
                            "Свойство": "Стационарность",
                            "Метод": "ADF Test (Augmented Dickey-Fuller)",
                            "Описание": "Проверяет наличие единичного корня (нестационарности). H₀: Ряд нестационарен.",
                            "Результат": "✅ Стационарен" if is_stationary else "❌ Нестационарен"
                        })

                        # ── 1.1. Детерминированность (R² тренда) ──────────────────
                        from scipy.stats import linregress
                        slope, intercept, r_value, p_value, std_err = linregress(range(len(analysis_series)), analysis_series)
                        r_squared = r_value**2
                        is_deterministic = r_squared >= 0.7

                        results_data.append({
                            "Свойство": "Детерминированность",
                            "Метод": "R² тренда + комбинация тестов",
                            "Описание": "Доля дисперсии, объяснённая детерминированным трендом. R² ≥ 0.7 → сильный детерминированный компонент.",
                            "Результат": f"{'✅ Детерминированный' if is_deterministic else '⚠️ Стохастический/Смешанный'} (R²={r_squared:.3f})"
                        })

                        # ── 2. Автокорреляция (Ljung-Box) ─────────────────────────
                        from statsmodels.stats.diagnostic import acorr_ljungbox
                        lb_res = acorr_ljungbox(analysis_series, lags=[10])
                        if isinstance(lb_res, pd.DataFrame):
                            lb_p = lb_res['lb_pvalue'].iloc[0]
                        else:
                            lb_p = lb_res[1][0]

                        is_white_noise = lb_p > 0.05
                        results_data.append({
                            "Свойство": "Автокорреляция",
                            "Метод": "Ljung-Box Test (Lag=10)",
                            "Описание": "Проверяет гипотезу о том, что значения ряда независимы (белый шум). H₀: Автокорреляция равна 0.",
                            "Результат": "✅ Белый шум (Нет АК)" if is_white_noise else "⚠️ Есть автокорреляция"
                        })

                        # ── 3. Нормальность (Jarque-Bera) ─────────────────────────
                        from scipy import stats
                        jb_res = stats.jarque_bera(analysis_series)
                        jb_p = jb_res.pvalue if hasattr(jb_res, 'pvalue') else jb_res[1]
                        is_normal = jb_p > 0.05

                        results_data.append({
                            "Свойство": "Нормальность",
                            "Метод": "Jarque-Bera Test",
                            "Описание": "Проверяет соответствие распределения нормальному (асимметрия и эксцесс). H₀: Распределение нормально.",
                            "Результат": "✅ Нормально" if is_normal else "⚠️ Отклонение от нормы"
                        })

                        # ── 4. Направление тренда ─────────────────────────
                        trend_dir = "📈 Восходящий" if slope > 0 else "📉 Нисходящий" if slope < 0 else "➡️ Горизонтальный"

                        results_data.append({
                            "Свойство": "Направление тренда",
                            "Метод": "OLS Linear Regression (Slope)",
                            "Описание": "Определяет угол наклона линии тренда через метод наименьших квадратов.",
                            "Результат": f"{trend_dir} (Slope={slope:.4f})"
                        })

                        # ── 📈 КОРРЕЛЯЦИЯ ЧИСЛОВЫХ ПРИЗНАКОВ (новая метрика) ─────
                        num_cols = ct_f.get("num", [])
                        if len(num_cols) >= 2 and target_col in num_cols:
                            # Рассчитываем корреляции только для числовых колонок в отфильтрованных данных
                            corr_df = df_filtered[num_cols].corr()
                            target_corr = corr_df[target_col].drop(target_col).sort_values(key=abs, ascending=False)

                            # Формируем строку с топ-3 корреляциями
                            top_corrs = []
                            for col, val in target_corr.head(3).items():
                                sign = "🟢" if val > 0.5 else ("🔴" if val < -0.5 else "🟡")
                                top_corrs.append(f"{sign} {col} ({val:.2f})")

                            corr_result = ", ".join(top_corrs) if top_corrs else "⚪ Нет сильных связей (|r|<0.5)"

                            results_data.append({
                                "Свойство": "Корреляция признаков",
                                "Метод": "Pearson correlation matrix",
                                "Описание": "Линейная связь целевой метрики с другими числовыми признаками. 🟢>0.5, 🔴<-0.5, 🟡 слабая",
                                "Результат": corr_result
                            })

                        # ── 5. СЕЗОННОСТЬ (группировка: Strength + ACF периоды) ──
                        from statsmodels.tsa.seasonal import STL
                        period = 7 if (inferred_freq and 'D' in inferred_freq) else 12
                        try:
                            stl_res = STL(analysis_series, period=period, robust=True).fit()
                            var_total = analysis_series.var()
                            var_resid = stl_res.resid.var()
                            var_detrended = var_total - stl_res.trend.var()
                            strength_seasonality = max(0, 1 - var_resid / var_detrended) if var_detrended > 0 else 0
                            is_seasonal = strength_seasonality > 0.6
                        except:
                            strength_seasonality = 0.0
                            is_seasonal = False

                        results_data.append({
                            "Свойство": "Сезонность (сила)",
                            "Метод": "STL Decomposition (Strength)",
                            "Описание": "Доля дисперсии, объясняемая сезонной компонентой (0..1). >0.6 = сильная сезонность.",
                            "Результат": f"{'✅ Сильная' if is_seasonal else '⚠️ Слабая/Нет'} (S={strength_seasonality:.2f})"
                        })

                        # ── 6. Сезонные периоды (из ACF) — СРАЗУ ПОСЛЕ СИЛЫ ───────
                        from statsmodels.tsa.stattools import acf as acf_func
                        max_lag = min(60, len(analysis_series) // 4)
                        acf_values = acf_func(analysis_series, nlags=max_lag)
                        confidence = 1.96 / np.sqrt(len(analysis_series))
                        significant_lags = np.where(np.abs(acf_values) > confidence)[0][1:]

                        seasonal_periods_acf = []
                        for i, lag in enumerate(significant_lags):
                            if i > 0 and lag - significant_lags[i-1] < 3:
                                continue
                            if lag > 2:
                                seasonal_periods_acf.append(lag)

                        acf_seasonality = seasonal_periods_acf[:3] if seasonal_periods_acf else []
                        results_data.append({
                            "Свойство": "Сезонные периоды (ACF)",
                            "Метод": "Автокорреляционная функция + порог значимости",
                            "Описание": "Лаги с корреляцией выше 95% доверительного интервала. Показывает периодичность ряда.",
                            "Результат": f"{'✅ ' + ', '.join(map(str, acf_seasonality)) if acf_seasonality else '⚠️ Не обнаружены'}"
                        })

                        # ── 7. Долгая память (Hurst Exponent) ─────────────────────
                        def hurst_exponent(series, max_lag=20):
                            lags = range(2, max_lag)
                            tau = [max(np.std(np.subtract(series[lag:], series[:-lag])), 1e-8) for lag in lags]
                            try:
                                return np.polyfit(np.log(lags), np.log(tau), 1)[0]
                            except: return 0.5

                        hurst_val = hurst_exponent(analysis_series.values)
                        memory_type = "🔵 Антиперсистентность" if hurst_val < 0.45 else ("🔴 Устойчивый тренд" if hurst_val > 0.55 else "⚪ Случайное блуждание")

                        results_data.append({
                            "Свойство": "Долгая память",
                            "Метод": "Hurst Exponent (R/S Analysis)",
                            "Описание": "Характеризует персистентность ряда. H=0.5 (Random Walk), H>0.5 (Trend), H<0.5 (Mean Reverting).",
                            "Результат": f"{memory_type} (H={hurst_val:.2f})"
                        })

                        # ═══════════════════════════════════════════════════════
                        #  🆕 СПЕКТРАЛЬНЫЕ СВОЙСТВА (из спектрального анализа)
                        # ═══════════════════════════════════════════════════════

                        # ── 8. Доминирующие частоты (FFT) ───────────────────────

                        n = len(analysis_series)
                        y = analysis_series.values - analysis_series.mean()
                        yf = fft(y)
                        xf = fftfreq(n, 1)[:n//2]
                        amplitude = 2.0/n * np.abs(yf[0:n//2])

                        peaks, _ = find_peaks(amplitude, height=np.mean(amplitude) + np.std(amplitude))
                        fft_periods = [1/xf[p] for p in peaks if xf[p] > 0 and xf[p] < 0.5]
                        fft_dominant = sorted(fft_periods)[:3] if fft_periods else []

                        results_data.append({
                            "Свойство": "Доминирующие частоты (FFT)",
                            "Метод": "Быстрое преобразование Фурье + поиск пиков",
                            "Описание": "Периоды с максимальной амплитудой в частотной области. Для создания Fourier features.",
                            "Результат": f"{'✅ ' + ', '.join([f'{p:.1f}' for p in fft_dominant]) if fft_dominant else '⚠️ Не обнаружены'}"
                        })

                        # ── 9. Спектральная плотность (Periodogram) ─────────────
                        from scipy.signal import periodogram
                        freq_per, pxx_per = periodogram(analysis_series.values, fs=1.0, window='hann')
                        peaks_per, _ = find_peaks(pxx_per, height=np.median(pxx_per)*2)
                        periodogram_periods = sorted([1/freq_per[p] for p in peaks_per if freq_per[p] > 0])[:3]

                        results_data.append({
                            "Свойство": "Значимые периоды (Периодограмма)",
                            "Метод": "Periodogram с окном Hann + порог мощности",
                            "Описание": "Частоты с мощностью выше медианы × 2. Показывает вклад разных циклов в дисперсию.",
                            "Результат": f"{'✅ ' + ', '.join([f'{p:.1f}' for p in periodogram_periods]) if periodogram_periods else '⚠️ Не обнаружены'}"
                        })

                        # ── 10. Вейвлет-масштабы (опционально) ─────────────────
                        wavelet_scales = []
                        try:
                            import pywt
                            widths = np.arange(1, min(128, len(analysis_series)//4))
                            cwtmatr, _ = pywt.cwt(analysis_series.values - analysis_series.mean(), widths, 'morl', sampling_period=1)
                            mean_power = np.mean(np.abs(cwtmatr), axis=1)
                            wavelet_peaks, _ = find_peaks(mean_power, height=np.mean(mean_power))
                            wavelet_scales = widths[wavelet_peaks][:3].tolist() if len(wavelet_peaks) > 0 else []

                            results_data.append({
                                "Свойство": "Доминирующие масштабы (Wavelet)",
                                "Метод": "Continuous Wavelet Transform (Morlet)",
                                "Описание": "Масштабы с максимальной средней мощностью. Показывает изменение циклов во времени.",
                                "Результат": f"{'✅ ' + ', '.join(map(str, wavelet_scales)) if wavelet_scales else '⚠️ Не обнаружены'}"
                            })
                        except ImportError:
                            results_data.append({
                                "Свойство": "Доминирующие масштабы (Wavelet)",
                                "Метод": "Continuous Wavelet Transform (Morlet)",
                                "Описание": "Требует библиотеку PyWavelets (`pip install PyWavelets`)",
                                "Результат": "⚠️ Библиотека не установлена"
                            })
                        except:
                            results_data.append({
                                "Свойство": "Доминирующие масштабы (Wavelet)",
                                "Метод": "Continuous Wavelet Transform (Morlet)",
                                "Описание": "Анализ частот во времени для нестационарных рядов",
                                "Результат": "⚠️ Ошибка вычисления"
                            })

                        # ── ВЫВОД ТАБЛИЦЫ ─────────────────────────────────────────
                        df_results = pd.DataFrame(results_data)

                        # 🔧 СОХРАНЕНИЕ ПАСПОРТА v1.0 В SESSION_STATE
                        if target_col and len(analysis_series) >= 30:
                            props_v10 = calculate_ts_passport(
                                analysis_series,
                                df_filtered=df_filtered,
                                ct_f=ct_f,
                                target_col=target_col
                            )
                            props_v10['version'] = 'v1.0 (сырые данные)'
                            st.session_state.ts_props_v10 = props_v10
                            st.session_state.ts_props_v10_target_col = target_col


                        # ── ВЫВОД ТАБЛИЦЫ ─────────────────────────────────────────
                        df_results = pd.DataFrame(results_data)

                        # Безопасный расчёт высоты (только если датафрейм не пустой)
                        if not df_results.empty:
                            n_rows = len(df_results)
                            # ~40px на строку + 45px на шапку таблицы, ограничиваем диапазон 200–600px
                            table_height = min(600, max(200, 45 + n_rows * 40))
                        else:
                            table_height = 200  # Дефолтная высота для пустой/ошибочной таблицы

                        st.dataframe(
                            df_results,
                            use_container_width=True,
                            hide_index=True,
                            height=table_height,
                            column_config={
                                "Свойство": st.column_config.TextColumn("СВОЙСТВО", width="small"),
                                "Метод": st.column_config.TextColumn("МЕТОД", width="medium"),
                                "Описание": st.column_config.TextColumn("ОПИСАНИЕ", width="large"),
                                "Результат": st.column_config.TextColumn("✅ РЕЗУЛЬТАТ", width="medium")
                            }
                        )

                        # ────────────────────────────────────────────────────────────
                        #  📊 ПРЕДВАРИТЕЛЬНЫЕ РЕКОМЕНДАЦИИ ПО МОДЕЛИРОВАНИЮ (Объединённый блок)
                        # ────────────────────────────────────────────────────────────
                        st.divider()
                        st.markdown("###  Предварительные рекомендации по моделированию")

                        recommendations = []
                        model_suggestions = []

                        # ═══════════════════════════════════════════════════════════
                        #  1. БАЗОВЫЕ СВОЙСТВА (стационарность, белый шум, тренд)
                        # ═══════════════════════════════════════════════════════════

                        # Стационарность + белый шум
                        if 'is_stationary' in locals() and 'is_white_noise' in locals():
                            if is_stationary and is_white_noise:
                                recommendations.append("• Ряд похож на белый шум → рассмотрите внешние факторы или агрегацию")
                                model_suggestions.append("Exponential Smoothing, Naive, External regressors")
                            elif is_stationary and not is_white_noise:
                                recommendations.append("• Ряд стационарен с автокорреляцией → подходит ARIMA/SARIMA")
                                model_suggestions.append("ARIMA(p,d,q), SARIMA с подбором порядков")
                            elif not is_stationary:
                                recommendations.append("• Ряд нестационарен → примените дифференцирование (diff) или детрендирование")
                                model_suggestions.append("ARIMA (с d≥1), Detrending + ARMA, Prophet")

                        # Детерминированность тренда
                        if 'r_squared' in locals():
                            if r_squared >= 0.7:
                                recommendations.append(f"• Сильный детерминированный тренд (R²={r_squared:.2f}) → учтите тренд в модели")
                                model_suggestions.append("Linear/Polynomial Trend + ARMA, Prophet with trend")

                        # Гетероскедастичность
                        if 'is_heteroscedastic' in locals() and is_heteroscedastic:
                            recommendations.append("• Обнаружена гетероскедастичность → рассмотрите модели с изменяющейся дисперсией")
                            model_suggestions.append("GARCH, ARIMA-GARCH, Log-transform")

                        # ═══════════════════════════════════════════════════════════
                        #  2. СПЕКТРАЛЬНЫЕ СВОЙСТВА (сезонность, частоты, циклы)
                        # ═══════════════════════════════════════════════════════════

                        # Сезонность из ACF
                        if 'acf_seasonality' in locals() and acf_seasonality:
                            m_val = acf_seasonality[0]
                            recommendations.append(f"• **Сезонность из ACF:** используйте SARIMA с **m={m_val}**")
                            model_suggestions.append(f"SARIMA(..., seasonal_order=(..., m={m_val}))")

                        # Доминирующие частоты из FFT
                        if 'fft_dominant' in locals() and fft_dominant:
                            periods_str = ', '.join([f'{p:.1f}' for p in fft_dominant[:3]])
                            recommendations.append(f"• **Fourier features:** добавьте гармоники с периодами [{periods_str}]")
                            model_suggestions.append(f"ML-модели с признаками: sin(2πt/P), cos(2πt/P) для P∈[{periods_str}]")

                        # Дополнительные периоды из периодограммы
                        if 'periodogram_periods' in locals() and periodogram_periods:
                            unique_periods = [p for p in periodogram_periods if p not in (fft_dominant if 'fft_dominant' in locals() else [])]
                            if unique_periods:
                                recommendations.append(f"• **Доп. периоды (Periodogram):** {unique_periods[:2]}")

                        # Изменение циклов во времени (Wavelet)
                        if 'wavelet_scales' in locals() and wavelet_scales:
                            recommendations.append("• **Нестационарность частот:** вейвлет показал изменение циклов во времени")
                            model_suggestions.append("Time-Varying Parameter (TVP) models, State Space, Adaptive filtering")

                        # Сила сезонности из STL
                        if 'is_seasonal' in locals() and is_seasonal:
                            recommendations.append(f"• **Сильная сезонность (STL):** S={strength_seasonality:.2f} → явно моделируйте сезонную компоненту")
                            model_suggestions.append("STL decomposition + ARIMA, Prophet with seasonality")

                        # ── ЧАСТОТА РЯДА ────────────────────────────────────────
                        if 'inferred_freq' in locals():
                            if inferred_freq:
                                freq_code = inferred_freq.split('-')[0] if '-' in inferred_freq else inferred_freq
                                recommendations.append(f"• **Частота ряда:** {freq_code} → подходит для классических TS-моделей")

                                # Специфичные рекомендации по частоте
                                if freq_code in ['D', 'B', 'H']:  # Дневные/часовые
                                    model_suggestions.append("Prophet (учёт праздников), LSTM для высокочастотных данных")
                                elif freq_code in ['W', 'M', 'Q']:  # Недельные/месячные/квартальные
                                    model_suggestions.append("SARIMA, ETS, TBATS для сезонных рядов")
                                elif freq_code == 'Y':  # Годовые
                                    model_suggestions.append("Простые трендовые модели, сравнение годовых значений")
                            else:
                                recommendations.append("• **Нерегулярная частота** → требуется ресемплинг или модели для неравномерных рядов")
                                model_suggestions.append("Interpolation + ARIMA, Gaussian Processes, State Space Models")

                        # ── КОРРЕЛЯЦИЯ ЧИСЛОВЫХ ПРИЗНАКОВ ─────────────────────
                        if 'target_corr' in locals() and target_corr is not None:
                            # Анализ сильных корреляций целевой метрики
                            strong_pos = target_corr[target_corr > 0.7]
                            strong_neg = target_corr[target_corr < -0.7]
                            multicollinear = []

                            # Проверка мультиколлинеарности между признаками
                            if len(ct_f["num"]) >= 2:
                                corr_matrix = df_filtered[ct_f["num"]].corr()
                                for i in range(len(corr_matrix.columns)):
                                    for j in range(i+1, len(corr_matrix.columns)):
                                        if abs(corr_matrix.iloc[i, j]) > 0.85:
                                            multicollinear.append(f"{corr_matrix.columns[i]} ↔ {corr_matrix.columns[j]}")

                            if len(strong_pos) > 0 or len(strong_neg) > 0:
                                top_feat = list(strong_pos.index) + list(strong_neg.index)
                                recommendations.append(f"• **Сильные предикторы:** {', '.join(top_feat[:3])} (|r|>0.7) → используйте как основные фичи")
                                model_suggestions.append("Linear Regression, Random Forest, XGBoost с отбором признаков")

                            if multicollinear:
                                recommendations.append(f"• **Мультиколлинеарность:** {multicollinear[0]} (|r|>0.85) → риск нестабильности оценок")
                                model_suggestions.append("PCA, Ridge/Lasso регуляризация, удаление одного из коррелированных признаков")

                            # Если все корреляции слабые
                            if len(strong_pos) == 0 and len(strong_neg) == 0 and len(multicollinear) == 0:
                                recommendations.append("• **Слабые линейные связи** → рассмотрите нелинейные модели или инженерные признаки")
                                model_suggestions.append("Polynomial features, Interaction terms, Neural Networks, Gradient Boosting")

                        # ═══════════════════════════════════════════════════════════
                        #  4. ОБЪЕДИНЁННЫЙ ВЫВОД РЕКОМЕНДАЦИЙ
                        # ═══════════════════════════════════════════════════════════

                        if recommendations:
                            st.markdown("**Список рекомендаций:**")
                            for i, rec in enumerate(recommendations, 1):
                                # Цветовая индикация приоритета
                                if "Сильная" in rec or "мультиколлинеарность" in rec.lower():
                                    st.warning(rec)  # Важные предупреждения
                                elif "Частота" in rec or "предикторы" in rec:
                                    st.info(rec)  # Информационные
                                else:
                                    st.success(rec)  # Общие рекомендации

                            if model_suggestions:
                                st.markdown("**Предлагаемые модели (по приоритету):**")
                                unique_models = list(dict.fromkeys(model_suggestions))  # Убираем дубли
                                for i, model in enumerate(unique_models, 1):
                                    st.markdown(f"{i}. {model}")
                        else:
                            st.success("✅ Ряд не имеет выраженных паттернов → начните с простых моделей")
                            st.info("• Exponential Smoothing (Holt-Winters)\n• Naive / Seasonal Naive\n• Linear Regression с лагами")

                        # ═══════════════════════════════════════════════════════════
                        #  5. МЕТОДОЛОГИЧЕСКОЕ ПОЯСНЕНИЕ (обновлённое)
                        # ═══════════════════════════════════════════════════════════
                        st.markdown("""
                        <div style='color: #000000; font-size: 14px; background: #f8fafc; padding: 12px; border-radius: 6px; border-left: 3px solid #3b82f6;'>

                        **🔄 Методология рекомендаций:**
                        - **Статистические тесты** (ADF, Ljung-Box, Jarque-Bera) → выбор класса моделей
                        - **Спектральный анализ** (ACF, FFT, Periodogram, Wavelet) → параметры сезонности и признаки
                        - **Декомпозиция** (STL) → сила сезонной компоненты
                        - **Частота ряда** → выбор частотно-зависимых моделей (Prophet для дневных, SARIMA для месячных)
                        - **Корреляционный анализ** → отбор признаков, борьба с мультиколлинеарностью, инженерия фич

                        **🔄 Порядок действий:**
                        1. Примените преобразования (diff, detrend, log) на основе стационарности
                        2. Добавьте спектральные признаки (Fourier terms, seasonal dummies) при наличии сезонности
                        3. При сильных корреляциях: используйте PCA или регуляризацию
                        4. Подберите параметры модели (p,d,q,m) через AIC/BIC или кросс-валидацию
                        5. Валидируйте на тестовой выборке (MAPE, RMSE, MASE)

                        </div>
                        """, unsafe_allow_html=True)


                except Exception as e:
                    st.error(f"Ошибка при анализе свойств: {e}")
                    st.exception(e)
    else:
        st.warning("⚠️ Нет числовых колонок для анализа.")


    # ─────────────────────────────────────────────────────────────
    # 📋 ГЕНЕРАТОР ПОЛНОГО ОТЧЕТА О СВОЙСТВАХ ВРЕМЕННОГО РЯДА
    # ─────────────────────────────────────────────────────────────
    st.divider()
    st.markdown('###  Генератор отчета о входящих свойствах временного ряда')
    st.caption("Выберите признак для формирования верифицированного отчета.")

    if ct_f["num"]:
        report_col = st.selectbox(
            "Исследуемый признак:",
            options=ct_f["num"],
            index=0,
            key="report_col_select"
        )

        # ── КНОПКА ГЕНЕРАЦИИ ПОЛНОГО ОТЧЕТА ───────────────────
        if st.button(" Сформировать полный отчет", type="primary", key="btn_generate_report"):
            with st.spinner("🔍 Выполняется комплексный анализ..."):
                try:
                    import io
                    from datetime import datetime as dt_now
                    from scipy.stats import jarque_bera, linregress
                    from statsmodels.tsa.stattools import adfuller
                    try:
                        from statsmodels.tsa.stattools import acorr_ljungbox
                    except ImportError:
                        from statsmodels.stats.diagnostic import acorr_ljungbox
                    from scipy import stats as sp_stats

                    # Подготовка данных
                    if ts_mode_active and report_col in df_ts.columns:
                        analysis_series = df_ts[report_col].resample('D').mean().dropna().astype(float)
                    else:
                        analysis_series = df_filtered[report_col].dropna().astype(float)

                    if len(analysis_series) < 30:
                        st.warning("⚠️ Недостаточно данных для формирования отчета (требуется > 30 точек).")
                    else:
                        # ── 1. РАСЧЕТ МЕТОДОВ ──────────────────────────────────────
                        # Статистики распределения
                        mean_val = analysis_series.mean()
                        median_val = analysis_series.median()
                        std_val = analysis_series.std()
                        skew_val = analysis_series.skew()
                        kurt_val = analysis_series.kurtosis()

                        # Авто-детект типа распределения
                        candidates = {"Нормальное": sp_stats.norm, "Логнормальное": sp_stats.lognorm, "Экспоненциальное": sp_stats.expon}
                        best_name, best_ks = "Эмпирическое", 1.0
                        for name, dist in candidates.items():
                            try:
                                if name in ["Логнормальное", "Экспоненциальное"] and analysis_series.min() <= 0: continue
                                params = dist.fit(analysis_series)
                                ks_stat, _ = sp_stats.kstest(analysis_series, dist.name, args=params)
                                if ks_stat < best_ks: best_ks, best_name = ks_stat, name
                            except: continue
                        dist_type = f"Непрерывное - {best_name}" if best_ks < 0.15 else "Непрерывное - Эмпирическое (сложная форма)"

                        # Свойства временного ряда
                        adf_res = adfuller(analysis_series, autolag='AIC')
                        is_stationary = adf_res[1] < 0.05

                        lb_res = acorr_ljungbox(analysis_series, lags=[10])
                        lb_p = lb_res['lb_pvalue'].iloc[0] if isinstance(lb_res, pd.DataFrame) else lb_res[1][0]
                        is_white_noise = lb_p > 0.05

                        jb_res = jarque_bera(analysis_series)
                        jb_p = jb_res.pvalue if hasattr(jb_res, 'pvalue') else jb_res[1]
                        is_normal = jb_p > 0.05

                        # Тренд и Детерминированность
                        slope, intercept, r_value, p_value, std_err = linregress(range(len(analysis_series)), analysis_series)
                        r_squared = r_value**2
                        trend_dir = "📈 Восходящий" if slope > 0 else "📉 Нисходящий" if slope < 0 else "➡️ Горизонтальный"

                        inferred_freq = pd.infer_freq(analysis_series.index)
                        period = 7 if (inferred_freq and 'D' in inferred_freq) else 12
                        try:
                            stl_res = STL(analysis_series, period=period, robust=True).fit()
                            var_detrended = analysis_series.var() - stl_res.trend.var()
                            strength_seasonality = max(0, 1 - stl_res.resid.var() / var_detrended) if var_detrended > 0 else 0
                            is_seasonal = strength_seasonality > 0.6
                        except:
                            strength_seasonality, is_seasonal = 0.0, False

                        def calc_hurst(series, max_lag=20):
                            lags = range(2, max_lag)
                            tau = [max(np.std(np.subtract(series[lag:], series[:-lag])), 1e-8) for lag in lags]
                            try: return np.polyfit(np.log(lags), np.log(tau), 1)[0]
                            except: return 0.5
                        hurst_val = calc_hurst(analysis_series.values)
                        memory_type = "🔵 Антиперсистентность" if hurst_val < 0.45 else ("🔴 Устойчивый тренд" if hurst_val > 0.55 else "⚪ Случайное блуждание")

                        # ── 2. АНАЛИЗ ГЕТЕРОСКЕДАСТИЧНОСТИ (ARCH-LM) ────────────────────
                        is_heteroscedastic = False
                        arch_p_val = None
                        try:
                            from statsmodels.stats.diagnostic import het_arch
                            lags = min(5, max(1, len(analysis_series) // 10))
                            arch_res = het_arch(analysis_series, lags=lags)
                            arch_p_val = arch_res[1]
                            is_heteroscedastic = arch_p_val < 0.05
                        except Exception:
                            pass

                        # ── 📈 КОРРЕЛЯЦИЯ ЧИСЛОВЫХ ПРИЗНАКОВ ─────────────────────
                        num_cols = ct_f.get("num", [])
                        target_corr_str = "N/A"
                        if len(num_cols) >= 2 and report_col in num_cols:
                            corr_df = df_filtered[num_cols].corr()
                            target_corr = corr_df[report_col].drop(report_col).sort_values(key=abs, ascending=False)
                            top_corrs = []
                            for col, val in target_corr.head(3).items():
                                sign = "🟢" if val > 0.5 else ("🔴" if val < -0.5 else "🟡")
                                top_corrs.append(f"{sign} {col} ({val:.2f})")
                            target_corr_str = ", ".join(top_corrs) if top_corrs else "⚪ Нет сильных связей (|r|<0.5)"

                        # ── 📅 СЕЗОННЫЕ ПЕРИОДЫ (из ACF) ────────────────────────
                        from statsmodels.tsa.stattools import acf as acf_func
                        max_lag = min(60, len(analysis_series) // 4)
                        acf_values = acf_func(analysis_series, nlags=max_lag)
                        confidence = 1.96 / np.sqrt(len(analysis_series))
                        significant_lags = np.where(np.abs(acf_values) > confidence)[0][1:]

                        seasonal_periods_acf = []
                        for i, lag in enumerate(significant_lags):
                            if i > 0 and lag - significant_lags[i-1] < 3:
                                continue
                            if lag > 2:
                                seasonal_periods_acf.append(lag)
                        acf_seasonality_str = ', '.join(map(str, seasonal_periods_acf[:3])) if seasonal_periods_acf else "Не обнаружены"

                        # ── ⚡ ДОМИНИРУЮЩИЕ ЧАСТОТЫ (FFT) ───────────────────────
                        n = len(analysis_series)
                        y = analysis_series.values - analysis_series.mean()
                        yf = fft(y)
                        xf = fftfreq(n, 1)[:n//2]
                        amplitude = 2.0/n * np.abs(yf[0:n//2])
                        peaks, _ = find_peaks(amplitude, height=np.mean(amplitude) + np.std(amplitude))
                        fft_periods = [1/xf[p] for p in peaks if xf[p] > 0 and xf[p] < 0.5]
                        fft_dominant_str = ', '.join([f'{p:.1f}' for p in sorted(fft_periods)[:3]]) if fft_periods else "Не обнаружены"

                        # ── 📊 ЗНАЧИМЫЕ ПЕРИОДЫ (Периодограмма) ─────────────
                        from scipy.signal import periodogram
                        freq_per, pxx_per = periodogram(analysis_series.values, fs=1.0, window='hann')
                        peaks_per, _ = find_peaks(pxx_per, height=np.median(pxx_per)*2)
                        periodogram_periods = sorted([1/freq_per[p] for p in peaks_per if freq_per[p] > 0])[:3]
                        periodogram_str = ', '.join([f'{p:.1f}' for p in periodogram_periods]) if periodogram_periods else "Не обнаружены"

                        # ── 🌊 ВЕЙВЛЕТ-МАСШТАБЫ ───────────────────────────────
                        wavelet_scales_str = "N/A"
                        try:
                            import pywt
                            widths = np.arange(1, min(128, len(analysis_series)//4))
                            cwtmatr, _ = pywt.cwt(analysis_series.values - analysis_series.mean(), widths, 'morl', sampling_period=1)
                            mean_power = np.mean(np.abs(cwtmatr), axis=1)
                            wavelet_peaks, _ = find_peaks(mean_power, height=np.mean(mean_power))
                            wavelet_scales = widths[wavelet_peaks][:3].tolist() if len(wavelet_peaks) > 0 else []
                            wavelet_scales_str = ', '.join(map(str, wavelet_scales)) if wavelet_scales else "Не обнаружены"
                        except:
                            wavelet_scales_str = "Ошибка/Не установлено"

                        # ── 3. СТРУКТУРИРОВАНИЕ РЕЗУЛЬТАТОВ ────────────────────────
                        tech_info = {
                            "Всего строк": f"{len(df):,}".replace(",", " "),
                            "Всего колонок": len(df.columns),
                            "Объем памяти": f"{df.memory_usage(deep=True).sum() / 1024**2:.2f} MB",
                            "Числовых / Категорий": f"{len(ct['num'])} / {len(ct['cat'])}",
                            "Диапазон дат": f"{df_ts.index.min().date()} — {df_ts.index.max().date()}" if ts_mode_active else "N/A",
                            "Inferred частота": pd.infer_freq(df_ts.index) if ts_mode_active else "N/A"
                        }

                        dist_stats = {
                            "Mean (среднее)": f"{mean_val:,.2f}",
                            "Median (медиана)": f"{median_val:,.2f}",
                            "Std (стандартное отклонение)": f"{std_val:,.2f}",
                            "Skewness (асимметрия)": f"{skew_val:.3f}",
                            "Kurtosis (эксцесс)": f"{kurt_val:.3f}",
                            "Тип распределения": dist_type
                        }

                        freq_label = inferred_freq if inferred_freq else "Нерегулярная"
                        freq_result = f"✅ {freq_label}" if inferred_freq else "⚠️ Нерегулярная (требуется ресемплинг)"

                        # 🔧 РАСШИРЕННЫЙ СПИСОК СВОЙСТВ (все метрики из Паспорта)
                        ts_passport = [
                            {"property": "Стационарность", "method": "ADF Test (autolag='AIC')", "result": "✅ Стационарен" if is_stationary else "❌ Нестационарен"},
                            {"property": "Детерминированность", "method": "R² тренда + комбинация тестов", "result": f"{'✅ Детерминированный' if r_squared >= 0.7 else '⚠️ Стохастический/Смешанный'} (R²={r_squared:.3f})"},
                            {"property": "Частота ряда", "method": "pd.infer_freq() + автодетект", "result": freq_result},
                            {"property": "Гетероскедастичность", "method": "ARCH-LM Test (Engle, 1982)", "result": "⚠️ Гетероскедастичность (p < 0.05)" if is_heteroscedastic else "✅ Нет (Гомоскедастичность)"},
                            {"property": "Автокорреляция", "method": "Ljung-Box Test (Lag=10)", "result": "✅ Белый шум" if is_white_noise else "⚠️ Есть АК"},
                            {"property": "Нормальность", "method": "Jarque-Bera Test", "result": "✅ Нормально" if is_normal else "⚠️ Отклонение"},
                            {"property": "Направление тренда", "method": "OLS Linear Regression", "result": f"{trend_dir} (Slope={slope:.4f})"},
                            {"property": "Сезонность (сила)", "method": "STL Decomposition (Strength)", "result": f"{'✅ Сильная' if is_seasonal else '⚠️ Слабая/Нет'} (S={strength_seasonality:.2f})"},
                            {"property": "Сезонные периоды (ACF)", "method": "Автокорреляция + порог значимости", "result": f"✅ {acf_seasonality_str}" if acf_seasonality_str != "Не обнаружены" else "⚠️ Не обнаружены"},
                            {"property": "Долгая память", "method": "Hurst Exponent (R/S)", "result": f"{memory_type} (H={hurst_val:.2f})"},
                            {"property": "Корреляция признаков", "method": "Pearson correlation matrix", "result": target_corr_str},
                            {"property": "Доминирующие частоты (FFT)", "method": "Быстрое преобразование Фурье", "result": f"✅ {fft_dominant_str}" if fft_dominant_str != "Не обнаружены" else "⚠️ Не обнаружены"},
                            {"property": "Значимые периоды (Периодограмма)", "method": "Periodogram с окном Hann", "result": f"✅ {periodogram_str}" if periodogram_str != "Не обнаружены" else "⚠️ Не обнаружены"},
                            {"property": "Доминирующие масштабы (Wavelet)", "method": "Continuous Wavelet Transform", "result": f"✅ {wavelet_scales_str}" if wavelet_scales_str not in ["Не обнаружены", "Ошибка/Не установлено"] else f"⚠️ {wavelet_scales_str}"}
                        ]

                        df_report = pd.DataFrame(ts_passport)

                        # ── 4. ОТОБРАЖЕНИЕ ТАБЛИЦЫ НА ЭКРАНЕ ──
                        st.divider()
                        st.markdown(f"#### 📄 Готовый отчет: {report_col}")
                        st.dataframe(
                            df_report,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "property": st.column_config.TextColumn("Свойство", width="small"),
                                "method": st.column_config.TextColumn("Метод", width="medium"),
                                "desc": st.column_config.TextColumn("Описание", width="large"),
                                "result": st.column_config.TextColumn("Результат", width="medium")
                            }
                        )

                        # ── 5. ГЕНЕРАЦИЯ EXCEL (с листом рекомендаций) ─────────────────

                        wb = Workbook()
                        ws = wb.active
                        ws.title = "1_Паспорт свойств"

                        header_font = Font(bold=True, size=12, color="FFFFFF")
                        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        title_font = Font(bold=True, size=16)
                        footer_font = Font(bold=True, color="0369A1")
                        footer_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
                        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                        yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                        row = 1
                        ws.merge_cells(f"A{row}:D{row}")
                        cell = ws.cell(row=row, column=1, value=f"Предварительный отчет о свойствах признака: {report_col}")
                        cell.font = title_font
                        cell.alignment = Alignment(horizontal='center')
                        row += 2

                        ws.cell(row=row, column=1, value=f"Исследуемый параметр: {report_col}").font = header_font
                        ws.cell(row=row, column=3, value=f"Дата: {dt_now.now().strftime('%d.%m.%Y %H:%M')}").font = header_font
                        row += 2

                        def write_table(data_dict, start_row, title):
                            ws.merge_cells(f"A{start_row}:B{start_row}")
                            cell = ws.cell(row=start_row, column=1, value=title)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center')
                            start_row += 1
                            ws.cell(row=start_row, column=1, value="Параметр").font = Font(bold=True)
                            ws.cell(row=start_row, column=2, value="Значение").font = Font(bold=True)
                            start_row += 1
                            for k, v in data_dict.items():
                                c1, c2 = ws.cell(row=start_row, column=1, value=k), ws.cell(row=start_row, column=2, value=v)
                                c1.border, c2.border = thin_border, thin_border
                                start_row += 1
                            return start_row + 1

                        row = write_table(tech_info, row, "🔧 Техническая информация")
                        row = write_table(dist_stats, row, "📈 Статистики распределения")

                        ws.merge_cells(f"A{row}:C{row}")
                        cell = ws.cell(row=row, column=1, value="📋 Итоговый паспорт свойств")
                        cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center')
                        row += 1

                        ws.cell(row=row, column=1, value="Свойство").font = Font(bold=True)
                        ws.cell(row=row, column=2, value="Метод").font = Font(bold=True)
                        ws.cell(row=row, column=3, value="Результат").font = Font(bold=True)
                        row += 1

                        for item in ts_passport:
                            c1, c2, c3 = ws.cell(row=row, column=1, value=item["property"]), ws.cell(row=row, column=2, value=item["method"]), ws.cell(row=row, column=3, value=item["result"])
                            c1.border, c2.border, c3.border = thin_border, thin_border, thin_border
                            if "✅" in item["result"]: c3.fill = green_fill
                            elif "⚠️" in item["result"]: c3.fill = yellow_fill
                            elif "❌" in item["result"]: c3.fill = red_fill
                            row += 1

                        # 🔧 НОВЫЙ ЛИСТ: РЕКОМЕНДАЦИИ ПО МОДЕЛЯМ
                        ws_rec = wb.create_sheet("2_Рекомендации")
                        ws_rec.title = "2_Рекомендации"

                        # Заголовок листа рекомендаций
                        row_rec = 1
                        ws_rec.merge_cells(f"A{row_rec}:C{row_rec}")
                        cell = ws_rec.cell(row=row_rec, column=1, value="Рекомендуемые модели и обоснование")
                        cell.font = title_font
                        cell.alignment = Alignment(horizontal='center')
                        row_rec += 2

                        ws_rec.cell(row=row_rec, column=1, value="Модель").font = header_font
                        ws_rec.cell(row=row_rec, column=2, value="Условие применения").font = header_font
                        ws_rec.cell(row=row_rec, column=3, value="Обоснование").font = header_font
                        row_rec += 1

                        # Формирование рекомендаций на основе свойств
                        recommendations = []

                        # Базовые свойства
                        if is_stationary and is_white_noise:
                            recommendations.append(("Exponential Smoothing / Naive", "Ряд похож на белый шум", "Отсутствие автокорреляции → внешние факторы важнее истории"))
                        elif is_stationary and not is_white_noise:
                            recommendations.append(("ARIMA(p,d,q)", "Стационарен + есть автокорреляция", "Классический выбор для стационарных рядов с АК"))
                        elif not is_stationary:
                            recommendations.append(("ARIMA с дифференцированием / Prophet", "Нестационарный ряд", "Требуется удаление тренда (diff) или модель с трендом"))

                        if r_squared >= 0.7:
                            recommendations.append(("Linear Trend + ARMA", f"Сильный детерминированный тренд (R²={r_squared:.2f})", "Явный тренд лучше моделировать отдельно"))

                        # Сезонность
                        if is_seasonal or acf_seasonality_str != "Не обнаружены":
                            m_val = seasonal_periods_acf[0] if seasonal_periods_acf else period
                            recommendations.append((f"SARIMA(..., m={m_val}) / Prophet", "Обнаружена сезонность", "Явное моделирование сезонной компоненты улучшает точность"))

                        # Частота
                        if inferred_freq and inferred_freq in ['D', 'H', 'T']:
                            recommendations.append(("Prophet / LSTM", f"Высокочастотные данные ({inferred_freq})", "Учет праздников и внутридневных паттернов"))
                        elif inferred_freq in ['W', 'M', 'Q']:
                            recommendations.append(("SARIMA / TBATS", f"Сезонные данные ({inferred_freq})", "Классические методы для недельной/месячной сезонности"))

                        # Корреляции
                        if target_corr_str != "N/A" and "🟢" in target_corr_str:
                            recommendations.append(("Linear Regression / XGBoost", "Есть сильные предикторы (|r|>0.7)", "Использовать коррелирующие признаки как регрессоры"))

                        # Спектральные особенности
                        if fft_dominant_str != "Не обнаружены":
                            recommendations.append(("ML + Fourier features", f"Доминирующие частоты: {fft_dominant_str}", "Добавить sin/cos гармоники как признаки для улучшения прогноза"))

                        # Если нет явных паттернов
                        if not recommendations:
                            recommendations.append(("Naive / Simple Exponential Smoothing", "Нет выраженных паттернов", "Начать с простых базовых моделей для оценки"))

                        # Запись рекомендаций в лист
                        for model, condition, justification in recommendations:
                            ws_rec.cell(row=row_rec, column=1, value=model).font = Font(bold=True)
                            ws_rec.cell(row=row_rec, column=2, value=condition)
                            ws_rec.cell(row=row_rec, column=3, value=justification)
                            # Применяем границы
                            for col in range(1, 4):
                                ws_rec.cell(row=row_rec, column=col).border = thin_border
                            row_rec += 1

                        # Методологическое пояснение в конце листа
                        row_rec += 1
                        ws_rec.merge_cells(f"A{row_rec}:C{row_rec}")
                        cell = ws_rec.cell(row=row_rec, column=1, value="📚 Методология выбора моделей")
                        cell.font = Font(bold=True, size=11, color="0369A1")
                        cell.alignment = Alignment(horizontal='left')
                        row_rec += 1
                        ws_rec.cell(row=row_rec, column=1, value="• Статистические тесты (ADF, Ljung-Box) → выбор класса моделей")
                        ws_rec.cell(row=row_rec+1, column=1, value="• Спектральный анализ (ACF, FFT) → параметры сезонности и признаки")
                        ws_rec.cell(row=row_rec+2, column=1, value="• Корреляционный анализ → отбор признаков, борьба с мультиколлинеарностью")
                        ws_rec.cell(row=row_rec+3, column=1, value="• Порядок действий: 1) Преобразования → 2) Признаки → 3) Подбор параметров → 4) Валидация")

                        # Настройка ширины колонок
                        ws.column_dimensions['A'].width = 35
                        ws.column_dimensions['B'].width = 30
                        ws.column_dimensions['C'].width = 50
                        ws_rec.column_dimensions['A'].width = 30
                        ws_rec.column_dimensions['B'].width = 35
                        ws_rec.column_dimensions['C'].width = 60

                        # Футер
                        row += 1
                        ws.merge_cells(f"A{row}:C{row}")
                        cell = ws.cell(row=row, column=1, value=" Исследовано платформой CISStat TS Analytics | ✅ Верифицировано СтатКомитетом СНГ")
                        cell.font, cell.fill, cell.alignment = footer_font, footer_fill, Alignment(horizontal='center')

                        output = io.BytesIO()
                        wb.save(output)
                        output.seek(0)

                        # ── 6. КНОПКА ВЫГРУЗКИ ──────────
                        st.divider()
                        st.download_button(
                            label="📥 Выгрузить отчет в Excel",
                            data=output,
                            file_name=f"CISStat_TS_Report_{report_col}_{dt_now.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="secondary"
                        )
                        st.success("✅ Отчет сформирован! Включает: Паспорт свойств + Рекомендации по моделям.")

                except Exception as e:
                    st.error(f"❌ Ошибка при формировании отчета: {e}")
                    st.exception(e)

        # 2️⃣ ОТОБРАЖЕНИЕ ОТЧЕТА НА ЭКРАНЕ
        if st.session_state.get("report_ready"):
            st.divider()
            st.markdown(f"#### 📄 Готовый отчет: {report_col}")
            st.dataframe(
                st.session_state.report_table,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "property": st.column_config.TextColumn("Свойство", width="small"),
                    "method": st.column_config.TextColumn("Метод", width="medium"),
                    "desc": st.column_config.TextColumn("Описание", width="large"),
                    "result": st.column_config.TextColumn("Результат", width="medium")
                }
            )
            if st.button("📥 Выгрузить отчет в Excel", type="secondary", key="btn_download_excel_report"):
                st.download_button(
                    label="📥 Скачать Excel-файл",
                    data=st.session_state.report_excel_data,
                    file_name=f"CISStat_TS_Report_{report_col}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_final_download_excel"
                )
                st.session_state.report_ready = False
    else:
        st.warning("⚠️ Нет числовых колонок для анализа.")

# ─────────────────────────────────────────────────────────────
#  ВКЛАДКА 2: ВАЛИДАЦИЯ
# ─────────────────────────────────────────────────────────────
with tab_validation:
    st.markdown("""
    <div style="padding-left: 20px; margin: 20px 0; text-align: right;">
        <p style="margin: 0 0 10px 0; color: #1e293b; line-height: 1.6; font-size: 18px; font-weight: 400;">
            "Garbage in, garbage out <br>
            Мусор на входе — мусор на выходе".
        </p>
        <p style="margin: 0; color: #64748B; font-style: italic; font-size: 16px; line-height: 1.5;">
            — в газетной статье 1957 года о работе математиков армии США Уильям Д. Меллин объяснял,<br>
            что «небрежно запрограммированные» входные данные приводят к неверным выходам.
        </p>
    </div>
    """, unsafe_allow_html=True)

    with st.expander(" Цели модуля и результаты его прохождения", expanded=False):
        st.markdown("""
        ######  Цели модуля "Валидация"
                            
        **Цель раздела**. Модуль проводит комплексную автоматическую диагностику данных по 10 критериям: проверка форматов (regex), диапазонов значений, уникальности, 
        логической согласованности, качества метаданных и равномерности временного шага. На основе выявленных нарушений вычисляется интегральный показатель Data Quality Score (DQ)
        для оценки пригодности ряда к применению статистических моделей и машинному обучению.           
        
                    
        **Что мы получим на выходе?** Интерактивный Data Quality Dashboard с детальным отчетом о проблемах, рекомендацией по доступным моделям прогнозирования и 
        сравнительным паспортом свойств ряда (v1.0 до валидации vs v1.1 после). Пользователь получает обновленный датафрейм в session_state.df, готовый к передаче 
        в блок «Предобработка».
        """)

    # ── ℹ️ ИНФОРМАЦИЯ О СОХРАНЕНИИ ДАННЫХ ────────────────────────────
    with st.expander("Как работает сохранение результатов?", expanded=False):
        st.markdown("""
        ###### Где хранятся обработанные данные?

        **Все изменения сохраняются в `st.session_state.df`** — это основной датафрейм,
        который используется во **всех вкладках** приложения.

        ###### Порядок работы:

        1. **Загрузите данные** → сохраняются в `session_state.df`
        2. **Запустите валидацию** → находятся проблемы
        3. **Примените стратегию** в любом блоке (пропуски, выбросы, форматы...)
        → изменения сохраняются в `session_state.df`
        4. **Запустите валидацию заново** → проверьте результат

        ###### Дополнительные рабочие копии:

        Каждая рабочая копия — это изолированная среда для безопасной обработки данных
        **до подтверждения стратегии**. После применения изменений копия синхронизируется
        с основным `session_state.df`.

        | Переменная | Модуль | Что проверяет |
        |------------|--------|---------------|
        | `df_missing_work` | Пропуски (Missing) | NaN, пустые значения |
        | `df_pattern_work` | Форматы (Regex) | Шаблоны email, телефон, дата |
        | `df_range_work` | Диапазоны (Ranges) | Мин/макс значения |
        | `df_outlier_work` | Выбросы (Outliers) | Аномальные значения |
        | `df_consistency_work` | Согласованность (Consistency) | Логика и хронология |
        | `df_uniqueness_work` | Уникальность (Uniqueness) | Дубликаты строк и дат |
        | `df_inclusion_work` | Принадлежность (Inclusion) | Справочники и домены |
        | `df_referential_work` | Ссылочная целостность (Referential) | Внешние ключи |
        | `df_text_work` | Качество текста (Text Quality) | Мусорные символы, кодировка |
        | `df_regularity_work` | Регулярность шага (Regularity) | Равномерность временного ряда |

        ###### Служебные копии:
        - `df_before_validation` — snapshot данных **до** валидации (для сравнения v1.0 → v1.1)
        - `df_after_fixes` — данные **после** применения стратегии (до синхронизации)

        ###### Как это работает:
        1. При обнаружении нарушений создаётся рабочая копия (например, `df_range_work`)
        2. Вы редактируете данные в интерактивной таблице или выбираете стратегию
        3. Нажимаете **"Сохранить"** или **"Подтвердить изменения"**
        4. Рабочая копия синхронизируется: `session_state.df = df_range_work`
        5. Флаг `validation_ready` сбрасывается → нужно перезапустить валидацию

        ###### Экспорт результатов:

        В конце каждого блока есть кнопка **"💾 Скачать результат (CSV)"** —
        используйте её для сохранения очищенных данных.
        """)

    # Сохраняем snapshot ДО валидации (для сравнения ДО/ПОСЛЕ в паспорте)
    if "df_before_validation" not in st.session_state:
        st.session_state.df_before_validation = df.copy()

    # ── КНОПКА ЗАПУСКА ─────────────────────────────────────
    if st.button(" Запустить валидацию", type="primary"):
        with st.spinner("🔍 Комплексный анализ данных..."):
            progress = st.progress(0)

            # 1. Базовая валидация схемы и типов
            progress.progress(0.10, text="Загрузка правил и базовая валидация...")
            rules = load_rules()
            st.session_state.rules = rules
            val = validate_dataframe(df, rules)

            # 2. Проверка форматов (Regex)
            progress.progress(0.20, text="Проверка форматов и шаблонов (Regex)...")
            pattern_results = validate_formats(df, rules)
            st.session_state.pattern_results = pattern_results

            # 3. Проверка диапазонов значений
            progress.progress(0.30, text="Проверка диапазонов значений...")
            range_results, range_masks, range_bounds = validate_ranges(df, rules)
            st.session_state.range_results = range_results
            st.session_state.range_masks = range_masks
            st.session_state.range_bounds = range_bounds

            # 4. Проверка согласованности (логика + хронология)
            progress.progress(0.40, text="Проверка согласованности данных...")
            consistency_results = validate_consistency(df, rules)
            consistency_masks = {}

            # 5. Проверка уникальности (дубликаты)
            progress.progress(0.50, text="Проверка уникальности записей...")
            uniqueness_results = []
            if st.session_state.col_types.get("date"):
                date_col = st.session_state.col_types["date"][0]
                dup_count = int(df.duplicated(subset=[date_col], keep=False).sum())
            else:
                dup_count = int(df.duplicated(keep=False).sum())
            if dup_count > 0:
                uniqueness_results.append({
                    "Правило": "Уникальность записей",
                    "Дубликатов": dup_count,
                    "Статус": "⚠️ Нарушено"
                })
            else:
                uniqueness_results.append({
                    "Правило": "Уникальность записей",
                    "Дубликатов": 0,
                    "Статус": "✅ Соблюдено"
                })

            # 6. Проверка принадлежности к набору (Inclusion)
            progress.progress(0.60, text="Проверка принадлежности к справочникам...")
            inclusion_results = []
            inclusion_masks = {}
            inclusion_rules = rules.get("inclusion", {})
            for col, allowed_vals in inclusion_rules.items():
                if col in df.columns and allowed_vals:
                    invalid_mask = ~df[col].isin(allowed_vals) & df[col].notna()
                    violations = int(invalid_mask.sum())
                    if violations > 0:
                        inclusion_masks[col] = invalid_mask
                        inclusion_results.append({
                            "Правило": f"Inclusion: {col}",
                            "Колонка": col,
                            "Нарушений": violations,
                            "% брака": f"{(violations / len(df)) * 100:.2f}%",
                            "Статус": "⚠️ Нарушено"
                        })

            # 7. Проверка ссылочной целостности
            progress.progress(0.70, text="Проверка ссылочной целостности...")
            ref_results, ref_masks = validate_referential(df, rules)

            # 8. Проверка качества текста
            progress.progress(0.80, text="Проверка качества текста...")
            text_results, text_masks = validate_text_quality(df, rules)

            # 9. Проверка равномерности временного шага
            progress.progress(0.90, text="Проверка равномерности временного шага...")
            regularity_results, regularity_masks, regularity_freq_info = validate_regular_step(
                df, rules,
                date_col=st.session_state.primary_date_col if st.session_state.get('primary_date_col') else None
            )

            # 10. Проверка достаточности числа наблюдений
            progress.progress(0.95, text="Проверка достаточности числа наблюдений...")
            sufficiency_results, sufficiency_recommendations = validate_sufficiency(
                df,
                rules,
                date_col=st.session_state.primary_date_col if st.session_state.get('primary_date_col') else None
            )

            # 11. Пропуски и выбросы
            progress.progress(0.93, text="Анализ пропусков и выбросов...")
            miss = analyze_missing(df, rules.get("missing", {}))
            outl = detect_outliers(df, rules.get("outliers", {}))

            # 12. TS-специфичные проверки
            progress.progress(0.96, text="TS-специфичные проверки...")
            ts_checks = {}
            if st.session_state.col_types.get("date") and st.session_state.col_types.get("num"):
                date_col = st.session_state.col_types["date"][0]
                num_col = st.session_state.col_types["num"][0]
                df_ts = df.sort_values(date_col).set_index(date_col)[[num_col]].dropna()
                if len(df_ts) >= 10:
                    try:
                        from statsmodels.tsa.stattools import adfuller
                        adf_result = adfuller(df_ts[num_col].dropna())
                        ts_checks["adf_pvalue"] = adf_result[1]
                        ts_checks["is_stationary"] = adf_result[1] < 0.05
                    except:
                        ts_checks["adf_pvalue"] = None
                        ts_checks["is_stationary"] = None

                    freq = pd.infer_freq(df_ts.index)
                    ts_checks["frequency"] = freq if freq else "Нерегулярная"

                    try:
                        if pd.api.types.is_datetime64_any_dtype(df_ts.index):
                            gaps = df_ts.index.to_series().diff().dropna()
                            ts_checks["max_gap"] = gaps.max() if len(gaps) > 0 else pd.Timedelta(0)
                        else:
                            ts_checks["max_gap"] = pd.Timedelta(0)
                    except Exception:
                        ts_checks["max_gap"] = pd.Timedelta(0)
                else:
                    ts_checks = {"error": "Недостаточно данных для TS-проверок"}
            else:
                ts_checks = {"error": "Не найдены колонки с датами и числовыми данными"}

            # Сохраняем всё в сессию для отрисовки дашборда
            st.session_state.val_results = {
                "val": val,
                "miss": miss,
                "outl": outl,
                "ts": ts_checks,
                "consistency": consistency_results,
                "consistency_masks": consistency_masks,
                "range_results": range_results,
                "range_masks": range_masks,
                "range_bounds": range_bounds,
                "uniqueness": uniqueness_results,
                "inclusion": inclusion_results,
                "inclusion_masks": inclusion_masks,
                "referential": ref_results,
                "referential_masks": ref_masks,
                "text": text_results,
                "text_masks": text_masks,
                "regularity": regularity_results,
                "regularity_masks": regularity_masks,
                "regularity_freq_info": regularity_freq_info,
                "sufficiency": sufficiency_results,
                "sufficiency_recommendations": sufficiency_recommendations
            }
            st.session_state.validation_ready = True
            progress.progress(1.0, text="✅ Валидация проведена!")
            st.success("Комплексная валидация завершена!")
            st.rerun()

    
        # ─────────────────────────────────────────────────────────
        # 📊 СОХРАНЕНИЕ ПАСПОРТА СВОЙСТВ v1.1 (после валидации)
        # ─────────────────────────────────────────────────────────
        if st.session_state.validation_ready:
            if (st.session_state.get('primary_date_col') and
                st.session_state.col_types.get('num')):

                date_col = st.session_state.primary_date_col
                num_col = st.session_state.col_types['num'][0]

                # Используем ту же целевую колонку, что и в v1.0
                target_col = st.session_state.get('ts_props_v10_target_col', num_col)

                df_validated = st.session_state.df.copy()

                # Формируем временной ряд (аналогично вкладке "Загрузка")
                try:
                    df_validated[date_col] = pd.to_datetime(df_validated[date_col])
                    df_ts_validated = df_validated.set_index(date_col)

                    if target_col in df_ts_validated.columns:
                        analysis_series_v11 = (
                            df_ts_validated[target_col]
                            .resample('D').mean()
                            .dropna()
                            .astype(float)
                        )

                        if len(analysis_series_v11) >= 30:
                            # 🔧 ИСПОЛЬЗУЕМ ТУ ЖЕ ФУНКЦИЮ, ЧТО И В v1.0!
                            props_v11 = calculate_ts_passport(
                                analysis_series_v11,
                                df_filtered=df_validated,
                                ct_f=st.session_state.col_types,
                                target_col=target_col
                            )
                            props_v11['version'] = 'v1.1 (после валидации)'

                            st.session_state.ts_props_v11 = props_v11

                            # Сравниваем с v1.0
                            if 'ts_props_v10' in st.session_state:
                                st.session_state.ts_props_comparison_v10_v11 = _compare_ts_props(
                                    st.session_state.ts_props_v10,
                                    props_v11
                                )
                except Exception as e:
                    st.warning(f"⚠️ Не удалось рассчитать свойства v1.1: {e}")

        # ─────────────────────────────────────────────────────────────
        # УНИВЕРСАЛЬНАЯ ФУНКЦИЯ РАСЧЁТА ПАСПОРТА СВОЙСТВ РЯДА
        # ─────────────────────────────────────────────────────────────
        def calculate_ts_passport(analysis_series: pd.Series,
                                df_filtered: pd.DataFrame = None,
                                ct_f: dict = None,
                                target_col: str = None) -> dict:
            """
            Рассчитывает полный паспорт свойств временного ряда (13 метрик).
            Структура идентична паспорту во вкладке "Загрузка".

            Returns:
                dict с ключами, соответствующими названиям свойств в таблице
            """
            from scipy.stats import linregress, jarque_bera
            from scipy.signal import periodogram
            from scipy.fft import fft, fftfreq
            from scipy.signal import find_peaks
            from statsmodels.tsa.stattools import adfuller, acf as acf_func
            from statsmodels.tsa.seasonal import STL
            from statsmodels.stats.diagnostic import acorr_ljungbox
            import numpy as np

            props = {}

            if len(analysis_series) < 30:
                return {"error": "Недостаточно данных (нужно > 30 точек)"}

            try:
                # 0. ЧАСТОТА РЯДА
                inferred_freq = pd.infer_freq(analysis_series.index.drop_duplicates().sort_values())
                props['freq'] = {
                    'value': inferred_freq if inferred_freq else 'Нерегулярная',
                    'is_ok': inferred_freq is not None
                }

                # 1. СТАЦИОНАРНОСТЬ (ADF)
                adf_res = adfuller(analysis_series, autolag='AIC')
                adf_p = adf_res[1]
                is_stationary = adf_p < 0.05
                props['stationarity'] = {
                    'value': adf_p,
                    'is_stationary': is_stationary,
                    'is_ok': is_stationary
                }

                # 2. ДЕТЕРМИНИРОВАННОСТЬ (R² тренда)
                slope, intercept, r_value, p_value, std_err = linregress(
                    range(len(analysis_series)), analysis_series
                )
                r_squared = r_value**2
                is_deterministic = r_squared >= 0.7
                props['determinism'] = {
                    'value': r_squared,
                    'slope': slope,
                    'is_deterministic': is_deterministic
                }

                # 3. АВТОКОРРЕЛЯЦИЯ (Ljung-Box)
                lb_res = acorr_ljungbox(analysis_series, lags=[10])
                if isinstance(lb_res, pd.DataFrame):
                    lb_p = lb_res['lb_pvalue'].iloc[0]
                else:
                    lb_p = lb_res[1][0]
                is_white_noise = lb_p > 0.05
                props['autocorrelation'] = {
                    'value': lb_p,
                    'is_white_noise': is_white_noise,
                    'is_ok': is_white_noise
                }

                # 4. НОРМАЛЬНОСТЬ (Jarque-Bera)
                jb_res = jarque_bera(analysis_series)
                jb_p = jb_res.pvalue if hasattr(jb_res, 'pvalue') else jb_res[1]
                is_normal = jb_p > 0.05
                props['normality'] = {
                    'value': jb_p,
                    'is_normal': is_normal,
                    'is_ok': is_normal
                }

                # 5. НАПРАВЛЕНИЕ ТРЕНДА
                if slope > 0:
                    trend_dir = 'up'
                elif slope < 0:
                    trend_dir = 'down'
                else:
                    trend_dir = 'flat'
                props['trend'] = {
                    'slope': slope,
                    'direction': trend_dir
                }

                # 6. КОРРЕЛЯЦИЯ ПРИЗНАКОВ (если есть другие числовые колонки)
                props['correlations'] = {}
                if df_filtered is not None and ct_f is not None and target_col:
                    num_cols = ct_f.get("num", [])
                    if len(num_cols) >= 2 and target_col in num_cols:
                        try:
                            corr_df = df_filtered[num_cols].corr()
                            target_corr = corr_df[target_col].drop(target_col).sort_values(key=abs, ascending=False)
                            top_corrs = {col: float(val) for col, val in target_corr.head(3).items()}
                            props['correlations'] = {
                                'top3': top_corrs,
                                'max_abs_corr': float(target_corr.iloc[0]) if len(target_corr) > 0 else 0.0
                            }
                        except:
                            pass

                # 7. СЕЗОННОСТЬ (STL strength)
                period = 7 if (inferred_freq and 'D' in inferred_freq) else 12
                try:
                    stl_res = STL(analysis_series, period=period, robust=True).fit()
                    var_total = analysis_series.var()
                    var_resid = stl_res.resid.var()
                    var_detrended = var_total - stl_res.trend.var()
                    strength_seasonality = max(0, 1 - var_resid / var_detrended) if var_detrended > 0 else 0
                    is_seasonal = strength_seasonality > 0.6
                except:
                    strength_seasonality = 0.0
                    is_seasonal = False
                props['seasonality'] = {
                    'strength': strength_seasonality,
                    'is_seasonal': is_seasonal
                }

                # 8. СЕЗОННЫЕ ПЕРИОДЫ (ACF)
                max_lag = min(60, len(analysis_series) // 4)
                acf_values = acf_func(analysis_series, nlags=max_lag)
                confidence = 1.96 / np.sqrt(len(analysis_series))
                significant_lags = np.where(np.abs(acf_values) > confidence)[0][1:]

                seasonal_periods_acf = []
                for i, lag in enumerate(significant_lags):
                    if i > 0 and lag - significant_lags[i-1] < 3:
                        continue
                    if lag > 2:
                        seasonal_periods_acf.append(int(lag))

                props['seasonal_periods'] = {
                    'periods': seasonal_periods_acf[:3],
                    'count': len(seasonal_periods_acf[:3])
                }

                # 9. ДОЛГАЯ ПАМЯТЬ (Hurst)
                def hurst_exponent(series, max_lag=20):
                    lags = range(2, max_lag)
                    tau = [max(np.std(np.subtract(series[lag:], series[:-lag])), 1e-8) for lag in lags]
                    try:
                        return np.polyfit(np.log(lags), np.log(tau), 1)[0]
                    except:
                        return 0.5

                hurst_val = hurst_exponent(analysis_series.values)
                if hurst_val < 0.45:
                    memory_type = 'anti_persistent'
                elif hurst_val > 0.55:
                    memory_type = 'persistent'
                else:
                    memory_type = 'random_walk'
                props['hurst'] = {
                    'value': hurst_val,
                    'type': memory_type
                }

                # 10. ДОМИНИРУЮЩИЕ ЧАСТОТЫ (FFT)
                n = len(analysis_series)
                y = analysis_series.values - analysis_series.mean()
                yf = fft(y)
                xf = fftfreq(n, 1)[:n//2]
                amplitude = 2.0/n * np.abs(yf[0:n//2])

                peaks, _ = find_peaks(amplitude, height=np.mean(amplitude) + np.std(amplitude))
                fft_periods = [1/xf[p] for p in peaks if xf[p] > 0 and xf[p] < 0.5]
                fft_dominant = sorted(fft_periods)[:3] if fft_periods else []
                props['fft'] = {
                    'dominant_periods': fft_dominant,
                    'count': len(fft_dominant)
                }

                # 11. ПЕРИОДОГРАММА
                freq_per, pxx_per = periodogram(analysis_series.values, fs=1.0, window='hann')
                peaks_per, _ = find_peaks(pxx_per, height=np.median(pxx_per)*2)
                periodogram_periods = sorted([1/freq_per[p] for p in peaks_per if freq_per[p] > 0])[:3]
                props['periodogram'] = {
                    'periods': periodogram_periods,
                    'count': len(periodogram_periods)
                }

                # 12. ВЕЙВЛЕТ-МАСШТАБЫ
                try:
                    import pywt
                    widths = np.arange(1, min(128, len(analysis_series)//4))
                    cwtmatr, _ = pywt.cwt(
                        analysis_series.values - analysis_series.mean(),
                        widths, 'morl', sampling_period=1
                    )
                    mean_power = np.mean(np.abs(cwtmatr), axis=1)
                    wavelet_peaks, _ = find_peaks(mean_power, height=np.mean(mean_power))
                    wavelet_scales = widths[wavelet_peaks][:3].tolist() if len(wavelet_peaks) > 0 else []
                    props['wavelet'] = {
                        'scales': wavelet_scales,
                        'count': len(wavelet_scales)
                    }
                except:
                    props['wavelet'] = {'scales': [], 'count': 0}

                # 13. БАЗОВЫЕ СТАТИСТИКИ (для числового сравнения)
                props['basic_stats'] = {
                    'n': len(analysis_series),
                    'mean': float(analysis_series.mean()),
                    'std': float(analysis_series.std()),
                    'min': float(analysis_series.min()),
                    'max': float(analysis_series.max())
                }

                props['timestamp'] = pd.Timestamp.now().isoformat()

            except Exception as e:
                props['error'] = str(e)

            return props

    # ───────────────────────────────────────────────────────────
    # 🛡️ DATA QUALITY DASHBOARD
    # ───────────────────────────────────────────────────────────
    if st.session_state.validation_ready:
        # Достаем правила из сессии (иначе NameError при перезапуске скрипта)
        rules = st.session_state.get("rules", {})

        # Безопасное извлечение (избегаем ValueError: too many values to unpack)
        val = st.session_state.val_results.get("val")
        miss = st.session_state.val_results.get("miss")
        outl = st.session_state.val_results.get("outl")
        ts = st.session_state.val_results.get("ts")
        consistency_results = st.session_state.val_results.get("consistency", [])

        st.divider()
        st.markdown("###  Панель контроля качества данных (data quality)")
        st.caption("Автоматическая проверка по стандартам. Проблемные блоки раскрываются автоматически.")

        # ── ФУНКЦИЯ-ПОМОЩНИК ДЛЯ СОЗДАНИЯ КАРТОЧКИ ─────────────
        def make_card(title, has_issues, methodology, result_text, result_status, fix_action=None):
            with st.container(border=True):
                st.markdown(f"##### {title}")
                with st.expander("Метрики и алгоритм", expanded=has_issues):
                    st.markdown(methodology)
                st.markdown(result_text)
                if fix_action and has_issues:
                    if st.button(fix_action['label'], key=f"fix_{title}", use_container_width=True):
                        fix_action['callback']()
                return has_issues

        # 🔹 СПРАВКА ПО СТАНДАРТАМ КАЧЕСТВА ДАННЫХ
        with st.expander(" Справка по стандартам качества данных", expanded=False):
            st.markdown("""
            ##### 🌍 Международные стандарты Data Quality

            Платформа CISStat TS Analysis реализует **10 измерений качества данных** в соответствии с международными стандартами:

            ---

            ###### 1️⃣ **DAMA DMBOK v2** (Data Management Body of Knowledge)
            **Организация:** DAMA International (Data Management Association)
            **Год:** 2017 (2-е издание)
            **Раздел:** Chapter 13 — Data Quality

            **10 измерений качества данных:**

            | Измерение | Описание | Пример проверки в CISStat |
            |-----------|----------|---------------------------|
            | **Completeness** | Полнота данных | Модуль "Пропуски (Missing)" |
            | **Uniqueness** | Уникальность записей | Модуль "Уникальность (дубликаты)" |
            | **Validity** | Соответствие домену/формату | Модули "Regex-форматы", "Диапазоны" |
            | **Timeliness** | Актуальность данных | Модуль "Хронологический порядок" |
            | **Accuracy** | Точность значений | Модуль "Выбросы (Outliers)" |
            | **Consistency** | Согласованность между источниками | Модуль "Логическая согласованность" |
            | **Integrity** | Целостность связей | Модули "Inclusion", "Referential" |
            | **Conformity** | Соответствие стандартам | Модуль "Типы данных" |

            **Применение в платформе:**
            Каждая из 10 проверок валидации соответствует одному или нескольким измерениям DAMA DMBOK. Результаты фиксируются в Data Quality Dashboard с метриками `% valid = valid/total`.

            **⚠️ Почитать**
            - первоисточник: https://dama.org/
            - статья на Хабре: https://habr.com/ru/companies/rshb/articles/767440/
            ---

            ###### 2️⃣ **ISO 8000-61** (Data Quality — Part 61: Process Reference Model)
            **Организация:** ISO/IEC JTC 1/SC 32
            **Год:** 2020
            **Стандарт:** ISO 8000-61:2020

            **Модель процесса управления качеством данных:**

            ```
            ┌─────────────────────────────────────────────────────────┐
            │  1. SPECIFY (Спецификация требований к качеству)        │
            │     → Правила валидации (YAML-конфиг)                   │
            └─────────────────────────────────────────────────────────┘
                              ↓
            ┌─────────────────────────────────────────────────────────┐
            │  2. MEASURE (Измерение качества)                        │
            │     → Автоматическая проверка (10 модулей)              │
            │     → Метрики: completeness, accuracy, consistency      │
            └─────────────────────────────────────────────────────────┘
                              ↓
            ┌─────────────────────────────────────────────────────────┐
            │  3. ANALYZE (Анализ отклонений)                         │
            │     → Data Quality Dashboard                            │
            │     → Выявление root cause (причины проблем)            │
            └─────────────────────────────────────────────────────────┘
                              ↓
            ┌─────────────────────────────────────────────────────────┐
            │  4. IMPROVE (Улучшение данных)                          │
            │     → Вкладка "Предобработка" (fill, drop, transform)   │
            │     → Стратегии исправления                             │
            └─────────────────────────────────────────────────────────┘
            ```

            **Применение в платформе:**
            CISStat реализует полный цикл PDCA (Plan-Do-Check-Act) по ISO 8000-61:
            - **Plan:** YAML-правила валидации
            - **Do:** Автоматическая проверка при загрузке
            - **Check:** Визуализация нарушений в дашборде
            - **Act:** Интерактивные пайплайны исправления

            ---

            ###### 3️⃣ **TDQM** (Total Data Quality Management)
            **Авторы:** Richard Wang, Diane Strong (MIT, 1996)
            **Концепция:** Расширение TQM (Total Quality Management) для данных

            **4 категории качества данных по TDQM:**

            | Категория | Измерения | Пример в CISStat |
            |-----------|-----------|------------------|
            | **Intrinsic DQ** | Accuracy, Objectivity, Believability, Reputation | Модули "Диапазоны", "Выбросы" (проверка на реалистичность) |
            | **Contextual DQ** | Relevancy, Value-added, Timeliness, Completeness, Amount of data | Модуль "Достаточность числа наблюдений" (n ≥ 50 для ARIMA) |
            | **Representational DQ** | Interpretability, Ease of understanding, Representational consistency, Concise representation | Модули "Типы данных", "Regex-форматы" |
            | **Accessibility DQ** | Accessibility, Access security | Авторизация, логирование действий |

            **Применение в платформе:**
            TDQM подчёркивает, что качество данных **зависит от контекста использования**. В CISStat это реализовано через:
            - **Профилирование по типу задачи:** TS-модели требуют больше наблюдений (n ≥ 100), чем простой тренд (n ≥ 10)
            - **Адаптивные пороги:** Для финансовых данных stricter диапазоны (цена ≥ 0), для демографии — мягче
            - **Обоснование методов:** Каждый вывод сопровождается ссылкой на стандарт (например, "ARIMA требует n ≥ 50 по TDQM Contextual DQ")

            ---

            ###### 📊 Сводная таблица соответствия

            | № | Проверка в CISStat | DAMA DMBOK | ISO 8000-61 | TDQM |
            |---|-------------------|------------|-------------|------|
            | 1 | Типы данных | Conformity | SPECIFY | Representational |
            | 2 | Regex-форматы | Validity | SPECIFY | Representational |
            | 3 | Диапазоны значений | Validity, Accuracy | MEASURE | Intrinsic |
            | 4 | Логическая согласованность | Consistency | MEASURE | Contextual |
            | 5 | Уникальность | Uniqueness | MEASURE | Intrinsic |
            | 6 | Inclusion (справочники) | Integrity | MEASURE | Contextual |
            | 7 | Referential integrity | Integrity | MEASURE | Accessibility |
            | 8 | Хронологический порядок | Consistency, Timeliness | MEASURE | Contextual |
            | 9 | Равномерность шага | Consistency | MEASURE | Contextual |
            | 10 | Достаточность наблюдений | Completeness | ANALYZE | Contextual |
            | + | Пропуски (Missing) | Completeness | MEASURE | Contextual |
            | + | Выбросы (Outliers) | Accuracy | ANALYZE | Intrinsic |

            ---

            ###### 🎯 Целевые показатели качества (Target DQ Levels)

            Согласно ISO 8000-61 и DAMA DMBOK, платформа устанавливает следующие пороги:

            | Метрика | Отлично (✅) | Хорошо (⚠️) | Критично (❌) |
            |---------|--------------|-------------|---------------|
            | Completeness | ≥ 99% | 95-99% | < 95% |
            | Accuracy | ≥ 98% | 90-98% | < 90% |
            | Consistency | ≥ 99% | 95-99% | < 95% |
            | Uniqueness | 100% | 99-100% | < 99% |
            | Validity | ≥ 98% | 95-98% | < 95% |

            **Интерпретация для временных рядов:**
            - **n < 50:** Недостаточно для ARIMA/SARIMA (TDQM Contextual DQ)
            - **n < 100:** Недостаточно для ML-моделей (LSTM, XGBoost)
            - **Пропуски > 5%:** Требуется интерполяция или удаление (ISO 8000-61 IMPROVE)
            - **Выбросы > 3σ:** Проверить на ошибки ввода (DAMA Accuracy)

            ---

            ###### 📚 Источники и литература

            1. **DAMA International.** DAMA-DMBOK: Data Management Body of Knowledge, 2nd Edition. Technics Publications, 2017. ISBN: 978-1634622974
            2. **ISO/IEC 8000-61:2020.** Data quality — Part 61: Process reference model. International Organization for Standardization, 2020.
            3. **Wang, R.Y., Strong, D.M.** Beyond Accuracy: What Data Quality Means to Data Consumers. Journal of Management Information Systems, 1996, Vol. 12, No. 4, pp. 5-33.
            4. **Redman, T.C.** Data Quality: The Field Guide. Digital Press, 2001.
            5. **CISStat Internal Standard.** Стандарт качества данных для временных рядов, 2026.

            ---

            💡 **Как использовать эту информацию:**
            - При обнаружении нарушения смотрите, какому стандарту оно противоречит
            - Используйте целевые показатели (Target DQ Levels) для приоритизации исправлений
            - Ссылайтесь на стандарты при обосновании необходимости очистки данных перед руководством
            """)

        st.divider()

        # ─────────────────────────────────────────────────────
        # ── 1. ПРОВЕРКА ТИПОВ ДАННЫХ ─────────────────────────

        if "classification_result" not in st.session_state:
            st.session_state.classification_result = None

        known_cols = set(
            st.session_state.col_types.get("date", []) +
            st.session_state.col_types.get("num", []) +
            st.session_state.col_types.get("cat", [])
        )
        undefined_cols = [col for col in df.columns if col not in known_cols]
        type_issues = len(undefined_cols) > 0

        if type_issues:
            cols_preview = ", ".join(f"`{c}`" for c in undefined_cols[:5])
            if len(undefined_cols) > 5:
                cols_preview += f" и ещё {len(undefined_cols) - 5}..."
            type_result = f"⚠️ Найдено **{len(undefined_cols)}** колонок с неопределённым типом:\n{cols_preview}"
        else:
            type_result = "✅ Все типы корректны"

        # Функция авто-классификации
        def auto_classify_types():
            known_cols = set(
                st.session_state.col_types.get("date", []) +
                st.session_state.col_types.get("num", []) +
                st.session_state.col_types.get("cat", [])
            )
            undefined_cols = [col for col in df.columns if col not in known_cols]

            classified = {"date": [], "num": [], "cat": []}
            classification_log = []

            for col in undefined_cols:
                sample = df[col].dropna()
                if len(sample) == 0:
                    continue
                if df[col].dtype == 'object' or str(df[col].dtype) == 'string':
                    try:
                        pd.to_datetime(sample.head(100), errors='raise')
                        classified["date"].append(col)
                        classification_log.append(f"**`{col}`** → 📅 datetime")
                        continue
                    except:
                        pass

                if pd.api.types.is_numeric_dtype(df[col]):
                    if df[col].nunique() < 50:
                        classified["cat"].append(col)
                        classification_log.append(f"**`{col}`** → 📋 categorical")
                    else:
                        classified["num"].append(col)
                        classification_log.append(f"**`{col}`** → 🔢 numeric")
                else:
                    classified["cat"].append(col)
                    classification_log.append(f"**`{col}`** → 📋 categorical")

            for cat, cols in classified.items():
                st.session_state.col_types[cat].extend(cols)

            if classification_log:
                st.session_state.classification_result = {"success": True, "count": len(classification_log), "log": classification_log}
            else:
                st.session_state.classification_result = {"success": False, "message": "⚠️ Не удалось классифицировать."}

        make_card(
            " Проверка типов данных (data types)",
            has_issues=type_issues,
            methodology=(
                "**◻️ Метрики:** Фактический `dtype`, автоклассификация `numeric/datetime/categorical`.  \n"
                "**◻️ Алгоритм:** `df.dtypes` → `select_dtypes()` → сверка с `detect_and_convert_year_columns`.  \n"
                "**◻️ Влияние на TS:** Некорректный тип ломает `DatetimeIndex` → критично для ARIMA/Prophet.  \n"
                "**◻️ Как работает авто-классификация:**  \n"
                "| Шаг | Логика | Пример |  \n"
                "|-----|--------|--------|  \n"
                "| **1. Детекция дат** | Пробует `pd.to_datetime()` на первых 100 значениях. Если успешно → datetime | `Date_Column` → datetime |  \n"
                "| **2. Числовые** | Проверяет `is_numeric_dtype()`. Если `<50` уникальных → категория, иначе → numeric | `Region_ID` → categorical (12 значений) |  \n"
                "| **3. Строковые** | Всё остальное → категория | `Product_Name` → categorical (45 уникальных) |"
            ),
            result_text=type_result,
            result_status="✅" if not type_issues else "⚠️",
            fix_action={"label": "Авто-классифицировать типы", "callback": auto_classify_types} if type_issues else None
        )

        # Показываем результат классификации
        if st.session_state.classification_result is not None:
            result = st.session_state.classification_result
            if result["success"]:
                st.success(f"✅ **Авто-классификация завершена!** Распределено **{result['count']}** колонок.")
                st.info("💡 **Следующий шаг:** Нажмите **Запустить валидацию** снова.")
                if st.button("🔄 Скрыть результат", key="clear_classification"):
                    st.session_state.classification_result = None
                    st.rerun()
            else:
                st.warning(result["message"])

        # ───────────────────────────────────────────────────────────
        # 2. ПРОВЕРКА ФОРМАТОВ И ШАБЛОНОВ (REGEX)
        # ───────────────────────────────────────────────────────────

        # Получаем результаты из session_state
        pattern_results_local = st.session_state.get("pattern_results", [])

        # ФИЛЬТРУЕМ только РЕАЛЬНЫЕ нарушения (игнорируем N/A и числовые колонки)
        real_violations = []
        for r in pattern_results_local:
            invalid_count = r.get('invalid_count', 'N/A')
            col_name = r.get('Колонка', '')

            # Пропускаем если:
            # 1. Нет данных о нарушениях (N/A)
            # 2. Нарушений = 0
            # 3. Колонка числовая (не должна проверяться regex)
            if (invalid_count == 'N/A' or invalid_count is None or invalid_count == 0):
                continue

            # Дополнительная проверка: пропускаем числовые колонки
            if col_name in df.select_dtypes(include=['int64', 'float64']).columns:
                continue

            real_violations.append(r)

        pattern_issues = len(real_violations) > 0

        # Сохраняем отфильтрованный список для дальнейшего использования
        st.session_state.pattern_results_filtered = real_violations

        # Формируем текст результата
        if pattern_issues:
            total_violations = sum(r.get('invalid_count', 0) for r in real_violations)
            res_text = f"⚠️ Найдено **{total_violations}** нарушений в **{len(real_violations)}** колонках"
        else:
            res_text = "✅ Все шаблоны соблюдены"

        # Отображаем карточку-резюме
        make_card(
            " Проверка форматов и шаблонов (regex)",
            has_issues=pattern_issues,
            methodology=(
                "**◻️ Метрики:** `% match = valid/total`, `invalid_count`.  \n"
                "**◻️ Алгоритм:** `df[col].str.fullmatch(regex)`.  \n"
                "**◻️ Влияние на TS:** Ошибочные ключи ломают `groupby`, `pivot`, `resample`.  \n"
                "**◻️ Описание:** Модуль проверяет данные на соответствие регулярным выражениям из справочника платформы "
                "(email, телефон, дата, код и т.д.). Значения, не проходящие валидацию, фиксируются для исправления и предлагаются стратегии их обработки."
            ),
            result_text=res_text,
            result_status="✅" if not pattern_issues else "⚠️",
            fix_action=None
        )

        # ── 🔧 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при нарушениях) ──
        if pattern_issues:
            with st.expander("📋 Показать детали нарушений", expanded=True):
                # Инициализация рабочего датафрейма для модуля форматов
                if "df_pattern_work" not in st.session_state:
                    st.session_state.df_pattern_work = df.copy()
                df_work = st.session_state.df_pattern_work

                for v in real_violations:
                    col_name = v.get("Колонка")
                    pattern = v.get("pattern", ".*")
                    invalid_count = v.get("invalid_count", 0)
                    invalid_pct = v.get("% invalid", "N/A")
                    valid_count = v.get("valid_count", 0)
                    total_count = v.get("total_count", 0)

                    # Заголовок нарушения
                    st.error(f"**🔴 Колонка:** `{col_name}`")
                    st.markdown(f"- **Нарушений:** `{invalid_count}` из `{total_count}`")
                    st.markdown(f"- **% брака:** `{invalid_pct}`")
                    st.markdown(f"- **Соответствий:** `{valid_count}`")

                    # 🔍 Показать примеры некорректных значений
                    if col_name and col_name in df_work.columns:
                        try:
                            # Создаём маску нарушений ТОЛЬКО для этой колонки
                            mask_invalid = ~df_work[col_name].astype(str).str.fullmatch(pattern, na=False)
                            invalid_values = df_work.loc[mask_invalid, col_name].dropna().unique()[:10]

                            if len(invalid_values) > 0:
                                st.markdown("##### 🔴 Примеры нарушений:")
                                st.code("\n".join(f"• {val}" for val in invalid_values), language="text")
                            else:
                                st.info("ℹ️ Все значения соответствуют шаблону (возможно, тип данных или пустые значения)")
                        except Exception as e:
                            st.warning(f"⚠️ Не удалось показать примеры: `{e}`")

                    st.divider()

            # ── 🧹 СТРАТЕГИИ ОБРАБОТКИ НАРУШЕНИЙ ─────────────────────
            with st.expander("🧹 Стратегии обработки нарушений форматов", expanded=True):
                # Выбор стратегии
                strategy = st.radio(
                    "Выберите действие:",
                    [
                        "Заменить на NaN (удалить значения)",
                        "Заменить на шаблонное значение (умное)",
                        "Применить regex-замену (нормализация)",
                        "Только отметить флагом (не менять данные)"
                    ],
                    key="format_fix_strategy",
                    horizontal=False
                )

                # Динамические подсказки
                if "NaN" in strategy:
                    st.warning("⚠️ Нарушающие значения будут заменены на `NaN` (пропущенные)")
                elif "шаблонное" in strategy:
                    st.info("🔄 **Умная замена:** email → `unknown@example.com`, телефон → `+0000000000`, остальное → `N/A`")
                elif "regex-замену" in strategy:
                    st.info("✏️ **Нормализация:** `strip()` + `lower()` + удаление спецсимволов")
                elif "флагом" in strategy:
                    st.info("🚩 Будет добавлена колонка `{имя}_format_valid` с True/False")

                # Кнопка применения стратегии
                if st.button("💾 Применить стратегию к данным", type="primary", key="btn_apply_format_fix"):
                    try:
                        import numpy as np
                        df_work = st.session_state.df_pattern_work.copy()
                        cols_fixed = 0

                        for v in real_violations:
                            col_name = v.get("Колонка")
                            pattern = v.get("pattern", ".*")

                            if not col_name or col_name not in df_work.columns:
                                continue

                            # Создаём маску нарушений для текущей колонки
                            mask_invalid = ~df_work[col_name].astype(str).str.fullmatch(pattern, na=False)
                            count_invalid = mask_invalid.sum()

                            if count_invalid == 0:
                                continue

                            # ── ПРИМЕНЕНИЕ СТРАТЕГИИ ──────────────────
                            if "Заменить на NaN" in strategy:
                                df_work.loc[mask_invalid, col_name] = np.nan
                                st.success(f"✅ `{col_name}`: **{count_invalid}** значений → `NaN`")

                            elif "шаблонное значение" in strategy:
                                col_lower = col_name.lower()
                                # Умная замена в зависимости от типа колонки
                                if "email" in col_lower:
                                    replacement = "unknown@example.com"
                                elif "phone" in col_lower or "телефон" in col_lower or "tel" in col_lower:
                                    replacement = "+0000000000"
                                elif "date" in col_lower or "дата" in col_lower:
                                    replacement = pd.NaT
                                elif df_work[col_name].dtype in ["int64", "float64"]:
                                    replacement = df_work[col_name].median()
                                else:
                                    replacement = "N/A"
                                df_work.loc[mask_invalid, col_name] = replacement
                                st.success(f"✅ `{col_name}`: **{count_invalid}** значений → `{replacement}`")

                            elif "regex-замену" in strategy:
                                # Нормализация ТОЛЬКО для нарушителей
                                invalid_mask = mask_invalid & df_work[col_name].notna()
                                if invalid_mask.any():
                                    df_work.loc[invalid_mask, col_name] = (
                                        df_work.loc[invalid_mask, col_name]
                                        .astype(str)
                                        .str.strip()                    # Убрать пробелы
                                        .str.lower()                    # Нижний регистр
                                        .str.replace(r"[^\w\s@.\-]", "", regex=True)  # Убрать спецсимволы
                                        .str.replace(r"\s+", " ", regex=True)         # Убрать множ. пробелы
                                    )
                                    st.success(f"✅ `{col_name}`: нормализовано **{invalid_mask.sum()}** значений")

                            elif "флагом" in strategy:
                                # Добавить колонку-индикатор
                                flag_col = f"{col_name}_format_valid"
                                df_work[flag_col] = ~mask_invalid
                                st.success(f"✅ `{col_name}`: добавлен флаг `{flag_col}`")

                            cols_fixed += 1

                        # ГЛАВНОЕ: Сохраняем изменения в session_state
                        st.session_state.df_pattern_work = df_work
                        st.session_state.df = df_work  # Синхронизация с основным датафреймом!
                        st.session_state.validation_ready = False  # Сброс валидации
                        st.session_state.pattern_results = []  # Сброс старых результатов форматов

                        st.divider()
                        st.success(f"✅ **Стратегия применена!** Обработано колонок: **{cols_fixed}**")
                        st.info("🔄 **Следующий шаг:** Нажмите **Запустить валидацию** сверху для проверки результата")

                        # Кнопка быстрого перезапуска
                        if st.button("🔄 Перезапустить валидацию сейчас", type="secondary", key="btn_rerun_validation_fmt"):
                            st.rerun()

                    except Exception as e:
                        st.error(f"❌ Ошибка при применении стратегии: `{e}`")
                        import traceback
                        with st.expander("🔍 Stack trace (для разработчика)"):
                            st.code(traceback.format_exc(), language="python")

                # ──  ЭКСПОРТ РЕЗУЛЬТАТА ───────────────────────────
                st.divider()
                st.download_button(
                    label="💾 Скачать результат (очищенные данные, CSV)",
                    data=df_work.to_csv(index=False, encoding="utf-8-sig"),
                    file_name=f"format_fixed_{st.session_state.get('original_filename', 'data').rsplit('.', 1)[0]}.csv",
                    mime="text/csv",
                    type="secondary",
                    key="btn_export_format_fixed"
                )

        # ======================================================================================
        # ── 3. ПРОВЕРКА ДИАПАЗОНОВ ЗНАЧЕНИЙ (с интерактивным пайплайном + кнопка сохранения) ──

        # 1. Загрузка правил диапазонов из YAML
        range_rules_config = rules.get("ranges", [])
        range_rule_bounds = {}      # col -> (min, max)
        range_violation_masks = {}  # col -> boolean mask
        range_results = []

        num_cols = df.select_dtypes(include='number').columns.tolist()

        for col in num_cols:
            col_lower = col.lower()
            matched_rule = None

            # Ищем правило по частичному совпадению ключевых слов из YAML
            for rule in range_rules_config:
                keywords = rule.get("keywords", [])
                if any(kw in col_lower for kw in keywords):
                    matched_rule = (rule.get("min"), rule.get("max"))
                    break

            if matched_rule:
                min_val, max_val = matched_rule
                range_rule_bounds[col] = (min_val, max_val)

                # Создаем маску нарушений для всей колонки
                mask = pd.Series(False, index=df.index)
                if min_val is not None: mask |= (df[col] < min_val)
                if max_val is not None: mask |= (df[col] > max_val)

                if mask.any():
                    range_violation_masks[col] = mask
                    violations = mask.sum()
                    range_results.append({
                        "Колонка": col,
                        "Правило": f"{min_val if min_val is not None else '-∞'} < x < {max_val if max_val is not None else '∞'}",
                        "Нарушений": int(violations),
                        "% брака": f"{(violations / len(df) * 100):.2f}%",
                        "Min факт": df[col].min(),
                        "Max факт": df[col].max()
                    })

        range_issues = len(range_results) > 0
        range_violations = sum(r.get('Нарушений', 0) for r in range_results)

        # 3. Карточка-резюме (Лаконичная версия)
        make_card(
            " Проверка диапазонов значений (value ranges)",
            has_issues=range_issues,
            methodology="**◻️ Метрики:** Доля значений вне границ `% invalid`, фактические `min/max`.  \n"
            "**◻️ Алгоритм:** Векторные маски `(df[col] < min) | (df[col] > max)` → агрегация нарушений.  \n"
            "**◻️ Влияние на TS:** Отрицательные цены ломают `log()`, проценты >100 искажают нормализацию.  \n"
            "**◻️ Описание:** Модуль анализирует числовые столбцы на соответствие границам допустимых значений (например, для цен, возраста и объемов < 0 или процентов > 100), автоматически определяя применимые правила по ключевым словам в названиях колонок. " \
            "Значения, нарушающие установленные пределы, зафиксируются и автоматически раскроется секция для их дальнейшего исправления или удаления.",
            result_text="✅ Все значения в допустимых диапазонах" if not range_issues else f"⚠️ Найдено {range_violations} нарушений в {len(range_results)} колонках",
            result_status="✅" if not range_issues else "⚠️",
            fix_action=None
        )

        # 4. 🔽 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при проблемах)
        if range_issues:
            with st.expander("Полный пайплайн обработки нарушений диапазонов", expanded=True):
                st.markdown("###  Таблица нарушений с ручной корректировкой")

                # Инициализация состояний
                if "df_range_work" not in st.session_state:
                    st.session_state.df_range_work = df.copy()

                df_work = st.session_state.df_range_work

                # Формируем общую маску нарушений
                combined_mask = pd.Series(False, index=df_work.index)
                for mask in range_violation_masks.values():
                    combined_mask |= mask

                # Фильтр отображения (автоматический, как в других блоках)
                view_filter = st.radio(
                    "Фильтр строк:",
                    ["⚠️ Только с нарушениями", "✅ Показать всё"],
                    horizontal=True,
                    key="range_view_filter"
                )

                if view_filter == "⚠️ Только с нарушениями":
                    df_view = df_work[combined_mask].copy() if combined_mask.any() else df_work.iloc[:0].copy()
                else:
                    df_view = df_work.copy()

                # Добавляем статус-колонку
                df_view = df_view.copy()
                df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Нарушение" if combined_mask.loc[idx] else "🟢 Норма"))

                # Интерактивная таблица (Ручная правка)
                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=300,  # Фиксированная высота для стабильности
                    num_rows="dynamic",
                    disabled=['_STATUS'],
                    key="range_editor",
                    column_config={
                        "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                    }
                )

                # 💾 КНОПКА СОХРАНЕНИЯ (Унифицирована с пропусками и выбросами)
                c_save1, c_save2 = st.columns([4, 1])
                with c_save1:
                    st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                with c_save2:
                    if st.button("💾 Сохранить", key="btn_save_manual_range", use_container_width=True):
                        if '_STATUS' in edited_df.columns:
                            edited_df = edited_df.drop(columns=['_STATUS'])
                        df_work.update(edited_df)
                        st.session_state.df_range_work = df_work
                        st.session_state.df = df_work  # Синхронизируем с основным датасетом
                        st.session_state.validation_ready = False
                        st.toast("✅ Правки сохранены!", icon="✅")
                        st.rerun()

                st.divider()

                # Панель стратегий (Автоматическая обработка)
                st.markdown("### 🧹 Стратегии обработки диапазонов")
                c1, c2 = st.columns([2, 1])
                with c1:
                    range_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Кэпировать до границ правил (min/max)",
                        "Заменить на медиану (числовые) / моду",
                        "Удалить строки с нарушениями",
                        "Заменить на 0 или NaN",
                        "Только отметить флагом (не менять данные)"],
                        key="range_fill_strategy"
                    )
                    if "Кэпировать" in range_strategy:
                        st.warning("⚠️ Выбросы будут заменены на границы: нижний на min, верхний на max")
                    elif "медиану" in range_strategy:
                        st.warning("⚠️ Все нарушения будут заменены на медиану.")
                    elif "Удалить" in range_strategy:
                        st.warning(f"⚠️ Будет удалено **{combined_mask.sum()} строк** ({combined_mask.sum()/len(df_work)*100:.1f}% данных).")
                    elif "0️⃣" in range_strategy:
                        st.warning("⚠️ Все нарушения будут заменены на 0 или NaN.")
                with c2:
                    # Пустая колонка для выравнивания
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка запуска превью (унифицирована)
                if "show_range_preview" not in st.session_state:
                    st.session_state.show_range_preview = False

                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button("📊 Показать прогноз влияния на статистики", type="secondary", use_container_width=True, key="btn_show_range_preview"):
                    st.session_state.show_range_preview = True
                    st.rerun()

                # Блок превью
                if st.session_state.show_range_preview:
                    strategy = range_strategy
                    st.markdown("##### 📈 Прогноз влияния на статистику:")

                    df_preview = df_work.copy()
                    cols_to_fix = list(range_rule_bounds.keys())

                    if "Кэпировать" in strategy:
                        for col in cols_to_fix:
                            # 🔧 ИСПРАВЛЕНИЕ: Приводим к float перед клипированием, чтобы избежать ошибки int64
                            df_preview[col] = df_preview[col].astype(float)
                            min_v, max_v = range_rule_bounds.get(col, (None, None))
                            if min_v is not None: df_preview[col] = df_preview[col].clip(lower=min_v)
                            if max_v is not None: df_preview[col] = df_preview[col].clip(upper=max_v)
                        note = "(кэпирование по границам)"
                    elif "медиану" in strategy:
                        for col in cols_to_fix:
                            if df_preview[col].dtype in ['int64', 'float64']:
                                df_preview[col] = df_preview[col].fillna(df_preview[col].median())
                        note = "(замена медианой)"
                    elif "Удалить" in strategy:
                        df_preview = df_preview[~combined_mask].reset_index(drop=True)
                        note = "(удаление строк)"
                    elif "0️⃣" in strategy:
                        for col in cols_to_fix:
                            df_preview[col] = df_preview[col].fillna(0)
                        note = "(замена на 0)"
                    else:
                        note = "(без изменений)"

                    # Метрики (4 колонки)
                    c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                    c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}", delta=f"{len(df_preview)-len(df_work):+}")

                    if cols_to_fix:
                        col = cols_to_fix[0]
                        def safe_stat(d, c, f):
                            return f(d[c]) if not d.empty and c in d.columns and d[c].notna().any() else 0.0
                        m_b, s_b, d_b = safe_stat(df_work, col, np.mean), safe_stat(df_work, col, np.std), safe_stat(df_work, col, np.median)
                        m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)
                        fmt = lambda x: f"{x:.2f}" if pd.notnull(x) and x != 0.0 else "N/A"
                        delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"
                        c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                        c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                        c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

                    # Кнопки подтверждения
                    st.divider()
                    c_ok, c_cancel = st.columns(2)
                    with c_ok:
                        if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_range"):
                            st.session_state.df_after_fixes = df_preview.copy()
                            st.session_state.df = df_preview
                            st.session_state.validation_ready = False
                            st.session_state.show_range_preview = False
                            st.success("✅ Стратегия применена! Перезапустите валидацию.")
                            st.rerun()
                    with c_cancel:
                        if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_range"):
                            st.session_state.show_range_preview = False
                            st.rerun()

        # ───────────────────────────────────────────────────────────
        # 4. ПРОВЕРКА СОГЛАСОВАННОСТИ (с интерактивным пайплайном)
        # ───────────────────────────────────────────────────────────

        # 4. Согласованность
        consistency_results_local = st.session_state.val_results.get("consistency", [])
        #  ИСПРАВЛЕНИЕ: Считаем именно нарушения, а не просто наличие результатов проверки
        consistency_violations = sum(r.get('Нарушений', 0) for r in consistency_results_local)
        consistency_issues = consistency_violations > 0

        make_card(" Проверка логики и хронологии (consistency)", consistency_issues,
            "**◻️ Метрики:** `% invalid = violations/valid`, кол-во логических противоречий.  \n"
            "**◻️ Алгоритм:** Векторизованные pandas-маски `(df[A] OP df[B])` → расчет % нарушений.  \n"
            "**◻️ Влияние на TS:** Несоответствия в хронологии искажают лаги, ломают `cumsum` и VAR/VECM.  \n"
            "**◻️ Описание:** Модуль автоматически выявляет логические и временные противоречия в данных, " \
            "сверяя значения между связанными колонками и хронологией событий на основе заданных бизнес-правил (прибыль больше выручки, потребление на освещение больше общего потребления, записи идут не по хронологии — после 10:05:00 идёт 10:03:00 и т.д.). " \
            "При обнаружении нарушений система рассчитает долю брака и предложит стратегии по исправлению аномалий.",
            "✅ Все бизнес-правила соблюдены" if not consistency_issues else f"️ Найдено {consistency_violations} нарушений",
            "✅" if not consistency_issues else "⚠️", None)

        # ── 🔧 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при проблемах) ──
        if consistency_issues:
            with st.expander("🔧 Полный пайплайн обработки нарушений согласованности", expanded=consistency_violations > 0):
                st.markdown("### 📋 Таблица нарушений с ручной корректировкой")

                # Инициализация состояний
                if "df_consistency_work" not in st.session_state:
                    st.session_state.df_consistency_work = df.copy()
                df_work = st.session_state.df_consistency_work

                # Формируем общую маску нарушений из словаря масок
                combined_mask = pd.Series(False, index=df_work.index)
                for mask in consistency_masks.values():
                    combined_mask |= mask

                # Фильтр отображения
                view_filter = st.radio(
                    "Фильтр строк:",
                    ["⚠️ Только с нарушениями", "✅ Показать всё"],
                    horizontal=True,
                    key="consistency_view_filter"
                )

                if view_filter == "⚠️ Только с нарушениями":
                    df_view = df_work[combined_mask].copy() if combined_mask.any() else df_work.iloc[:0].copy()
                else:
                    df_view = df_work.copy()

                # Добавляем статус-колонку
                df_view = df_view.copy()
                df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Нарушение" if combined_mask.loc[idx] else "🟢 Норма"))

                # Интерактивная таблица (Ручная правка)
                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=300,
                    num_rows="dynamic",
                    disabled=['_STATUS'],
                    key="consistency_editor",
                    column_config={
                        "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                    }
                )

                # 💾 КНОПКА СОХРАНЕНИЯ
                c_save1, c_save2 = st.columns([4, 1])
                with c_save1:
                    st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                with c_save2:
                    if st.button("💾 Сохранить", key="btn_save_manual_consistency", use_container_width=True):
                        if '_STATUS' in edited_df.columns:
                            edited_df = edited_df.drop(columns=['_STATUS'])
                        df_work.update(edited_df)
                        st.session_state.df_consistency_work = df_work
                        st.session_state.df = df_work  # Синхронизируем с основным датасетом
                        st.session_state.validation_ready = False
                        st.toast("✅ Правки сохранены!", icon="✅")
                        st.rerun()

                st.divider()

                # Панель стратегий (Автоматическая обработка)
                st.markdown("### 🧹 Стратегии обработки нарушений согласованности")
                c1, c2 = st.columns([2, 1])
                with c1:
                    consistency_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Удалить строки с нарушениями",
                        "Заменить на медиану (числовые) / моду",
                        "Исправить хронологию (сортировка по дате)",
                        "Заменить на 0 или NaN",
                        "Только отметить флагом (не менять данные)"],
                        key="consistency_fill_strategy"
                    )
                    if "Удалить" in consistency_strategy:
                        st.warning(f"⚠️ Будет удалено **{combined_mask.sum()} строк** ({combined_mask.sum()/len(df_work)*100:.1f}% данных).")
                    elif "медиану" in consistency_strategy:
                        st.warning("⚠️ Нарушающие значения будут заменены на медиану/моду.")
                    elif "хронологию" in consistency_strategy:
                        st.info("ℹ️ Данные будут отсортированы по временной колонке для исправления хронологии.")
                    elif "0️⃣" in consistency_strategy:
                        st.warning("⚠️ Все нарушения будут заменены на 0 или NaN.")
                with c2:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка запуска превью
                if "show_consistency_preview" not in st.session_state:
                    st.session_state.show_consistency_preview = False

                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button("📊 Показать прогноз влияния на статистики", type="secondary", use_container_width=True, key="btn_show_consistency_preview"):
                    st.session_state.show_consistency_preview = True
                    st.rerun()

                # Блок превью
                if st.session_state.show_consistency_preview:
                    strategy = consistency_strategy
                    st.markdown("##### 📈 Прогноз влияния на статистику:")

                    df_preview = df_work.copy()
                    num_cols_to_fix = df_preview.select_dtypes(include=['number']).columns.tolist()

                    if "Удалить" in strategy:
                        df_preview = df_preview[~combined_mask].reset_index(drop=True)
                        note = "(удаление строк)"
                    elif "медиану" in strategy:
                        for col in num_cols_to_fix:
                            if df_preview[col].isnull().any() or combined_mask.any():
                                df_preview.loc[combined_mask, col] = df_preview[col].median()
                        note = "(замена медианой)"
                    elif "хронологию" in strategy:
                        date_cols = [c for c in df_preview.columns if 'date' in c.lower() or 'дата' in c.lower()]
                        if date_cols:
                            df_preview = df_preview.sort_values(date_cols[0]).reset_index(drop=True)
                            note = f"(сортировка по {date_cols[0]})"
                        else:
                            st.warning("⚠️ Временные колонки не найдены")
                            note = "(без изменений)"
                    elif "0️⃣" in strategy:
                        for col in num_cols_to_fix:
                            df_preview.loc[combined_mask, col] = 0
                        note = "(замена на 0)"
                    else:
                        note = "(без изменений)"

                    # Метрики (4 колонки)
                    c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                    c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}", delta=f"{len(df_preview)-len(df_work):+}")

                    if num_cols_to_fix:
                        col = num_cols_to_fix[0]
                        def safe_stat(d, c, f):
                            return f(d[c]) if not d.empty and c in d.columns and d[c].notna().any() else 0.0
                        m_b, s_b, d_b = safe_stat(df_work, col, np.mean), safe_stat(df_work, col, np.std), safe_stat(df_work, col, np.median)
                        m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)
                        fmt = lambda x: f"{x:.2f}" if pd.notnull(x) and x != 0.0 else "N/A"
                        delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"
                        c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                        c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                        c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

                    # Кнопки подтверждения
                    st.divider()
                    c_ok, c_cancel = st.columns(2)
                    with c_ok:
                        if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_consistency"):
                            st.session_state.df_after_fixes = df_preview.copy()
                            st.session_state.df = df_preview
                            st.session_state.validation_ready = False
                            st.session_state.show_consistency_preview = False
                            st.success("✅ Стратегия применена! Перезапустите валидацию.")
                            st.rerun()
                    with c_cancel:
                        if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_consistency"):
                            st.session_state.show_consistency_preview = False
                            st.rerun()

        # ───────────────────────────────────────────────────────────
        # 5. ПРОВЕРКА УНИКАЛЬНОСТИ (с интерактивным пайплайном)
        # ───────────────────────────────────────────────────────────

        uniqueness_results_local = locals().get('uniqueness_results', [])
        uniqueness_issues = len(uniqueness_results_local) > 0
        uniqueness_dup_count = sum(r.get('Дубликатов', 0) for r in uniqueness_results_local) if uniqueness_issues else 0

        make_card(" Проверка уникальности (uniqueness checks)", uniqueness_issues,
            "**◻️ Метрики:** `% dup = duplicated/total`, дубли временных меток.  \n"
            "**◻️ Алгоритм:** `df.duplicated()`, аудит `datetime.duplicated()`.  \n"
            "**◻️ Влияние на TS:** Дубли дат ломают `DatetimeIndex`, `resample()`, STL и ARIMA.  \n"
            "**◻️ Описание:** Модуль проводит проверку на двух уровнях:  \n"
            " 1. Полные дубликаты строк. Используется метод df.duplicated() с параметром keep=False, "
            "помечаются все копии дублирующихся строк, возвращает булеву маску для последующей обработки.  \n"
            " 2. Дубликаты по ключу времени (для TS-режима) — "
            "проверяет уникальность по колонке с датой (критически важно для построения DatetimeIndex). Если активен режим панели данных — проверка по (Country, Date).",
            "✅ Дубликаты отсутствуют" if not uniqueness_issues else f"⚠️ Найдено {uniqueness_dup_count} дубликатов",
            "✅" if not uniqueness_issues else "⚠️", None)

        # ── 🔧 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при проблемах) ──
        if uniqueness_issues:
            with st.expander("🔧 Полный пайплайн обработки дубликатов", expanded=True):
                st.markdown("### 📋 Таблица дубликатов с ручной корректировкой")

                # Инициализация состояний
                if "df_uniqueness_work" not in st.session_state:
                    st.session_state.df_uniqueness_work = df.copy()
                df_work = st.session_state.df_uniqueness_work

                # Поиск дубликатов
                if st.session_state.col_types.get("date"):
                    date_col = st.session_state.col_types["date"][0]
                    # Для временных рядов: дубли по дате
                    duplicate_mask = df_work.duplicated(subset=[date_col], keep=False)
                    duplicate_indices = df_work[duplicate_mask].index.tolist()
                else:
                    # Общий случай: полные дубликаты строк
                    duplicate_mask = df_work.duplicated(keep=False)
                    duplicate_indices = df_work[duplicate_mask].index.tolist()

                # Фильтр отображения
                view_filter = st.radio(
                    "Фильтр строк:",
                    ["⚠️ Только дубликаты", "✅ Показать всё"],
                    horizontal=True,
                    key="uniqueness_view_filter"
                )

                if view_filter == "⚠️ Только дубликаты":
                    df_view = df_work.loc[duplicate_indices].copy() if duplicate_indices else df_work.iloc[:0].copy()
                else:
                    df_view = df_work.copy()

                # Добавляем статус-колонку
                df_view = df_view.copy()
                df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Дубликат" if idx in duplicate_indices else "🟢 Уникально"))

                # Интерактивная таблица (Ручная правка)
                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=300,
                    num_rows="dynamic",
                    disabled=['_STATUS'],
                    key="uniqueness_editor",
                    column_config={
                        "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                    }
                )

                # 💾 КНОПКА СОХРАНЕНИЯ
                c_save1, c_save2 = st.columns([4, 1])
                with c_save1:
                    st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                with c_save2:
                    if st.button("💾 Сохранить", key="btn_save_manual_uniqueness", use_container_width=True):
                        if '_STATUS' in edited_df.columns:
                            edited_df = edited_df.drop(columns=['_STATUS'])
                        df_work.update(edited_df)
                        st.session_state.df_uniqueness_work = df_work
                        st.session_state.df = df_work
                        st.session_state.validation_ready = False
                        st.toast("✅ Правки сохранены!", icon="✅")
                        st.rerun()

                st.divider()

                # Панель стратегий (Автоматическая обработка)
                st.markdown("### 🧹 Стратегии обработки дубликатов")
                c1, c2 = st.columns([2, 1])
                with c1:
                    uniqueness_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Удалить дубликаты (оставить первый)",
                        "Удалить дубликаты (оставить последний)",
                        "Удалить все дубликаты полностью",
                        "Агрегировать дубликаты (mean/sum)",
                        "Только отметить флагом (не менять данные)"],
                        key="uniqueness_fill_strategy"
                    )
                    if "Удалить дубликаты" in uniqueness_strategy:
                        st.warning(f"⚠️ Будет удалено **{len(duplicate_indices)} строк** с дубликатами ({len(duplicate_indices)/len(df_work)*100:.1f}% данных).")
                    elif "Агрегировать" in uniqueness_strategy:
                        st.info("ℹ️ Дубликаты будут агрегированы через mean (числовые) / first (категории).")
                    elif "флагом" in uniqueness_strategy:
                        st.info("🚩 Будет добавлена колонка `_is_duplicate` с True/False")
                with c2:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка запуска превью
                if "show_uniqueness_preview" not in st.session_state:
                    st.session_state.show_uniqueness_preview = False

                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button("📊 Показать прогноз влияния на статистики", type="secondary", use_container_width=True, key="btn_show_uniqueness_preview"):
                    st.session_state.show_uniqueness_preview = True
                    st.rerun()

                # Блок превью
                if st.session_state.show_uniqueness_preview:
                    strategy = uniqueness_strategy
                    st.markdown("##### 📈 Прогноз влияния на статистику:")

                    df_preview = df_work.copy()
                    num_cols_to_check = df_preview.select_dtypes(include=['number']).columns.tolist()

                    if "Удалить дубликаты (оставить первый)" in strategy:
                        if st.session_state.col_types.get("date"):
                            df_preview = df_preview.drop_duplicates(subset=[st.session_state.col_types["date"][0]], keep='first')
                        else:
                            df_preview = df_preview.drop_duplicates(keep='first')
                        note = "(удаление, keep='first')"
                    elif "Удалить дубликаты (оставить последний)" in strategy:
                        if st.session_state.col_types.get("date"):
                            df_preview = df_preview.drop_duplicates(subset=[st.session_state.col_types["date"][0]], keep='last')
                        else:
                            df_preview = df_preview.drop_duplicates(keep='last')
                        note = "(удаление, keep='last')"
                    elif "Удалить все дубликаты полностью" in strategy:
                        if st.session_state.col_types.get("date"):
                            df_preview = df_preview.drop_duplicates(subset=[st.session_state.col_types["date"][0]], keep=False)
                        else:
                            df_preview = df_preview.drop_duplicates(keep=False)
                        note = "(удаление всех дубликатов)"
                    elif "Агрегировать дубликаты" in strategy:
                        if st.session_state.col_types.get("date"):
                            date_col = st.session_state.col_types["date"][0]
                            # Агрегация: mean для числовых, first для категориальных
                            agg_dict = {col: 'mean' if col in num_cols_to_check else 'first' for col in df_preview.columns if col != date_col}
                            df_preview = df_preview.groupby(date_col).agg(agg_dict).reset_index()
                        note = "(агрегация)"
                    else:
                        note = "(без изменений)"

                    # Метрики (4 колонки)
                    c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                    c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}", delta=f"{len(df_preview)-len(df_work):+}")

                    if num_cols_to_check:
                        col = num_cols_to_check[0]
                        def safe_stat(d, c, f):
                            return f(d[c]) if not d.empty and c in d.columns and d[c].notna().any() else 0.0
                        m_b, s_b, d_b = safe_stat(df_work, col, np.mean), safe_stat(df_work, col, np.std), safe_stat(df_work, col, np.median)
                        m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)
                        fmt = lambda x: f"{x:.2f}" if pd.notnull(x) and x != 0.0 else "N/A"
                        delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"
                        c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                        c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                        c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

                    # Кнопки подтверждения
                    st.divider()
                    c_ok, c_cancel = st.columns(2)
                    with c_ok:
                        if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_uniqueness"):
                            st.session_state.df_after_fixes = df_preview.copy()
                            st.session_state.df = df_preview
                            st.session_state.validation_ready = False
                            st.session_state.show_uniqueness_preview = False
                            st.success("✅ Стратегия применена! Перезапустите валидацию.")
                            st.rerun()
                    with c_cancel:
                        if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_uniqueness"):
                            st.session_state.show_uniqueness_preview = False
                            st.rerun()

        # ───────────────────────────────────────────────────────────
        # 6. ПРОВЕРКА ПРИНАДЛЕЖНОСТИ К НАБОРУ (inclusion/lookup)
        # ───────────────────────────────────────────────────────────

        # 6. Inclusion/Lookup
        inclusion_results_local = locals().get('inclusion_results', [])
        inclusion_issues = len(inclusion_results_local) > 0
        inclusion_viol_count = sum(r.get('Нарушений', 0) for r in inclusion_results_local) if inclusion_issues else 0

        make_card(" Проверка принадлежности к набору (inclusion/lookup)", inclusion_issues,
            "**◻️ Метрики:** `% invalid = not_in_set/total`, число значений вне справочника.  \n"
            "**◻️ Алгоритм:** `df[col].isin(allowed_values)`, проверка по внешним справочникам.  \n"
            "**◻️ Влияние на TS:** Неизвестные категории ломают `groupby`, `pivot`, `resample` по категориям.  \n"
            "**◻️ Описание:** Модуль проверяет, что значения в категориальных колонках принадлежат утверждённому справочнику (домену). Для типовых датасетов используются готовые шаблоны YAML.  \n"
            "Для новых типов данных платформа автогенерирует базовые правила, аналитик их корректирует. Выявляются опечатки, устаревшие коды, несуществующие категории. Критично для агрегации и фильтрации данных.",
            "✅ Все значения соответствуют справочникам" if not inclusion_issues else f"⚠️ Найдено {inclusion_viol_count} нарушений",
            "✅" if not inclusion_issues else "⚠️", None)

        # ── 🔧 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при проблемах) ──
        if inclusion_issues:
            with st.expander("🔧 Полный пайплайн обработки нарушений принадлежности", expanded=True):
                st.markdown("### 📋 Таблица нарушений с ручной корректировкой")

                # Инициализация состояний
                if "df_inclusion_work" not in st.session_state:
                    st.session_state.df_inclusion_work = df.copy()
                df_work = st.session_state.df_inclusion_work

                # Поиск нарушений inclusion
                inclusion_violations = []
                for col, allowed_vals in inclusion_rules.items():  # inclusion_rules из rules.yaml
                    if col in df_work.columns:
                        invalid_mask = ~df_work[col].isin(allowed_vals) & df_work[col].notna()
                        if invalid_mask.any():
                            invalid_values = df_work.loc[invalid_mask, col].unique()
                            inclusion_violations.append({
                                'column': col,
                                'invalid_values': invalid_values,
                                'count': invalid_mask.sum(),
                                'mask': invalid_mask
                            })

                # Общая маска всех нарушений
                combined_mask = pd.Series(False, index=df_work.index)
                for v in inclusion_violations:
                    combined_mask |= v['mask']

                # Фильтр отображения
                view_filter = st.radio(
                    "Фильтр строк:",
                    ["⚠️ Только с нарушениями", "✅ Показать всё"],
                    horizontal=True,
                    key="inclusion_view_filter"
                )

                if view_filter == "⚠️ Только с нарушениями":
                    df_view = df_work[combined_mask].copy() if combined_mask.any() else df_work.iloc[:0].copy()
                else:
                    df_view = df_work.copy()

                # Добавляем статус-колонку
                df_view = df_view.copy()
                df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Нарушение" if combined_mask.loc[idx] else "🟢 Норма"))

                # Интерактивная таблица (Ручная правка)
                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=300,
                    num_rows="dynamic",
                    disabled=['_STATUS'],
                    key="inclusion_editor",
                    column_config={
                        "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                    }
                )

                # 💾 КНОПКА СОХРАНЕНИЯ
                c_save1, c_save2 = st.columns([4, 1])
                with c_save1:
                    st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                with c_save2:
                    if st.button("💾 Сохранить", key="btn_save_manual_inclusion", use_container_width=True):
                        if '_STATUS' in edited_df.columns:
                            edited_df = edited_df.drop(columns=['_STATUS'])
                        df_work.update(edited_df)
                        st.session_state.df_inclusion_work = df_work
                        st.session_state.df = df_work
                        st.session_state.validation_ready = False
                        st.toast("✅ Правки сохранены!", icon="✅")
                        st.rerun()

                st.divider()

                # Панель стратегий (Автоматическая обработка)
                st.markdown("### 🧹 Стратегии обработки нарушений принадлежности")
                c1, c2 = st.columns([2, 1])
                with c1:
                    inclusion_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Заменить на наиболее частое допустимое значение",
                        "Заменить на NaN (удалить значения)",
                        "Удалить строки с нарушениями",
                        "Заменить на значение по умолчанию (из справочника)",
                        "Только отметить флагом (не менять данные)"],
                        key="inclusion_fill_strategy"
                    )
                    if "наиболее частое" in inclusion_strategy:
                        st.info("ℹ️ Нарушающие значения будут заменены на моду (наиболее частое допустимое значение).")
                    elif "NaN" in inclusion_strategy:
                        st.warning("⚠️ Нарушающие значения будут заменены на `NaN` (пропущенные).")
                    elif "Удалить" in inclusion_strategy:
                        st.warning(f"⚠️ Будет удалено **{combined_mask.sum()} строк** ({combined_mask.sum()/len(df_work)*100:.1f}% данных).")
                    elif "по умолчанию" in inclusion_strategy:
                        st.info("ℹ️ Нарушения будут заменены на значение по умолчанию из справочника.")
                with c2:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка запуска превью
                if "show_inclusion_preview" not in st.session_state:
                    st.session_state.show_inclusion_preview = False

                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button("📊 Показать прогноз влияния на статистики", type="secondary", use_container_width=True, key="btn_show_inclusion_preview"):
                    st.session_state.show_inclusion_preview = True
                    st.rerun()

                # Блок превью
                if st.session_state.show_inclusion_preview:
                    strategy = inclusion_strategy
                    st.markdown("##### 📈 Прогноз влияния на статистику:")

                    df_preview = df_work.copy()
                    num_cols_to_check = df_preview.select_dtypes(include=['number']).columns.tolist()

                    if "наиболее частое" in strategy:
                        # Для каждой колонки с нарушениями заменяем на моду
                        for v in inclusion_violations:
                            col = v['column']
                            allowed = inclusion_rules.get(col, [])
                            mode_val = df_preview[df_preview[col].isin(allowed)][col].mode()
                            if len(mode_val) > 0:
                                df_preview.loc[v['mask'], col] = mode_val[0]
                        note = "(замена на моду)"
                    elif "NaN" in strategy:
                        for v in inclusion_violations:
                            df_preview.loc[v['mask'], v['column']] = np.nan
                        note = "(замена на NaN)"
                    elif "Удалить" in strategy:
                        df_preview = df_preview[~combined_mask].reset_index(drop=True)
                        note = "(удаление строк)"
                    elif "по умолчанию" in strategy:
                        for v in inclusion_violations:
                            col = v['column']
                            default_val = inclusion_defaults.get(col, np.nan)  # inclusion_defaults из конфига
                            df_preview.loc[v['mask'], col] = default_val
                        note = "(замена на default)"
                    else:
                        note = "(без изменений)"

                    # Метрики (4 колонки)
                    c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                    c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}", delta=f"{len(df_preview)-len(df_work):+}")

                    if num_cols_to_check:
                        col = num_cols_to_check[0]
                        def safe_stat(d, c, f):
                            return f(d[c]) if not d.empty and c in d.columns and d[c].notna().any() else 0.0
                        m_b, s_b, d_b = safe_stat(df_work, col, np.mean), safe_stat(df_work, col, np.std), safe_stat(df_work, col, np.median)
                        m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)
                        fmt = lambda x: f"{x:.2f}" if pd.notnull(x) and x != 0.0 else "N/A"
                        delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"
                        c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                        c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                        c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

                    # Кнопки подтверждения
                    st.divider()
                    c_ok, c_cancel = st.columns(2)
                    with c_ok:
                        if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_inclusion"):
                            st.session_state.df_after_fixes = df_preview.copy()
                            st.session_state.df = df_preview
                            st.session_state.validation_ready = False
                            st.session_state.show_inclusion_preview = False
                            st.success("✅ Стратегия применена! Перезапустите валидацию.")
                            st.rerun()
                    with c_cancel:
                        if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_inclusion"):
                            st.session_state.show_inclusion_preview = False
                            st.rerun()

        # ───────────────────────────────────────────────────────────
        # 7. ПРОВЕРКА ССЫЛОЧНОЙ ЦЕЛОСТНОСТИ (с интерактивным пайплайном)
        # ───────────────────────────────────────────────────────────

        # 7. Referential Integrity
        ref_results_local = locals().get('ref_results', [])
        ref_masks = locals().get('ref_masks', {})  # Маски нарушений (если есть)
        referential_issues = len(ref_results_local) > 0
        ref_violations = sum(r.get('Нарушений', 0) for r in ref_results_local) if referential_issues else 0

        make_card(" Проверка ссылочной целостности (referential integrity)", referential_issues,
            "**◻️ Метрики:** `% invalid = violations/valid`, кол-во 'сиротских' записей.  \n"
            "**◻️ Алгоритм:** Векторизованные маски `(df[A]==cond) & df[B].isna()`, проверка внешних ключей.  \n"
            "**◻️ Влияние на TS:** Нарушение связей искажает агрегацию, ломает VAR/VECM.  \n"
            "**◻️ Описание:** Модуль выявляет 'сиротские' записи — строки, где значение в дочерней колонке "
            "не имеет соответствия в родительской таблице/справочнике. Например, цена для несуществующей страны, "
            "операция без связанного клиента, или разрыв между `supply_id` и таблицей поставщиков. "
            "Нарушения ссылочной целостности критичны для JOIN-операций и панельного анализа.",
            "✅ Ссылочная целостность не нарушена" if not referential_issues else f"⚠️ Найдено {ref_violations} нарушений связей",
            "✅" if not referential_issues else "⚠️", None)

        # ── 🔧 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при проблемах) ──
        if referential_issues:
            with st.expander("🔧 Полный пайплайн обработки нарушений ссылочной целостности", expanded=True):
                st.markdown("### 📋 Таблица нарушений с ручной корректировкой")

                # Инициализация состояний
                if "df_referential_work" not in st.session_state:
                    st.session_state.df_referential_work = df.copy()
                df_work = st.session_state.df_referential_work

                # Формируем общую маску нарушений
                combined_mask = pd.Series(False, index=df_work.index)

                # Если есть готовые маски из validate_referential()
                if ref_masks:
                    for mask in ref_masks.values():
                        combined_mask |= mask
                else:
                    # Fallback: определяем нарушения из результатов
                    for r in ref_results_local:
                        col = r.get('Колонка') or r.get('child_column')
                        allowed_values = r.get('allowed_values', [])
                        if col and col in df_work.columns and allowed_values:
                            mask = ~df_work[col].isin(allowed_values) & df_work[col].notna()
                            combined_mask |= mask

                # Фильтр отображения
                view_filter = st.radio(
                    "Фильтр строк:",
                    ["⚠️ Только с нарушениями", "✅ Показать всё"],
                    horizontal=True,
                    key="referential_view_filter"
                )

                if view_filter == "⚠️ Только с нарушениями":
                    df_view = df_work[combined_mask].copy() if combined_mask.any() else df_work.iloc[:0].copy()
                else:
                    df_view = df_work.copy()

                # Добавляем статус-колонку
                df_view = df_view.copy()
                df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Сирота" if combined_mask.loc[idx] else "🟢 Норма"))

                # Интерактивная таблица (Ручная правка)
                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=300,
                    num_rows="dynamic",
                    disabled=['_STATUS'],
                    key="referential_editor",
                    column_config={
                        "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                    }
                )

                # 💾 КНОПКА СОХРАНЕНИЯ
                c_save1, c_save2 = st.columns([4, 1])
                with c_save1:
                    st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                with c_save2:
                    if st.button("💾 Сохранить", key="btn_save_manual_referential", use_container_width=True):
                        if '_STATUS' in edited_df.columns:
                            edited_df = edited_df.drop(columns=['_STATUS'])
                        df_work.update(edited_df)
                        st.session_state.df_referential_work = df_work
                        st.session_state.df = df_work
                        st.session_state.validation_ready = False
                        st.toast("✅ Правки сохранены!", icon="✅")
                        st.rerun()

                st.divider()

                # Панель стратегий (Автоматическая обработка)
                st.markdown("### 🧹 Стратегии обработки нарушений ссылочной целостности")
                c1, c2 = st.columns([2, 1])
                with c1:
                    ref_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Удалить сиротские записи",
                        "Заменить на значение по умолчанию (из справочника)",
                        "Заменить на наиболее частое валидное значение (мода)",
                        "Заменить на NaN (пометить как пропуск)",
                        "Только отметить флагом (не менять данные)"],
                        key="referential_fill_strategy"
                    )
                    if "Удалить" in ref_strategy:
                        st.warning(f"⚠️ Будет удалено **{combined_mask.sum()} строк** ({combined_mask.sum()/len(df_work)*100:.1f}% данных).")
                    elif "по умолчанию" in ref_strategy:
                        st.info("ℹ️ Нарушения будут заменены на значение по умолчанию из справочника (если задано).")
                    elif "моду" in ref_strategy:
                        st.info("ℹ️ Нарушения будут заменены на наиболее частое валидное значение в колонке.")
                    elif "NaN" in ref_strategy:
                        st.warning("⚠️ Нарушения будут заменены на `NaN` (потребует дальнейшей обработки пропусков).")
                    elif "флагом" in ref_strategy:
                        st.info("🚩 Будет добавлена колонка `_ref_valid` с True/False.")
                with c2:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка запуска превью
                if "show_referential_preview" not in st.session_state:
                    st.session_state.show_referential_preview = False

                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button("📊 Показать прогноз влияния на статистики", type="secondary", use_container_width=True, key="btn_show_referential_preview"):
                    st.session_state.show_referential_preview = True
                    st.rerun()

                # Блок превью
                if st.session_state.show_referential_preview:
                    strategy = ref_strategy
                    st.markdown("##### 📈 Прогноз влияния на статистику:")

                    df_preview = df_work.copy()
                    num_cols_to_check = df_preview.select_dtypes(include=['number']).columns.tolist()

                    if "Удалить" in strategy:
                        df_preview = df_preview[~combined_mask].reset_index(drop=True)
                        note = "(удаление сиротских записей)"
                    elif "по умолчанию" in strategy:
                        # Заменяем на значение по умолчанию из правил (если задано)
                        for r in ref_results_local:
                            col = r.get('Колонка') or r.get('child_column')
                            default_val = r.get('default_value', 'Unknown')
                            if col and col in df_preview.columns:
                                df_preview.loc[combined_mask, col] = default_val
                        note = "(замена на default)"
                    elif "моду" in strategy:
                        # Заменяем на моду валидных значений
                        for r in ref_results_local:
                            col = r.get('Колонка') or r.get('child_column')
                            allowed = r.get('allowed_values', [])
                            if col and col in df_preview.columns and allowed:
                                valid_data = df_preview[df_preview[col].isin(allowed)][col]
                                if not valid_data.empty:
                                    mode_val = valid_data.mode()
                                    if len(mode_val) > 0:
                                        df_preview.loc[combined_mask, col] = mode_val[0]
                        note = "(замена на моду)"
                    elif "NaN" in strategy:
                        for r in ref_results_local:
                            col = r.get('Колонка') or r.get('child_column')
                            if col and col in df_preview.columns:
                                df_preview.loc[combined_mask, col] = np.nan
                        note = "(замена на NaN)"
                    else:
                        note = "(без изменений)"

                    # Метрики (4 колонки)
                    c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                    c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}", delta=f"{len(df_preview)-len(df_work):+}")

                    if num_cols_to_check:
                        col = num_cols_to_check[0]
                        def safe_stat(d, c, f):
                            return f(d[c]) if not d.empty and c in d.columns and d[c].notna().any() else 0.0
                        m_b, s_b, d_b = safe_stat(df_work, col, np.mean), safe_stat(df_work, col, np.std), safe_stat(df_work, col, np.median)
                        m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)
                        fmt = lambda x: f"{x:.2f}" if pd.notnull(x) and x != 0.0 else "N/A"
                        delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"
                        c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                        c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                        c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

                    # Кнопки подтверждения
                    st.divider()
                    c_ok, c_cancel = st.columns(2)
                    with c_ok:
                        if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_referential"):
                            st.session_state.df_after_fixes = df_preview.copy()
                            st.session_state.df = df_preview
                            st.session_state.validation_ready = False
                            st.session_state.show_referential_preview = False
                            st.success("✅ Стратегия применена! Перезапустите валидацию.")
                            st.rerun()
                    with c_cancel:
                        if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_referential"):
                            st.session_state.show_referential_preview = False
                            st.rerun()

        # ───────────────────────────────────────────────────────────
        # 8. ПРОВЕРКА ЦЕЛОСТНОСТИ ТЕКСТА (text quality)
        # ───────────────────────────────────────────────────────────

        # 8. Text Quality
        text_results_local = locals().get('text_results', [])
        text_masks = locals().get('text_masks', {})  # Маски нарушений (если есть)
        text_issues = len(text_results_local) > 0
        text_problem_cols = len([r for r in text_results_local if r.get('Нарушений', 0) > 0])

        make_card(" Проверка целостности текста (text quality)", text_issues,
            "**◻️ Метрики:** `% issues = invalid/total`, наличие 'мусорных' символов, длина строк, кодировка.  \n"
            "**◻️ Алгоритм:** `str.len()`, Regex-паттерны (спецсимволы, нечитаемые символы, mixed case).  \n"
            "**◻️ Влияние на TS:** 'Грязный' текст фрагментирует категории при `groupby`, ломает справочники.  \n"
            "**◻️ Описание:** Модуль анализирует строковые колонки на наличие проблем качества: нечитаемых символов "
            "(последовательности типа '', '\\x00', 'ï¿½'), лишних пробелов, смешанного регистра, подозрительно коротких "
            "или длинных значений, а также технических артефактов кодировки. Критично для колонок Country, Category, "
            "Name — любые искажения приводят к дублированию категорий и ошибочной агрегации.",
            "✅ Качество текста соответствует стандартам" if not text_issues else f"⚠️ Найдено {text_problem_cols} текстовых колонок с проблемами",
            "✅" if not text_issues else "️", None)

        # ── ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при проблемах) ──
        if text_issues:
            with st.expander("🔧 Полный пайплайн обработки текстовых нарушений", expanded=True):
                st.markdown("### 📋 Таблица нарушений с ручной корректировкой")

                # Инициализация состояний
                if "df_text_work" not in st.session_state:
                    st.session_state.df_text_work = df.copy()
                df_work = st.session_state.df_text_work

                # Формируем общую маску нарушений
                combined_mask = pd.Series(False, index=df_work.index)

                if text_masks:
                    for mask in text_masks.values():
                        combined_mask |= mask
                else:
                    # Fallback: определяем нарушения по результатам проверки
                    text_cols = df_work.select_dtypes(include=['object', 'string']).columns.tolist()
                    for r in text_results_local:
                        col = r.get('Колонка')
                        if col and col in df_work.columns:
                            issue_type = r.get('Тип', '').lower()
                            if 'мусор' in issue_type or 'символ' in issue_type:
                                # Паттерн мусорных символов
                                mask = df_work[col].astype(str).str.contains(
                                    r'[\x00\x01\x02ï¿½\ufffd]', na=False, regex=True
                                )
                            elif 'длина' in issue_type or 'len' in issue_type:
                                # Слишком короткие или длинные значения
                                mask = (df_work[col].astype(str).str.len() < 1) | \
                                       (df_work[col].astype(str).str.len() > 200)
                            elif 'регистр' in issue_type or 'case' in issue_type:
                                # Смешанный регистр
                                mask = df_work[col].astype(str).str.contains(
                                    r'[A-ZА-ЯЁ]+[a-zа-яё]+[A-ZА-ЯЁ]', na=False, regex=True
                                )
                            else:
                                # Общий паттерн: non-printable chars
                                mask = df_work[col].astype(str).str.contains(
                                    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', na=False, regex=True
                                )
                            combined_mask |= mask

                # Фильтр отображения
                view_filter = st.radio(
                    "Фильтр строк:",
                    ["⚠️ Только с нарушениями", "✅ Показать всё"],
                    horizontal=True,
                    key="text_view_filter"
                )

                if view_filter == "⚠️ Только с нарушениями":
                    df_view = df_work[combined_mask].copy() if combined_mask.any() else df_work.iloc[:0].copy()
                else:
                    df_view = df_work.copy()

                # Добавляем статус-колонку
                df_view = df_view.copy()
                df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Проблема" if combined_mask.loc[idx] else "🟢 Норма"))

                # Интерактивная таблица (Ручная правка)
                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=300,
                    num_rows="dynamic",
                    disabled=['_STATUS'],
                    key="text_editor",
                    column_config={
                        "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                    }
                )

                # 💾 КНОПКА СОХРАНЕНИЯ
                c_save1, c_save2 = st.columns([4, 1])
                with c_save1:
                    st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                with c_save2:
                    if st.button("💾 Сохранить", key="btn_save_manual_text", use_container_width=True):
                        if '_STATUS' in edited_df.columns:
                            edited_df = edited_df.drop(columns=['_STATUS'])
                        df_work.update(edited_df)
                        st.session_state.df_text_work = df_work
                        st.session_state.df = df_work
                        st.session_state.validation_ready = False
                        st.toast("✅ Правки сохранены!", icon="✅")
                        st.rerun()

                st.divider()

                # Панель стратегий (Автоматическая обработка)
                st.markdown("### 🧹 Стратегии обработки текстовых нарушений")
                c1, c2 = st.columns([2, 1])
                with c1:
                    text_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Очистить спецсимволы и нормализовать (strip + lower)",
                        "Удалить строки с нарушениями текста",
                        "Заменить на NaN (пометить как пропуск)",
                        "Заменить на 'Неизвестно' (default для категорий)",
                        "Только отметить флагом (не менять данные)"],
                        key="text_fill_strategy"
                    )
                    if "Очистить" in text_strategy:
                        st.info("ℹ️ Будет выполнена нормализация: удаление спецсимволов, strip(), lower().")
                    elif "Удалить" in text_strategy:
                        st.warning(f"⚠️ Будет удалено **{combined_mask.sum()} строк** ({combined_mask.sum()/len(df_work)*100:.1f}% данных).")
                    elif "NaN" in text_strategy:
                        st.warning("⚠️ Нарушения будут заменены на `NaN` (потребует дальнейшей обработки пропусков).")
                    elif "Неизвестно" in text_strategy:
                        st.info("ℹ️ Нарушения будут заменены на строку `'Неизвестно'` (для категорий).")
                    elif "флагом" in text_strategy:
                        st.info("🚩 Будет добавлена колонка `_text_valid` с True/False.")
                with c2:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка запуска превью
                if "show_text_preview" not in st.session_state:
                    st.session_state.show_text_preview = False

                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button(" Показать прогноз влияния на текст", type="secondary", use_container_width=True, key="btn_show_text_preview"):
                    st.session_state.show_text_preview = True
                    st.rerun()

                # Блок превью (специфичный для текста)
                if st.session_state.show_text_preview:
                    strategy = text_strategy
                    st.markdown("##### 📈 Прогноз влияния на текстовые колонки:")

                    df_preview = df_work.copy()
                    text_cols = df_preview.select_dtypes(include=['object', 'string']).columns.tolist()

                    if "Очистить" in strategy:
                        for col in text_cols:
                            # Нормализация только проблемных значений
                            df_preview.loc[combined_mask, col] = (
                                df_preview.loc[combined_mask, col]
                                .astype(str)
                                .str.strip()
                                .str.lower()
                                .str.replace(r'[^\w\sа-яёa-z0-9\-]', '', regex=True)
                                .str.replace(r'\s+', ' ', regex=True)
                            )
                        note = "(нормализация текста)"
                    elif "Удалить" in strategy:
                        df_preview = df_preview[~combined_mask].reset_index(drop=True)
                        note = "(удаление строк)"
                    elif "NaN" in strategy:
                        for col in text_cols:
                            df_preview.loc[combined_mask, col] = np.nan
                        note = "(замена на NaN)"
                    elif "Неизвестно" in strategy:
                        for col in text_cols:
                            df_preview.loc[combined_mask, col] = "Неизвестно"
                        note = "(замена на 'Неизвестно')"
                    else:
                        note = "(без изменений)"

                    # Метрики для текста (4 колонки)
                    c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                    c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}", delta=f"{len(df_preview)-len(df_work):+}")

                    if text_cols:
                        col = text_cols[0]
                        # Для текста считаем другие метрики
                        def safe_text_stat(d, c):
                            if c not in d.columns or d.empty:
                                return 0, 0, 0
                            s = d[c].astype(str)
                            avg_len = s.str.len().mean()
                            unique = d[c].nunique()
                            invalid = combined_mask.sum() if len(d) == len(df_work) else 0
                            return avg_len, unique, invalid

                        al_b, u_b, inv_b = safe_text_stat(df_work, col)
                        al_a, u_a, inv_a = safe_text_stat(df_preview, col)

                        fmt = lambda x: f"{x:.1f}" if pd.notnull(x) else "N/A"

                        c_p2.metric("📏 Avg длина", f"{fmt(al_b)} → {fmt(al_a)}")
                        c_p3.metric(" Уникальных", f"{u_b} → {u_a}")
                        c_p4.metric("⚠️ Нарушений", f"{inv_b} → {inv_a}")

                    st.caption(f"📝 Стратегия: {note}")

                    # Кнопки подтверждения
                    st.divider()
                    c_ok, c_cancel = st.columns(2)
                    with c_ok:
                        if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_text"):
                            st.session_state.df_after_fixes = df_preview.copy()
                            st.session_state.df = df_preview
                            st.session_state.validation_ready = False
                            st.session_state.show_text_preview = False
                            st.success("✅ Стратегия применена! Перезапустите валидацию.")
                            st.rerun()
                    with c_cancel:
                        if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_text"):
                            st.session_state.show_text_preview = False
                            st.rerun()

        # ───────────────────────────────────────────────────────────
        # 9. ПРОВЕРКА РАВНОМЕРНОСТИ ВРЕМЕННОГО ШАГА (regularity)
        # ───────────────────────────────────────────────────────────

        # 9. Regularity Check
        regularity_results_local = locals().get('regularity_results', [])
        regularity_masks = locals().get('regularity_masks', {})
        regularity_freq_info = locals().get('regularity_freq_info', {})

        regularity_issues = len(regularity_results_local) > 0
        regularity_gaps = sum(r.get('Пропусков', r.get('Всего пропусков', 0)) for r in regularity_results_local) if regularity_issues else 0

        make_card(" Проверка равномерности временного шага (regularity)", regularity_issues,
            "**◻️ Метрики:** `inferred_freq` (определенная частота), `gap_count` (число пропусков), `interval_variance`.  \n"
            "**◻️ Алгоритм:** `pd.infer_freq()`, вычисление интервалов `diff()`, детекция аномалий (>1.5×моды).  \n"
            "**◻️ Влияние на TS:** Неравномерный шаг ломает ARIMA/SARIMA (требуют регулярный индекс), искажает FFT/спектральный анализ.  \n"
            "**◻️ Описание:** Модуль проверяет, что временные метки следуют с постоянным интервалом (день, месяц, год). "
            "Для панельных данных (страны × время) проверка выполняется внутри каждой группы. Обнаруживаются пропущенные периоды, "
            "нерегулярные интервалы, смешанные частоты. Критично для моделей, требующих DatetimeIndex с постоянной частотой.",
            "✅ Временной шаг равномерный" if not regularity_issues else f"⚠️ Обнаружено {regularity_gaps} пропусков в {len(regularity_results_local)} группах",
            "✅" if not regularity_issues else "⚠️", None)

        # ── 🔧 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН ОБРАБОТКИ (раскрывается при проблемах) ──
        if regularity_issues:
            with st.expander("🔧 Полный пайплайн обработки нарушений регулярности", expanded=True):
                st.markdown("### 📋 Таблица нарушений с ручной корректировкой")

                # Инициализация состояний
                if "df_regularity_work" not in st.session_state:
                    st.session_state.df_regularity_work = df.copy()
                df_work = st.session_state.df_regularity_work

                # Формируем общую маску нарушений
                combined_mask = pd.Series(False, index=df_work.index)

                if regularity_masks:
                    for mask in regularity_masks.values():
                        combined_mask |= mask
                else:
                    # Fallback: определяем нарушения из результатов
                    date_col = None
                    for r in regularity_results_local:
                        if 'Временная колонка' in r:
                            date_col = r['Временная колонка']
                            break

                    if date_col and date_col in df_work.columns:
                        if pd.api.types.is_datetime64_any_dtype(df_work[date_col]):
                            ts_sorted = df_work.sort_values(date_col)
                            intervals = ts_sorted[date_col].diff()
                            modal_interval = intervals.mode().iloc[0] if len(intervals.mode()) > 0 else intervals.median()
                            combined_mask.loc[intervals[intervals > modal_interval * 1.5].index] = True

                # Фильтр отображения
                view_filter = st.radio(
                    "Фильтр строк:",
                    ["⚠️ Только с нарушениями (пропуски)", "✅ Показать всё"],
                    horizontal=True,
                    key="regularity_view_filter"
                )

                if view_filter == "⚠️ Только с нарушениями (пропуски)":
                    df_view = df_work[combined_mask].copy() if combined_mask.any() else df_work.iloc[:0].copy()
                else:
                    df_view = df_work.copy()

                # Добавляем статус-колонку
                df_view = df_view.copy()
                df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Пропуск" if combined_mask.loc[idx] else "🟢 Норма"))

                # Интерактивная таблица (Ручная правка)
                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=300,
                    num_rows="dynamic",
                    disabled=['_STATUS'],
                    key="regularity_editor",
                    column_config={
                        "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                    }
                )

                # 💾 КНОПКА СОХРАНЕНИЯ
                c_save1, c_save2 = st.columns([4, 1])
                with c_save1:
                    st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                with c_save2:
                    if st.button("💾 Сохранить", key="btn_save_manual_regularity", use_container_width=True):
                        if '_STATUS' in edited_df.columns:
                            edited_df = edited_df.drop(columns=['_STATUS'])
                        df_work.update(edited_df)
                        st.session_state.df_regularity_work = df_work
                        st.session_state.df = df_work
                        st.session_state.validation_ready = False
                        st.toast("✅ Правки сохранены!", icon="✅")
                        st.rerun()

                st.divider()

                # Панель стратегий (Автоматическая обработка)
                st.markdown("### 🧹 Стратегии обработки пропусков временного ряда")
                c1, c2 = st.columns([2, 1])
                with c1:
                    regularity_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Resample + Interpolate (линейная интерполяция)",
                        "Resample + Forward Fill (последнее значение)",
                        "Resample + Backward Fill (следующее значение)",
                        "AsFreq (обозначить пропуски как NaN)",
                        "Добавить фиктивные записи с нулевыми значениями",
                        "Только отметить флагом (не менять данные)"],
                        key="regularity_fill_strategy"
                    )
                    if "Interpolate" in regularity_strategy:
                        st.info("ℹ️ Пропущенные значения будут заполнены линейной интерполяцией между соседними точками.")
                    elif "Forward Fill" in regularity_strategy:
                        st.info("ℹ️ Пропуски будут заполнены последним доступным значением (метод LOCF).")
                    elif "Backward Fill" in regularity_strategy:
                        st.info("ℹ️ Пропуски будут заполнены следующим значением (метод NOCB).")
                    elif "AsFreq" in regularity_strategy:
                        st.warning("⚠️ Ряд будет приведен к регулярной частоте, пропуски обозначены как NaN.")
                    elif "фиктивные" in regularity_strategy:
                        st.warning("⚠️ Будут добавлены строки с пропущенными датами и значениями 0/NaN.")
                    elif "флагом" in regularity_strategy:
                        st.info("🚩 Будет добавлена колонка `_has_gap` с отметкой о пропусках.")
                with c2:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка запуска превью
                if "show_regularity_preview" not in st.session_state:
                    st.session_state.show_regularity_preview = False

                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button("📊 Показать прогноз влияния на статистику", type="secondary", use_container_width=True, key="btn_show_regularity_preview"):
                    st.session_state.show_regularity_preview = True
                    st.rerun()

                # Блок превью
                if st.session_state.show_regularity_preview:
                    strategy = regularity_strategy
                    st.markdown("##### 📈 Прогноз влияния на временной ряд:")

                    df_preview = df_work.copy()

                    # Определяем временную колонку
                    date_col = None
                    for r in regularity_results_local:
                        if 'Временная колонка' in r:
                            date_col = r['Временная колонка']
                            break

                    if date_col and date_col in df_preview.columns:
                        # Определяем частоту
                        freq = regularity_freq_info.get('inferred_freq', 'D')
                        if not freq:
                            freq = 'D'  # Default to daily

                        num_cols = df_preview.select_dtypes(include=['number']).columns.tolist()

                        if "Interpolate" in strategy:
                            # Resample + Interpolate
                            if len(num_cols) > 0:
                                df_preview = df_preview.set_index(date_col)
                                for col in num_cols:
                                    df_preview[col] = df_preview[col].resample(freq).interpolate(method='linear')
                                df_preview = df_preview.reset_index()
                            note = "(resample + interpolate)"

                        elif "Forward Fill" in strategy:
                            # Resample + FFill
                            if len(num_cols) > 0:
                                df_preview = df_preview.set_index(date_col)
                                for col in num_cols:
                                    df_preview[col] = df_preview[col].resample(freq).ffill()
                                df_preview = df_preview.reset_index()
                            note = "(resample + ffill)"

                        elif "Backward Fill" in strategy:
                            # Resample + BFill
                            if len(num_cols) > 0:
                                df_preview = df_preview.set_index(date_col)
                                for col in num_cols:
                                    df_preview[col] = df_preview[col].resample(freq).bfill()
                                df_preview = df_preview.reset_index()
                            note = "(resample + bfill)"

                        elif "AsFreq" in strategy:
                            # AsFreq (просто обозначить пропуски)
                            if len(num_cols) > 0:
                                df_preview = df_preview.set_index(date_col)
                                for col in num_cols:
                                    df_preview[col] = df_preview[col].resample(freq).asfreq()
                                df_preview = df_preview.reset_index()
                            note = "(asfreq, NaN в пропусках)"

                        elif "фиктивные" in strategy:
                            # Добавить фиктивные записи
                            df_preview = df_preview.set_index(date_col)
                            full_range = pd.date_range(start=df_preview.index.min(),
                                                      end=df_preview.index.max(),
                                                      freq=freq)
                            df_preview = df_preview.reindex(full_range)
                            df_preview = df_preview.reset_index()
                            df_preview = df_preview.rename(columns={'index': date_col})
                            note = "(добавлены фиктивные записи)"
                        else:
                            note = "(без изменений)"

                        # Метрики
                        c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                        c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}",
                                   delta=f"{len(df_preview)-len(df_work):+}")

                        if num_cols:
                            col = num_cols[0]
                            def safe_stat(d, c, f):
                                if c not in d.columns or d.empty or d[c].notna().sum() == 0:
                                    return 0.0
                                return f(d[c].dropna())

                            m_b, s_b, d_b = safe_stat(df_work, col, np.mean), safe_stat(df_work, col, np.std), safe_stat(df_work, col, np.median)
                            m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)

                            fmt = lambda x: f"{x:.2f}" if pd.notnull(x) and x != 0.0 else "N/A"
                            delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"

                            c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                            c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                            c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

                        st.caption(f"📅 Частота: {freq} | Стратегия: {note}")

                        # Кнопки подтверждения
                        st.divider()
                        c_ok, c_cancel = st.columns(2)
                        with c_ok:
                            if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_regularity"):
                                st.session_state.df_after_fixes = df_preview.copy()
                                st.session_state.df = df_preview
                                st.session_state.validation_ready = False
                                st.session_state.show_regularity_preview = False
                                st.success("✅ Стратегия применена! Перезапустите валидацию.")
                                st.rerun()
                        with c_cancel:
                            if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_regularity"):
                                st.session_state.show_regularity_preview = False
                                st.rerun()
                    else:
                        st.error("❌ Не найдена временная колонка для применения стратегии")

        # ───────────────────────────────────────────────────────────
        # 10. ПРОВЕРКА ДОСТАТОЧНОСТИ ЧИСЛА НАБЛЮДЕНИЙ (sufficiency)
        # ───────────────────────────────────────────────────────────

        sufficiency_results_local = locals().get('sufficiency_results', [])
        sufficiency_recommendations = locals().get('sufficiency_recommendations', {})

        sufficiency_issues = len([r for r in sufficiency_results_local if r.get('Нарушений', 0) > 0]) > 0
        sufficiency_groups_with_issues = len([r for r in sufficiency_results_local if r.get('Нарушений', 0) > 0])

        make_card(" Проверка достаточности числа наблюдений (sufficiency)", sufficiency_issues,
            "**◻️ Метрики:** `n_total` (число наблюдений), `n_seasons` (полных сезонов), `frequency` (частота).  \n"
            "**◻️ Алгоритм:** Сравнение с порогами для тренда (≥10), ARIMA (≥50), FFT (≥64), ML (≥100), сезонности (≥2 полных сезона).  \n"
            "**◻️ Влияние на TS:** Недостаточный объём данных приводит к переобучению, нестабильным оценкам параметров, невозможности выделить сезонность.  \n"
            "**◻️ Описание:** Модуль оценивает, хватает ли данных для применения различных TS-моделей. Проверяет как общее число наблюдений, "
            "так и количество полных сезонных циклов. Для панельных данных (страны × время) проверка выполняется внутри каждой группы. "
            "Результат — список доступных и недоступных моделей с рекомендациями по сбору дополнительных данных.",
            "✅ Достаточность обеспечена для всех моделей" if not sufficiency_issues else f"⚠️ {sufficiency_groups_with_issues} групп с недостаточным объёмом данных",
            "✅" if not sufficiency_issues else "⚠️", None)

        # ── 🔧 ИНТЕРАКТИВНЫЙ ПАЙПЛАЙН (раскрывается при проблемах) ──
        if sufficiency_issues:
            with st.expander("🔧 Полный пайплайн анализа достаточности", expanded=True):
                st.markdown("### 📋 Детальный отчёт по группам")

                # Таблица с результатами (только для чтения, т.к. это аналитическая проверка)
                results_df = pd.DataFrame(sufficiency_results_local)

                # Фильтр: только проблемные группы
                view_filter = st.radio(
                    "Фильтр групп:",
                    ["⚠️ Только с проблемами", "✅ Показать все"],
                    horizontal=True,
                    key="sufficiency_view_filter"
                )

                if view_filter == "⚠️ Только с проблемами":
                    df_view = results_df[results_df['Нарушений'] > 0]
                else:
                    df_view = results_df

                # Отображение таблицы
                st.dataframe(
                    df_view[['Тип', 'Группа', 'Всего наблюдений', 'Частота', 'Полных сезонов (лет)', 'Нарушений', 'Статус']],
                    use_container_width=True,
                    height=250
                )

                # Детализация по каждой проблемной группе
                st.markdown("### 🔍 Детализация нарушений")
                for r in sufficiency_results_local:
                    if r.get('Нарушений', 0) > 0:
                        with st.container(border=True):
                            st.markdown(f"** {r.get('Группа', 'Весь датасет')}** ({r.get('Всего наблюдений', 0)} наблюдений)")
                            st.markdown("#### Нарушения:")
                            for detail in r.get('Детали', []):
                                st.markdown(f"- {detail}")
                            st.markdown("#### 💡 Рекомендации:")
                            st.markdown(r.get('Рекомендации', 'Нет рекомендаций'))

                st.divider()

                # Блок доступных моделей
                st.markdown("### 🤖 Доступные модели по группам")

                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("#### ✅ Доступные модели")
                    for group_name, rec in sufficiency_recommendations.items():
                        if rec.get('available_models'):
                            st.markdown(f"**{group_name}:**")
                            for model in rec['available_models']:
                                st.markdown(f"- ✅ {model}")

                with c2:
                    st.markdown("#### ⛔ Недоступные модели")
                    for group_name, rec in sufficiency_recommendations.items():
                        if rec.get('unavailable_models'):
                            st.markdown(f"**{group_name}:**")
                            for model in rec['unavailable_models']:
                                st.markdown(f"- ⛔ {model}")

                st.divider()

                # Стратегии действий
                st.markdown("### 🧹 Стратегии решения проблемы недостаточности данных")
                c1, c2 = st.columns([2, 1])
                with c1:
                    sufficiency_strategy = st.radio(
                        "Выберите стратегию:",
                        ["Продолжить анализ с доступными моделями (рекомендуется)",
                        "Агрегировать данные по более крупным периодам (например, месяц → квартал)",
                        "Исключить группы с недостаточным объёмом данных",
                        "Добавить синтетические данные (bootstrap/augmentation)",
                        "Только отметить в отчёте (без изменений)"],
                        key="sufficiency_fill_strategy"
                    )
                    if "Продолжить" in sufficiency_strategy:
                        st.info("ℹ️ Платформа автоматически ограничит выбор моделей только доступными.")
                    elif "Агрегировать" in sufficiency_strategy:
                        st.warning("⚠️ Данные будут агрегированы (mean/sum), что снизит детализацию, но увеличит n.")
                    elif "Исключить" in sufficiency_strategy:
                        st.warning(f"⚠️ Будет удалено **{sufficiency_groups_with_issues} групп** с недостаточным объёмом.")
                    elif "синтетические" in sufficiency_strategy:
                        st.warning("⚠️ Синтетические данные могут исказить реальные закономерности.")
                with c2:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                # Кнопка применения стратегии
                st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)
                if st.button("📊 Применить выбранную стратегию", type="primary", use_container_width=True, key="btn_apply_sufficiency"):
                    strategy = sufficiency_strategy

                    if "Агрегировать" in strategy:
                        # Определяем временную колонку
                        date_col = None
                        for r in sufficiency_results_local:
                            if 'Тип' in r and 'Группа' in r:
                                # Ищем колонку с датами
                                for col in df.columns:
                                    if 'year' in col.lower() or 'date' in col.lower():
                                        date_col = col
                                        break
                                break

                        if date_col:
                            df_aggregated = df.copy()
                            # Агрегация по году (упрощённый вариант)
                            if pd.api.types.is_datetime64_any_dtype(df_aggregated[date_col]):
                                df_aggregated['_year'] = df_aggregated[date_col].dt.year
                            else:
                                df_aggregated['_year'] = df_aggregated[date_col]

                            num_cols = df_aggregated.select_dtypes(include='number').columns.tolist()
                            group_cols = [c for c in df_aggregated.columns if c != date_col and c != '_year' and df_aggregated[c].dtype == 'object']

                            if group_cols:
                                agg_dict = {col: 'mean' for col in num_cols if col in df_aggregated.columns}
                                df_aggregated = df_aggregated.groupby(group_cols + ['_year']).agg(agg_dict).reset_index()
                            else:
                                agg_dict = {col: 'mean' for col in num_cols if col in df_aggregated.columns}
                                df_aggregated = df_aggregated.groupby('_year').agg(agg_dict).reset_index()

                            df_aggregated = df_aggregated.drop(columns=['_year'], errors='ignore')

                            st.session_state.df = df_aggregated
                            st.session_state.validation_ready = False
                            st.success(f"✅ Данные агрегированы. Новое число наблюдений: {len(df_aggregated)}")
                            st.rerun()
                        else:
                            st.error(" Не найдена временная колонка для агрегации")

                    elif "Исключить" in strategy:
                        # Удаляем группы с недостаточным объёмом
                        groups_to_remove = [r['Группа'] for r in sufficiency_results_local if r.get('Нарушений', 0) > 0]

                        # Определяем группирующую колонку
                        group_col = None
                        for c in df.columns:
                            if c != date_col and df[c].dtype == 'object' and df[c].nunique() < 100:
                                if 'country' in c.lower() or 'стран' in c.lower() or 'region' in c.lower():
                                    group_col = c
                                    break

                        if group_col:
                            df_filtered = df[~df[group_col].isin(groups_to_remove)]
                            st.session_state.df = df_filtered
                            st.session_state.validation_ready = False
                            st.success(f"✅ Удалено {len(groups_to_remove)} групп. Осталось: {len(df_filtered)} строк")
                            st.rerun()
                        else:
                            st.error("❌ Не найдена группирующая колонка")

                    elif "Продолжить" in strategy or "отметить" in strategy:
                        # Сохраняем рекомендации в session_state для использования в моделировании
                        st.session_state.sufficiency_recommendations = sufficiency_recommendations
                        st.session_state.validation_ready = False
                        st.success("✅ Рекомендации сохранены. Используйте их при выборе моделей.")
                        st.info("💡 Перейдите во вкладку 'Моделирование' — система предложит только доступные модели.")

                    else:
                        st.info("️ Стратегия не применена")

            st.session_state.ts_props_v10 = props_v10  # Рассчитывается в "Загрузка"
            st.session_state.ts_props_v11 = props_v11  # Рассчитывается после валидации
            st.session_state.dq_score = dq_score       # Возвращается из generate_validation_passport()
            st.session_state.val_results = val_results  # Полный dict результатов валидации

        # ═══════════════════════════════════════════════════════
        # ПАСПОРТ ВАЛИДАЦИИ ВРЕМЕННОГО РЯДА
        # ═══════════════════════════════════════════════════════
        st.divider()
        st.markdown("###  Паспорт валидации временного ряда")
        st.caption(
            "Сводная таблица результатов всех проверок качества данных. "
            "Структура соответствует DAMA DMBOK v2, ISO 8000-61 и TDQM."
        )

        # Получаем название датасета из session_state
        dataset_name = st.session_state.get("original_filename", "Неизвестный датасет")
        df_before = st.session_state.get("df_before_validation", df)
        df_after = st.session_state.get("df_after_fixes", None)

        from validation.engine import generate_validation_passport
        df_passport, dq_score, metadata = generate_validation_passport(
            df_before,
            st.session_state.val_results,
            df_after,
            dataset_name=dataset_name
        )

        # ─── ШАПКА ПАСПОРТА (с названием датасета) ─────────
        st.markdown(f"""
        <div style='background: linear-gradient(135deg, #2E4057 0%, #048A81 100%);
                    padding: 20px; border-radius: 12px; color: white; margin: 15px 0;'>
            <h3 style='margin: 0 0 10px 0;'>{metadata['document_title']}</h3>
            <p style='margin: 0; font-size: 16px;'>
                <b>Датасет:</b> <span style='background: rgba(255,255,255,0.2);
                padding: 3px 8px; border-radius: 4px; font-family: monospace;'>
                {metadata['dataset_name']}</span>
            </p>
            <p style='margin: 5px 0 0 0; font-size: 14px; opacity: 0.9;'>
                 {metadata['n_rows']} строк × {metadata['n_cols']} колонок |
                🕐 {metadata['generated_at']}
            </p>
        </div>
        """, unsafe_allow_html=True)

        # ── KPI-БЛОК ─────────────────────────────────────
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("ℹ️ Composite DQ Score", f"{dq_score:.1f}%",
                  delta=None if dq_score >= 80 else f"{80 - dq_score:+.1f}% до нормы")
        c2.metric("✅ Пройдено проверок",
                  f"{metadata['checks_passed']}/{metadata['checks_total']}")
        c3.metric("⚠️ Требуют внимания",
                  len([p for p in df_passport.to_dict('records') if p['Статус'] == "⚠️"]))
        c4.metric("❌ Критические",
                  len([p for p in df_passport.to_dict('records') if p['Статус'] == "❌"]))

        # ── ТАБЛИЦА ПАСПОРТА ──────────────────────────────
        st.dataframe(
            df_passport,
            use_container_width=True,
            height=600,
            column_config={
                "Вид проверки": st.column_config.TextColumn("Вид проверки", width="large"),
                "Измерение DQ": st.column_config.TextColumn("Измерение", width="small"),
                "Метрика": st.column_config.TextColumn("Метрика", width="medium"),
                "Алгоритм": st.column_config.TextColumn("Алгоритм", width="medium"),
                "Значение ДО": st.column_config.TextColumn("Значение ДО", width="medium"),
                "Значение ПОСЛЕ": st.column_config.TextColumn("Значение ПОСЛЕ", width="medium"),
                "Δ": st.column_config.TextColumn("Δ", width="small"),
                "Влияние на TS": st.column_config.TextColumn("Влияние на TS-модели", width="large"),
                "Статус": st.column_config.TextColumn("Статус", width="small"),
            }
        )

        # ═════════════════════════════════════════════════════════
        #  СОЗДАНИЕ СЛОВАРЯ recommendations ДЛЯ ЭКСПОРТА
        # ══════════════════════════════════════════════════════════
        # Этот блок должен быть ДО кнопок экспорта паспорта

        recommender = st.session_state.get("recommender")

        if recommender:
            # Построение профиля данных
            profile = recommender.build_profile_from_session_state(st.session_state)

            # Классификация моделей
            available_models_list = []
            limited_models_list = []
            unavailable_models_list = []

            for model in recommender.catalog.models:
                req = model.requirements
                issues = []

                if profile['n_observations'] < req.min_observations:
                    issues.append(f"Нужно ≥{req.min_observations} наблюдений")
                if req.stationarity == "required" and not profile['is_stationary']:
                    issues.append("Требуется стационарность")
                elif req.stationarity == "optional" and not profile['is_stationary']:
                    issues.append("Нестационарность")
                if req.seasonality == "required" and not profile['has_seasonality']:
                    issues.append("Требуется сезонность")
                if req.regularity == "required" and not profile['is_regular']:
                    issues.append("Требуется регулярная частота")
                elif req.regularity == "optional" and not profile['is_regular']:
                    issues.append("Нерегулярная частота")
                if req.exogenous == "required" and not profile['has_exogenous']:
                    issues.append("Требуются экзогенные признаки")

                if not issues:
                    available_models_list.append(model)
                else:
                    has_preprocessing = False
                    if model.preprocessing:
                        if not profile['is_stationary'] and model.preprocessing.if_not_stationary:
                            has_preprocessing = True
                        if not profile['is_regular'] and model.preprocessing.if_not_regular:
                            has_preprocessing = True

                    if has_preprocessing:
                        limited_models_list.append(model)
                    else:
                        unavailable_models_list.append(model)

            # Сортировка
            category_priority = {c.id: c.priority for c in recommender.catalog.categories}
            available_models_list.sort(key=lambda m: category_priority.get(m.category, 99))
            limited_models_list.sort(key=lambda m: category_priority.get(m.category, 99))
            unavailable_models_list.sort(key=lambda m: category_priority.get(m.category, 99))

            # Определение уровня качества
            if dq_score >= 80:
                tier_val = 'high'
            elif dq_score >= 50:
                tier_val = 'medium'
            else:
                tier_val = 'low'

            # Создание словаря recommendations
            recommendations = {
                'primary_recommendation': available_models_list[0].name if available_models_list else "Нет доступных моделей",
                'tier': tier_val,
                'available': [m.name for m in available_models_list],
                'limited': [m.name for m in limited_models_list],
                'unavailable': [m.name for m in unavailable_models_list],
                'explanation': f"DQ Score: {dq_score:.1f}%, Уровень качества: {tier_val}"
            }
        else:
            # Fallback если рекомендатель не загружен
            recommendations = {
                'primary_recommendation': "Нет данных",
                'tier': 'low',
                'available': [],
                'limited': [],
                'unavailable': [],
                'explanation': "Справочник моделей не загружен"
            }

        # ═══════════════════════════════════════════════════════
        # КНОПКИ ЭКСПОРТА ПАСПОРТА
        # ═══════════════════════════════════════════════════════
        c_export1, c_export2 = st.columns(2)

        with c_export1:
            if st.button("📥 Скачать паспорт (CSV)", use_container_width=True, key="btn_export_passport_csv"):
                csv_comments = [
                    f"# {metadata['document_title']}",
                    f"# Датасет: {metadata['dataset_name']}",
                    f"# {metadata['platform_tagline']}",
                    f"# {metadata['verification']}",
                    f"# Дата генерации: {metadata['generated_at']}",
                    f"# DQ Score: {dq_score:.1f}%",
                    ""
                ]
                csv_data = df_passport.to_csv(index=False, encoding="utf-8-sig")
                full_csv = "\n".join(csv_comments) + csv_data

                st.download_button(
                    label="📥 Скачать CSV",
                    data=full_csv.encode("utf-8-sig"),
                    file_name=f"passport_{dataset_name.replace('.xlsx','').replace('.csv','')}.csv",
                    mime="text/csv",
                    key="btn_download_csv_passport"
                )

        with c_export2:
            if st.button("📊 Скачать паспорт (Excel)", use_container_width=True, key="btn_export_passport_excel"):
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                import io

                wb = Workbook()
                ws = wb.active
                ws.title = "Паспорт валидации"

                ws.merge_cells('A1:I1')
                title_cell = ws['A1']
                title_cell.value = metadata['document_title']
                title_cell.font = Font(bold=True, size=16, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.merge_cells('A2:I2')
                ds_cell = ws['A2']
                ds_cell.value = f"Датасет: {metadata['dataset_name']}"
                ds_cell.font = Font(bold=True, size=12)
                ds_cell.alignment = Alignment(horizontal="left")

                ws.merge_cells('A3:I3')
                info_cell = ws['A3']
                info_cell.value = f"Строк: {metadata['n_rows']} | Колонок: {metadata['n_cols']} | Дата: {metadata['generated_at']}"
                info_cell.font = Font(size=10, color="666666")
                ws.row_dimensions[4].height = 8

                headers = list(df_passport.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="048A81", end_color="048A81", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(bottom=Side(style='thin'))

                for row_idx, row in enumerate(df_passport.to_dict('records'), 6):
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        if header == "Статус":
                            if "✅" in str(cell.value):
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif "❌" in str(cell.value):
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            elif "⚠️" in str(cell.value):
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

                last_data_row = 5 + len(df_passport)
                ws.row_dimensions[last_data_row + 1].height = 8

                ws.merge_cells(f'A{last_data_row + 2}:I{last_data_row + 2}')
                sign_cell_1 = ws.cell(row=last_data_row + 2, column=1)
                sign_cell_1.value = metadata['platform_tagline'] + "."
                sign_cell_1.font = Font(bold=True, size=11, color="1D3557")
                sign_cell_1.alignment = Alignment(horizontal="left")

                ws.merge_cells(f'A{last_data_row + 3}:I{last_data_row + 3}')
                sign_cell_2 = ws.cell(row=last_data_row + 3, column=1)
                sign_cell_2.value = metadata['verification'] + "."
                sign_cell_2.font = Font(bold=True, size=11, color="1D3557")
                sign_cell_2.alignment = Alignment(horizontal="left")

                ws.merge_cells(f'A{last_data_row + 4}:I{last_data_row + 4}')
                date_cell = ws.cell(row=last_data_row + 4, column=1)
                date_cell.value = f"Дата генерации: {metadata['generated_at']}"
                date_cell.font = Font(italic=True, size=10, color="666666")

                for col_idx in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 25

                # Лист 2: Рекомендации по моделям
                ws2 = wb.create_sheet(title="Рекомендации по моделям")

                ws2.merge_cells('A1:D1')
                title2 = ws2['A1']
                title2.value = "РЕКОМЕНДАЦИИ ПО ВЫБОРУ МОДЕЛЕЙ"
                title2.font = Font(bold=True, size=14, color="FFFFFF")
                title2.fill = PatternFill(start_color="048A81", end_color="048A81", fill_type="solid")
                title2.alignment = Alignment(horizontal="center", vertical="center")

                ws2.merge_cells('A2:D2')
                info2 = ws2['A2']
                info2.value = f"Датасет: {metadata['dataset_name']} | DQ Score: {dq_score:.1f}%"
                info2.font = Font(size=11)
                info2.alignment = Alignment(horizontal="left")

                ws2.merge_cells('A3:D3')
                primary_cell = ws2['A3']
                primary_cell.value = f" Первичная рекомендация: {recommendations['primary_recommendation']}"
                primary_cell.font = Font(bold=True, size=12, color="1D3557")
                primary_cell.fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
                primary_cell.alignment = Alignment(horizontal="left", vertical="center")
                ws2.row_dimensions[3].height = 25

                headers2 = ["Категория", "Модель", "Статус применимости", "Комментарий"]
                for col_idx, h in enumerate(headers2, 1):
                    cell = ws2.cell(row=5, column=col_idx, value=h)
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                row = 6
                for m in recommendations["available"]:
                    ws2.cell(row=row, column=1, value="Доступно")
                    ws2.cell(row=row, column=2, value=m)
                    ws2.cell(row=row, column=3, value="✅ Рекомендуется")
                    ws2.cell(row=row, column=4, value="Можно применять без ограничений")
                    for c in range(1, 5):
                        ws2.cell(row=row, column=c).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    row += 1

                for m in recommendations["limited"]:
                    ws2.cell(row=row, column=1, value="Ограничено")
                    ws2.cell(row=row, column=2, value=m)
                    ws2.cell(row=row, column=3, value="️ С осторожностью")
                    ws2.cell(row=row, column=4, value="Требует предобработки / валидации")
                    for c in range(1, 5):
                        ws2.cell(row=row, column=c).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    row += 1

                for m in recommendations["unavailable"]:
                    ws2.cell(row=row, column=1, value="Недоступно")
                    ws2.cell(row=row, column=2, value=m)
                    ws2.cell(row=row, column=3, value="❌ Не применимо")
                    ws2.cell(row=row, column=4, value="Недостаточно данных / низкое качество")
                    for c in range(1, 5):
                        ws2.cell(row=row, column=c).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    row += 1

                row += 1
                ws2.merge_cells(f'A{row}:D{row}')
                expl_cell = ws2.cell(row=row, column=1, value="Обоснование:")
                expl_cell.font = Font(bold=True, size=11, color="1D3557")
                row += 1
                ws2.merge_cells(f'A{row}:D{row}')
                expl_text = ws2.cell(row=row, column=1, value=recommendations["explanation"])
                expl_text.font = Font(size=10, italic=True)
                expl_text.alignment = Alignment(wrap_text=True)
                ws2.row_dimensions[row].height = 50

                ws2.column_dimensions['A'].width = 15
                ws2.column_dimensions['B'].width = 35
                ws2.column_dimensions['C'].width = 20
                ws2.column_dimensions['D'].width = 45

                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    label="📥 Скачать Excel",
                    data=buffer,
                    file_name=f"passport_{dataset_name.replace('.xlsx','').replace('.csv','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_download_excel_passport"
                )


        # ══════════════════════════════════════════════════════════
        # ТАБЛИЦА РЕКОМЕНДАЦИЙ ПО МОДЕЛЯМ (5 СТОЛБЦОВ, ВСЕ 20 МОДЕЛЕЙ)
        # ═══════════════════════════════════════════════════════════
        st.divider()
        st.markdown("###  Рекомендации по выбору моделей")

        # Получаем рекомендатель
        recommender = st.session_state.get("recommender")

        if recommender:
            # ─ УРОВЕНЬ КАЧЕСТВА ДАННЫХ ──────────────────────────
            quality_level = ""
            if dq_score >= 80:
                quality_level = f"✅ **Высокое качество (DQ ≥ 80%)** — все модели применимы"
            elif dq_score >= 50:
                quality_level = f"️ **Среднее качество (DQ 50-80%)** — рекомендована предобработка"
            else:
                quality_level = f"❌ **Низкое качество (DQ < 50%)** — только базовые модели"

            st.info(quality_level)

            # ── ПОСТРОЕНИЕ ПРОФИЛЯ ДАННЫХ ────────────────────────
            profile = recommender.build_profile_from_session_state(st.session_state)

            # ── КЛАССИФИКАЦИЯ ВСЕХ 20 МОДЕЛЕЙ ────────────────────
            available_models = []      # Все требования выполнены
            limited_models = []        # Есть preprocessing шаги
            unavailable_models = []    # Нет preprocessing или критические нарушения

            for model in recommender.catalog.models:
                req = model.requirements
                issues = []

                # Объём данных
                if profile['n_observations'] < req.min_observations:
                    issues.append(f"Нужно ≥{req.min_observations} наблюдений (есть {profile['n_observations']})")

                # Стационарность
                if req.stationarity == "required" and not profile['is_stationary']:
                    issues.append("Требуется стационарность")
                elif req.stationarity == "optional" and not profile['is_stationary']:
                    issues.append("Нестационарность (можно дифференцировать)")

                # Сезонность
                if req.seasonality == "required" and not profile['has_seasonality']:
                    issues.append("Требуется сезонность")

                # Регулярность
                if req.regularity == "required" and not profile['is_regular']:
                    issues.append("Требуется регулярная частота")
                elif req.regularity == "optional" and not profile['is_regular']:
                    issues.append("Нерегулярная частота (можно ресемплинг)")

                # Экзогенные признаки
                if req.exogenous == "required" and not profile['has_exogenous']:
                    issues.append("Требуются экзогенные признаки")

                # Классификация модели
                if not issues:
                    available_models.append(model)
                else:
                    # Проверяем, есть ли preprocessing для решения проблем
                    has_preprocessing = False
                    preprocessing_steps = []

                    if model.preprocessing:
                        if not profile['is_stationary'] and model.preprocessing.if_not_stationary:
                            has_preprocessing = True
                            preprocessing_steps.append(f"Нестационарность: {model.preprocessing.if_not_stationary.description}")

                        if not profile['is_regular'] and model.preprocessing.if_not_regular:
                            has_preprocessing = True
                            preprocessing_steps.append(f"Нерегулярность: {model.preprocessing.if_not_regular.description}")

                    if has_preprocessing:
                        limited_models.append({
                            'model': model,
                            'issues': issues,
                            'preprocessing': preprocessing_steps
                        })
                    else:
                        unavailable_models.append({
                            'model': model,
                            'issues': issues
                        })

            # Сортировка по приоритету категории
            category_priority = {c.id: c.priority for c in recommender.catalog.categories}
            available_models.sort(key=lambda m: category_priority.get(m.category, 99))
            limited_models.sort(key=lambda x: category_priority.get(x['model'].category, 99))
            unavailable_models.sort(key=lambda x: category_priority.get(x['model'].category, 99))

            # ── ФОРМИРОВАНИЕ ТАБЛИЦЫ (5 СТОЛБЦОВ) ────────────────
            table_rows = []
            max_rows = max(len(available_models), len(limited_models), len(unavailable_models))

            for i in range(max_rows):
                row = {}

                # Столбец 1: ✅ Доступные модели
                if i < len(available_models):
                    model = available_models[i]
                    row['available'] = f"✅ {model.name}"
                else:
                    row['available'] = ""

                # Столбец 2: ⚠️ С ограничениями
                if i < len(limited_models):
                    model = limited_models[i]['model']
                    row['limited'] = f"⚠️ {model.name}"
                else:
                    row['limited'] = ""

                # Столбец 3: Сделать доступными (для "С ограничениями")
                if i < len(limited_models):
                    steps = limited_models[i]['preprocessing']
                    row['limited_rec'] = "; ".join(steps) if steps else "Требуется предобработка"
                else:
                    row['limited_rec'] = ""

                # Столбец 4: ❌ Недоступные модели
                if i < len(unavailable_models):
                    model = unavailable_models[i]['model']
                    row['unavailable'] = f"❌ {model.name}"
                else:
                    row['unavailable'] = ""

                # Столбец 5: Сделать доступными (для "Недоступные")
                if i < len(unavailable_models):
                    issues = unavailable_models[i]['issues']

                    # Формируем понятное обоснование
                    if any("Нужно ≥" in issue for issue in issues):
                        rec_text = "Недостаточно данных для этой модели"
                    elif any("Требуется сезонность" in issue for issue in issues):
                        rec_text = "Модель требует явную сезонность в данных"
                    elif any("Требуются экзогенные" in issue for issue in issues):
                        rec_text = "Модель требует дополнительные признаки (экзогенные переменные)"
                    elif any("Требуется стационарность" in issue for issue in issues):
                        rec_text = "Модель требует стационарный ряд (дифференцирование не поможет)"
                    elif any("Требуется регулярная" in issue for issue in issues):
                        rec_text = "Модель требует регулярный временной шаг (ресемплинг не применим)"
                    else:
                        rec_text = "Модель не поддерживает текущий формат данных"

                    row['unavailable_rec'] = rec_text
                else:
                    row['unavailable_rec'] = ""

                table_rows.append(row)

            # Создаём DataFrame
            df_table = pd.DataFrame(table_rows)

            # Отображаем таблицу с 5 столбцами
            st.dataframe(
                df_table,
                use_container_width=True,
                hide_index=True,
                height=500,
                column_config={
                    "available": st.column_config.TextColumn("✅ Доступные модели", width="medium"),
                    "limited": st.column_config.TextColumn("⚠️ С ограничениями", width="medium"),
                    "limited_rec": st.column_config.TextColumn("Сделать доступными:", width="large"),
                    "unavailable": st.column_config.TextColumn("❌ Недоступные модели", width="medium"),
                    "unavailable_rec": st.column_config.TextColumn("Сделать доступными:", width="large"),
                }
            )

            # Сводка по количеству моделей
            st.caption(f"📊 **Всего моделей в каталоге: 20** | "
                    f"✅ Доступно: {len(available_models)} | "
                    f"⚠️ С ограничениями: {len(limited_models)} | "
                    f"❌ Недоступно: {len(unavailable_models)}")

            # Легенда
            st.markdown("""
            <div style='background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 15px 0;'>
            <strong> Как использовать таблицу:</strong><br>
            ◻️ <strong>Доступные модели</strong> — можно применять сразу без дополнительной предобработки<br>
            ◻️ <strong>С ограничениями</strong> — требуют предобработки (см. колонку "Сделать доступными")<br>
            ◻️ <strong>Недоступные</strong> — требуют значительной предобработки или не подходят для данных<br>
            <br>
            <strong> Рекомендации:</strong> Выполните указанные преобразования во вкладке "Предобработка",
            затем перезапустите валидацию для обновления списка доступных моделей.
            </div>
            """, unsafe_allow_html=True)

            # Главная рекомендация (первая доступная модель)
            if available_models:
                primary_model = available_models[0]
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, #048A81 0%, #1D3557 100%);
                            padding: 15px 20px; border-radius: 10px; color: white; margin-top: 15px;'>
                    <p style='margin: 0; font-size: 16px;'>
                        <b> Первичная рекомендация:</b>
                        <span style='font-size: 18px; font-weight: bold;'>
                        {primary_model.name}</span>
                    </p>
                    <p style='margin: 5px 0 0 0; font-size: 13px; opacity: 0.9;'>
                        На основе DQ Score ({dq_score:.1f}%), регулярности шага и достаточности наблюдений
                    </p>
                </div>
                """, unsafe_allow_html=True)

            # ─── ПОДПИСЬ (платформа + верификация) ──────────────
            st.divider()
            st.markdown(f"""
            <div style='background: #F1FAEE; border-left: 4px solid #1D3557;
                        padding: 15px 20px; border-radius: 8px; margin: 10px 0;'>
                <p style='margin: 0; font-size: 14px; color: #1D3557;'>
                    <b>{metadata['platform_tagline']}.</b><br>
                    <b>{metadata['verification']}.</b><br>
                    Дата генерации: <i>{metadata['generated_at']}</i>
                </p>
            </div>
            """, unsafe_allow_html=True)

        else:
            st.warning("⚠️ Справочник моделей не загружен. Проверьте файл `config/models/ts_models_catalog.yaml`")


        # ═══════════════════════════════════════════════════════
        # ШАГ 1: СОХРАНЕНИЕ КЛЮЧЕВЫХ ДАННЫХ В SESSION_STATE
        # ═══════════════════════════════════════════════════════
        # Сохраняем dq_score и метаданные СРАЗУ, чтобы они были доступны для сравнения
        st.session_state.dq_score = dq_score
        st.session_state.validation_metadata = metadata
        st.session_state.validation_passport_df = df_passport

        # Рассчитываем сравнение паспортов, если есть оба (v1.0 и v1.1)
        if "ts_props_v10" in st.session_state and "ts_props_v11" in st.session_state:
            if "ts_props_comparison_v10_v11" not in st.session_state:
                from validation.engine import _compare_ts_props
                st.session_state.ts_props_comparison_v10_v11 = _compare_ts_props(
                    st.session_state.ts_props_v10,
                    st.session_state.ts_props_v11
                )


        # ═══════════════════════════════════════════════════════
        # 🔍 ОТЛАДКА: Проверка состояния session_state
        # ═══════════════════════════════════════════════════════
        with st.expander("🔍 Отладка: состояние session_state", expanded=False):
            st.markdown("### Ключевые переменные:")
            checks = {
                "ts_props_v10": "ts_props_v10" in st.session_state,
                "ts_props_v11": "ts_props_v11" in st.session_state,
                "validation_ready": st.session_state.get("validation_ready", False),
                "primary_date_col": st.session_state.get("primary_date_col"),
                "num_cols_count": len(st.session_state.col_types.get("num", [])),
                "dq_score": st.session_state.get("dq_score"),
                "comparison_ready": "ts_props_comparison_v10_v11" in st.session_state
            }
            for key, value in checks.items():
                status = "✅" if value else "❌"
                st.markdown(f"{status} **{key}**: `{value}`")

            st.divider()
            st.markdown("### Лог событий:")
            if st.session_state.error_log:
                for entry in st.session_state.error_log[-5:]:
                    st.markdown(f"- {entry}")
            else:
                st.info("Лог пуст")

        # ═══════════════════════════════════════════════════════
        # 🔍 ОТЛАДКА СТРУКТУРЫ ПАСПОРТОВ
        # ═══════════════════════════════════════════════════════
        if "ts_props_v10" in st.session_state and "ts_props_v11" in st.session_state:
            with st.expander("🔍 Отладка: структура паспортов v1.0 и v1.1", expanded=False):
                st.markdown("#### Паспорт v1.0 (ключи верхнего уровня):")
                st.json(list(st.session_state.ts_props_v10.keys()))

                st.markdown("#### Паспорт v1.1 (ключи верхнего уровня):")
                st.json(list(st.session_state.ts_props_v11.keys()))

                st.markdown("#### Пример метрик из v1.0:")
                for section in ['basic_stats', 'stationarity', 'determinism', 'autocorrelation']:
                    if section in st.session_state.ts_props_v10:
                        st.code(f"{section}: {st.session_state.ts_props_v10[section]}")

                st.markdown("#### Пример метрик из v1.1:")
                for section in ['basic_stats', 'stationarity', 'determinism', 'autocorrelation']:
                    if section in st.session_state.ts_props_v11:
                        st.code(f"{section}: {st.session_state.ts_props_v11[section]}")

        # ═══════════════════════════════════════════════════════════
        # 📊 СРАВНЕНИЕ СВОЙСТВ РЯДА: v1.0 → v1.1 (ПО СТРУКТУРЕ EXCEL)
        # ═══════════════════════════════════════════════════════════
        if st.session_state.validation_ready:
            st.divider()
            st.markdown("###  Сравнение свойств ряда: v1.0 (сырые) → v1.1 (после валидации)")
            st.caption("На основе паспорта свойств из вкладки «Загрузка» и результатов этапа «Валидация»")

            # ── ВЫБОР ИССЛЕДУЕМОГО ПРИЗНАКА ──────────────────────
            num_cols = st.session_state.col_types.get("num", [])
            if num_cols:
                target_col = st.selectbox(
                    "🔢 Исследуемый признак:",
                    options=num_cols,
                    index=0 if num_cols else None,
                    key="validation_target_col_select",
                    help="Выберите числовую колонку для анализа свойств временного ряда"
                )

                # ── РАСЧЁТ ПАСПОРТА v1.1 (после валидации) ───────
                if target_col and st.session_state.primary_date_col:
                    try:
                        date_col = st.session_state.primary_date_col
                        df_validated = st.session_state.df.copy()

                        # Формируем временной ряд
                        df_validated[date_col] = pd.to_datetime(df_validated[date_col])
                        df_ts_validated = df_validated.set_index(date_col)

                        if target_col in df_ts_validated.columns:
                            analysis_series_v11 = (
                                df_ts_validated[target_col]
                                .resample('D').mean()
                                .dropna()
                                .astype(float)
                            )

                            if len(analysis_series_v11) >= 30:
                                # Рассчитываем паспорт v1.1
                                props_v11 = calculate_ts_passport(
                                    analysis_series_v11,
                                    df_filtered=df_validated,
                                    ct_f=st.session_state.col_types,
                                    target_col=target_col
                                )
                                props_v11['version'] = 'v1.1 (после валидации)'
                                st.session_state.ts_props_v11 = props_v11

                                # Сравниваем с v1.0 (если есть)
                                if 'ts_props_v10' in st.session_state:
                                    props_v10 = st.session_state.ts_props_v10

                                    # ── ФОРМИРОВАНИЕ ТАБЛИЦЫ СРАВНЕНИЯ ───────
                                    comparison_rows = []

                                    # 1. Стационарность
                                    stat_v10 = "✅ Стационарен" if props_v10.get('stationarity', {}).get('is_stationary', False) else "❌ Нестационарен"
                                    stat_v11 = "✅ Стационарен" if props_v11.get('stationarity', {}).get('is_stationary', False) else "❌ Нестационарен"
                                    adf_p_v10 = props_v10.get('stationarity', {}).get('p_value', 0)
                                    adf_p_v11 = props_v11.get('stationarity', {}).get('p_value', 0)
                                    delta_adf = adf_p_v11 - adf_p_v10
                                    change_stat = f"{delta_adf:+.4f}" if abs(delta_adf) > 0.001 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Стационарность",
                                        "Метод": "ADF Test (autolag='AIC')",
                                        "Паспорт v1.0 (Загрузка)": stat_v10,
                                        "Паспорт v1.1 (Валидация)": stat_v11,
                                        "Изменение v1.1/v1.0": change_stat,
                                        "Вывод": "✅ Без изменений" if stat_v10 == stat_v11 else "⚠️ Изменилось"
                                    })

                                    # 2. Детерминированность
                                    r2_v10 = props_v10.get('determinism', {}).get('value', 0)
                                    r2_v11 = props_v11.get('determinism', {}).get('value', 0)
                                    det_v10 = f"✅ Детерминированный (R²={r2_v10:.3f})" if r2_v10 >= 0.7 else f"⚠️ Стохастический/Смешанный (R²={r2_v10:.3f})"
                                    det_v11 = f"✅ Детерминированный (R²={r2_v11:.3f})" if r2_v11 >= 0.7 else f"⚠️ Стохастический/Смешанный (R²={r2_v11:.3f})"
                                    delta_r2 = r2_v11 - r2_v10
                                    change_det = f"{delta_r2:+.4f}" if abs(delta_r2) > 0.001 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Детерминированность",
                                        "Метод": "R² тренда + комбинация тестов",
                                        "Паспорт v1.0 (Загрузка)": det_v10,
                                        "Паспорт v1.1 (Валидация)": det_v11,
                                        "Изменение v1.1/v1.0": change_det,
                                        "Вывод": "✅ Без изменений" if abs(r2_v11 - r2_v10) < 0.1 else "⚠️ Изменилось"
                                    })

                                    # 3. Частота ряда
                                    freq_v10 = props_v10.get('freq', {}).get('value', 'Нерегулярная')
                                    freq_v11 = props_v11.get('freq', {}).get('value', 'Нерегулярная')
                                    freq_str_v10 = f"✅ {freq_v10}" if freq_v10 and freq_v10 != 'Нерегулярная' else "⚠️ Нерегулярная (требуется ресемплинг)"
                                    freq_str_v11 = f"✅ {freq_v11}" if freq_v11 and freq_v11 != 'Нерегулярная' else "⚠️ Нерегулярная (требуется ресемплинг)"
                                    change_freq = "—" if freq_v10 == freq_v11 else f"{freq_v10} → {freq_v11}"
                                    comparison_rows.append({
                                        "Свойство": "Частота ряда",
                                        "Метод": "pd.infer_freq() + автодетект",
                                        "Паспорт v1.0 (Загрузка)": freq_str_v10,
                                        "Паспорт v1.1 (Валидация)": freq_str_v11,
                                        "Изменение v1.1/v1.0": change_freq,
                                        "Вывод": "✅ Без изменений" if freq_v10 == freq_v11 else "⚠️ Изменилось"
                                    })

                                    # 4. Гетероскедастичность
                                    arch_p_v10 = props_v10.get('heteroskedasticity', {}).get('p_value', 1)
                                    arch_p_v11 = props_v11.get('heteroskedasticity', {}).get('p_value', 1)
                                    arch_str_v10 = "✅ Нет (Гомоскедастичность)" if arch_p_v10 > 0.05 else "⚠️ Есть (Гетероскедастичность)"
                                    arch_str_v11 = "✅ Нет (Гомоскедастичность)" if arch_p_v11 > 0.05 else "⚠️ Есть (Гетероскедастичность)"
                                    delta_arch = arch_p_v11 - arch_p_v10
                                    change_arch = f"{delta_arch:+.4f}" if abs(delta_arch) > 0.001 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Гетероскедастичность",
                                        "Метод": "ARCH-LM Test (Engle, 1982)",
                                        "Паспорт v1.0 (Загрузка)": arch_str_v10,
                                        "Паспорт v1.1 (Валидация)": arch_str_v11,
                                        "Изменение v1.1/v1.0": change_arch,
                                        "Вывод": "✅ Без изменений" if arch_str_v10 == arch_str_v11 else "⚠️ Изменилось"
                                    })

                                    # 5. Автокорреляция
                                    lb_p_v10 = props_v10.get('autocorrelation', {}).get('value', 1)
                                    lb_p_v11 = props_v11.get('autocorrelation', {}).get('value', 1)
                                    ac_v10 = "✅ Белый шум" if lb_p_v10 > 0.05 else "⚠️ Есть автокорреляция"
                                    ac_v11 = "✅ Белый шум" if lb_p_v11 > 0.05 else "⚠️ Есть автокорреляция"
                                    delta_lb = lb_p_v11 - lb_p_v10
                                    change_ac = f"{delta_lb:+.4f}" if abs(delta_lb) > 0.001 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Автокорреляция",
                                        "Метод": "Ljung-Box Test (Lag=10)",
                                        "Паспорт v1.0 (Загрузка)": ac_v10,
                                        "Паспорт v1.1 (Валидация)": ac_v11,
                                        "Изменение v1.1/v1.0": change_ac,
                                        "Вывод": "✅ Без изменений" if (lb_p_v10 > 0.05) == (lb_p_v11 > 0.05) else "⚠️ Изменилось"
                                    })

                                    # 6. Нормальность
                                    jb_p_v10 = props_v10.get('normality', {}).get('value', 1)
                                    jb_p_v11 = props_v11.get('normality', {}).get('value', 1)
                                    norm_v10 = "✅ Нормально" if jb_p_v10 > 0.05 else "⚠️ Отклонение"
                                    norm_v11 = "✅ Нормально" if jb_p_v11 > 0.05 else "⚠️ Отклонение"
                                    delta_jb = jb_p_v11 - jb_p_v10
                                    change_norm = f"{delta_jb:+.4f}" if abs(delta_jb) > 0.001 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Нормальность",
                                        "Метод": "Jarque-Bera Test",
                                        "Паспорт v1.0 (Загрузка)": norm_v10,
                                        "Паспорт v1.1 (Валидация)": norm_v11,
                                        "Изменение v1.1/v1.0": change_norm,
                                        "Вывод": "✅ Без изменений" if (jb_p_v10 > 0.05) == (jb_p_v11 > 0.05) else "⚠️ Изменилось"
                                    })

                                    # 7. Направление тренда
                                    slope_v10 = props_v10.get('determinism', {}).get('slope', 0)
                                    slope_v11 = props_v11.get('determinism', {}).get('slope', 0)
                                    trend_v10 = ("📈 Восходящий" if slope_v10 > 0 else "📉 Нисходящий" if slope_v10 < 0 else "➡️ Горизонтальный") + f" (Slope={slope_v10:.4f})"
                                    trend_v11 = ("📈 Восходящий" if slope_v11 > 0 else "📉 Нисходящий" if slope_v11 < 0 else "➡️ Горизонтальный") + f" (Slope={slope_v11:.4f})"
                                    delta_slope = slope_v11 - slope_v10
                                    change_trend = f"{delta_slope:+.4f}" if abs(delta_slope) > 0.0001 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Направление тренда",
                                        "Метод": "OLS Linear Regression",
                                        "Паспорт v1.0 (Загрузка)": trend_v10,
                                        "Паспорт v1.1 (Валидация)": trend_v11,
                                        "Изменение v1.1/v1.0": change_trend,
                                        "Вывод": "✅ Без изменений" if (slope_v10 > 0) == (slope_v11 > 0) else "⚠️ Изменилось"
                                    })

                                    # 8. Сезонность (сила)
                                    season_v10 = props_v10.get('seasonality', {}).get('strength', 0)
                                    season_v11 = props_v11.get('seasonality', {}).get('strength', 0)
                                    seas_v10 = f"✅ Сильная (S={season_v10:.2f})" if season_v10 > 0.6 else f"⚠️ Слабая/Нет (S={season_v10:.2f})"
                                    seas_v11 = f"✅ Сильная (S={season_v11:.2f})" if season_v11 > 0.6 else f"⚠️ Слабая/Нет (S={season_v11:.2f})"
                                    delta_season = season_v11 - season_v10
                                    change_season = f"{delta_season:+.4f}" if abs(delta_season) > 0.01 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Сезонность (сила)",
                                        "Метод": "STL Decomposition (Strength)",
                                        "Паспорт v1.0 (Загрузка)": seas_v10,
                                        "Паспорт v1.1 (Валидация)": seas_v11,
                                        "Изменение v1.1/v1.0": change_season,
                                        "Вывод": "✅ Без изменений" if abs(season_v11 - season_v10) < 0.1 else "⚠️ Изменилось"
                                    })

                                    # 9. Сезонные периоды (ACF)
                                    periods_v10 = props_v10.get('seasonal_periods', {}).get('periods', [])
                                    periods_v11 = props_v11.get('seasonal_periods', {}).get('periods', [])
                                    per_v10 = f"✅ {', '.join(map(str, periods_v10))}" if periods_v10 else "⚠️ Не обнаружены"
                                    per_v11 = f"✅ {', '.join(map(str, periods_v11))}" if periods_v11 else "⚠️ Не обнаружены"
                                    if periods_v10 == periods_v11:
                                        change_periods = "—"
                                    elif not periods_v10 and periods_v11:
                                        change_periods = f"Новые: {', '.join(map(str, periods_v11))}"
                                    elif periods_v10 and not periods_v11:
                                        change_periods = "Исчезли"
                                    else:
                                        added = set(periods_v11) - set(periods_v10)
                                        removed = set(periods_v10) - set(periods_v11)
                                        parts = []
                                        if added:
                                            parts.append(f"+{', '.join(map(str, added))}")
                                        if removed:
                                            parts.append(f"-{', '.join(map(str, removed))}")
                                        change_periods = "; ".join(parts) if parts else "Изменились"
                                    comparison_rows.append({
                                        "Свойство": "Сезонные периоды (ACF)",
                                        "Метод": "Автокорреляция + порог значимости",
                                        "Паспорт v1.0 (Загрузка)": per_v10,
                                        "Паспорт v1.1 (Валидация)": per_v11,
                                        "Изменение v1.1/v1.0": change_periods,
                                        "Вывод": "✅ Без изменений" if periods_v10 == periods_v11 else "⚠️ Изменилось"
                                    })

                                    # 10. Долгая память (Hurst)
                                    hurst_v10 = props_v10.get('hurst', {}).get('value', 0.5)
                                    hurst_v11 = props_v11.get('hurst', {}).get('value', 0.5)
                                    h_v10 = ("🔵 Антиперсистентность" if hurst_v10 < 0.45 else "🔴 Устойчивый тренд" if hurst_v10 > 0.55 else "⚪ Случайное блуждание") + f" (H={hurst_v10:.2f})"
                                    h_v11 = ("🔵 Антиперсистентность" if hurst_v11 < 0.45 else "🔴 Устойчивый тренд" if hurst_v11 > 0.55 else "⚪ Случайное блуждание") + f" (H={hurst_v11:.2f})"
                                    delta_hurst = hurst_v11 - hurst_v10
                                    change_hurst = f"{delta_hurst:+.4f}" if abs(delta_hurst) > 0.01 else "—"
                                    comparison_rows.append({
                                        "Свойство": "Долгая память",
                                        "Метод": "Hurst Exponent (R/S)",
                                        "Паспорт v1.0 (Загрузка)": h_v10,
                                        "Паспорт v1.1 (Валидация)": h_v11,
                                        "Изменение v1.1/v1.0": change_hurst,
                                        "Вывод": "✅ Без изменений" if abs(hurst_v11 - hurst_v10) < 0.05 else "⚠️ Изменилось"
                                    })

                                    # 11. Корреляция признаков
                                    corr_v10 = props_v10.get('correlations', {}).get('top3', {})
                                    corr_v11 = props_v11.get('correlations', {}).get('top3', {})
                                    if corr_v10:
                                        top_corr_v10 = list(corr_v10.items())[0]
                                        corr_str_v10 = f"🟡 {top_corr_v10[0]} ({top_corr_v10[1]:.2f})"
                                    else:
                                        corr_str_v10 = "⚪ Нет сильных связей"
                                    if corr_v11:
                                        top_corr_v11 = list(corr_v11.items())[0]
                                        corr_str_v11 = f"🟡 {top_corr_v11[0]} ({top_corr_v11[1]:.2f})"
                                    else:
                                        corr_str_v11 = "⚪ Нет сильных связей"
                                    change_corr = "—" if corr_str_v10 == corr_str_v11 else "Изменилась"
                                    comparison_rows.append({
                                        "Свойство": "Корреляция признаков",
                                        "Метод": "Pearson correlation matrix",
                                        "Паспорт v1.0 (Загрузка)": corr_str_v10,
                                        "Паспорт v1.1 (Валидация)": corr_str_v11,
                                        "Изменение v1.1/v1.0": change_corr,
                                        "Вывод": "✅ Без изменений" if corr_str_v10 == corr_str_v11 else "⚠️ Изменилось"
                                    })

                                    # 12. Доминирующие частоты (FFT)
                                    fft_v10 = props_v10.get('fft', {}).get('dominant_periods', [])
                                    fft_v11 = props_v11.get('fft', {}).get('dominant_periods', [])
                                    fft_str_v10 = f"✅ {', '.join([f'{p:.1f}' for p in fft_v10])}" if fft_v10 else "⚠️ Не обнаружены"
                                    fft_str_v11 = f"✅ {', '.join([f'{p:.1f}' for p in fft_v11])}" if fft_v11 else "⚠️ Не обнаружены"
                                    if fft_v10 == fft_v11:
                                        change_fft = "—"
                                    elif not fft_v10 and fft_v11:
                                        change_fft = f"Новые: {', '.join([f'{p:.1f}' for p in fft_v11])}"
                                    elif fft_v10 and not fft_v11:
                                        change_fft = "Исчезли"
                                    else:
                                        added = set(fft_v11) - set(fft_v10)
                                        removed = set(fft_v10) - set(fft_v11)
                                        parts = []
                                        if added:
                                            parts.append(f"+{', '.join([f'{p:.1f}' for p in added])}")
                                        if removed:
                                            parts.append(f"-{', '.join([f'{p:.1f}' for p in removed])}")
                                        change_fft = "; ".join(parts) if parts else "Изменились"
                                    comparison_rows.append({
                                        "Свойство": "Доминирующие частоты (FFT)",
                                        "Метод": "Быстрое преобразование Фурье",
                                        "Паспорт v1.0 (Загрузка)": fft_str_v10,
                                        "Паспорт v1.1 (Валидация)": fft_str_v11,
                                        "Изменение v1.1/v1.0": change_fft,
                                        "Вывод": "✅ Без изменений" if fft_v10 == fft_v11 else "⚠️ Изменилось"
                                    })

                                    # 13. Значимые периоды (Периодограмма)
                                    per_v10 = props_v10.get('periodogram', {}).get('periods', [])
                                    per_v11 = props_v11.get('periodogram', {}).get('periods', [])
                                    per_str_v10 = f"✅ {', '.join([f'{p:.1f}' for p in per_v10])}" if per_v10 else "⚠️ Не обнаружены"
                                    per_str_v11 = f"✅ {', '.join([f'{p:.1f}' for p in per_v11])}" if per_v11 else "⚠️ Не обнаружены"
                                    if per_v10 == per_v11:
                                        change_per = "—"
                                    elif not per_v10 and per_v11:
                                        change_per = f"Новые: {', '.join([f'{p:.1f}' for p in per_v11])}"
                                    elif per_v10 and not per_v11:
                                        change_per = "Исчезли"
                                    else:
                                        added = set(per_v11) - set(per_v10)
                                        removed = set(per_v10) - set(per_v11)
                                        parts = []
                                        if added:
                                            parts.append(f"+{', '.join([f'{p:.1f}' for p in added])}")
                                        if removed:
                                            parts.append(f"-{', '.join([f'{p:.1f}' for p in removed])}")
                                        change_per = "; ".join(parts) if parts else "Изменились"
                                    comparison_rows.append({
                                        "Свойство": "Значимые периоды (Периодограмма)",
                                        "Метод": "Periodogram с окном Hann",
                                        "Паспорт v1.0 (Загрузка)": per_str_v10,
                                        "Паспорт v1.1 (Валидация)": per_str_v11,
                                        "Изменение v1.1/v1.0": change_per,
                                        "Вывод": "✅ Без изменений" if per_v10 == per_v11 else "⚠️ Изменилось"
                                    })

                                    # 14. Доминирующие масштабы (Wavelet)
                                    wave_v10 = props_v10.get('wavelet', {}).get('scales', [])
                                    wave_v11 = props_v11.get('wavelet', {}).get('scales', [])
                                    wave_str_v10 = f"✅ {', '.join(map(str, wave_v10))}" if wave_v10 else "⚠️ Не обнаружены"
                                    wave_str_v11 = f"✅ {', '.join(map(str, wave_v11))}" if wave_v11 else "⚠️ Не обнаружены"
                                    if wave_v10 == wave_v11:
                                        change_wave = "—"
                                    elif not wave_v10 and wave_v11:
                                        change_wave = f"Новые: {', '.join(map(str, wave_v11))}"
                                    elif wave_v10 and not wave_v11:
                                        change_wave = "Исчезли"
                                    else:
                                        added = set(wave_v11) - set(wave_v10)
                                        removed = set(wave_v10) - set(wave_v11)
                                        parts = []
                                        if added:
                                            parts.append(f"+{', '.join(map(str, added))}")
                                        if removed:
                                            parts.append(f"-{', '.join(map(str, removed))}")
                                        change_wave = "; ".join(parts) if parts else "Изменились"
                                    comparison_rows.append({
                                        "Свойство": "Доминирующие масштабы (Wavelet)",
                                        "Метод": "Continuous Wavelet Transform",
                                        "Паспорт v1.0 (Загрузка)": wave_str_v10,
                                        "Паспорт v1.1 (Валидация)": wave_str_v11,
                                        "Изменение v1.1/v1.0": change_wave,
                                        "Вывод": "✅ Без изменений" if wave_v10 == wave_v11 else "⚠️ Изменилось"
                                    })

                                    # Создаём DataFrame и отображаем
                                    df_comparison = pd.DataFrame(comparison_rows)

                                    # Отображаем таблицу
                                    st.dataframe(
                                        df_comparison,
                                        use_container_width=True,
                                        hide_index=True,
                                        height=600,
                                        column_config={
                                            "Свойство": st.column_config.TextColumn("Свойство", width="medium"),
                                            "Метод": st.column_config.TextColumn("Метод", width="large"),
                                            "Паспорт v1.0 (Загрузка)": st.column_config.TextColumn("Паспорт v1.0", width="medium"),
                                            "Паспорт v1.1 (Валидация)": st.column_config.TextColumn("Паспорт v1.1", width="medium"),
                                            "Изменение v1.1/v1.0": st.column_config.TextColumn("Изменение", width="small"),
                                            "Вывод": st.column_config.TextColumn("Вывод", width="medium")
                                        }
                                    )

                                    # ═══════════════════════════════════════════════════════
                                    # КНОПКИ ЭКСПОРТА ТАБЛИЦЫ СРАВНЕНИЯ
                                    # ═══════════════════════════════════════════════════════

                                    # st.markdown("### 📥 Экспорт таблицы сравнения свойств")

                                    c_export3, c_export4 = st.columns(2)

                                    with c_export3:
                                        if st.button("📥 Скачать таблицу (CSV)", use_container_width=True, key="btn_export_comparison_csv"):
                                            csv_data = df_comparison.to_csv(index=False, encoding="utf-8-sig")
                                            st.download_button(
                                                label="📥 Скачать CSV",
                                                data=csv_data.encode("utf-8-sig"),
                                                file_name=f"comparison_{target_col}_v1.0_vs_v1.1.csv",
                                                mime="text/csv",
                                                key="btn_download_comparison_csv"
                                            )

                                    with c_export4:
                                        if st.button("📊 Скачать таблицу (Excel)", use_container_width=True, key="btn_export_comparison_excel"):
                                            from openpyxl import Workbook
                                            from openpyxl.styles import Font, PatternFill, Alignment
                                            from openpyxl.utils import get_column_letter
                                            import io

                                            wb = Workbook()
                                            ws = wb.active
                                            ws.title = "Сравнение свойств"

                                            headers = list(df_comparison.columns)
                                            for col_idx, header in enumerate(headers, 1):
                                                cell = ws.cell(row=1, column=col_idx, value=header)
                                                cell.font = Font(bold=True, color="FFFFFF", size=11)
                                                cell.fill = PatternFill(start_color="048A81", end_color="048A81", fill_type="solid")
                                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                                            for row_idx, row in enumerate(df_comparison.to_dict('records'), 2):
                                                for col_idx, header in enumerate(headers, 1):
                                                    cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
                                                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                                            for col_idx in range(1, len(headers) + 1):
                                                ws.column_dimensions[get_column_letter(col_idx)].width = 20

                                            buffer = io.BytesIO()
                                            wb.save(buffer)
                                            buffer.seek(0)

                                            st.download_button(
                                                label="📥 Скачать Excel",
                                                data=buffer,
                                                file_name=f"comparison_{target_col}_v1.0_vs_v1.1.xlsx",
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key="btn_download_comparison_excel"
                                            )

                                    # Итоговое резюме
                                    n_changes = len([r for r in comparison_rows if "⚠️" in r.get("Вывод", "")])
                                    if n_changes > 0:
                                        st.warning(f"⚠️ **Итог:** Валидация повлияла на {n_changes} свойств ряда")
                                    else:
                                        st.success("✅ **Итог:** Валидация незначительно повлияла на свойства ряда. Данные стабильны.")

                                else:
                                    st.info("ℹ️ Сначала рассчитайте паспорт v1.0 во вкладке «Загрузка»")
                            else:
                                st.warning(f"⚠️ Недостаточно данных для анализа: {len(analysis_series_v11)} точек (минимум 30)")
                        else:
                            st.error(f"❌ Колонка '{target_col}' не найдена в данных")
                    except Exception as e:
                        st.error(f"❌ Ошибка при расчёте паспорта v1.1: {e}")
                        import traceback
                        st.code(traceback.format_exc(), language="python")
                else:
                    st.info("ℹ️ Выберите признак и убедитесь, что активирован режим временных рядов")
            else:
                st.warning("⚠️ В датасете нет числовых колонок для анализа")


# ────────────────────────────────────────────────────────────
#  ВКЛАДКА 3: ПРЕДОБРАБОТКА
# ────────────────────────────────────────────────────────────
with tab_preprocessing:
    st.markdown("""
    <div style="padding-left: 20px; margin: 20px 0; text-align: right;">
        <p style="margin: 0 0 10px 0; color: #1e293b; line-height: 1.6; font-size: 18px; font-weight: 400;">
            "Если бы Кеплер опирался на точные данные, учитывающие все сложности взаимного влияния планет,<br>
            он никогда бы не сформулировал свои законы. Именно огрубление данных Тихо Браге позволило увидеть главное".
        </p>
        <p style="margin: 0; color: #64748B; font-style: italic; font-size: 16px; line-height: 1.5;">
            — Арнольд Зоммерфельд, немецкий физик-теоретик и математик
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.expander(" Цели модуля и результаты его прохождения", expanded=False):
        st.markdown("""
        ######  Цели модуля "Предобработка"

        Большинство классических моделей временных рядов и нейросетей предъявляют строгие требования к данным:
        - отсутствие пропусков
        - стационарность
        - гомоскедастичность
        - нормальность распределения и др.
                    
        **Цель раздела**. Применить математические преобразования, чтобы удовлетворить эти требования, сохранив при этом полезный сигнал (тренд, цикличность, сезонность).
        Предобработка решает задачу превращения данных в формат, пригодный для машинного обучения.           
        
                    
        **Что мы получим на выходе?** Применив обратные преобразования после предобработки, мы имеем трансформированный датасет, готовый к загрузке в блок «Моделирование».
        Пользователь получает рекомендации по доступным моделям прогнозирования и сравнительные паспорта свойств ряда для анализа их изменения:
        - v1.0 до валидации vs v1.3 после предобработки
        - v1.2 до предобработки vs v1.3 после предобработки. 
        """)


    # ───────────────────────────────────────────────────────────
    # ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ СРАВНЕНИЯ СВОЙСТВ РЯДА
    # ───────────────────────────────────────────────────────────
    def _calc_ts_props(series: pd.Series) -> dict:
        """Быстрый расчёт ключевых метрик для сравнения До/После."""
        props = {}
        if len(series) < 30:
            return {"error": "Недостаточно данных (<30 точек)"}
        try:
            from statsmodels.tsa.stattools import adfuller
            from statsmodels.stats.diagnostic import acorr_ljungbox
            from scipy import stats
            from scipy.stats import linregress

            # 1. ADF Test (стационарность)
            adf_res = adfuller(series, autolag='AIC')
            props['adf_p'] = adf_res[1]
            props['adf_stat'] = "✅ Стационарен" if adf_res[1] < 0.05 else "❌ Нестационарен"

            # 2. Ljung-Box Test (автокорреляция)
            lb_res = acorr_ljungbox(series, lags=[10])
            lb_p = lb_res['lb_pvalue'].iloc[0] if isinstance(lb_res, pd.DataFrame) else lb_res[1][0]
            props['lb_p'] = lb_p
            props['lb_stat'] = "✅ Белый шум" if lb_p > 0.05 else "⚠️ Есть АК"

            # 3. Jarque-Bera Test (нормальность)
            jb_res = stats.jarque_bera(series)
            jb_p = jb_res.pvalue if hasattr(jb_res, 'pvalue') else jb_res[1]
            props['jb_p'] = jb_p
            props['jb_stat'] = "✅ Нормально" if jb_p > 0.05 else "⚠️ Отклонение"

            # 4. R² тренда (детерминированность)
            slope, _, r_val, _, _ = linregress(range(len(series)), series)
            props['r2'] = r_val**2
            props['r2_stat'] = "Детерминированный" if r_val**2 >= 0.7 else "Стохастический"
        except Exception as e:
            props['error'] = str(e)
        return props

    def _show_comparison_table(props_before: dict, props_after: dict, stage: str):
        """Отображает таблицу сравнения До/После с цветовыми индикаторами."""
        metrics = [
            ('adf_p', 'ADF p-value', 'lower', 'Стационарность'),
            ('lb_p', 'Ljung-Box p-value', 'higher', 'Автокорреляция'),
            ('jb_p', 'Jarque-Bera p-value', 'higher', 'Нормальность'),
            ('r2', 'R² тренда', 'neutral', 'Детерминированность')
        ]
        comp_data = []
        for key, name, goal, desc in metrics:
            b_val = props_before.get(key)
            a_val = props_after.get(key)
            if b_val is None or a_val is None:
                continue

            diff = a_val - b_val
            status, rec = "⚠️ Нейтрально", "Нет существенного влияния"

            if goal == 'lower':  # Чем меньше, тем лучше (ADF)
                if a_val < b_val and a_val < 0.05:
                    status, rec = "✅ Улучшение", "Ряд стал стационарнее. Подходит для ARIMA/ML."
                elif a_val < b_val:
                    status, rec = "✅ Динамика+", "Значение снизилось, но ещё не стационарно."
                else:
                    status, rec = "❌ Ухудшение", "Тренд/нестационарность усилились."
            elif goal == 'higher':  # Чем больше, тем лучше (LB, JB)
                if a_val > b_val and a_val > 0.05:
                    status, rec = "✅ Улучшение", "Автокорреляция/нормальность улучшились."
                elif a_val > b_val:
                    status, rec = "✅ Динамика+", "Показатель вырос к целевому порогу."
                else:
                    status, rec = "❌ Ухудшение", "Зависимость/асимметрия усилились."
            else:  # Нейтральная метрика (R²)
                if abs(diff) < 0.1:
                    status, rec = "⚠️ Нейтрально", "Детерминированность не изменилась существенно."
                elif a_val > b_val:
                    status, rec = "✅ Усиление тренда", "Детерминированная компонента стала сильнее."
                else:
                    status, rec = "✅ Стохастичность+", "Ряд стал менее предсказуемым по тренду."

            comp_data.append({
                "📊 Свойство": f"{desc} ({name})",
                "📉 До": f"{b_val:.4f}",
                "📈 После": f"{a_val:.4f}",
                "📐 Изменение": f"{diff:+.4f}",
                "💡 Рекомендация": rec,
                "🏷️ Статус": status
            })

        df_comp = pd.DataFrame(comp_data)
        if df_comp.empty:
            st.warning(f"⚠️ {stage}: Нет метрик для сравнения.")
            return

        def color_status(val):
            if "Улучшение" in val:
                return 'color: #16a34a; font-weight: bold; background: #f0fdf4;'
            if "Ухудшение" in val:
                return 'color: #dc2626; font-weight: bold; background: #fef2f2;'
            return ''

        st.markdown(f"#### 📋 Сравнение свойств после этапа: `{stage}`")
        st.dataframe(
            df_comp.style.map(color_status, subset=['🏷️ Статус']),
            use_container_width=True, hide_index=True, height=220
        )

    # ───────────────────────────────────────────────────────────
    # 🔹 ИНИЦИАЛИЗАЦИЯ БАЗОВЫХ СВОЙСТВ ПРИ ВХОДЕ ВО ВКЛАДКУ
    # ───────────────────────────────────────────────────────────
    if "prep_props_baseline" not in st.session_state:
        df_curr = st.session_state.df
        num_cols = st.session_state.col_types.get("num", [])
        if num_cols and not df_curr.empty:
            target = num_cols[0]
            s = df_curr[target].dropna()
            st.session_state.prep_props_baseline = _calc_ts_props(s)
            st.session_state.prep_target_col = target
        else:
            st.session_state.prep_props_baseline = {}
            st.session_state.prep_target_col = None

    # ───────────────────────────────────────────────────────────
    # 🔹 ПОДГОТОВКА РЕЗУЛЬТАТОВ ВАЛИДАЦИИ (для перенесенных блоков)
    # ───────────────────────────────────────────────────────────
    if "val_results" not in st.session_state or not st.session_state.val_results.get("miss"):
        from validation.missing import analyze_missing
        from validation.outliers import detect_outliers
        rules = st.session_state.get("rules", {})
        st.session_state.val_results = {
            "miss": analyze_missing(st.session_state.df, rules.get("missing", {})),
            "outl": detect_outliers(st.session_state.df, rules.get("outliers", {}))
        }
    miss = st.session_state.val_results.get("miss", {})
    outl = st.session_state.val_results.get("outl", {})
    df = st.session_state.df  # Алиас для совместимости с оригинальным кодом

    st.markdown("""
    <style>
        .plotly .main-svg .subplot-title {
            font-weight: normal !important;
            fill: #6b7280 !important;
            font-size: 13px !important;
        }
    </style>
    """, unsafe_allow_html=True)


    # ═══════════════════════════════════════════════════════
    # 🔹 1. ПРОВЕРКА НА ПРОПУСКИ

    missing_issues = miss["summary"]["total_missing"] > 0
    missing_result = "✅ Пропуски отсутствуют" if not missing_issues else f"⚠️ Найдено {miss['summary']['total_missing']} пропусков ({miss['summary']['missing_rate_pct']:.1f}%)"

    st.markdown("#####  Проверка на пропуски")
    st.caption("Пропуски нарушают `DatetimeIndex`, делают невозможной STL-декомпозицию, "
           "искажают ACF/PACF и ломают ARIMA/SARIMA.")

    # Метрики и алгоритм — всегда в экспандере
    with st.expander("Метрики и алгоритм", expanded=missing_issues):
        st.markdown("**◻️ Метрики:** `% NaN = nulls/total`, макс. непрерывный разрыв.  \n"
        "**◻️ Алгоритм:** `isnull().sum()`, кумулятивная сумма разрывов, сверка с `infer_freq()`.  \n"
        "**◻️ Влияние на TS:** Пропуски нарушают `DatetimeIndex` → невозможна STL, искажает ACF/PACF.")

    # Краткий результат — всегда виден
    st.markdown(missing_result)

    # 🔽 ПОЛНЫЙ ПАЙПЛАЙН — внутри экспандера с правильной логикой
    with st.expander(" Полный пайплайн обработки пропусков", expanded=missing_issues):
        if not missing_issues:
            st.info("ℹ️ Пропуски не обнаружены. Разверните секцию для детального анализа механизма пропусков или превентивной настройки стратегий.")

        st.markdown("###  Статистическая панель анализа пропусков")

        # ── Инициализация состояний ─────────────────────────────────
        if "df_missing_orig" not in st.session_state:
            st.session_state.df_missing_orig = df.copy()
        if "df_missing_work" not in st.session_state:
            st.session_state.df_missing_work = df.copy()
        if "missing_strategy" not in st.session_state:
            st.session_state.missing_strategy = {}

        df_work = st.session_state.df_missing_work
        df_orig = st.session_state.df_missing_orig

        # ── 1️⃣ СТАТИСТИКА И АНАЛИЗ МЕХАНИЗМА ───────────────────────
        col_stats = []
        for c in df_work.columns:
            dtype = str(df_work[c].dtype)
            n_miss = df_work[c].isnull().sum()
            pct_miss = (n_miss / len(df_work) * 100) if len(df_work) > 0 else 0
            n_uniq = df_work[c].nunique()
            if pct_miss == 0:
                rec = "✅ Чисто"
            elif pct_miss > 50:
                rec = "🗑️ Обработать столбец"
            elif df_work[c].dtype in ['object', 'string', 'category']:
                rec = "🔄 Заполнить модой"
            elif pct_miss < 5:
                rec = "🗑️ Обработать строки"
            else:
                rec = "⚡ Заполнить медианой"
            col_stats.append({"Столбец": c, "Тип": dtype[:10], "Пропуски": int(n_miss), "%": f"{pct_miss:.1f}%", "Уник.": n_uniq, "Рекомендация": rec})

        stats_df = pd.DataFrame(col_stats).sort_values("%", key=lambda x: x.str.replace('%','').astype(float), ascending=False)

        def color_miss(val):
            if isinstance(val, str) and '%' in val:
                try:
                    pct = float(val.replace('%',''))
                    if pct > 20: return 'color: #dc2626; font-weight: bold;'
                    elif pct >= 5: return 'color: #d97706; font-weight: bold;'
                    else: return 'color: #16a34a;'
                except: pass
            return ''

        st.dataframe(stats_df.style.map(color_miss, subset=['%']), use_container_width=True, height=250, hide_index=True)

        rows_with_miss = df_work.isnull().any(axis=1).sum()
        rows_empty = df_work.isnull().all(axis=1).sum()
        miss_per_row = df_work.isnull().sum(axis=1)

        c1, c2 = st.columns(2)
        with c1:
            st.metric("📉 Строк с пропусками", f"{rows_with_miss} ({rows_with_miss/len(df_work)*100:.1f}%)")
            st.metric("️ Полностью пустых строк", rows_empty)
        with c2:
            if rows_with_miss > 0:
                fig_hist = px.histogram(x=miss_per_row[miss_per_row>0], nbins=int(min(20, miss_per_row.max())), title="📊 Распределение пропусков по строкам", labels={'x': 'Кол-во пропусков в строке', 'y': 'Кол-во строк'})
                fig_hist.update_layout(margin=dict(l=20, r=20, t=30, b=20), height=200)
                st.plotly_chart(fig_hist, use_container_width=True)
            else:
                st.info("ℹ️ Пропусков нет")

        # Анализ механизма (MCAR/MAR)
        with st.expander(" Анализ механизма пропусков (MCAR/MAR/MNAR) помогает выбрать корректную стратегию восстановления", expanded=False):
            miss_ind = df_work.isnull().astype(int)
            if miss_ind.sum().sum() > 0:
                corr_miss = miss_ind.corr()
                fig_heat = px.imshow(corr_miss, text_auto='.2f', color_continuous_scale='Blues', title="🔥 Корреляция пропусков между столбцами (MAR-диагностика)")
                st.plotly_chart(fig_heat, use_container_width=True)
                mar_pairs = []
                for c1 in miss_ind.columns:
                    for c2 in miss_ind.columns:
                        if c1 != c2 and corr_miss.loc[c1, c2] > 0.2:
                            mar_pairs.append(f"`{c1}` ↔ `{c2}` (ρ={corr_miss.loc[c1,c2]:.2f})")
                if mar_pairs:
                    st.warning(f"⚠️ **Обнаружены признаки MAR:** {', '.join(mar_pairs[:3])}. Рекомендуется множественное восстановление или индикаторные переменные.")
                else:
                    st.info("ℹ️ Сильных корреляций между пропусками нет. Возможно MCAR или MNAR. Для точного теста Литтла установите библиотеку `pingouin`.")
            else:
                st.success("✅ Пропусков нет, механизм не применим")

        st.divider()

        # ── 2️⃣ ИНТЕРАКТИВНЫЕ ВИЗУАЛИЗАЦИИ ──────────────────────────
        st.markdown("###  Визуализация пропусков")
        viz_type = st.selectbox("Тип графика", ["🔲 Матрица пропусков", "🔥 Тепловая карта корреляции", "📦 Сравнение распределений (Boxplot)"], index=1, key="miss_viz_type")

        if viz_type == "🔲 Матрица пропусков":
            miss_mat = df_work.isnull().astype(int)
            fig_mat = px.imshow(miss_mat.T, aspect='auto', color_continuous_scale=['#2563EB', '#FCA5A5'], labels=dict(x="Строка", y="Столбец", color="Пропуск"), title="🔲 Матрица пропусков (Красный = пропуск)")
            fig_mat.update_xaxes(showticklabels=False)
            fig_mat.update_layout(height=400, margin=dict(l=40, r=20, t=40, b=20))
            st.plotly_chart(fig_mat, use_container_width=True)
        elif viz_type == "🔥 Тепловая карта корреляции":
            corr_miss = df_work.isnull().astype(int).corr()
            fig_heat = px.imshow(corr_miss, text_auto='.2f', color_continuous_scale='RdBu_r', title="🔥 Корреляция наличия пропусков")
            st.plotly_chart(fig_heat, use_container_width=True)
        elif viz_type == "📦 Сравнение распределений (Boxplot)":
            col_box = st.selectbox("Столбец для сравнения", df_work.select_dtypes(include='number').columns.tolist(), key="miss_col_box")
            miss_col = st.selectbox("Столбец-индикатор", [c for c in df_work.columns if c != col_box], key="miss_col_ind")
            if col_box and miss_col:
                df_plot = df_work.copy()
                df_plot['Has_Miss'] = df_plot[miss_col].isnull().map({True: "С пропуском", False: "Без пропуска"})
                fig_box = px.box(df_plot, x='Has_Miss', y=col_box, color='Has_Miss', color_discrete_map={"С пропуском": "#ef4444", "Без пропуска": "#94a3b8"}, title=f"📦 Влияние пропусков в `{miss_col}` на распределение `{col_box}`")
                st.plotly_chart(fig_box, use_container_width=True)

        st.divider()

        # ── 3️⃣ ТАБЛИЦА С МАРКИРОВКОЙ (СТАБИЛЬНАЯ ВЕРСИЯ) ───────────
        st.markdown("###  Таблица данных с маркировкой пропусков")

        # 🔧 ИСПРАВЛЕНИЕ: Убираем st.columns(), оставляем только radio с horizontal=True
        show_filter = st.radio(
            "Фильтр:",
            ["Все", "Пропуски", "Без пропусков"],
            horizontal=True,
            key="miss_table_filter",
            label_visibility="collapsed"
        )

        # Подготовка данных
        mask_miss = df_work.isnull().any(axis=1)
        if show_filter == "Пропуски":
            df_view = df_work[mask_miss].copy() if mask_miss.any() else df_work.iloc[:0].copy()
        elif show_filter == "Без пропусков":
            df_view = df_work[~mask_miss].copy() if (~mask_miss).any() else df_work.iloc[:0].copy()
        else:
            df_view = df_work.copy()

        # Добавление статуса
        df_view = df_view.copy()
        df_view.insert(0, '_STATUS', df_view.index.map(lambda idx: "🔴 Пусто" if df_work.loc[idx].isnull().any() else "🟢 Норма"))

        # 🔧 ГЛАВНЫЙ ФИКС: Таблица в контейнере с ФИКСИРОВАННОЙ высотой
        with st.container(border=True):
            st.markdown("""
            <style>
                div[data-testid="stVerticalBlock"] { gap: 0.5rem; }
            </style>
            """, unsafe_allow_html=True)

            edited_df = st.data_editor(
                df_view,
                use_container_width=True,
                height=300,
                num_rows="dynamic",
                disabled=['_STATUS'],
                key="miss_editor",
                column_config={
                    "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small", help="🔴 Пусто / 🟢 Норма")
                }
            )

            # Кнопка сохранения
            c_save1, c_save2 = st.columns([4, 1])
            with c_save1:
                st.caption("💡 Отредактируйте вручную или используйте стратегии ниже")
            with c_save2:
                if st.button("💾 Сохранить", key="btn_save_manual_missing", use_container_width=True):
                    if '_STATUS' in edited_df.columns:
                        edited_df = edited_df.drop(columns=['_STATUS'])
                    df_work.update(edited_df)
                    st.session_state.df_missing_work = df_work
                    st.session_state.df = df_work
                    st.session_state.validation_ready = False
                    st.toast("✅ Правки сохранены!", icon="✅")
                    st.rerun()

        st.divider()

        # ── 4️⃣ ПАНЕЛЬ УПРАВЛЕНИЯ ОБРАБОТКОЙ ────────────────────────
        st.markdown("###  Стратегии обработки пропусков")
        c1, c2 = st.columns([2, 1])
        with c1:
            fill_strategy = st.radio(
                "Выберите стратегию:",
                [" Удалить строки", " Медиана/мода", " Среднее/мода", " Ноль/Unknown", " Интерполяция", " Индикатор"],
                key="fill_strategy",
                horizontal=False,
                label_visibility="collapsed"
            )
            # Динамические подсказки
            if "Удалить" in fill_strategy:
                st.warning(f"⚠️ Будет удалено **{rows_with_miss} строк** ({rows_with_miss/len(df_work)*100:.1f}%)")
            elif "медианой" in fill_strategy or "Медиана" in fill_strategy:
                st.info("📊 **Медиана (числовые)/мода (категориальные)** — устойчиво к выбросам")
            elif "средним" in fill_strategy or "Среднее" in fill_strategy:
                st.info("📈 **Среднее** — для нормального распределения")
            elif "нулём" in fill_strategy or "Ноль" in fill_strategy:
                st.info("0️⃣ **Константа** — просто, но может исказать данные")
            elif "Интерполяция" in fill_strategy:
                st.info("↕️ **Интерполяция** — для временных рядов")
            elif "индикатор" in fill_strategy:
                st.info("🚩 **Индикатор** — добавит колонки miss_* с флагом 0/1")
        with c2:
            # 🔧 ИСПРАВЛЕНИЕ: Кнопка "Применить" удалена.
            # Пустой блок для визуального выравнивания с другими модулями
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

        # Кнопка-триггер превью (отдельно от логики превью)
        if st.button(" Показать прогноз влияния на статистики", type="secondary", use_container_width=True, key="btn_show_fill_preview"):
            st.session_state.show_fill_preview = True
            st.rerun()

        # ── 🔍 ПРЕВЬЮ ВЛИЯНИЯ (с безопасным доступом к session_state) ────────────────────
        # 🔧 ГЛАВНОЕ ИСПРАВЛЕНИЕ: Используем .get() вместо прямого доступа
        show_preview = st.session_state.get("show_fill_preview", False)

        if show_preview:
            strategy = fill_strategy
            st.markdown("#####  Прогноз влияния на статистику:")

            df_preview = df_work.copy()
            num_cols = df_preview.select_dtypes(include='number').columns.tolist()
            cat_cols = df_preview.select_dtypes(include=['object', 'string', 'category']).columns.tolist()

            # Применение стратегии для превью
            if "Удалить" in strategy:
                df_preview = df_preview.dropna()
                note = "(удаление)"
            elif "медианой" in strategy or "Медиана" in strategy:
                for col in num_cols:
                    if df_preview[col].isnull().any():
                        df_preview[col] = df_preview[col].fillna(df_preview[col].median())
                for col in cat_cols:
                    if df_preview[col].isnull().any():
                        mode = df_preview[col].mode()[0] if not df_preview[col].mode().empty else "Unknown"
                        df_preview[col] = df_preview[col].fillna(mode)
                note = "(медиана/мода)"
            elif "средним" in strategy or "Среднее" in strategy:
                for col in num_cols:
                    if df_preview[col].isnull().any():
                        df_preview[col] = df_preview[col].fillna(df_preview[col].mean())
                for col in cat_cols:
                    if df_preview[col].isnull().any():
                        mode = df_preview[col].mode()[0] if not df_preview[col].mode().empty else "Unknown"
                        df_preview[col] = df_preview[col].fillna(mode)
                note = "(среднее/мода)"
            elif "нулём" in strategy or "Ноль" in fill_strategy:
                for col in num_cols: df_preview[col] = df_preview[col].fillna(0)
                for col in cat_cols: df_preview[col] = df_preview[col].fillna("Unknown")
                note = "(константа)"
            elif "Интерполяция" in strategy:
                for col in num_cols:
                    if df_preview[col].isnull().any():
                        df_preview[col] = df_preview[col].interpolate(method='linear')
                note = "(интерполяция)"
            elif "индикатор" in strategy:
                for col in df_preview.columns:
                    if df_preview[col].isnull().any():
                        df_preview[f"miss_{col}"] = df_preview[col].isnull().astype(int)
                note = "(индикатор)"
            else:
                note = "(без изменений)"

            # Метрики (4 колонки)
            c_p1, c_p2, c_p3, c_p4 = st.columns(4)
            c_p1.metric("📊 Записей", f"{len(df_work)} → {len(df_preview)}", delta=f"{len(df_preview)-len(df_work):+}")

            if num_cols:
                col = num_cols[0]
                def safe_stat(df, c, func):
                    return func(df[c]) if not df.empty and c in df.columns and df[c].notna().any() else 0.0
                m_b, s_b, d_b = safe_stat(df_work, col, np.mean), safe_stat(df_work, col, np.std), safe_stat(df_work, col, np.median)
                m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)
                fmt = lambda x: f"{x:,.2f}".replace(',', ' ') if pd.notnull(x) and x != 0.0 else "N/A"
                delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"
                c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

            # Кнопки подтверждения
            st.divider()
            c_ok, c_cancel = st.columns(2)
            with c_ok:
                if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_fill"):
                    # 🔧 СИНХРОНИЗАЦИЯ ВСЕХ РАБОЧИХ КОПИЙ
                    st.session_state.df = df_preview.copy()
                    st.session_state.validation_ready = False
                    st.session_state.show_fill_preview = False
                    
                    # 🔥 Удаляем все рабочие копии - они пересоздадутся из обновлённого df
                    work_dfs = [
                        "df_missing_work", "df_pattern_work", "df_range_work",
                        "df_outlier_work", "df_inclusion_work", "df_referential_work",
                        "df_text_work", "df_consistency_work", "df_uniqueness_work",
                        "df_regularity_work"
                    ]
                    for work_df_name in work_dfs:
                        if work_df_name in st.session_state:
                            del st.session_state[work_df_name]
                    
                    # Сбрасываем результаты валидации
                    if "val_results" in st.session_state:
                        del st.session_state.val_results
                    
                    st.success("✅ Стратегия применена! Перезапустите валидацию.")
                    st.rerun()
            with c_cancel:
                if st.button(" Отмена", use_container_width=True, key="btn_cancel_fill"):
                    st.session_state.show_fill_preview = False
                    st.rerun()

    # === КОНЕЦ ПОЛНОГО ПАЙПЛАЙНА ПРОПУСКОВ ===

    # 🔹 КНОПКА И ТАБЛИЦА СРАВНЕНИЯ ПОСЛЕ ПРОПУСКОВ
    if st.button(" Пересчитать свойства после преобразования (пропуски)", type="primary", key="btn_compare_missing"):
        target = st.session_state.get("prep_target_col")
        if target and not st.session_state.df.empty:
            s_after = st.session_state.df[target].dropna()
            props_after = _calc_ts_props(s_after)
            _show_comparison_table(st.session_state.prep_props_baseline, props_after, "Очистка пропусков")
            st.session_state.prep_props_baseline = props_after  # Обновляем baseline для следующего этапа

    # ═══════════════════════════════════════════════════════
    # 🔹 2. ПРОВЕРКА НА ВЫБРОСЫ
    # ═══════════════════════════════════════════════════════
    st.divider()
    outlier_issues = outl["summary"]["total_outliers"] > 0
    outlier_result = "✅ Выбросы не обнаружены" if not outlier_issues else f"⚠️ Найдено {outl['summary']['total_outliers']} выбросов"

    st.markdown("#####  Проверка на выбросы")
    st.caption("Выбросы завышают дисперсию, искажают оценки тренда и ломают тесты стационарности "
           "(ADF/KPSS).")

    with st.expander("Метрики и алгоритм", expanded=outlier_issues):
        st.markdown("**◻️ Метрики:** Z-score (`|x-μ|/σ>3`), IQR-границы, `% outliers`.  \n"
        "**◻️ Алгоритм:** Квантили `Q1/Q3`, стандартное отклонение, анализ остатков STL.  \n"
        "**◻️ Влияние на TS:** Завышают дисперсию, ломают тесты стационарности (ADF/KPSS).  \n"
        "**◻️ Описание:** Поддерживаются 4 метода обнаружения: IQR, Z-score, Modified Z-score (MAD) и процентильный — "
        "с автоматической рекомендацией по асимметрии и объёму выборки.")
    st.markdown(outlier_result)

    # ───────────────────────────────────────────────────────────
    # 🔽 ПОЛНЫЙ ПАЙПЛАЙН ОБРАБОТКИ ВЫБРОСОВ (Унифицированная версия)
    # ───────────────────────────────────────────────────────────
    # CSS для уменьшения заголовков
    st.markdown("""
    <style>
    .st-emotion-cache-3o718f h3 {
        font-size: 1.25rem !important;
        font-weight: 600;
        padding: 0.5rem 0px 0.75rem;
    }
    </style>
    """, unsafe_allow_html=True)


    if outlier_issues:
        with st.expander(" Полный пайплайн обработки выбросов", expanded=True):
            st.markdown("###  Работа с выбросами")

            # Инициализация session_state
            if "outlier_mask" not in st.session_state:
                st.session_state.outlier_mask = pd.Series([False] * len(df))
            if "show_outlier_preview" not in st.session_state:
                st.session_state.show_outlier_preview = False

            num_cols = df.select_dtypes(include='number').columns.tolist()

            if not num_cols:
                st.info("ℹ️ Нет числовых колонок для анализа выбросов.")
            else:
                # ── 1️⃣ НАСТРОЙКА ОБНАРУЖЕНИЯ ──────────────────────
                with st.expander(" Настройка обнаружения выбросов", expanded=True):
                    c1, c2 = st.columns(2)
                    with c1:
                        selected_out_cols = st.multiselect(
                            "Выберите столбец для анализа",
                            options=num_cols,
                            default=[],
                            placeholder="Выберите столбец",
                            key="sel_out_cols"
                        )
                    with c2:
                        skewness_vals = df[selected_out_cols].skew() if selected_out_cols else pd.Series([])
                        avg_skewness = abs(skewness_vals.mean()) if len(skewness_vals) > 0 else 0
                        has_extreme_skew = any(abs(s) > 2 for s in skewness_vals)
                        small_sample = len(df) < 100

                        if has_extreme_skew:
                            recommended_method = "Modified Z-score (MAD)"
                            rec_reason = "Обнаружена сильная асимметрия распределения (skewness > 2)"
                        elif small_sample:
                            recommended_method = "IQR (Межквартильный)"
                            rec_reason = "Малый объём выборки (< 100 записей)"
                        else:
                            recommended_method = "Z-score"
                            rec_reason = "Распределение близко к нормальному"

                        # ── МЕТОДЫ С ПОДСКАЗКАМИ ──────────────────────
                        st.markdown("<h5 style='margin: 0 0 8px 0; font-size: 15px; color: #1e293b; font-weight: 600;'> Метод обнаружения</h5>", unsafe_allow_html=True)

                        # Создаем кастомные radio-кнопки с подсказками
                        method_options = {
                            "IQR (Межквартильный)": "Основан на межквартильном размахе (IQR = Q3 - Q1). Выбросы: значения за пределами [Q1 - k×IQR; Q3 + k×IQR]. Устойчив к выбросам, не требует нормальности распределения.",
                            "Z-score": "Классический метод стандартных отклонений. Выбросы: |Z| > k, где Z = (x - μ)/σ. Требует нормального распределения данных. Чувствителен к выбросам.",
                            "Modified Z-score (MAD)": "Модифицированный Z-score на основе медианы и MAD (Median Absolute Deviation). Устойчив к выбросам и асимметричным распределениям. Рекомендуется при skewness > 2.",
                            "Процентильный": "Основан на процентилях распределения. Выбросы: значения ниже p1 или выше p2 процентиля. Гибкий метод для любых распределений."
                        }

                        method = st.radio(
                            "Выберите метод:",
                            list(method_options.keys()),
                            key="radio_out_method",
                            horizontal=False,
                            label_visibility="collapsed"
                        )

                        # Отображение подсказки для выбранного метода
                        st.markdown(
                            f'<div style="background: #f0f9ff; border-left: 3px solid #0284c7; padding: 8px 12px; margin: 8px 0; border-radius: 4px;">'
                            f'<span style="color: #0369a1; font-size: 13px;">'
                            f'💡 <strong>{method}:</strong> {method_options[method]}'
                            f'</span></div>',
                            unsafe_allow_html=True
                        )

                        if method == recommended_method:
                            st.success(f"✅ **Рекомендовано:** {rec_reason}")
                        else:
                            st.info(f"💡 **Рекомендация:** {recommended_method} ({rec_reason})")

                    # Параметры метода
                    c3, _ = st.columns(2)
                    param_val = 1.5
                    if method == "IQR":
                        param_val = c3.number_input("Множитель IQR", min_value=0.1, max_value=5.0, value=1.5, step=0.1,
                                                   key="param_iqr",
                                                   help="Стандартное значение: 1.5 (умеренные выбросы) или 3.0 (экстремальные)")
                    elif method == "Z-score":
                        param_val = c3.number_input("Порог Z-score", min_value=1.0, max_value=10.0, value=3.0, step=0.5,
                                                   key="param_z",
                                                   help="Стандартное значение: 3.0 (99.7% данных в пределах ±3σ)")
                    elif method == "Modified Z-score (MAD)":
                        param_val = c3.number_input("Порог MAD", min_value=1.0, max_value=10.0, value=3.5, step=0.5,
                                                   key="param_mad",
                                                   help="Рекомендуемое значение: 3.5 для обнаружения выбросов")
                    elif method == "Процентильный":
                        p_low = c3.number_input("Нижний процентиль (%)", value=1.0, key="p_low_out",
                                               help="Значения ниже этого процентиля считаются выбросами")
                        p_high = c3.number_input("Верхний процентиль (%)", value=99.0, key="p_high_out",
                                                help="Значения выше этого процентиля считаются выбросами")
                        param_val = (p_low, p_high)

                    # Кнопка запуска обнаружения
                    if st.button("🔍 Применить метод", type="primary", key="btn_apply_outlier_detect"):
                        if selected_out_cols:
                            global_mask = pd.Series([False] * len(df))
                            for col in selected_out_cols:
                                col_mask = pd.Series([False] * len(df))
                                vals = df[col]
                                if method == "IQR":
                                    Q1, Q3 = vals.quantile(0.25), vals.quantile(0.75)
                                    IQR = Q3 - Q1
                                    lower, upper = Q1 - param_val * IQR, Q3 + param_val * IQR
                                    col_mask = (vals < lower) | (vals > upper)
                                elif method == "Z-score":
                                    z = np.abs((vals - vals.mean()) / vals.std())
                                    col_mask = z > param_val
                                elif method == "Modified Z-score (MAD)":
                                    median = vals.median()
                                    mad = np.median(np.abs(vals - median))
                                    mod_z = 0.6745 * (vals - median) / mad if mad > 0 else 0
                                    col_mask = np.abs(mod_z) > param_val
                                elif method == "Процентильный":
                                    low, high = vals.quantile(param_val[0]/100), vals.quantile(param_val[1]/100)
                                    col_mask = (vals < low) | (vals > high)
                                global_mask = global_mask | col_mask
                            st.session_state.outlier_mask = global_mask
                            st.rerun()
                        else:
                            st.warning("Выберите хотя бы одну колонку")

                    # 🔧 Отображение результата
                    if st.session_state.outlier_mask.any():
                        outlier_count = int(st.session_state.outlier_mask.sum())
                        outlier_pct = outlier_count / len(df) * 100
                        st.success(f"✅ Найдено **{outlier_count}** выбросов ({outlier_pct:.1f}% данных)")
                        st.caption(f"Метод: {method} | Колонки: {', '.join(selected_out_cols)}")
                    elif selected_out_cols and st.session_state.outlier_mask is not None:
                        st.info("ℹ️ Выбросы не обнаружены при текущих настройках")

                # ── 2️⃣ ВИЗУАЛИЗАЦИЯ (если маска есть) ───────────
                if st.session_state.outlier_mask.any():
                    with st.expander("📊 Визуализация распределения", expanded=False):
                        viz_col = st.selectbox("Столбец для графика", selected_out_cols if selected_out_cols else num_cols, key="viz_col_out")
                        tab_v1, tab_v2 = st.tabs(["Boxplot", "Гистограмма + Плотность"])
                        with tab_v1:
                            df_plot = df[[viz_col]].copy()
                            df_plot['Status'] = ['Выброс' if x else 'Норма' for x in st.session_state.outlier_mask]
                            fig_box = px.box(df_plot, y=viz_col, color='Status', color_discrete_map={'Выброс': '#ef4444', 'Норма': '#94a3b8'}, title=f"Boxplot: {viz_col}")
                            st.plotly_chart(fig_box, use_container_width=True)
                        with tab_v2:
                            fig_hist = px.histogram(df, x=viz_col, nbins=50, marginal="box", title=f"Распределение: {viz_col}")
                            if method == "IQR" and viz_col in selected_out_cols:
                                Q1, Q3 = df[viz_col].quantile(0.25), df[viz_col].quantile(0.75)
                                IQR = Q3 - Q1
                                fig_hist.add_vline(x=Q1 - param_val*IQR, line_dash="dash", line_color="red")
                                fig_hist.add_vline(x=Q3 + param_val*IQR, line_dash="dash", line_color="red")
                            st.plotly_chart(fig_hist, use_container_width=True)

                    # ── 3️⃣ ТАБЛИЦА С МАРКИРОВКОЙ (Автоматический фильтр) ──
                    st.divider()
                    st.markdown("###  Таблица данных с маркировкой выбросов")

                    # ✅ Автоматический фильтр (срабатывает сразу при клике)
                    filter_mode = st.radio(
                        "Фильтр отображения:",
                        ["Все", "Только выбросы", "Только норма"],
                        horizontal=True,
                        key="filter_radio_out"
                    )

                    # Логика фильтрации
                    display_df = df.copy()
                    outlier_indices = df.index[st.session_state.outlier_mask]
                    display_df['_IS_OUTLIER'] = display_df.index.isin(outlier_indices)
                    display_df['_STATUS'] = np.where(display_df['_IS_OUTLIER'], "🔴 Выброс", "🟢 Норма")

                    # Перемещаем статус в начало
                    cols = ['_STATUS'] + [c for c in display_df.columns if c != '_STATUS']
                    display_df = display_df[cols]

                    # Применяем фильтр
                    if filter_mode == "Только выбросы":
                        display_df = display_df[display_df['_STATUS'] == "🔴 Выброс"].copy() if (display_df['_STATUS'] == "🔴 Выброс").any() else display_df.iloc[:0].copy()
                    elif filter_mode == "Только норма":
                        display_df = display_df[display_df['_STATUS'] == "🟢 Норма"].copy() if (display_df['_STATUS'] == "🟢 Норма").any() else display_df.iloc[:0].copy()

                    # ✅ Таблица с фиксированной высотой (чтобы минимизировать скачки)
                    edited_df = st.data_editor(
                        display_df,
                        use_container_width=True,
                        height=300,
                        num_rows="dynamic",
                        disabled=['_STATUS'],
                        key="outlier_editor",
                        column_config={
                            "_IS_OUTLIER": None,
                            "_STATUS": st.column_config.TextColumn("Статус", disabled=True, width="small")
                        }
                    )

                    # Кнопка сохранения (как в пропусках)
                    c_save1, c_save2 = st.columns([4, 1])
                    with c_save1:
                        st.caption("💡 Отредактируйте значения вручную или выберите стратегию ниже")
                    with c_save2:
                        if st.button("💾 Сохранить", key="btn_save_manual_outliers", use_container_width=True):
                            if '_STATUS' in edited_df.columns:
                                edited_df = edited_df.drop(columns=['_STATUS'])
                            df.update(edited_df)
                            st.session_state.df = df
                            st.toast("✅ Правки сохранены!", icon="✅")
                            st.rerun()

                    st.divider()

                    # ── 4️⃣ СТРАТЕГИИ ОБРАБОТКИ (раскрыта по умолчанию) ──
                    with st.expander(" Стратегии обработки выбросов (всегда анализируем природу выбросов!)", expanded=True):
                        st.markdown("###  Стратегии обработки выбросов")
                        c1, c2 = st.columns([2, 1])
                        with c1:
                            clean_strategy = st.radio(
                                "Выберите стратегию:",
                                [" Удаление строк", " Кэпирование", " Замена на медиану", " Только флаг"],
                                key="radio_outlier_strategy",
                                label_visibility="collapsed"
                            )
                            # Компактные подсказки
                            if "Кэпирование" in clean_strategy:
                                st.warning("⚠️ Выбросы будут заменены на границы по правилу 1.5×IQR: нижний Q1 - 1.5×IQR, верхний Q3 + 1.5×IQR")
                            elif "Удаление" in clean_strategy:
                                st.warning("⚠️ Строки с выбросами будут полностью удалены.")
                            elif "медиану" in clean_strategy or "Замена" in clean_strategy:
                                st.warning("⚠️ Все выбросы будут заменены на медиану.")
                        with c2:
                            # Пустая колонка для выравнивания с пропусками
                            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)

                        # ── 5️⃣ КНОПКА ПРОГНОЗА (унифицирована: белый фон, остаётся видимой) ──
                        st.markdown("<div style='margin: 15px 0;'></div>", unsafe_allow_html=True)

                        if st.button(" Показать прогноз влияния на статистики", type="secondary", use_container_width=True, key="btn_show_outlier_preview"):
                            st.session_state.show_outlier_preview = True
                            st.rerun()

                        # ── 6️⃣ БЛОК ПРОГНОЗА (отображается при активном флаге) ──
                        if st.session_state.show_outlier_preview:
                            st.markdown("#####  Прогноз влияния на статистику:")

                            df_preview = df.copy()
                            mask = st.session_state.outlier_mask

                            if "Удаление" in clean_strategy:
                                df_preview = df_preview[~mask].reset_index(drop=True)
                                note = "(удаление)"
                            elif "Кэпирование" in clean_strategy:
                                for col in num_cols:
                                    # 🔧 ИСПРАВЛЕНИЕ: Используем df_preview вместо df_final
                                    df_preview[col] = df_preview[col].astype(float)
                                    Q1, Q3 = df_preview[col].quantile(0.25), df_preview[col].quantile(0.75)
                                    IQR = Q3 - Q1
                                    df_preview.loc[mask, col] = df_preview.loc[mask, col].clip(Q1 - 1.5*IQR, Q3 + 1.5*IQR)
                                note = "(кэпирование)"
                            elif "медиану" in clean_strategy or "Замена" in clean_strategy:
                                for col in num_cols:
                                    df_preview[col] = df_preview[col].astype(float)
                                    df_preview.loc[mask, col] = df_preview[col].median()
                                note = "(медиана)"
                            else:
                                note = "(без изменений)"

                            # Метрики (4 колонки, как в пропусках)
                            c_p1, c_p2, c_p3, c_p4 = st.columns(4)
                            c_p1.metric("📊 Записей", f"{len(df):,} → {len(df_preview):,}".replace(',', ' '), delta=f"{len(df_preview)-len(df):+}")

                            cols_to_check = selected_out_cols if selected_out_cols else num_cols
                            if cols_to_check:
                                col = cols_to_check[0]
                                def safe_stat(d, c, f):
                                    return f(d[c]) if not d.empty and c in d.columns and d[c].notna().any() else 0.0
                                m_b, s_b, d_b = safe_stat(df, col, np.mean), safe_stat(df, col, np.std), safe_stat(df, col, np.median)
                                m_a, s_a, d_a = safe_stat(df_preview, col, np.mean), safe_stat(df_preview, col, np.std), safe_stat(df_preview, col, np.median)
                                fmt = lambda x: f"{x:,.2f}".replace(',', ' ') if pd.notnull(x) and x != 0.0 else "N/A"
                                delta = lambda b, a: f"{((a-b)/abs(b)*100):+.1f}%" if b != 0 and pd.notnull(b) else "0%"
                                c_p2.metric("📈 Mean", f"{fmt(m_b)} → {fmt(m_a)}", delta=delta(m_b, m_a))
                                c_p3.metric("📉 Std", f"{fmt(s_b)} → {fmt(s_a)}", delta=delta(s_b, s_a))
                                c_p4.metric("📊 Median", f"{fmt(d_b)} → {fmt(d_a)}", delta=delta(d_b, d_a))

                            # Кнопки подтверждения (унифицированы)
                            st.divider()
                            c_ok, c_cancel = st.columns(2)
                            with c_ok:
                                if st.button("💾 Подтвердить изменения", type="primary", use_container_width=True, key="btn_confirm_outlier"):
                                    try:
                                        df_final = st.session_state.df.copy()
                                        msk = st.session_state.outlier_mask
                                        num_cols_final = df_final.select_dtypes(include='number').columns.tolist()

                                        if "Удаление" in clean_strategy:
                                            df_final = df_final[~msk].reset_index(drop=True)
                                        elif "Кэпирование" in clean_strategy:
                                            for c in num_cols_final:
                                                df_final[c] = df_final[c].astype(float)
                                                Q1, Q3 = df_final[c].quantile(0.25), df_final[c].quantile(0.75)
                                                IQR = Q3 - Q1
                                                lower_bound = Q1 - 1.5 * IQR
                                                upper_bound = Q3 + 1.5 * IQR
                                                df_final.loc[msk & (df_final[c] < lower_bound), c] = Q1
                                                df_final.loc[msk & (df_final[c] > upper_bound), c] = Q3
                                        elif "медиану" in clean_strategy or "Замена" in clean_strategy:
                                            for c in num_cols_final:
                                                df_final[c] = df_final[c].astype(float)
                                                df_final.loc[msk, c] = df_final[c].median()

                                        # 🔧 СИНХРОНИЗАЦИЯ ВСЕХ РАБОЧИХ КОПИЙ
                                        st.session_state.df = df_final.copy()
                                        st.session_state.validation_ready = False

                                        # 🔥 Удаляем маску и все рабочие копии
                                        if "outlier_mask" in st.session_state:
                                            del st.session_state.outlier_mask
                                        
                                        work_dfs = [
                                            "df_missing_work", "df_pattern_work", "df_range_work",
                                            "df_outlier_work", "df_inclusion_work", "df_referential_work",
                                            "df_text_work", "df_consistency_work", "df_uniqueness_work",
                                            "df_regularity_work"
                                        ]
                                        for work_df_name in work_dfs:
                                            if work_df_name in st.session_state:
                                                del st.session_state[work_df_name]
                                        
                                        # Сбрасываем результаты валидации
                                        if "val_results" in st.session_state:
                                            del st.session_state.val_results

                                        st.success("✅ Выбросы успешно обработаны! Перезапустите валидацию.")
                                        st.rerun()

                                    except Exception as e:
                                        st.error(f"❌ Ошибка при обработке выбросов: {e}")
                                        st.exception(e)
                            with c_cancel:
                                if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_outlier"):
                                    st.session_state.show_outlier_preview = False
                                    st.rerun()

        # 🔹 КНОПКА И ТАБЛИЦА СРАВНЕНИЯ ПОСЛЕ ВЫБРОСОВ
        if st.button(" Пересчитать свойства после преобразования (выбросы)", type="primary", key="btn_compare_outliers"):
            target = st.session_state.get("prep_target_col")
            if target and not st.session_state.df.empty:
                s_after = st.session_state.df[target].dropna()
                props_after = _calc_ts_props(s_after)
                _show_comparison_table(st.session_state.prep_props_baseline, props_after, "Очистка выбросов")
                st.session_state.prep_props_baseline = props_after  # Обновляем baseline

    
    # ═══════════════════════════════════════════════════════
    # 🔹 3. ПРОВЕРКА РЕГУЛЯРНОСТИ ЧАСТОТЫ
    # ═══════════════════════════════════════════════════════
    st.divider()
    st.markdown("### Регулярность временного ряда")
    st.caption("Критично для ARIMA, STL, FFT, сезонного дифференцирования. Нерегулярный шаг ломает модели.")

    
    # ── ДИАГНОСТИКА ТЕКУЩЕЙ РЕГУЛЯРНОСТИ ──────────────────
    if st.session_state.primary_date_col:
        date_col = st.session_state.primary_date_col
        df_reg = st.session_state.df.copy()
        df_reg[date_col] = pd.to_datetime(df_reg[date_col])
        df_reg = df_reg.sort_values(date_col)
        
        # Индексы для анализа
        df_reg_ts = df_reg.set_index(date_col)
        
        # Определение частоты
        inferred_freq = pd.infer_freq(df_reg_ts.index.drop_duplicates().sort_values())
        is_regular = inferred_freq is not None
        
        # Интервалы между наблюдениями
        intervals = df_reg_ts.index.to_series().diff().dropna()
        intervals_seconds = intervals.dt.total_seconds()
        interval_std = intervals_seconds.std() if len(intervals_seconds) > 1 else 0
        modal_interval = intervals.mode().iloc[0] if len(intervals.mode()) > 0 else pd.Timedelta(hours=24)
        gap_count = sum(1 for i in intervals if i > modal_interval * 1.5)
        
        # ── ОТОБРАЖЕНИЕ МЕТРИК С УМЕНЬШЕННЫМ ШРИФТОМ ──────
        c_diag1, c_diag2, c_diag3, c_diag4 = st.columns(4)
        
        with c_diag1:
            st.markdown("**Обнаруженная частота**")
            st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: #1e293b;'>{inferred_freq if inferred_freq else '❌ Нерегулярная'}</div>", 
                    unsafe_allow_html=True)
        
        with c_diag2:
            st.markdown("**Пропущенных периодов**")
            # Красный цвет если есть пропуски, зелёный если нет
            gap_color = "#dc2626" if gap_count > 0 else "#16a34a"
            # Форматируем число с разделителями тысяч (пробел вместо запятой)
            gap_count_formatted = f"{gap_count:,}".replace(",", " ")
            st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {gap_color};'>{gap_count_formatted}</div>", 
                    unsafe_allow_html=True)
        
        with c_diag3:
            st.markdown("**Std интервалов**")
            # Форматируем число с разделителями тысяч
            interval_std_formatted = f"{interval_std:,.0f}".replace(",", " ")
            
            # 🔧 Эмодзи в ту же строку, что и значение
            icon = "⚠️" if interval_std > 3600 else "✅"
            icon_color = "#dc2626" if interval_std > 3600 else "#16a34a"
            
            st.markdown(
                f"<div style='font-size: 14px; font-weight: 600; color: #1e293b;'>"
                f"{interval_std_formatted} сек "
                f"<span style='color: {icon_color};'>{icon}</span>"
                f"</div>",
                unsafe_allow_html=True
            )
        
        with c_diag4:
            st.markdown("**Статус**")
            status_text = "Регулярная" if is_regular else "Нерегулярная"
            status_color = "#16a34a" if is_regular else "#dc2626"
            st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {status_color};'>{status_text}</div>", 
                    unsafe_allow_html=True)
        
        # Добавляем пустую строку для отступа
        st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)

        # ── ТЕХНИЧЕСКАЯ СПРАВКА (ПЕРЕНЕСЕНА В НАЧАЛО) ─────────
        with st.expander(" Цели субмодуля ⁞ Регулярность частоты", expanded=False):
            st.markdown("""
            **Зачем нужна регулярность:**
            -  **ARIMA/SARIMA:** Требуют регулярный `DatetimeIndex` для расчёта лагов
            -  **STL-декомпозиция:** Не работает с пропусками во временной оси
            -  **FFT/Спектральный анализ:** Требуют равномерную сетку частот
            -  **Сезонное дифференцирование:** Нужен фиксированный сезонный период
            
            **Методы обработки:**
            
            | Метод | Когда использовать | Преимущества | Недостатки |
            |-------|-------------------|--------------|------------|
            | **Interpolate** | Плавные процессы (цены, температура) | Сохраняет тренд | Может создавать ложные значения |
            | **Forward Fill** | Финансовые данные, котировки | Реалистично для рынков | Не заполняет первый пропуск |
            | **Backward Fill** | Отчётные данные | Заполняет с конца | Может искажать динамику |
            | **AsFreq** | Когда пропуски = реальные отсутствия | Честно отражает данные | Создаёт NaN для моделей |
            | **Агрегация** | Шумные данные, высокая частота | Сглаживает шум | Теряется детализация |
            
            **Контроль качества:**
            - ✅ `pd.infer_freq()` — определяет частоту автоматически
            - ✅ Тест на постоянство интервалов (`std < 10%` от медианы)
            - ✅ Визуальная проверка на графике
                        
            ⚠️ **Почитать:**
            - регулярность частоты и экономические временные ряды: https://c0ldness.quarto.pub/dynamics/economicdynamicsindicators.html
            """) 
        
        # Если ряд уже регулярный — показываем только информацию
        if is_regular and gap_count == 0:
            st.success("✅ Временной ряд имеет регулярную частоту. Преобразование не требуется.")
        else:
            # ────────────────────────────────────────────────
            # 🎮 ПЕСОЧНИЦА: РЕГУЛЯРНОСТЬ ЧАСТОТЫ
            # ────────────────────────────────────────────────
            
            # Инициализация session_state
            if "df_regular_work" not in st.session_state:
                st.session_state.df_regular_work = df.copy()
            if "regularity_config" not in st.session_state:
                st.session_state.regularity_config = {}
            if "show_regular_preview" not in st.session_state:
                st.session_state.show_regular_preview = False
            
            # ── ЛЕВАЯ КОЛОНКА: ПАНЕЛЬ УПРАВЛЕНИЯ ───────────
            c1, c2, c3 = st.columns([27, 52, 21])
            
            with c1:
                st.markdown("######  Панель управления")
                
                # Выбор метода обработки
                reg_method = st.radio(
                    "Метод обработки:",
                    ["Resample + Interpolate (линейная)", "Resample + Forward Fill (LOCF)", 
                    "Resample + Backward Fill (NOCB)", "AsFreq (обозначить NaN)",
                    "Агрегация (день → неделя/месяц)"],
                    key="regularity_method",
                    label_visibility="collapsed"
                )
                
                # Выбор целевой частоты
                freq_options = {
                    "Дневная": "D",
                    "Недельная": "W",
                    "Месячная": "ME",
                    "Квартальная": "QE",
                    "Годовая": "YE"
                }
                
                current_freq_display = inferred_freq if inferred_freq else "Не определена"
                st.markdown(f"**Текущая частота:** `{current_freq_display}`")
                
                target_freq = st.selectbox(
                    "Целевая частота:",
                    options=list(freq_options.keys()),
                    index=0 if inferred_freq == "D" else 2,  # По умолчанию месяц для нерегулярных
                    key="regularity_target_freq"
                )
                freq_code = freq_options[target_freq]
                
                # Настройки для interpolate
                if "Interpolate" in reg_method:
                    interp_method = st.selectbox(
                        "Метод интерполяции:",
                        ["linear", "spline", "quadratic", "polynomial"],
                        index=0,
                        key="regularity_interp_method"
                    )
                    if interp_method in ["spline", "polynomial"]:
                        order = st.slider("Порядок полинома:", 2, 5, 2, key="regularity_interp_order")
                
                st.divider()
                
                # Кнопки действий
                if st.button("▶ Применить преобразование", type="primary", use_container_width=True, key="btn_apply_regularity"):
                    st.session_state.show_regular_preview = True
                    st.rerun()
                
                if st.button("↶ Откатить всё", use_container_width=True, key="btn_reset_regularity"):
                    st.session_state.df_regular_work = df.copy()
                    st.session_state.show_regular_preview = False
                    st.rerun()
                
                if st.button("💾 Сохранить конфигурацию", use_container_width=True, key="btn_save_regularity"):
                    st.session_state.regularity_config = {
                        "method": reg_method,
                        "target_freq": freq_code,
                        "interpolation": interp_method if "Interpolate" in reg_method else None
                    }
                    st.toast("✅ Конфигурация сохранена!", icon="💾")
            
            # ── ЦЕНТРАЛЬНАЯ КОЛОНКА: ВИЗУАЛИЗАЦИЯ ──────────
            with c2:
                st.markdown("######  Визуализация: До / После")
                
                # Выбор колонки для отображения
                num_cols = st.session_state.col_types.get("num", [])
                if num_cols:
                    target_col = st.selectbox(
                        "Числовой признак для анализа:",
                        options=num_cols,
                        index=0,
                        key="regularity_plot_col"
                    )
                    
                    if target_col in df_reg_ts.columns:
                        original_series = df_reg_ts[target_col].dropna()
                        
                        if st.session_state.show_regular_preview:
                            # Применяем преобразование для preview
                            df_preview_reg = df_reg.copy()
                            df_preview_reg[date_col] = pd.to_datetime(df_preview_reg[date_col])
                            df_preview_reg = df_preview_reg.sort_values(date_col)
                            df_preview_reg = df_preview_reg.set_index(date_col)
                            
                            if "Resample" in reg_method or "AsFreq" in reg_method:
                                if "AsFreq" in reg_method:
                                    df_resampled = df_preview_reg[target_col].resample(freq_code).mean()
                                elif "Interpolate" in reg_method:
                                    if interp_method == "linear":
                                        df_resampled = df_preview_reg[target_col].resample(freq_code).mean().interpolate(method="linear")
                                    elif interp_method == "spline":
                                        df_resampled = df_preview_reg[target_col].resample(freq_code).mean().interpolate(method="spline", order=order)
                                    elif interp_method == "quadratic":
                                        df_resampled = df_preview_reg[target_col].resample(freq_code).mean().interpolate(method="polynomial", order=2)
                                    elif interp_method == "polynomial":
                                        df_resampled = df_preview_reg[target_col].resample(freq_code).mean().interpolate(method="polynomial", order=order)
                                elif "Forward Fill" in reg_method:
                                    df_resampled = df_preview_reg[target_col].resample(freq_code).mean().ffill()
                                elif "Backward Fill" in reg_method:
                                    df_resampled = df_preview_reg[target_col].resample(freq_code).mean().bfill()
                            elif "Агрегация" in reg_method:
                                df_resampled = df_preview_reg[target_col].resample(freq_code).mean()
                            else:
                                df_resampled = df_preview_reg[target_col]
                            
                            # Метрики после преобразования
                            new_inferred_freq = pd.infer_freq(df_resampled.index.dropna())
                            new_is_regular = new_inferred_freq is not None
                            
                            # ── ГРАФИК СРАВНЕНИЯ ────────────
                            fig = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    f"Исходный ряд (частота: {current_freq_display})",
                                    f"После преобразования (частота: {new_inferred_freq or 'Нерегулярная'})"
                                ),
                                vertical_spacing=0.10
                            )
                            fig.update_annotations(font=dict(size=13, color="#6b7280"))  # Серый цвет

                            # Исходный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=original_series.index, y=original_series.values,
                                    mode='lines+markers',
                                    name='Исходные данные',
                                    line=dict(color='#048A81', width=2),
                                    marker=dict(size=5),
                                    showlegend=False
                                ),
                                row=1, col=1
                            )

                            # Преобразованный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=df_resampled.index, y=df_resampled.values,
                                    mode='lines+markers',
                                    name='После обработки',
                                    line=dict(color='#DC2626', width=2),
                                    marker=dict(size=5),
                                    showlegend=False
                                ),
                                row=2, col=1
                            )

                            # LAYOUT
                            fig.update_layout(
                                height=700,
                                margin=dict(l=50, r=20, t=80, b=40),
                                hovermode='x unified',
                            )

                            fig.update_xaxes(title_text="Дата", row=2, col=1)
                            fig.update_yaxes(title_text="Значение", row=1, col=1)
                            fig.update_yaxes(title_text="Значение", row=2, col=1)

                            # Стилизуем subplot_titles
                            fig.update_annotations(
                                font=dict(size=14, color="#1e293b"),
                                yshift=10
                            )

                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Статус преобразования
                            if new_is_regular:
                                st.success(f"✅ Преобразование успешно! Частота стала регулярной: `{new_inferred_freq}`")
                            else:
                                st.warning("⚠️ Частота всё ещё нерегулярная. Попробуйте другой метод.")
                            
                            # Кнопка подтверждения
                            st.divider()
                            c_ok_reg, c_cancel_reg = st.columns(2)
                            with c_ok_reg:
                                if st.button("✅ Применить к данным", type="primary", use_container_width=True, key="btn_confirm_regularity"):
                                    # Обновляем основной df
                                    df_final_reg = df.copy()
                                    df_final_reg[date_col] = pd.to_datetime(df_final_reg[date_col])
                                    df_final_reg = df_final_reg.sort_values(date_col)
                                    df_final_reg = df_final_reg.set_index(date_col)
                                    
                                    if "Resample" in reg_method or "AsFreq" in reg_method:
                                        if "AsFreq" in reg_method:
                                            df_resampled_full = df_final_reg[target_col].resample(freq_code).mean()
                                        elif "Interpolate" in reg_method:
                                            if interp_method == "linear":
                                                df_resampled_full = df_final_reg[target_col].resample(freq_code).mean().interpolate(method="linear")
                                            elif interp_method == "spline":
                                                df_resampled_full = df_final_reg[target_col].resample(freq_code).mean().interpolate(method="spline", order=order)
                                            elif interp_method == "quadratic":
                                                df_resampled_full = df_final_reg[target_col].resample(freq_code).mean().interpolate(method="polynomial", order=2)
                                            elif interp_method == "polynomial":
                                                df_resampled_full = df_final_reg[target_col].resample(freq_code).mean().interpolate(method="polynomial", order=order)
                                        elif "Forward Fill" in reg_method:
                                            df_resampled_full = df_final_reg[target_col].resample(freq_code).mean().ffill()
                                        elif "Backward Fill" in reg_method:
                                            df_resampled_full = df_final_reg[target_col].resample(freq_code).mean().bfill()
                                    elif "Агрегация" in reg_method:
                                        df_resampled_full = df_final_reg[target_col].resample(freq_code).mean()
                                    
                                    # Возвращаем в DataFrame
                                    df_final_reg = df_resampled_full.dropna().reset_index()
                                    
                                    # Синхронизация всех рабочих копий
                                    st.session_state.df = df_final_reg.copy()
                                    st.session_state.validation_ready = False
                                    st.session_state.show_regular_preview = False
                                    
                                    work_dfs = [
                                        "df_missing_work", "df_pattern_work", "df_range_work",
                                        "df_outlier_work", "df_inclusion_work", "df_referential_work",
                                        "df_text_work", "df_consistency_work", "df_uniqueness_work",
                                        "df_regularity_work", "df_regular_work"
                                    ]
                                    for work_df_name in work_dfs:
                                        if work_df_name in st.session_state:
                                            del st.session_state[work_df_name]
                                    
                                    if "val_results" in st.session_state:
                                        del st.session_state.val_results
                                    
                                    st.success("✅ Регулярность частоты восстановлена! Перезапустите валидацию.")
                                    st.rerun()
                            
                            with c_cancel_reg:
                                if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_regularity"):
                                    st.session_state.show_regular_preview = False
                                    st.rerun()
                        else:
                            # Показываем только исходный ряд
                            fig = px.line(x=original_series.index, y=original_series.values,
                                        labels={'x': 'Дата', 'y': target_col},
                            )
                            fig.update_layout(height=400, margin=dict(l=40, r=20, t=40, b=40))
                            st.plotly_chart(fig, use_container_width=True)
                            
                            st.info("💡 Выберите метод обработки и нажмите 'Применить преобразование' для просмотра результата.")
                else:
                    st.warning("⚠️ Нет числовых колонок для анализа.")
            
            # ── ПРАВАЯ КОЛОНКА: МЕТРИКИ КАЧЕСТВА ───────────
            with c3:
                st.markdown("######  Метрики качества")
                
                with st.container(border=True):
                    st.markdown("**До преобразования:**")
                    st.metric("Частота", current_freq_display)
                    st.metric("Пропусков в ряду", f"{gap_count:,}".replace(",", " "))
                    # Форматируем Std с разделителем тысяч
                    interval_std_formatted = f"{interval_std:,.0f}".replace(",", " ")
                    st.metric("Std интервалов", f"{interval_std_formatted} сек")
                    st.metric("Статус", "✅ Регулярная" if is_regular else "❌ Нерегулярная",
                            delta="Исправлено" if st.session_state.get('show_regular_preview') else None)
                
                if st.session_state.show_regular_preview:
                    # Метрики после преобразования
                    new_intervals = df_resampled.index.to_series().diff().dropna()
                    new_intervals_sec = new_intervals.dt.total_seconds()
                    new_std = new_intervals_sec.std() if len(new_intervals_sec) > 1 else 0
                    new_gaps = sum(1 for i in new_intervals if i > new_intervals.mode().iloc[0] * 1.5) if len(new_intervals.mode()) > 0 else 0
                    
                    with st.container(border=True):
                        st.markdown("**После преобразования:**")
                        st.metric("Частота", new_inferred_freq if new_inferred_freq else "❌ Нерегулярная")
                        st.metric("Пропусков в ряду", f"{new_gaps:,}".replace(",", " "))
                        # Форматируем Std с разделителем тысяч
                        new_std_formatted = f"{new_std:,.0f}".replace(",", " ")
                        # Форматируем delta с разделителем
                        delta_std = new_std - interval_std
                        if abs(delta_std) > 100:
                            delta_formatted = f"{delta_std:+,.0f}".replace(",", " ")
                        else:
                            delta_formatted = None
                        st.metric("Std интервалов", f"{new_std_formatted} сек", delta=delta_formatted)
                        st.metric("Статус", "✅ Регулярная" if new_is_regular else "❌ Нерегулярная")
                    
                    st.divider()
                    
                    # Рекомендации
                    if not new_is_regular:
                        st.warning("💡 **Рекомендация:** Попробуйте метод 'AsFreq' или увеличьте частоту агрегации.")
                    else:
                        st.success("✅ **Ряд готов** для ARIMA, STL, FFT, сезонного дифференцирования!")
                
                # Сохранённая конфигурация
                if st.session_state.regularity_config:
                    with st.expander("💾 Сохранённая конфигурация", expanded=False):
                        st.json(st.session_state.regularity_config)
    else:
        st.warning("⚠️ Не обнаружена колонка с датами. Убедитесь, что активирован режим временных рядов во вкладке 'Загрузка'.")

    
    # ═══════════════════════════════════════════════════════
    # 🔹 4. ДЕКОМПОЗИЦИЯ ВРЕМЕННОГО РЯДА
    # ═══════════════════════════════════════════════════════
    st.divider()
    st.markdown("###  Декомпозиция временного ряда")
    st.caption("Анализ структуры: Тренд + Сезонность + Цикличность + Остаток. Критично для понимания природы данных.")

    
    # ── ДИАГНОСТИКА ТЕКУЩЕГО СОСТОЯНИЯ ──────────────────
    if st.session_state.primary_date_col and st.session_state.col_types.get("num"):
        date_col = st.session_state.primary_date_col
        num_cols = st.session_state.col_types.get("num", [])
        
        if num_cols:
            df_decomp = st.session_state.df.copy()
            df_decomp[date_col] = pd.to_datetime(df_decomp[date_col])
            df_decomp = df_decomp.sort_values(date_col)
            df_decomp_ts = df_decomp.set_index(date_col)
            
            # Метрики текущего состояния
            c_diag1, c_diag2, c_diag3, c_diag4 = st.columns(4)
            
            with c_diag1:
                st.markdown("**Длина ряда**")
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: #1e293b;'>{len(df_decomp_ts)} наблюдений</div>", 
                        unsafe_allow_html=True)
            
            with c_diag2:
                st.markdown("**Частота**")
                inferred_freq = pd.infer_freq(df_decomp_ts.index.drop_duplicates().sort_values())
                freq_display = inferred_freq if inferred_freq else "❌ Нерегулярная"
                freq_color = "#16a34a" if inferred_freq else "#dc2626"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {freq_color};'>{freq_display}</div>", 
                        unsafe_allow_html=True)
            
            with c_diag3:
                st.markdown("**Пропусков**")
                total_missing = df_decomp_ts[num_cols[0]].isna().sum()
                miss_color = "#dc2626" if total_missing > 0 else "#16a34a"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {miss_color};'>{total_missing}</div>", 
                        unsafe_allow_html=True)
            
            with c_diag4:
                st.markdown("**Готовность к STL**")
                ready = inferred_freq is not None and total_missing == 0 and len(df_decomp_ts) >= 30
                ready_text = "✅ Готов" if ready else "❌ Не готов"
                ready_color = "#16a34a" if ready else "#dc2626"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {ready_color};'>{ready_text}</div>", 
                        unsafe_allow_html=True)
            
            # Пустая строка для отступа
            st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)

            # ── ТЕХНИЧЕСКАЯ СПРАВКА ─────────────────────────────
            with st.expander(" Цели субмодуля ⁞ Декомпозиция временных рядов", expanded=False):
                st.markdown("""
                **Зачем нужна декомпозиция:**
                -  **Понимание структуры:** Разделяет ряд на компоненты (Trend, Seasonal, Cycle, Residual)
                -  **Выбор модели:** Помогает определить, какие модели подходят (ARIMA, ETS, Prophet)
                -  **Прогнозирование:** Каждая компонента прогнозируется отдельно
                -  **Предобработка:** Позволяет удалить тренд/сезонность перед моделированием
                
                **Методы декомпозиции:**
                
                | Метод | Когда использовать | Преимущества | Недостатки |
                |-------|-------------------|--------------|------------|
                | **STL** | Универсальный метод | Устойчив к выбросам, гибкий | Требует регулярный ряд |
                | **Additive** | Постоянная амплитуда сезонности | Простой, быстрый | Не работает с растущей амплитудой |
                | **Multiplicative** | Растущая амплитуда сезонности | Учитывает масштаб | Требует положительных значений |
                
                **Компоненты ряда:**
                - **Trend:** Долгосрочное направление (рост/падение)
                - **Seasonal:** Периодические колебания (день/неделя/месяц/год)
                - **Cycle:** Долгосрочные циклы (>1 года, нерегулярные)
                - **Residual:** Случайный шум (должен быть белым шумом)
                
                **Диагностика остатков:**
                - ✅ **ADF Test:** Проверка стационарности (p < 0.05)
                - ✅ **Ljung-Box:** Проверка автокорреляции (p > 0.05 = белый шум)
                - ✅ **Jarque-Bera:** Проверка нормальности (p > 0.05)
                - ✅ **ARCH-LM:** Проверка гомоскедастичности (p > 0.05)
                
                **Контроль качества:**
                - Остатки должны быть **белым шумом** (нет автокорреляции)
                - Остатки должны быть **стационарны** (ADF test)
                - Остатки должны быть **нормально распределены** (JB test)
                """)
            
            # ────────────────────────────────────────────────
            # 🎮 ПЕСОЧНИЦА: ДЕКОМПОЗИЦИЯ
            # ────────────────────────────────────────────────
            
            # Инициализация session_state
            if "show_decomp_preview" not in st.session_state:
                st.session_state.show_decomp_preview = False
            
            # ── ЛЕВАЯ КОЛОНКА: ПАНЕЛЬ УПРАВЛЕНИЯ ───────────
            c1, c2, c3 = st.columns([27, 52, 21])
            
            with c1:
                st.markdown("######  Панель управления")
                
                # Выбор числовой колонки
                target_col = st.selectbox(
                    "Числовой признак:",
                    options=num_cols,
                    index=0,
                    key="decomp_target_col"
                )
                
                # Выбор метода декомпозиции
                decomp_method = st.radio(
                    "Метод декомпозиции:",
                    ["STL (рекомендуется)", "Additive (классический)", "Multiplicative (классический)"],
                    key="decomp_method",
                    label_visibility="collapsed"
                )
                
                # Параметры STL
                if "STL" in decomp_method:
                    st.markdown("**Параметры STL:**")
                    
                    # Сезонный период
                    if inferred_freq:
                        if 'D' in inferred_freq:
                            default_period = 7
                        elif 'W' in inferred_freq:
                            default_period = 52
                        elif 'M' in inferred_freq:
                            default_period = 12
                        elif 'Q' in inferred_freq:
                            default_period = 4
                        else:
                            default_period = 12
                    else:
                        default_period = 12
                    
                    seasonal_period = st.number_input(
                        "Сезонный период:",
                        min_value=2,
                        max_value=365,
                        value=default_period,
                        step=1,
                        key="stl_period",
                        help="Длина сезонного цикла (7 для дневных, 12 для месячных, 52 для недельных)"
                    )
                    
                    # Robust
                    robust = st.checkbox("Robust (устойчивость к выбросам)", value=True, key="stl_robust")
                    
                    # Seasonal window
                    seasonal_window = st.selectbox(
                        "Сглаживание сезонности:",
                        [7, 9, 11, 13, 15],
                        index=0,
                        key="stl_seasonal_window",
                        help="Нечётное число > 1. Больше = глаже сезонность"
                    )
                    
                    # Trend window
                    trend_window = st.selectbox(
                        "Сглаживание тренда:",
                        [None, 15, 21, 31, 51],
                        index=0,
                        key="stl_trend_window",
                        help="None = автоматический выбор. Нечётное число > 1."
                    )
                
                else:
                    # Параметры классической декомпозиции
                    st.markdown("**Параметры классической декомпозиции:**")
                    
                    if inferred_freq:
                        if 'D' in inferred_freq:
                            default_period = 7
                        elif 'W' in inferred_freq:
                            default_period = 52
                        elif 'M' in inferred_freq:
                            default_period = 12
                        elif 'Q' in inferred_freq:
                            default_period = 4
                        else:
                            default_period = 12
                    else:
                        default_period = 12
                    
                    classic_period = st.number_input(
                        "Сезонный период:",
                        min_value=2,
                        max_value=365,
                        value=default_period,
                        step=1,
                        key="classic_period"
                    )
                
                st.divider()
                
                # Кнопки действий
                if st.button("▶ Выполнить декомпозицию", type="primary", use_container_width=True, key="btn_apply_decomp"):
                    st.session_state.show_decomp_preview = True
                    st.rerun()
                
                if st.button("↶ Сбросить", use_container_width=True, key="btn_reset_decomp"):
                    st.session_state.show_decomp_preview = False
                    st.rerun()
            
            # ── ЦЕНТРАЛЬНАЯ КОЛОНКА: ВИЗУАЛИЗАЦИЯ ──────────
            with c2:
                st.markdown("###### Декомпозиция временного ряда")
                
                if target_col in df_decomp_ts.columns:
                    original_series = df_decomp_ts[target_col].dropna()
                    
                    if st.session_state.show_decomp_preview:
                        # Проверка готовности
                        if inferred_freq is None:
                            st.error("❌ Ряд нерегулярный. Сначала выполните модуль 'Регулярность частоты'.")
                        elif len(original_series) < 30:
                            st.error(f"❌ Недостаточно данных: {len(original_series)} наблюдений (минимум 30).")
                        elif total_missing > 0:
                            st.error(f"❌ Есть пропуски: {total_missing}. Сначала обработайте пропуски.")
                        else:
                            try:
                                # Выполняем декомпозицию
                                if "STL" in decomp_method:
                                    from statsmodels.tsa.seasonal import STL
                                    
                                    stl = STL(
                                        original_series,
                                        period=seasonal_period,
                                        seasonal=seasonal_window,
                                        trend=trend_window,
                                        robust=robust
                                    )
                                    result = stl.fit()
                                    
                                    trend = result.trend
                                    seasonal = result.seasonal
                                    residual = result.resid
                                    
                                    # Цикличность = Trend - сглаженный тренд (упрощённо)
                                    cycle = trend - trend.rolling(window=seasonal_period*2, center=True, min_periods=1).mean()
                                    
                                    decomp_type = "STL"
                                    
                                elif "Additive" in decomp_method:
                                    from statsmodels.tsa.seasonal import seasonal_decompose
                                    
                                    result = seasonal_decompose(original_series, model='additive', period=classic_period)
                                    trend = result.trend
                                    seasonal = result.seasonal
                                    residual = result.resid
                                    cycle = pd.Series(0, index=original_series.index)  # Нет явной цикличности
                                    
                                    decomp_type = "Additive"
                                    
                                elif "Multiplicative" in decomp_method:
                                    from statsmodels.tsa.seasonal import seasonal_decompose
                                    
                                    if (original_series <= 0).any():
                                        st.error("❌ Multiplicative декомпозиция требует положительных значений.")
                                    else:
                                        result = seasonal_decompose(original_series, model='multiplicative', period=classic_period)
                                        trend = result.trend
                                        seasonal = result.seasonal
                                        residual = result.resid
                                        cycle = pd.Series(1, index=original_series.index)  # Нет явной цикличности
                                        
                                        decomp_type = "Multiplicative"
                                
                                # ── ГРАФИКИ ДЕКОМПОЗИЦИИ ────────────
                                fig = make_subplots(
                                    rows=5, cols=1,
                                    subplot_titles=(
                                        f" Исходный ряд: {target_col}",
                                        f" Тренд (Trend)",
                                        f" Сезонность (Seasonal)",
                                        f" Цикличность (Cycle)",
                                        f" Остатки (Residual)"
                                    ),
                                    vertical_spacing=0.06
                                )
                                
                                # Исходный ряд
                                fig.add_trace(
                                    go.Scatter(
                                        x=original_series.index, y=original_series.values,
                                        mode='lines',
                                        name='Исходные данные',
                                        line=dict(color='#048A81', width=2),
                                        showlegend=False
                                    ),
                                    row=1, col=1
                                )
                                
                                # Тренд
                                fig.add_trace(
                                    go.Scatter(
                                        x=trend.index, y=trend.values,
                                        mode='lines',
                                        name='Trend',
                                        line=dict(color='#DC2626', width=2),
                                        showlegend=False
                                    ),
                                    row=2, col=1
                                )
                                
                                # Сезонность
                                fig.add_trace(
                                    go.Scatter(
                                        x=seasonal.index, y=seasonal.values,
                                        mode='lines',
                                        name='Seasonal',
                                        line=dict(color='#2563EB', width=2),
                                        showlegend=False
                                    ),
                                    row=3, col=1
                                )
                                
                                # Цикличность
                                fig.add_trace(
                                    go.Scatter(
                                        x=cycle.index, y=cycle.values,
                                        mode='lines',
                                        name='Cycle',
                                        line=dict(color='#9333EA', width=2),
                                        showlegend=False
                                    ),
                                    row=4, col=1
                                )
                                
                                # Остатки
                                fig.add_trace(
                                    go.Scatter(
                                        x=residual.index, y=residual.values,
                                        mode='lines',
                                        name='Residual',
                                        line=dict(color='#6B7280', width=1),
                                        showlegend=False
                                    ),
                                    row=5, col=1
                                )
                                
                                fig.update_layout(
                                    height=1000,
                                    margin=dict(l=50, r=20, t=80, b=40),
                                    hovermode='x unified',
                                )
                                
                                fig.update_xaxes(title_text="Дата", row=5, col=1)
                                
                                for i in range(1, 6):
                                    fig.update_yaxes(title_text="Значение", row=i, col=1)
                                
                                fig.update_annotations(
                                    font=dict(size=13, color="#1e293b"),
                                    yshift=10
                                )
                                
                                st.plotly_chart(fig, use_container_width=True, key="decomp_main_chart")
                                
                                st.success(f"✅ Декомпозиция ({decomp_type}) выполнена успешно!")
                                
                                # Сохраняем результаты в session_state
                                st.session_state.decomp_result = {
                                    'method': decomp_type,
                                    'trend': trend,
                                    'seasonal': seasonal,
                                    'cycle': cycle,
                                    'residual': residual,
                                    'target_col': target_col
                                }
                                
                            except Exception as e:
                                st.error(f"❌ Ошибка декомпозиции: {e}")
                                import traceback
                                st.code(traceback.format_exc(), language="python")
                    else:
                        # Показываем только исходный ряд
                        fig = px.line(
                            x=original_series.index,
                            y=original_series.values,
                            labels={'x': 'Дата', 'y': target_col},
                        )
                        fig.update_layout(height=400, margin=dict(l=40, r=20, t=40, b=20))
                        st.plotly_chart(fig, use_container_width=True, key="decomp_main_chart")
                        
                        st.info("💡 Выберите параметры и нажмите 'Выполнить декомпозицию' для анализа структуры ряда.")
                else:
                    st.warning("⚠️ Выбранная колонка не найдена в данных.")
            
            # ── ПРАВАЯ КОЛОНКА: МЕТРИКИ КАЧЕСТВА ───────────
            with c3:
                st.markdown("###### Диагностика остатков")
                
                if st.session_state.show_decomp_preview and 'decomp_result' in st.session_state:
                    residual = st.session_state.decomp_result['residual']
                    
                    # Диагностика остатков
                    try:
                        from statsmodels.tsa.stattools import adfuller
                        from statsmodels.stats.diagnostic import acorr_ljungbox
                        from scipy.stats import jarque_bera
                        from statsmodels.stats.diagnostic import het_arch
                        
                        # 1. ADF Test (стационарность)
                        adf_res = adfuller(residual.dropna(), autolag='AIC')
                        adf_p = adf_res[1]
                        adf_stat = "✅ Стационарны" if adf_p < 0.05 else "❌ Нестационарны"
                        
                        # 2. Ljung-Box (автокорреляция)
                        lb_res = acorr_ljungbox(residual.dropna(), lags=[10])
                        lb_p = lb_res['lb_pvalue'].iloc[0] if isinstance(lb_res, pd.DataFrame) else lb_res[1][0]
                        lb_stat = "✅ Белый шум" if lb_p > 0.05 else "❌ Есть АК"
                        
                        # 3. Jarque-Bera (нормальность)
                        jb_res = jarque_bera(residual.dropna())
                        jb_p = jb_res.pvalue if hasattr(jb_res, 'pvalue') else jb_res[1]
                        jb_stat = "✅ Нормальны" if jb_p > 0.05 else "❌ Отклонение"
                        
                        # 4. ARCH-LM (гетероскедастичность)
                        try:
                            arch_lm, arch_p, _, _ = het_arch(residual.dropna(), nlags=5)
                            arch_stat = "✅ Гомоскедастичны" if arch_p > 0.05 else "❌ Гетероскедастичны"
                        except:
                            arch_stat = "⚠️ Недостаточно данных"
                            arch_p = None
                        
                        # Отображаем метрики через st.metric
                        with st.container(border=True):
                            st.markdown("**Тесты на остатки:**")
                            
                            st.metric(
                                label="ADF TEST (стационарность)",
                                value=adf_stat,
                                delta=f"p-value: {adf_p:.4f}"
                            )
                            
                            st.metric(
                                label="LJUNG-BOX (автокорреляция)",
                                value=lb_stat,
                                delta=f"p-value: {lb_p:.4f}"
                            )
                            
                            st.metric(
                                label="JARQUE-BERA (нормальность)",
                                value=jb_stat,
                                delta=f"p-value: {jb_p:.4f}"
                            )
                            
                            if arch_p is not None:
                                st.metric(
                                    label="ARCH-LM (волатильность)",
                                    value=arch_stat,
                                    delta=f"p-value: {arch_p:.4f}"
                                )
                            else:
                                st.metric(
                                    label="ARCH-LM (волатильность)",
                                    value=arch_stat,
                                    delta="N/A"
                                )
                        
                        st.divider()
                        
                        # Общая оценка
                        tests_passed = sum([
                            adf_p < 0.05,
                            lb_p > 0.05,
                            jb_p > 0.05,
                            arch_p > 0.05 if arch_p else False
                        ])
                        
                        if tests_passed >= 3:
                            st.success(f"✅ **Качество декомпозиции:** Отличное ({tests_passed}/4 тестов)")
                        elif tests_passed >= 2:
                            st.warning(f"⚠️ **Качество декомпозиции:** Удовлетворительное ({tests_passed}/4 тестов)")
                        else:
                            st.error(f"❌ **Качество декомпозиции:** Плохое ({tests_passed}/4 тестов)")
                        
                        # Рекомендации
                        st.divider()
                        st.markdown("**💡 Рекомендации:**")
                        
                        if adf_p >= 0.05:
                            st.markdown("- Остатки нестационарны → попробуйте удалить тренд")
                        if lb_p <= 0.05:
                            st.markdown("- Есть автокорреляция → увеличьте сезонный период")
                        if jb_p <= 0.05:
                            st.markdown("- Остатки не нормальны → проверьте выбросы")
                        if arch_p and arch_p <= 0.05:
                            st.markdown("- Гетероскедастичность → используйте GARCH модели")
                        
                        if tests_passed >= 3:
                            st.markdown("- ✅ Остатки близки к белому шуму — декомпозиция успешна!")
                    
                    except Exception as e:
                        st.error(f"❌ Ошибка диагностики: {e}")
                        import traceback
                        with st.expander("🔍 Stack trace"):
                            st.code(traceback.format_exc(), language="python")
                
                else:
                    st.info("ℹ️ Выполните декомпозицию для просмотра диагностики остатков.")
        
        else:
            st.warning("⚠️ В датасете нет числовых колонок для анализа.")
    else:
        st.warning("⚠️ Не обнаружены колонки с датами или числовыми данными. Убедитесь, что активирован режим временных рядов.")


    # ═══════════════════════════════════════════════════════
    # 🔹 5. СТАБИЛИЗАЦИЯ ДИСПЕРСИИ (VARIANCE STABILIZATION)
    # ═══════════════════════════════════════════════════════
    st.divider()
    st.markdown("### Стабилизация дисперсии (трансформации)")
    st.caption("Стабилизация дисперсии временного ряда необходима для приведения ряда к стационарному виду - выравнивает затухание/усиление амплитуды колебаний.")

    # ── ДИАГНОСТИКА ТЕКУЩЕГО СОСТОЯНИЯ ──────────────────
    if st.session_state.primary_date_col and st.session_state.col_types.get("num"):
        date_col = st.session_state.primary_date_col
        num_cols = st.session_state.col_types.get("num", [])
        
        if num_cols:
            df_var = st.session_state.df.copy()
            df_var[date_col] = pd.to_datetime(df_var[date_col])
            df_var = df_var.sort_values(date_col)
            df_var_ts = df_var.set_index(date_col)
            
            # ── ФУНКЦИЯ ТЕСТА НА ГЕТЕРОСКЕДАСТИЧНОСТЬ ─────
            def test_heteroskedasticity(series, window=30):
                """Тест на гетероскедастичность через скользящее std"""
                if len(series) < window * 2:
                    return {
                        'bp_pvalue': None,
                        'is_hetero': None,
                        'rolling_std_corr': None,
                        'amplitude_ratio': None
                    }
                
                # 1. Тест Бройша-Пагана (если есть statsmodels)
                try:
                    from statsmodels.stats.diagnostic import het_breuschpagan
                    import statsmodels.api as sm
                    
                    # Регрессия на тренд для получения остатков
                    X = sm.add_constant(np.arange(len(series)))
                    model = sm.OLS(series, X).fit()
                    resid = model.resid
                    
                    # Тест Бройша-Пагана
                    _, bp_pvalue, _, _ = het_breuschpagan(resid, X)
                    is_hetero = bp_pvalue < 0.05
                except Exception:
                    bp_pvalue = None
                    is_hetero = None
                
                # 2. Корреляция между rolling_std и rolling_mean
                rolling_mean = series.rolling(window=window).mean()
                rolling_std = series.rolling(window=window).std()
                
                valid_mask = rolling_mean.notna() & rolling_std.notna() & (rolling_std > 0)
                if valid_mask.sum() > 10:
                    corr = rolling_mean[valid_mask].corr(rolling_std[valid_mask])
                else:
                    corr = None
                
                # 3. Отношение амплитуд (последняя треть / первая треть)
                n = len(series)
                first_third = series.iloc[:n//3]
                last_third = series.iloc[2*n//3:]
                if len(first_third) > 0 and len(last_third) > 0:
                    amplitude_ratio = last_third.std() / (first_third.std() + 1e-10)
                else:
                    amplitude_ratio = None
                
                return {
                    'bp_pvalue': bp_pvalue,
                    'is_hetero': is_hetero,
                    'rolling_std_corr': corr,
                    'amplitude_ratio': amplitude_ratio
                }
            
            # ── МЕТРИКИ ТЕКУЩЕГО СОСТОЯНИЯ ─────────────────
            c_diag1, c_diag2, c_diag3, c_diag4 = st.columns(4)
            
            target_col_var = st.session_state.get('ts_props_v10_target_col', num_cols[0])
            series_raw = df_var_ts[target_col_var].dropna().astype(float)
            
            hetero_raw = test_heteroskedasticity(series_raw)
            
            with c_diag1:
                st.markdown("**Анализ ряда**")
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: #1e293b;'>`{target_col_var}`</div>", 
                        unsafe_allow_html=True)
            
            with c_diag2:
                st.markdown("**Тест Бройша-Пагана**")
                if hetero_raw['bp_pvalue'] is not None:
                    bp_color = "#dc2626" if hetero_raw['is_hetero'] else "#16a34a"
                    bp_text = "Гетероскедастичность" if hetero_raw['is_hetero'] else "Гомоскедастичность"
                    st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {bp_color};'>"
                            f"{bp_text} (p={hetero_raw['bp_pvalue']:.4f})</div>", 
                            unsafe_allow_html=True)
                else:
                    st.markdown("<div style='font-size: 14px; color: #6b7280;'>Недоступно</div>", unsafe_allow_html=True)
            
            with c_diag3:
                st.markdown("**Корр. std vs mean**")
                if hetero_raw['rolling_std_corr'] is not None:
                    corr_val = hetero_raw['rolling_std_corr']
                    corr_color = "#dc2626" if abs(corr_val) > 0.5 else "#16a34a"
                    st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {corr_color};'>"
                            f"{corr_val:.3f}</div>", unsafe_allow_html=True)
                else:
                    st.markdown("<div style='font-size: 14px; color: #6b7280;'>N/A</div>", unsafe_allow_html=True)
            
            with c_diag4:
                st.markdown("**Амплитуда (конец/начало)**")
                if hetero_raw['amplitude_ratio'] is not None:
                    ratio = hetero_raw['amplitude_ratio']
                    ratio_color = "#dc2626" if ratio > 1.5 else "#16a34a"
                    st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {ratio_color};'>"
                            f"{ratio:.2f}x</div>", unsafe_allow_html=True)
                else:
                    st.markdown("<div style='font-size: 14px; color: #6b7280;'>N/A</div>", unsafe_allow_html=True)
            
            # Отступ
            st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)
            
            # ── ТЕХНИЧЕСКАЯ СПРАВКА ─────────────────────────
            with st.expander("Цели субмодуля ⁞ Стабилизация дисперсии", expanded=False):
                st.markdown("""
                **Зачем нужна стабилизация дисперсии.**
                Если амплитуда колебаний растёт вместе с уровнем ряда (гетероскедастичность), 
                модели ошибаются. Трансформации «сжимают» большие значения и «растягивают» маленькие, 
                обеспечивая гомоскедастичность — ключевое допущение ARIMA/регрессии.
                - **Гомоскедастичность** — ключевое допущение ARIMA, регрессии, Gaussian processes
                - **Стабильность прогноза** — постоянная ширина доверительных интервалов
                - **Корректность тестов** — ADF, Ljung-Box, Jarque-Bera требуют стабильной дисперсии
                - **Качество ML** — LSTM, XGBoost лучше обучаются на гомоскедастичных данных
                
                **Диагностика гетероскедастичности:**
                - ✅ **Тест Бройша-Пагана** (p < 0.05 → гетероскедастичность)
                - ✅ **Корреляция rolling_std vs rolling_mean** (|r| > 0.5 → проблема)
                - ✅ **Отношение амплитуд** (конец/начало > 1.5 → растущая дисперсия)
                
                **Методы трансформации:**
                
                | Метод | Формула | λ (Box-Cox) | Когда использовать |
                |-------|---------|-------------|-------------------|
                | **Box-Cox** | (y^λ - 1)/λ | Авто | Только y > 0, универсальный |
                | **Yeo-Johnson** | piecewise | Авто | Есть y ≤ 0, расширение Box-Cox |
                | **Log (ln)** | ln(y) | λ→0 | Экспоненциальный рост, y > 0 |
                | **Log1p** | ln(1+y) | — | Есть нули, y ≥ 0 |
                | **Square Root** | √y | λ=0.5 | Count data (Пуассон), y ≥ 0 |
                | **Reciprocal** | 1/y | λ=-1 | Сильно растущие ряды, y > 0 |
                
                **Как выбрать метод:**
                1. Если **все значения > 0** → Box-Cox (оптимальный λ подбирается автоматически)
                2. Если **есть отрицательные** → Yeo-Johnson
                3. Если **есть нули** → Log1p или Square Root
                4. Если **данные — количество событий** → Square Root
                5. Если **нужна интерпретируемость** → Log (коэффициенты = эластичность)
                
                **Обратимость:**
                - Все трансформации обратимы — параметры сохраняются в `session_state`
                - Перед прогнозом в исходной шкале применяется **обратное преобразование**
                - Для Box-Cox/Yeo-Johnson сохраняется λ, для Log — база (e)
                
                **⚠️ Почитать:**
                - Box-Cox transformation: https://en.wikipedia.org/wiki/Power_transform
                - Гомоскедастичность в TS: https://otexts.com/fpp3/transformations.html
                """)
            
            # Отступ
            st.markdown("<div style='height: 25px;'></div>", unsafe_allow_html=True)

            # ────────────────────────────────────────────────
            # 🎮 ПЕСОЧНИЦА: СТАБИЛИЗАЦИЯ ДИСПЕРСИИ
            # ────────────────────────────────────────────────
            
            # Инициализация session_state
            if "show_variance_preview" not in st.session_state:
                st.session_state.show_variance_preview = False
            if "variance_transform_params" not in st.session_state:
                st.session_state.variance_transform_params = {}
            
            # ── ЛЕВАЯ КОЛОНКА: ПАНЕЛЬ УПРАВЛЕНИЯ ───────────
            c1, c2, c3 = st.columns([27, 52, 21])
            
            with c1:
                st.markdown("###### Панель управления")
                
                # Выбор числовой колонки
                target_col = st.selectbox(
                    "Исследуемый признак:",
                    options=num_cols,
                    index=num_cols.index(target_col_var) if target_col_var in num_cols else 0,
                    key="variance_target_col"
                )
                
                series = df_var_ts[target_col].dropna().astype(float)
                
                # Проверка на отрицательные значения
                has_negative = (series <= 0).any()
                has_zero = (series == 0).any()
                
                # Формируем список доступных методов
                method_options = []
                method_descriptions = {}
                
                if not has_negative:
                    method_options.append("Box-Cox (авто λ)")
                    method_descriptions["Box-Cox (авто λ)"] = "Универсальный метод, только y > 0. Автоматически подбирает оптимальный λ."
                    
                    method_options.append("Log (натуральный)")
                    method_descriptions["Log (натуральный)"] = "ln(y), только y > 0. Интерпретируемый, для экспоненциального роста."
                    
                    method_options.append("Reciprocal (1/y)")
                    method_descriptions["Reciprocal (1/y)"] = "1/y, только y > 0. Агрессивное сжатие больших значений."
                
                if not has_negative:
                    method_options.append("Square Root")
                    method_descriptions["Square Root"] = "√y, только y ≥ 0. Для count data (распределение Пуассона)."
                    
                    if has_zero:
                        method_options.append("Log1p (ln(1+y))")
                        method_descriptions["Log1p (ln(1+y))"] = "Безопасный логарифм для данных с нулями."
                
                # Yeo-Johnson всегда доступен
                method_options.append("Yeo-Johnson (авто λ)")
                method_descriptions["Yeo-Johnson (авто λ)"] = "Работает с любыми значениями (включая отрицательные). Расширение Box-Cox."
                
                # Если нет доступных методов (все отрицательные) — только Yeo-Johnson
                if not method_options:
                    method_options = ["Yeo-Johnson (авто λ)"]
                    method_descriptions["Yeo-Johnson (авто λ)"] = "Единственный доступный метод для данных с отрицательными значениями."
                
                # Предупреждения о доступности
                if has_negative:
                    st.warning("⚠️ Обнаружены **отрицательные значения** — доступны только Yeo-Johnson")
                elif has_zero:
                    st.info("ℹ️ Обнаружены **нули** — Box-Cox/Log недоступны, используйте Log1p или Yeo-Johnson")
                
                # Выбор метода
                variance_method = st.radio(
                    "Метод трансформации:",
                    options=method_options,
                    index=0,
                    key="variance_method",
                    label_visibility="collapsed"
                )
                
                # Описание выбранного метода
                st.markdown(
                    f'<div style="background: #f0f9ff; border-left: 3px solid #0284c7; padding: 8px 12px; '
                    f'margin: 8px 0; border-radius: 4px;">'
                    f'<span style="color: #0369a1; font-size: 13px;">'
                    f'💡 <strong>{variance_method}:</strong> {method_descriptions[variance_method]}'
                    f'</span></div>',
                    unsafe_allow_html=True
                )
                
                # Ручной подбор λ (для Box-Cox/Yeo-Johnson)
                if "Box-Cox" in variance_method or "Yeo-Johnson" in variance_method:
                    auto_lambda = st.checkbox("Автоматический подбор λ (MLE)", value=True, key="auto_lambda_var")
                    if not auto_lambda:
                        lambda_value = st.slider(
                            "Ручной подбор λ:",
                            min_value=-2.0,
                            max_value=2.0,
                            value=0.0,
                            step=0.1,
                            key="manual_lambda_var",
                            help="λ=0 → логарифм, λ=0.5 → корень, λ=1 → без изменений, λ=-1 → обратное"
                        )
                    else:
                        lambda_value = None
                
                st.divider()
                
                # Кнопки действий
                if st.button("▶ Применить трансформацию", type="primary", use_container_width=True, key="btn_apply_variance"):
                    st.session_state.show_variance_preview = True
                    st.rerun()
                
                if st.button("↶ Сбросить", use_container_width=True, key="btn_reset_variance"):
                    st.session_state.show_variance_preview = False
                    st.rerun()
            
            # ── ЦЕНТРАЛЬНАЯ КОЛОНКА: ВИЗУАЛИЗАЦИЯ ──────────
            with c2:
                st.markdown("###### Визуализация: До / После")
                
                if target_col in df_var_ts.columns:
                    original_series = df_var_ts[target_col].dropna().astype(float)
                    
                    if st.session_state.show_variance_preview:
                        try:
                            from sklearn.preprocessing import PowerTransformer
                            from scipy import stats
                            
                            # Применяем трансформацию
                            values = original_series.values.reshape(-1, 1)
                            
                            if "Box-Cox" in variance_method:
                                if auto_lambda:
                                    transformed, lambda_opt = stats.boxcox(values.flatten() + 1e-10)
                                    transformed_series = pd.Series(transformed, index=original_series.index)
                                    lambda_used = lambda_opt
                                else:
                                    transformed, _ = stats.boxcox(values.flatten() + 1e-10, lmbda=lambda_value)
                                    transformed_series = pd.Series(transformed, index=original_series.index)
                                    lambda_used = lambda_value
                                method_name = "Box-Cox"
                            
                            elif "Yeo-Johnson" in variance_method:
                                pt = PowerTransformer(method='yeo-johnson', standardize=False)
                                if auto_lambda:
                                    transformed = pt.fit_transform(values)
                                    lambda_used = pt.lambdas_[0]
                                else:
                                    # Ручной λ для Yeo-Johnson — через кастомную реализацию
                                    def yeo_johnson_manual(y, lmbda):
                                        result = np.zeros_like(y, dtype=float)
                                        pos = y >= 0
                                        neg = ~pos
                                        if lmbda != 0:
                                            result[pos] = ((y[pos] + 1) ** lmbda - 1) / lmbda
                                        else:
                                            result[pos] = np.log1p(y[pos])
                                        if lmbda != 2:
                                            result[neg] = -((-y[neg] + 1) ** (2 - lmbda) - 1) / (2 - lmbda)
                                        else:
                                            result[neg] = -np.log1p(-y[neg])
                                        return result
                                    transformed = yeo_johnson_manual(values.flatten(), lambda_value)
                                    lambda_used = lambda_value
                                transformed_series = pd.Series(transformed.flatten(), index=original_series.index)
                                method_name = "Yeo-Johnson"
                            
                            elif "Log (натуральный)" in variance_method:
                                transformed_series = np.log(original_series)
                                lambda_used = 0
                                method_name = "Log"
                            
                            elif "Log1p" in variance_method:
                                transformed_series = np.log1p(original_series)
                                lambda_used = 0
                                method_name = "Log1p"
                            
                            elif "Square Root" in variance_method:
                                transformed_series = np.sqrt(original_series)
                                lambda_used = 0.5
                                method_name = "Square Root"
                            
                            elif "Reciprocal" in variance_method:
                                transformed_series = 1 / original_series
                                lambda_used = -1
                                method_name = "Reciprocal"
                            
                            # ── ГРАФИК СРАВНЕНИЯ (2 ряда) ───────
                            fig = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    f" Исходный ряд: {target_col}",
                                    f" После трансформации ({method_name}, λ={lambda_used:.3f})"
                                ),
                                vertical_spacing=0.12
                            )
                            
                            # Исходный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=original_series.index, y=original_series.values,
                                    mode='lines',
                                    name='Исходные данные',
                                    line=dict(color='#048A81', width=2),
                                    showlegend=False
                                ),
                                row=1, col=1
                            )
                            
                            # Трансформированный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=transformed_series.index, y=transformed_series.values,
                                    mode='lines',
                                    name='После трансформации',
                                    line=dict(color='#DC2626', width=2),
                                    showlegend=False
                                ),
                                row=2, col=1
                            )
                            
                            fig.update_layout(
                                height=600,
                                margin=dict(l=50, r=20, t=80, b=40),
                                hovermode='x unified',
                            )
                            
                            fig.update_xaxes(title_text="Дата", row=2, col=1)
                            fig.update_yaxes(title_text="Значение", row=1, col=1)
                            fig.update_yaxes(title_text="Трансформированное значение", row=2, col=1)
                            
                            fig.update_annotations(font=dict(size=13, color="#1e293b"), yshift=10)
                            
                            st.plotly_chart(fig, use_container_width=True, key="variance_main_chart")
                            
                            # ── ГРАФИК ROLLING STD (ключевой для гетероскедастичности) ──
                            st.markdown("##### Скользящее стандартное отклонение (ключевой тест)")
                            st.caption("Если линия идёт вверх → дисперсия растёт. После трансформации должна быть горизонтальной.")
                            
                            window = min(30, len(original_series) // 5)
                            rolling_std_before = original_series.rolling(window=window).std()
                            rolling_std_after = transformed_series.rolling(window=window).std()
                            
                            fig_rolling = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    f"Rolling Std (окно={window}) — ДО",
                                    f"Rolling Std (окно={window}) — ПОСЛЕ"
                                ),
                                vertical_spacing=0.12
                            )
                            
                            fig_rolling.add_trace(
                                go.Scatter(x=rolling_std_before.index, y=rolling_std_before.values,
                                        mode='lines', line=dict(color='#DC2626', width=2),
                                        name='Rolling Std (До)', showlegend=False),
                                row=1, col=1
                            )
                            
                            fig_rolling.add_trace(
                                go.Scatter(x=rolling_std_after.index, y=rolling_std_after.values,
                                        mode='lines', line=dict(color='#16a34a', width=2),
                                        name='Rolling Std (После)', showlegend=False),
                                row=2, col=1
                            )
                            
                            fig_rolling.update_layout(height=450, margin=dict(l=50, r=20, t=60, b=40))
                            st.plotly_chart(fig_rolling, use_container_width=True, key="variance_rolling_chart")
                            
                            # ── ГИСТОГРАММЫ СРАВНЕНИЯ ───────────
                            st.markdown("##### Распределение значений")
                            
                            fig_hist = make_subplots(
                                rows=1, cols=2,
                                subplot_titles=("Исходное распределение", "После трансформации")
                            )
                            
                            fig_hist.add_trace(
                                go.Histogram(x=original_series.values, nbinsx=40,
                                            marker_color='#048A81', name='До', showlegend=False),
                                row=1, col=1
                            )
                            
                            fig_hist.add_trace(
                                go.Histogram(x=transformed_series.values, nbinsx=40,
                                            marker_color='#DC2626', name='После', showlegend=False),
                                row=1, col=2
                            )
                            
                            fig_hist.update_layout(height=300, margin=dict(l=40, r=20, t=50, b=40), barmode='overlay')
                            st.plotly_chart(fig_hist, use_container_width=True, key="variance_hist_chart")
                            
                            # ── СОХРАНЕНИЕ РЕЗУЛЬТАТОВ ──────────
                            st.session_state.variance_transform_result = {
                                'method': method_name,
                                'lambda': lambda_used,
                                'original_series': original_series,
                                'transformed_series': transformed_series,
                                'target_col': target_col
                            }
                            
                            # Статус
                            hetero_after = test_heteroskedasticity(transformed_series)
                            
                            st.divider()
                            
                            if hetero_after['bp_pvalue'] is not None:
                                if not hetero_after['is_hetero']:
                                    st.success(f"✅ **Гетероскедастичность устранена!** Тест Бройша-Пагана: p={hetero_after['bp_pvalue']:.4f}")
                                else:
                                    st.warning(f"⚠️ **Гетероскедастичность сохраняется.** Попробуйте другой метод или ручной подбор λ.")
                            else:
                                st.info("ℹ️ Тест Бройша-Пагана недоступен. Оцените результат визуально.")
                            
                            # Кнопки подтверждения
                            c_ok_var, c_cancel_var = st.columns(2)
                            with c_ok_var:
                                if st.button("✅ Применить к данным", type="primary", use_container_width=True, key="btn_confirm_variance"):
                                    # Применяем трансформацию к основному df
                                    df_final_var = st.session_state.df.copy()
                                    
                                    # Сохраняем оригинальный столбец с суффиксом _original
                                    orig_col_name = f"{target_col}_original"
                                    if orig_col_name not in df_final_var.columns:
                                        df_final_var[orig_col_name] = df_final_var[target_col]
                                    
                                    # Применяем трансформацию
                                    if method_name == "Box-Cox":
                                        df_final_var[target_col] = stats.boxcox(
                                            df_final_var[target_col].astype(float).values + 1e-10, 
                                            lmbda=lambda_used
                                        )
                                    elif method_name == "Yeo-Johnson":
                                        if auto_lambda:
                                            pt = PowerTransformer(method='yeo-johnson', standardize=False)
                                            df_final_var[target_col] = pt.fit_transform(
                                                df_final_var[target_col].astype(float).values.reshape(-1, 1)
                                            ).flatten()
                                        else:
                                            df_final_var[target_col] = yeo_johnson_manual(
                                                df_final_var[target_col].astype(float).values, lambda_used
                                            )
                                    elif method_name == "Log":
                                        df_final_var[target_col] = np.log(df_final_var[target_col].astype(float))
                                    elif method_name == "Log1p":
                                        df_final_var[target_col] = np.log1p(df_final_var[target_col].astype(float))
                                    elif method_name == "Square Root":
                                        df_final_var[target_col] = np.sqrt(df_final_var[target_col].astype(float))
                                    elif method_name == "Reciprocal":
                                        df_final_var[target_col] = 1 / df_final_var[target_col].astype(float)
                                    
                                    # Сохраняем параметры для обратного преобразования
                                    st.session_state.variance_transform_params = {
                                        'method': method_name,
                                        'lambda': lambda_used,
                                        'column': target_col,
                                        'original_col_name': orig_col_name
                                    }
                                    
                                    # Синхронизация
                                    st.session_state.df = df_final_var.copy()
                                    st.session_state.validation_ready = False
                                    st.session_state.show_variance_preview = False
                                    
                                    # Удаляем рабочие копии
                                    work_dfs = [
                                        "df_missing_work", "df_pattern_work", "df_range_work",
                                        "df_outlier_work", "df_inclusion_work", "df_referential_work",
                                        "df_text_work", "df_consistency_work", "df_uniqueness_work",
                                        "df_regularity_work", "df_regular_work", "df_variance_work"
                                    ]
                                    for work_df_name in work_dfs:
                                        if work_df_name in st.session_state:
                                            del st.session_state[work_df_name]
                                    
                                    if "val_results" in st.session_state:
                                        del st.session_state.val_results
                                    
                                    st.success(f"✅ Трансформация **{method_name}** (λ={lambda_used:.3f}) применена!")
                                    st.info(f"💡 Оригинальные значения сохранены в колонке `{orig_col_name}`. "
                                        f"Перезапустите валидацию для обновления статистик.")
                                    st.rerun()
                            
                            with c_cancel_var:
                                if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_variance"):
                                    st.session_state.show_variance_preview = False
                                    st.rerun()
                        
                        except Exception as e:
                            st.error(f"❌ Ошибка трансформации: {e}")
                            import traceback
                            with st.expander("🔍 Stack trace"):
                                st.code(traceback.format_exc(), language="python")
                    
                    else:
                        # Показываем только исходный ряд
                        fig = px.line(
                            x=original_series.index,
                            y=original_series.values,
                            labels={'x': 'Дата', 'y': target_col},
                        )
                        fig.update_layout(height=400, margin=dict(l=40, r=20, t=40, b=40))
                        st.plotly_chart(fig, use_container_width=True, key="variance_main_chart")
                        
                        st.info("💡 Выберите метод трансформации и нажмите **'Применить трансформацию'** для просмотра результата.")
                else:
                    st.warning("⚠️ Выбранная колонка не найдена в данных.")
            
            # ── ПРАВАЯ КОЛОНКА: МЕТРИКИ КАЧЕСТВА ───────────
            with c3:
                st.markdown("###### Метрики качества")
                
                with st.container(border=True):
                    st.markdown("**До трансформации:**")
                    
                    # Тест Бройша-Пагана
                    if hetero_raw['bp_pvalue'] is not None:
                        bp_status = "❌ Гетеро" if hetero_raw['is_hetero'] else "✅ Гомо"
                        st.metric("Тест Бройша-Пагана", bp_status, delta=f"p={hetero_raw['bp_pvalue']:.4f}")
                    else:
                        st.metric("Тест Бройша-Пагана", "N/A")
                    
                    # Корреляция std vs mean
                    if hetero_raw['rolling_std_corr'] is not None:
                        st.metric("Корр. std/mean", f"{hetero_raw['rolling_std_corr']:.3f}",
                                delta="Сильная" if abs(hetero_raw['rolling_std_corr']) > 0.5 else "Слабая")
                    else:
                        st.metric("Корр. std/mean", "N/A")
                    
                    # Отношение амплитуд
                    if hetero_raw['amplitude_ratio'] is not None:
                        st.metric("Амплитуда (к/н)", f"{hetero_raw['amplitude_ratio']:.2f}x",
                                delta="Растёт" if hetero_raw['amplitude_ratio'] > 1.5 else "Стабильна")
                    else:
                        st.metric("Амплитуда (к/н)", "N/A")
                    
                    # Базовые статистики
                    st.divider()
                    st.markdown("**Статистики ряда:**")
                    st.metric("Среднее", f"{series.mean():.2f}")
                    st.metric("Стд. отклонение", f"{series.std():.2f}")
                    st.metric("Мин / Макс", f"{series.min():.2f} / {series.max():.2f}")
                
                # Метрики ПОСЛЕ трансформации
                if st.session_state.show_variance_preview and 'variance_transform_result' in st.session_state:
                    transformed_series = st.session_state.variance_transform_result['transformed_series']
                    hetero_after = test_heteroskedasticity(transformed_series)
                    
                    with st.container(border=True):
                        st.markdown("**После трансформации:**")
                        
                        if hetero_after['bp_pvalue'] is not None:
                            bp_status_after = "❌ Гетеро" if hetero_after['is_hetero'] else "✅ Гомо"
                            delta_bp = hetero_after['bp_pvalue'] - (hetero_raw['bp_pvalue'] or 0)
                            st.metric("Тест Бройша-Пагана", bp_status_after, 
                                    delta=f"Δp={delta_bp:+.4f}")
                        else:
                            st.metric("Тест Бройша-Пагана", "N/A")
                        
                        if hetero_after['rolling_std_corr'] is not None:
                            delta_corr = hetero_after['rolling_std_corr'] - (hetero_raw['rolling_std_corr'] or 0)
                            st.metric("Корр. std/mean", f"{hetero_after['rolling_std_corr']:.3f}",
                                    delta=f"{delta_corr:+.3f}")
                        else:
                            st.metric("Корр. std/mean", "N/A")
                        
                        if hetero_after['amplitude_ratio'] is not None:
                            delta_ratio = hetero_after['amplitude_ratio'] - (hetero_raw['amplitude_ratio'] or 1)
                            st.metric("Амплитуда (к/н)", f"{hetero_after['amplitude_ratio']:.2f}x",
                                    delta=f"{delta_ratio:+.2f}x")
                        else:
                            st.metric("Амплитуда (к/н)", "N/A")
                        
                        st.divider()
                        st.markdown("**Статистики ряда:**")
                        st.metric("Среднее", f"{transformed_series.mean():.2f}")
                        st.metric("Стд. отклонение", f"{transformed_series.std():.2f}")
                        st.metric("Мин / Макс", f"{transformed_series.min():.2f} / {transformed_series.max():.2f}")
                    
                    st.divider()
                    
                    # Рекомендации
                    if hetero_after['bp_pvalue'] is not None:
                        if not hetero_after['is_hetero']:
                            st.success("✅ **Ряд готов** для ARIMA, регрессии, Gaussian processes!")
                        else:
                            st.warning("💡 **Рекомендация:** Попробуйте другой метод или ручной подбор λ.")
                    
                    # Информация о параметрах
                    st.info(f"**Параметры:** λ = {st.session_state.variance_transform_result['lambda']:.3f}")
            
            # ── ИНФОРМАЦИЯ О СОХРАНЁННЫХ ТРАНСФОРМАЦИЯХ ────
            if st.session_state.variance_transform_params:
                st.divider()
                with st.expander("💾 Сохранённые параметры трансформации", expanded=False):
                    st.json(st.session_state.variance_transform_params)
                    st.caption("Эти параметры будут использованы для **обратного преобразования** прогноза в исходную шкалу.")
        
        else:
            st.warning("⚠️ В датасете нет числовых колонок для анализа.")
    else:
        st.warning("⚠️ Не обнаружены колонки с датами или числовыми данными. Убедитесь, что активирован режим временных рядов.")


    # ═══════════════════════════════════════════════════════
    # 🔹 6. СГЛАЖИВАНИЕ (SMOOTHING)
    # ═══════════════════════════════════════════════════════
    st.divider()
    st.markdown("### Сглаживание временного ряда")
    st.caption(""" Уменьшение шума и высокочастотных колебаний для выделения тренда и сезонности.
               Опциональный шаг — применяется, если ряд слишком "шумный" для моделей.""")

    # ── ДИАГНОСТИКА ТЕКУЩЕГО СОСТОЯНИЯ ──────────────────
    if st.session_state.primary_date_col and st.session_state.col_types.get("num"):
        date_col = st.session_state.primary_date_col
        num_cols = st.session_state.col_types.get("num", [])
        
        if num_cols:
            df_smooth = st.session_state.df.copy()
            df_smooth[date_col] = pd.to_datetime(df_smooth[date_col])
            df_smooth = df_smooth.sort_values(date_col)
            df_smooth_ts = df_smooth.set_index(date_col)
            
            # ── ФУНКЦИЯ МЕТРИК КАЧЕСТВА СГЛАЖИВАНИЯ ───────
            def calculate_smoothing_metrics(original: pd.Series, smoothed: pd.Series) -> dict:
                """Расчёт метрик качества сглаживания"""
                metrics = {}
                
                # 1. SNR (Signal-to-Noise Ratio)
                signal_power = smoothed.var()
                noise = original - smoothed
                noise_power = noise.var()
                metrics['snr'] = 10 * np.log10(signal_power / (noise_power + 1e-10)) if noise_power > 0 else np.inf
                
                # 2. Корреляция с исходным рядом
                metrics['correlation'] = original.corr(smoothed)
                
                # 3. Сглаженность (smoothness) — сумма квадратов вторых разностей
                # Чем МЕНЬШЕ, тем ГЛАЖЕ ряд
                second_diff_orig = original.diff().diff().dropna()
                second_diff_smooth = smoothed.diff().diff().dropna()
                metrics['roughness_orig'] = np.sum(second_diff_orig**2)
                metrics['roughness_smooth'] = np.sum(second_diff_smooth**2)
                metrics['smoothness_ratio'] = (metrics['roughness_orig'] / 
                                            (metrics['roughness_smooth'] + 1e-10))
                
                # 4. Сохранение тренда (R² линейного тренда)
                from scipy.stats import linregress
                slope_orig, _, r_orig, _, _ = linregress(range(len(original)), original)
                slope_smooth, _, r_smooth, _, _ = linregress(range(len(smoothed)), smoothed)
                metrics['r2_orig'] = r_orig**2
                metrics['r2_smooth'] = r_smooth**2
                metrics['trend_preservation'] = abs(r_smooth**2 - r_orig**2)
                
                # 5. Потеря информации (разница дисперсий)
                metrics['variance_loss_pct'] = ((original.var() - smoothed.var()) / 
                                            (original.var() + 1e-10)) * 100
                
                # 6. Amplitude reduction (ослабление амплитуды)
                metrics['amplitude_reduction'] = 1 - (smoothed.std() / (original.std() + 1e-10))
                
                return metrics
            
            # ── МЕТРИКИ ТЕКУЩЕГО СОСТОЯНИЯ ─────────────────
            c_diag1, c_diag2, c_diag3, c_diag4 = st.columns(4)
            
            target_col_smooth = st.session_state.get('ts_props_v10_target_col', num_cols[0])
            series_raw = df_smooth_ts[target_col_smooth].dropna().astype(float)
            
            # Базовая диагностика "шумности"
            second_diff = series_raw.diff().diff().dropna()
            roughness = np.sum(second_diff**2)
            noise_ratio = (series_raw.diff().std() / (series_raw.std() + 1e-10))
            
            with c_diag1:
                st.markdown("**Анализ ряда**")
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: #1e293b;'>`{target_col_smooth}`</div>", 
                        unsafe_allow_html=True)
            
            with c_diag2:
                st.markdown("**Шумность ряда**")
                noise_color = "#dc2626" if noise_ratio > 0.5 else "#d97706" if noise_ratio > 0.3 else "#16a34a"
                noise_text = "Высокая" if noise_ratio > 0.5 else "Средняя" if noise_ratio > 0.3 else "Низкая"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {noise_color};'>"
                        f"{noise_text} ({noise_ratio:.2f})</div>", 
                        unsafe_allow_html=True)
            
            with c_diag3:
                st.markdown("**Roughness (2-й diff)**")
                roughness_formatted = f"{roughness:,.0f}".replace(",", " ")
                roughness_color = "#dc2626" if roughness > 1e6 else "#d97706" if roughness > 1e5 else "#16a34a"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {roughness_color};'>"
                        f"{roughness_formatted}</div>", unsafe_allow_html=True)
            
            with c_diag4:
                st.markdown("**Рекомендация**")
                if noise_ratio > 0.5:
                    rec_text, rec_color = "Нужно сглаживание", "#dc2626"
                elif noise_ratio > 0.3:
                    rec_text, rec_color = "Желательно", "#d97706"
                else:
                    rec_text, rec_color = "Не требуется", "#16a34a"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {rec_color};'>"
                        f"{rec_text}</div>", unsafe_allow_html=True)
            
            # Отступ
            st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)
            
            # ── ТЕХНИЧЕСКАЯ СПРАВКА ─────────────────────────
            with st.expander("Цели субмодуля ⁞ Сглаживание временных рядов", expanded=False):
                st.markdown("""
                **Зачем нужно сглаживание:**
                - **Выделение тренда** — удаление высокочастотного шума
                - **Улучшение прогноза** — модели лучше работают на "чистом" сигнале
                - **Стабилизация оценок** — меньше ложных срабатываний на выбросах
                - **Подготовка для ML** — сглаженные признаки улучшают качество
                
                **Когда применять:**
                - ✅ Ряд содержит значительный шум (noise_ratio > 0.3)
                - ✅ Нужно выделить тренд для интерпретации
                - ✅ Перед декомпозицией (если STL не справляется с шумом)
                - ❌ **НЕ применять** перед дифференцированием (искажает стационарность)
                - ❌ **НЕ применять** для рядов с важными краткосрочными колебаниями
                
                **Методы сглаживания:**
                
                | Метод | Формула | Параметры | Когда использовать |
                |-------|---------|-----------|-------------------|
                | **SMA** | (x_t + ... + x_{t-w+1})/w | window | Простое усреднение, стабильные ряды |
                | **EMA** | α·x_t + (1-α)·s_{t-1} | span (α=2/(span+1)) | Реакция на последние изменения |
                | **WMA** | Σ(w_i · x_{t-i}) / Σw_i | window | Больший вес свежим данным |
                | **Медиана** | median(x_{t-w/2},...,x_{t+w/2}) | window | Устойчивость к выбросам |
                | **LOWESS** | Локальная регрессия | frac (0.1-0.5) | Нелинейные тренды, сложная структура |
                | **HP-filter** | min Σ(y_t-τ_t)² + λΣ(Δ²τ_t)² | λ (100-1600) | Выделение тренда/цикла (экономика) |
                
                **Выбор параметров:**
                - **SMA/WMA/Медиана:** window = период сезонности (7 для дневных, 12 для месячных)
                - **EMA:** span = 2·window - 1 (эквивалент SMA по "памяти")
                - **LOWESS:** frac = 0.2-0.3 (меньше = гибче, больше = глаже)
                - **HP-filter:** λ = 100 (квартальные), 1600 (годовые), 14400 (ежедневные)
                
                **Метрики качества сглаживания:**
                - ✅ **SNR (Signal-to-Noise Ratio)** — чем больше, тем лучше (>10 dB хорошо)
                - ✅ **Smoothness ratio** — во сколько раз снизилась "шероховатость"
                - ✅ **Trend preservation** — насколько сохранён тренд (R²)
                - ✅ **Variance loss** — сколько информации потеряно (<30% хорошо)
                
                **Обратимость:**
                - 🔁 **HP-filter** — единственный обратимый метод (сохраняет trend + cycle)
                - 🔁 Остальные методы **необратимы** — информация о шуме теряется
                - 🔁 Рекомендуется сохранять **исходный ряд** в отдельной колонке
                
                **⚠️ Почитать:**
                - Hodrick-Prescott filter: https://en.wikipedia.org/wiki/Hodrick%E2%80%93Prescott_filter
                - LOWESS regression: https://en.wikipedia.org/wiki/Local_regression
                - Сглаживание в TS: https://otexts.com/fpp3/moving-averages.html
                """)
            
            # ────────────────────────────────────────────────
            # 🎮 ПЕСОЧНИЦА: СГЛАЖИВАНИЕ
            # ────────────────────────────────────────────────
            
            # Инициализация session_state
            if "show_smooth_preview" not in st.session_state:
                st.session_state.show_smooth_preview = False
            if "smooth_transform_params" not in st.session_state:
                st.session_state.smooth_transform_params = {}
            
            # ── ЛЕВАЯ КОЛОНКА: ПАНЕЛЬ УПРАВЛЕНИЯ ───────────
            c1, c2, c3 = st.columns([27, 52, 21])
            
            with c1:
                st.markdown("###### Панель управления")
                
                # Выбор числовой колонки
                target_col = st.selectbox(
                    "Исследуемый признак:",
                    options=num_cols,
                    index=num_cols.index(target_col_smooth) if target_col_smooth in num_cols else 0,
                    key="smooth_target_col"
                )
                
                series = df_smooth_ts[target_col].dropna().astype(float)
                
                # Выбор метода
                smooth_method = st.radio(
                    "Метод сглаживания:",
                    ["SMA (простое скользящее среднее)",
                    "EMA (экспоненциально взвешенное)",
                    "WMA (взвешенное скользящее)",
                    "Скользящая медиана (устойчиво к выбросам)",
                    "LOWESS/LOESS (локальная регрессия)",
                    "HP-filter (Ходрика-Прескотта)"],
                    index=0,
                    key="smooth_method",
                    label_visibility="collapsed"
                )
                
                # Описание метода
                method_descriptions = {
                    "SMA (простое скользящее среднее)": 
                        "Равномерное усреднение последних `window` значений. Простой и интерпретируемый метод.",
                    "EMA (экспоненциально взвешенное)": 
                        "Больший вес последним наблюдениям. Быстрее реагирует на изменения, чем SMA.",
                    "WMA (взвешенное скользящее)": 
                        "Линейно возрастающие веса: свежим данным — больше внимания. Компромисс между SMA и EMA.",
                    "Скользящая медиана (устойчиво к выбросам)": 
                        "Медиана вместо среднего. Устойчива к выбросам, но менее гладкая.",
                    "LOWESS/LOESS (локальная регрессия)": 
                        "Локальная полиномиальная регрессия. Гибкий метод для нелинейных трендов.",
                    "HP-filter (Ходрика-Прескотта)": 
                        "Выделение тренда через минимизацию функции потерь. Стандарт в макроэкономике."
                }
                
                st.markdown(
                    f'<div style="background: #f0f9ff; border-left: 3px solid #0284c7; padding: 8px 12px; '
                    f'margin: 8px 0; border-radius: 4px;">'
                    f'<span style="color: #0369a1; font-size: 13px;">'
                    f'💡 <strong>{smooth_method.split(" ")[0]}:</strong> {method_descriptions[smooth_method]}'
                    f'</span></div>',
                    unsafe_allow_html=True
                )
                
                # Параметры метода
                st.divider()
                st.markdown("**Параметры:**")
                
                if "SMA" in smooth_method or "WMA" in smooth_method or "Медиана" in smooth_method:
                    window = st.slider(
                        "Окно сглаживания (window):",
                        min_value=3,
                        max_value=min(100, len(series) // 3),
                        value=7,
                        step=1,
                        key="smooth_window",
                        help="Чем больше окно, тем глаже ряд, но больше задержка"
                    )
                    param_value = window
                    
                elif "EMA" in smooth_method:
                    span = st.slider(
                        "Span (период полуспада):",
                        min_value=3,
                        max_value=min(100, len(series) // 3),
                        value=14,
                        step=1,
                        key="smooth_span",
                        help="α = 2/(span+1). Меньше span = больше реакция на новые данные"
                    )
                    param_value = span
                    
                elif "LOWESS" in smooth_method:
                    frac = st.slider(
                        "Доля данных для локальной регрессии (frac):",
                        min_value=0.05,
                        max_value=0.9,
                        value=0.2,
                        step=0.05,
                        key="smooth_frac",
                        help="Больше frac = глаже результат. Обычно 0.2-0.3"
                    )
                    param_value = frac
                    
                elif "HP-filter" in smooth_method:
                    hp_lambda = st.selectbox(
                        "Параметр λ (штраф за изменчивость тренда):",
                        options=[100, 400, 1600, 6400, 14400],
                        index=2,
                        key="smooth_hp_lambda",
                        help="100 — квартальные, 1600 — годовые, 14400 — ежедневные данные"
                    )
                    param_value = hp_lambda
                    
                    # Подсказка по λ
                    lambda_hints = {
                        100: "📅 Для квартальных данных",
                        400: "📅 Для месячных данных (альтернатива)",
                        1600: "📅 Стандарт для годовых данных (Hodrick-Prescott)",
                        6400: "📅 Для высокочастотных данных",
                        14400: "📅 Для ежедневных данных"
                    }
                    st.info(f"💡 {lambda_hints.get(hp_lambda, '')}")
                
                st.divider()
                
                # Кнопки действий
                if st.button("▶ Применить сглаживание", type="primary", use_container_width=True, key="btn_apply_smooth"):
                    st.session_state.show_smooth_preview = True
                    st.rerun()
                
                if st.button("↶ Сбросить", use_container_width=True, key="btn_reset_smooth"):
                    st.session_state.show_smooth_preview = False
                    st.rerun()
            
            # ── ЦЕНТРАЛЬНАЯ КОЛОНКА: ВИЗУАЛИЗАЦИЯ ──────────
            with c2:
                st.markdown("###### Визуализация: До / После")
                
                if target_col in df_smooth_ts.columns:
                    original_series = df_smooth_ts[target_col].dropna().astype(float)
                    
                    if st.session_state.show_smooth_preview:
                        try:
                            # Применяем сглаживание
                            if "SMA" in smooth_method:
                                smoothed = original_series.rolling(window=window, center=True, min_periods=1).mean()
                                method_name = "SMA"
                            
                            elif "EMA" in smooth_method:
                                smoothed = original_series.ewm(span=span, adjust=False).mean()
                                method_name = "EMA"
                            
                            elif "WMA" in smooth_method:
                                # Линейно-взвешенное скользящее среднее
                                weights = np.arange(1, window + 1)
                                smoothed = original_series.rolling(window=window).apply(
                                    lambda x: np.dot(x, weights) / weights.sum(), raw=True
                                )
                                smoothed = smoothed.fillna(method='bfill')
                                method_name = "WMA"
                            
                            elif "Медиана" in smooth_method:
                                smoothed = original_series.rolling(window=window, center=True, min_periods=1).median()
                                method_name = "Median"
                            
                            elif "LOWESS" in smooth_method:
                                from statsmodels.nonparametric.smoothers_lowess import lowess
                                
                                x = np.arange(len(original_series))
                                y = original_series.values
                                
                                # LOWESS возвращает массив [x, y_smooth]
                                lowess_result = lowess(y, x, frac=frac, return_sorted=False)
                                smoothed = pd.Series(lowess_result, index=original_series.index)
                                method_name = "LOWESS"
                            
                            elif "HP-filter" in smooth_method:
                                from statsmodels.tsa.filters.hp_filter import hpfilter
                                
                                cycle, trend = hpfilter(original_series, lamb=hp_lambda)
                                smoothed = trend
                                method_name = "HP-filter"
                                
                                # Сохраняем cycle для возможного использования
                                st.session_state.hp_cycle = cycle
                            
                            # ── ГРАФИК СРАВНЕНИЯ (2 ряда) ───────
                            fig = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    f" Исходный ряд: {target_col}",
                                    f" После сглаживания ({method_name}, параметр={param_value})"
                                ),
                                vertical_spacing=0.12
                            )
                            
                            # Исходный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=original_series.index, y=original_series.values,
                                    mode='lines',
                                    name='Исходные данные',
                                    line=dict(color='#048A81', width=1, dash='dot'),
                                    showlegend=False
                                ),
                                row=1, col=1
                            )
                            
                            # Сглаженный ряд (на верхнем графике — наложением)
                            fig.add_trace(
                                go.Scatter(
                                    x=smoothed.index, y=smoothed.values,
                                    mode='lines',
                                    name='Сглаженный',
                                    line=dict(color='#DC2626', width=2.5),
                                    showlegend=False
                                ),
                                row=1, col=1
                            )
                            
                            # Только сглаженный ряд (на нижнем графике)
                            fig.add_trace(
                                go.Scatter(
                                    x=smoothed.index, y=smoothed.values,
                                    mode='lines',
                                    name='Сглаженный',
                                    line=dict(color='#DC2626', width=2.5),
                                    showlegend=False
                                ),
                                row=2, col=1
                            )
                            
                            fig.update_layout(
                                height=600,
                                margin=dict(l=50, r=20, t=80, b=40),
                                hovermode='x unified',
                            )
                            
                            fig.update_xaxes(title_text="Дата", row=2, col=1)
                            fig.update_yaxes(title_text="Значение", row=1, col=1)
                            fig.update_yaxes(title_text="Сглаженное значение", row=2, col=1)
                            
                            fig.update_annotations(font=dict(size=13, color="#1e293b"), yshift=10)
                            
                            st.plotly_chart(fig, use_container_width=True, key="smooth_main_chart")
                            
                            # ── ГРАФИК ОСТАТКОВ (что было "отфильтровано") ──
                            st.markdown("##### Остатки (шум, удалённый при сглаживании)")
                            st.caption("Если остатки похожи на белый шум — сглаживание работает корректно.")
                            
                            residuals = original_series - smoothed
                            
                            fig_resid = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    "Остатки (оригинал − сглаженный)",
                                    "ACF остатков (должны быть в пределах синей зоны)"
                                ),
                                vertical_spacing=0.15
                            )
                            
                            # Остатки
                            fig_resid.add_trace(
                                go.Scatter(x=residuals.index, y=residuals.values,
                                        mode='lines', line=dict(color='#6B7280', width=1),
                                        name='Остатки', showlegend=False),
                                row=1, col=1
                            )
                            fig_resid.add_hline(y=0, line_dash="dash", line_color="red", row=1, col=1)
                            
                            # ACF остатков
                            from statsmodels.tsa.stattools import acf
                            max_lag = min(40, len(residuals) // 4)
                            acf_vals = acf(residuals.dropna(), nlags=max_lag)
                            conf_int = 1.96 / np.sqrt(len(residuals))
                            
                            fig_resid.add_trace(
                                go.Bar(x=list(range(len(acf_vals))), y=acf_vals,
                                    marker_color='#2563EB', name='ACF', showlegend=False),
                                row=2, col=1
                            )
                            fig_resid.add_hline(y=conf_int, line_dash="dash", line_color="red", row=2, col=1)
                            fig_resid.add_hline(y=-conf_int, line_dash="dash", line_color="red", row=2, col=1)
                            
                            fig_resid.update_layout(height=500, margin=dict(l=50, r=20, t=60, b=40))
                            st.plotly_chart(fig_resid, use_container_width=True, key="smooth_resid_chart")
                            
                            # ── ГИСТОГРАММЫ СРАВНЕНИЯ ───────────
                            st.markdown("##### Распределение значений")
                            
                            fig_hist = make_subplots(
                                rows=1, cols=2,
                                subplot_titles=("Исходное распределение", "После сглаживания")
                            )
                            
                            fig_hist.add_trace(
                                go.Histogram(x=original_series.values, nbinsx=40,
                                            marker_color='#048A81', name='До', showlegend=False),
                                row=1, col=1
                            )
                            
                            fig_hist.add_trace(
                                go.Histogram(x=smoothed.values, nbinsx=40,
                                            marker_color='#DC2626', name='После', showlegend=False),
                                row=1, col=2
                            )
                            
                            fig_hist.update_layout(height=300, margin=dict(l=40, r=20, t=50, b=40), barmode='overlay')
                            st.plotly_chart(fig_hist, use_container_width=True, key="smooth_hist_chart")
                            
                            # ── РАСЧЁТ МЕТРИК КАЧЕСТВА ──────────
                            metrics = calculate_smoothing_metrics(original_series, smoothed)
                            
                            # Сохранение результатов
                            st.session_state.smooth_transform_result = {
                                'method': method_name,
                                'param': param_value,
                                'original_series': original_series,
                                'smoothed_series': smoothed,
                                'target_col': target_col,
                                'metrics': metrics
                            }
                            
                            # ── HP-filter: дополнительно показываем цикл ──
                            if "HP-filter" in smooth_method:
                                st.markdown("##### Циклическая компонента (HP-filter)")
                                st.caption("HP-filter разделяет ряд на тренд и цикл. Ниже — выделенный цикл.")
                                
                                fig_cycle = go.Figure()
                                fig_cycle.add_trace(go.Scatter(
                                    x=original_series.index, y=st.session_state.hp_cycle.values,
                                    mode='lines', line=dict(color='#9333EA', width=2),
                                    name='Цикл'
                                ))
                                fig_cycle.add_hline(y=0, line_dash="dash", line_color="gray")
                                fig_cycle.update_layout(
                                    height=250, margin=dict(l=40, r=20, t=30, b=20),
                                    title="Циклическая компонента"
                                )
                                st.plotly_chart(fig_cycle, use_container_width=True, key="smooth_cycle_chart")
                            
                            st.divider()
                            
                            # Оценка качества сглаживания
                            if metrics['snr'] > 10:
                                st.success(f"✅ **Отличное сглаживание!** SNR = {metrics['snr']:.1f} dB")
                            elif metrics['snr'] > 5:
                                st.success(f"✅ **Хорошее сглаживание.** SNR = {metrics['snr']:.1f} dB")
                            elif metrics['snr'] > 0:
                                st.warning(f"⚠️ **Умеренное сглаживание.** SNR = {metrics['snr']:.1f} dB. Попробуйте другие параметры.")
                            else:
                                st.error(f"❌ **Слабое сглаживание.** SNR = {metrics['snr']:.1f} dB. Сигнал слабее шума.")
                            
                            # Кнопки подтверждения
                            c_ok_smooth, c_cancel_smooth = st.columns(2)
                            with c_ok_smooth:
                                if st.button("✅ Применить к данным", type="primary", use_container_width=True, key="btn_confirm_smooth"):
                                    # Применяем сглаживание к основному df
                                    df_final_smooth = st.session_state.df.copy()
                                    
                                    # Сохраняем оригинальный столбец с суффиксом _original
                                    orig_col_name = f"{target_col}_original"
                                    if orig_col_name not in df_final_smooth.columns:
                                        df_final_smooth[orig_col_name] = df_final_smooth[target_col]
                                    
                                    # Применяем сглаживание
                                    if method_name == "SMA":
                                        df_final_smooth[target_col] = df_final_smooth[target_col].astype(float).rolling(
                                            window=param_value, center=True, min_periods=1).mean()
                                    elif method_name == "EMA":
                                        df_final_smooth[target_col] = df_final_smooth[target_col].astype(float).ewm(
                                            span=param_value, adjust=False).mean()
                                    elif method_name == "WMA":
                                        weights = np.arange(1, param_value + 1)
                                        df_final_smooth[target_col] = df_final_smooth[target_col].astype(float).rolling(
                                            window=param_value).apply(
                                            lambda x: np.dot(x, weights) / weights.sum(), raw=True
                                        ).fillna(method='bfill')
                                    elif method_name == "Median":
                                        df_final_smooth[target_col] = df_final_smooth[target_col].astype(float).rolling(
                                            window=param_value, center=True, min_periods=1).median()
                                    elif method_name == "LOWESS":
                                        from statsmodels.nonparametric.smoothers_lowess import lowess
                                        x = np.arange(len(df_final_smooth))
                                        y = df_final_smooth[target_col].astype(float).values
                                        lowess_result = lowess(y, x, frac=param_value, return_sorted=False)
                                        df_final_smooth[target_col] = lowess_result
                                    elif method_name == "HP-filter":
                                        from statsmodels.tsa.filters.hp_filter import hpfilter
                                        cycle, trend = hpfilter(df_final_smooth[target_col].astype(float), lamb=param_value)
                                        df_final_smooth[target_col] = trend
                                    
                                    # Сохраняем параметры
                                    st.session_state.smooth_transform_params = {
                                        'method': method_name,
                                        'param': param_value,
                                        'column': target_col,
                                        'original_col_name': orig_col_name,
                                        'metrics': metrics
                                    }
                                    
                                    # Синхронизация
                                    st.session_state.df = df_final_smooth.copy()
                                    st.session_state.validation_ready = False
                                    st.session_state.show_smooth_preview = False
                                    
                                    # Удаляем рабочие копии
                                    work_dfs = [
                                        "df_missing_work", "df_pattern_work", "df_range_work",
                                        "df_outlier_work", "df_inclusion_work", "df_referential_work",
                                        "df_text_work", "df_consistency_work", "df_uniqueness_work",
                                        "df_regularity_work", "df_regular_work", "df_variance_work",
                                        "df_smooth_work"
                                    ]
                                    for work_df_name in work_dfs:
                                        if work_df_name in st.session_state:
                                            del st.session_state[work_df_name]
                                    
                                    if "val_results" in st.session_state:
                                        del st.session_state.val_results
                                    
                                    st.success(f"✅ Сглаживание **{method_name}** (параметр={param_value}) применено!")
                                    st.info(f"💡 Оригинальные значения сохранены в колонке `{orig_col_name}`. "
                                        f"Перезапустите валидацию для обновления статистик.")
                                    st.rerun()
                            
                            with c_cancel_smooth:
                                if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_smooth"):
                                    st.session_state.show_smooth_preview = False
                                    st.rerun()
                        
                        except Exception as e:
                            st.error(f"❌ Ошибка сглаживания: {e}")
                            import traceback
                            with st.expander("🔍 Stack trace"):
                                st.code(traceback.format_exc(), language="python")
                    
                    else:
                        # Показываем только исходный ряд
                        fig = px.line(
                            x=original_series.index,
                            y=original_series.values,
                            labels={'x': 'Дата', 'y': target_col},
                        )
                        fig.update_layout(height=400, margin=dict(l=40, r=20, t=40, b=40))
                        st.plotly_chart(fig, use_container_width=True, key="smooth_main_chart")
                        
                        st.info("💡 Выберите метод сглаживания и нажмите **'Применить сглаживание'** для просмотра результата.")
                else:
                    st.warning("⚠️ Выбранная колонка не найдена в данных.")
            
            # ── ПРАВАЯ КОЛОНКА: МЕТРИКИ КАЧЕСТВА ───────────
            with c3:
                st.markdown("###### Метрики качества")
                
                with st.container(border=True):
                    st.markdown("**Исходный ряд:**")
                    st.metric("Длина ряда", f"{len(series):,}".replace(",", " "))
                    st.metric("Среднее", f"{series.mean():.2f}")
                    st.metric("Стд. отклонение", f"{series.std():.2f}")
                    st.metric("Шумность (σ_diff/σ)", f"{noise_ratio:.3f}")
                
                # Метрики ПОСЛЕ сглаживания
                if st.session_state.show_smooth_preview and 'smooth_transform_result' in st.session_state:
                    metrics = st.session_state.smooth_transform_result['metrics']
                    smoothed_series = st.session_state.smooth_transform_result['smoothed_series']
                    
                    with st.container(border=True):
                        st.markdown("**После сглаживания:**")
                        
                        # SNR
                        snr_color = "normal" if metrics['snr'] > 10 else "off" if metrics['snr'] > 5 else "inverse"
                        st.metric("SNR (dB)", f"{metrics['snr']:.1f}",
                                delta="Отлично" if metrics['snr'] > 10 else "Хорошо" if metrics['snr'] > 5 else "Слабо")
                        
                        # Корреляция
                        st.metric("Корреляция с исх.", f"{metrics['correlation']:.3f}",
                                delta="Сильная" if metrics['correlation'] > 0.9 else "Средняя")
                        
                        # Smoothness ratio
                        st.metric("Smoothness ratio", f"{metrics['smoothness_ratio']:.1f}x",
                                delta=f"В {metrics['smoothness_ratio']:.1f} раз глаже")
                        
                        # Trend preservation
                        delta_trend = metrics['trend_preservation']
                        st.metric("Δ Тренд (R²)", f"{delta_trend:+.3f}",
                                delta="Сохранён" if abs(delta_trend) < 0.1 else "Изменён")
                        
                        # Variance loss
                        st.metric("Потеря дисперсии", f"{metrics['variance_loss_pct']:.1f}%",
                                delta="Мало" if metrics['variance_loss_pct'] < 30 else "Много")
                        
                        # Amplitude reduction
                        st.metric("Ослабление ампл.", f"{metrics['amplitude_reduction']*100:.1f}%")
                    
                    st.divider()
                    
                    # Рекомендации
                    if metrics['snr'] > 10 and metrics['correlation'] > 0.9:
                        st.success("✅ **Отличный результат!** Шум удалён, тренд сохранён.")
                    elif metrics['snr'] > 5:
                        st.success("✅ **Хороший результат.** Можно применять.")
                    elif metrics['variance_loss_pct'] > 50:
                        st.warning("💡 **Слишком агрессивное сглаживание.** Уменьшите window/span/frac.")
                    elif metrics['correlation'] < 0.7:
                        st.warning("💡 **Сильное искажение.** Ряд сильно изменился. Попробуйте другой метод.")
                    else:
                        st.info("💡 Попробуйте другие параметры для улучшения результата.")
                    
                    # Информация о параметрах
                    st.info(f"**Параметры:** {st.session_state.smooth_transform_result['method']} = {st.session_state.smooth_transform_result['param']}")
            
            # ── ИНФОРМАЦИЯ О СОХРАНЁННЫХ ТРАНСФОРМАЦИЯХ ────
            if st.session_state.smooth_transform_params:
                st.divider()
                with st.expander("💾 Сохранённые параметры сглаживания", expanded=False):
                    st.json(st.session_state.smooth_transform_params)
                    st.caption("Эти параметры описывают применённое сглаживание. "
                            "Оригинальные значения сохранены в колонке `_original`.")
        
        else:
            st.warning("⚠️ В датасете нет числовых колонок для анализа.")
    else:
        st.warning("⚠️ Не обнаружены колонки с датами или числовыми данными. Убедитесь, что активирован режим временных рядов.")


    # ═══════════════════════════════════════════════════════
    # 🔹 7. ОБЕСПЕЧЕНИЕ СТАЦИОНАРНОСТИ (STATIONARITY)
    # ═══════════════════════════════════════════════════════
    st.divider()
    st.markdown("### Обеспечение стационарности (дифференцирование)")
    st.caption("Большинство TS-моделей (ARIMA, VAR, линейная регрессия) требуют стационарного ряда — "
            "с постоянными mean, variance и autocorrelation. Дифференцирование удаляет тренд и сезонность, "
            "исключая смещение положения равновесия.")

    # ── ДИАГНОСТИКА ТЕКУЩЕГО СОСТОЯНИЯ ──────────────────
    if st.session_state.primary_date_col and st.session_state.col_types.get("num"):
        date_col = st.session_state.primary_date_col
        num_cols = st.session_state.col_types.get("num", [])
        
        if num_cols:
            df_stat = st.session_state.df.copy()
            df_stat[date_col] = pd.to_datetime(df_stat[date_col])
            df_stat = df_stat.sort_values(date_col)
            df_stat_ts = df_stat.set_index(date_col)
            
            # ── ФУНКЦИЯ МНОЖЕСТВЕННЫХ ТЕСТОВ СТАЦИОНАРНОСТИ ──
            def run_stationarity_tests(series: pd.Series, max_lag: int = None) -> dict:
                """
                Запускает 4 теста стационарности и возвращает консенсус.
                
                Returns:
                    dict с ключами:
                    - adf: {'stat': float, 'pvalue': float, 'is_stationary': bool}
                    - kpss: {'stat': float, 'pvalue': float, 'is_stationary': bool}
                    - pp: {'stat': float, 'pvalue': float, 'is_stationary': bool}
                    - za: {'stat': float, 'pvalue': float, 'is_stationary': bool, 'breakpoint': int} (если доступен)
                    - consensus: 'stationary' | 'non-stationary' | 'trend-stationary' | 'inconclusive'
                    - recommendation: str
                """
                results = {}
                n = len(series)
                
                if n < 30:
                    return {'error': 'Недостаточно данных (нужно ≥ 30)'}
                
                try:
                    from statsmodels.tsa.stattools import adfuller, kpss
                    from statsmodels.tsa.stattools import PhillipsPerron
                    
                    # ── 1. ADF TEST ──────────────────────────
                    # H0: ряд имеет единичный корень (нестационарен)
                    # H1: ряд стационарен
                    if max_lag is None:
                        max_lag_adf = min(int(12 * (n / 100) ** 0.25), n // 3)
                    else:
                        max_lag_adf = max_lag
                    
                    adf_result = adfuller(series.dropna(), autolag='AIC', maxlag=max_lag_adf)
                    results['adf'] = {
                        'stat': adf_result[0],
                        'pvalue': adf_result[1],
                        'lags': adf_result[2],
                        'is_stationary': adf_result[1] < 0.05,
                        'critical_values': adf_result[4]
                    }
                    
                    # ─ 2. KPSS TEST ─────────────────────────
                    # H0: ряд стационарен (вокруг уровня или тренда)
                    # H1: ряд нестационарен
                    # Тестируем два варианта: level (вокруг константы) и trend (вокруг тренда)
                    try:
                        kpss_level = kpss(series.dropna(), regression='c', nlags='auto')
                        kpss_trend = kpss(series.dropna(), regression='ct', nlags='auto')
                        results['kpss'] = {
                            'stat_level': kpss_level[0],
                            'pvalue_level': kpss_level[1],
                            'stat_trend': kpss_trend[0],
                            'pvalue_trend': kpss_trend[1],
                            'is_stationary_level': kpss_level[1] > 0.05,
                            'is_stationary_trend': kpss_trend[1] > 0.05
                        }
                    except Exception as e:
                        results['kpss'] = {'error': str(e)}
                    
                    # ── 3. PHILLIPS-PERRON TEST ──────────────
                    # Альтернатива ADF, устойчива к гетероскедастичности
                    try:
                        pp_result = PhillipsPerron(series.dropna(), lags=max_lag_adf)
                        results['pp'] = {
                            'stat': pp_result.stat,
                            'pvalue': pp_result.pvalue,
                            'is_stationary': pp_result.pvalue < 0.05
                        }
                    except Exception:
                        # Fallback: если PP недоступен, используем ADF как proxy
                        results['pp'] = {
                            'stat': adf_result[0],
                            'pvalue': adf_result[1],
                            'is_stationary': adf_result[1] < 0.05,
                            'note': 'PP недоступен, используется ADF'
                        }
                    
                    # ── 4. ZIVOT-ANDREWS TEST (опционально) ──
                    # Учитывает структурные разрывы
                    try:
                        from statsmodels.tsa.stattools import zivot_andrews
                        za_result = zivot_andrews(series.dropna(), model='c')
                        results['za'] = {
                            'stat': za_result[0],
                            'pvalue': za_result[1],  # Может быть недоступен в старых версиях
                            'breakpoint': za_result[2] if len(za_result) > 2 else None,
                            'is_stationary': za_result[0] < -4.8  # Приблизительный критический уровень
                        }
                    except Exception:
                        results['za'] = {'note': 'Тест Zivot-Andrews недоступен (statsmodels < 0.14)'}
                    
                    # ── 5. КОНСЕНСУС ─────────────────────────
                    adf_stat = results['adf']['is_stationary']
                    kpss_level_stat = results['kpss'].get('is_stationary_level', None)
                    kpss_trend_stat = results['kpss'].get('is_stationary_trend', None)
                    pp_stat = results['pp']['is_stationary']
                    
                    if adf_stat and kpss_level_stat:
                        consensus = 'stationary'
                        recommendation = '✅ Ряд стационарен. Дифференцирование не требуется.'
                    elif not adf_stat and kpss_trend_stat:
                        consensus = 'trend-stationary'
                        recommendation = '⚠️ Ряд стационарен вокруг тренда. Достаточно удалить тренд (детренд).'
                    elif not adf_stat and not kpss_level_stat:
                        consensus = 'non-stationary'
                        if pp_stat:
                            recommendation = '⚠️ ADF и KPSS противоречат PP. Попробуйте другое дифференцирование.'
                        else:
                            recommendation = ' Ряд нестационарен. Требуется дифференцирование.'
                    else:
                        consensus = 'inconclusive'
                        recommendation = '️ Результаты тестов противоречивы. Визуальный анализ + пробное дифференцирование.'
                    
                    results['consensus'] = consensus
                    results['recommendation'] = recommendation
                    
                except ImportError as e:
                    results['error'] = f'Не установлены необходимые библиотеки: {e}'
                except Exception as e:
                    results['error'] = str(e)
                
                return results
            
            # ── ФУНКЦИЯ ДИФФЕРЕНЦИРОВАНИЯ ───────────────────
            def apply_differencing(series: pd.Series, method: str, d: int = 1, s: int = None, 
                                frac_d: float = None) -> pd.Series:
                """
                Применяет дифференцирование к ряду.
                
                Args:
                    series: исходный ряд
                    method: 'first', 'seasonal', 'second', 'log', 'fractional', 'combined'
                    d: порядок первого различия
                    s: сезонный период (для seasonal/combined)
                    frac_d: дробный порядок (для fractional, 0 < d < 1)
                
                Returns:
                    дифференцированный ряд
                """
                s_clean = series.dropna()
                
                if method == 'first':
                    return s_clean.diff(d).dropna()
                
                elif method == 'seasonal':
                    if s is None:
                        s = 12  # default
                    return s_clean.diff(s).dropna()
                
                elif method == 'second':
                    return s_clean.diff(2).dropna()
                
                elif method == 'log':
                    if (s_clean <= 0).any():
                        raise ValueError("Логарифмическое различие требует положительных значений")
                    return np.log(s_clean).diff().dropna()
                
                elif method == 'fractional':
                    if frac_d is None or not (0 < frac_d < 1):
                        raise ValueError("Дробный порядок должен быть в диапазоне (0, 1)")
                    # Реализация дробного дифференцирования по López de Prado
                    # (1 - L)^d = Σ_{k=0}^{∞} (-1)^k * C(d,k) * L^k
                    # где C(d,k) = d*(d-1)*...*(d-k+1)/k!
                    from scipy.special import comb
                    
                    weights = []
                    for k in range(len(s_clean)):
                        weight = (-1) ** k * comb(frac_d, k)
                        weights.append(weight)
                        if abs(weight) < 1e-5:  # Обрезаем малые веса
                            break
                    
                    weights = np.array(weights[:len(s_clean)])
                    
                    # Применяем свёртку
                    result = np.zeros(len(s_clean))
                    values = s_clean.values
                    for i in range(len(s_clean)):
                        for j, w in enumerate(weights):
                            if i - j >= 0:
                                result[i] += w * values[i - j]
                    
                    return pd.Series(result, index=s_clean.index).dropna()
                
                elif method == 'combined':
                    # Сначала сезонное, потом первое различие
                    if s is None:
                        s = 12
                    result = s_clean.diff(s).dropna()
                    result = result.diff(d).dropna()
                    return result
                
                else:
                    raise ValueError(f"Неизвестный метод: {method}")
            
            # ── МЕТРИКИ ТЕКУЩЕГО СОСТОЯНИЯ ─────────────────
            c_diag1, c_diag2, c_diag3, c_diag4 = st.columns(4)
            
            target_col_stat = st.session_state.get('ts_props_v10_target_col', num_cols[0])
            series_raw = df_stat_ts[target_col_stat].dropna().astype(float)
            
            # Запускаем тесты
            tests_raw = run_stationarity_tests(series_raw)
            
            with c_diag1:
                st.markdown("**Анализ ряда**")
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: #1e293b;'>`{target_col_stat}`</div>", 
                        unsafe_allow_html=True)
            
            with c_diag2:
                st.markdown("**ADF Test**")
                if 'adf' in tests_raw:
                    adf_p = tests_raw['adf']['pvalue']
                    adf_color = "#16a34a" if adf_p < 0.05 else "#dc2626"
                    adf_text = "Стационарен" if adf_p < 0.05 else "Нестационарен"
                    st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {adf_color};'>"
                            f"{adf_text} (p={adf_p:.4f})</div>", 
                            unsafe_allow_html=True)
                else:
                    st.markdown("<div style='font-size: 14px; color: #6b7280;'>N/A</div>", unsafe_allow_html=True)
            
            with c_diag3:
                st.markdown("**KPSS Test**")
                if 'kpss' in tests_raw and 'pvalue_level' in tests_raw['kpss']:
                    kpss_p = tests_raw['kpss']['pvalue_level']
                    kpss_color = "#16a34a" if kpss_p > 0.05 else "#dc2626"
                    kpss_text = "Стационарен" if kpss_p > 0.05 else "Нестационарен"
                    st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {kpss_color};'>"
                            f"{kpss_text} (p={kpss_p:.4f})</div>", 
                            unsafe_allow_html=True)
                else:
                    st.markdown("<div style='font-size: 14px; color: #6b7280;'>N/A</div>", unsafe_allow_html=True)
            
            with c_diag4:
                st.markdown("**Консенсус**")
                consensus = tests_raw.get('consensus', 'unknown')
                consensus_colors = {
                    'stationary': '#16a34a',
                    'trend-stationary': '#d97706',
                    'non-stationary': '#dc2626',
                    'inconclusive': '#6b7280'
                }
                consensus_texts = {
                    'stationary': '✅ Стационарен',
                    'trend-stationary': '⚠️ Тренд-стационарен',
                    'non-stationary': '❌ Нестационарен',
                    'inconclusive': '⚠️ Неопределённость'
                }
                color = consensus_colors.get(consensus, '#6b7280')
                text = consensus_texts.get(consensus, 'Неизвестно')
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {color};'>{text}</div>", 
                        unsafe_allow_html=True)
            
            # Отступ
            st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)
            
            # ── ТЕХНИЧЕСКАЯ СПРАВКА ─────────────────────────
            with st.expander("Цели субмодуля ⁞ Обеспечение стационарности", expanded=False):
                st.markdown("""
                **Зачем нужна стационарность:**
                -  **ARIMA/SARIMA** — требуют стационарный ряд (после дифференцирования d, D)
                -  **VAR/VECM** — работают только со стационарными рядами
                -  **Линейная регрессия** — предположение о стационарных остатках
                -  **ML-модели** — стационарные признаки улучшают обобщающую способность
                
                **Что такое стационарность:**
                Ряд стационарен, если его **mean**, **variance** и **autocorrelation** постоянны во времени.
                - ✅ **Строгая стационарность** — распределение не меняется
                - ✅ **Слабая стационарность** — постоянны mean, variance, autocovariance
                
                **Методы дифференцирования:**
                
                | Метод | Формула | Порядок | Когда использовать |
                |-------|---------|---------|-------------------|
                | **Первое различие** | Δy_t = y_t - y_{t-1} | d=1 | Линейный тренд (базовый) |
                | **Сезонное различие** | Δ_s y_t = y_t - y_{t-s} | D=1, s=период | Сезонность (SARIMA) |
                | **Второе различие** | Δ²y_t = Δ(Δy_t) | d=2 | Квадратичный тренд (редко) |
                | **Логарифмическое** | Δln(y_t) | — | Темпы роста, экспоненциальный тренд |
                | **Дробное (Fractional)** | (1-L)^d, d∈(0,1) | d=0.3-0.7 | Сохранение долгосрочной памяти |
                | **Комбинированное** | Δ^d Δ_s^D y_t | d+D | Тренд + сезонность (SARIMA) |
                
                **Тесты стационарности (консенсус ADF + KPSS):**
                
                | ADF (p < 0.05) | KPSS (p < 0.05) | Вывод |
                |----------------|-----------------|-------|
                | ✅ Отвергает H0 | ❌ Не отвергает H0 | **Стационарен** ✅ |
                | ❌ Не отвергает H0 | ✅ Отвергает H0 | **Нестационарен** ❌ |
                | ✅ Отвергает H0 | ✅ Отвергает H0 | **Тренд-стационарен** ⚠️ |
                | ❌ Не отвергает H0 | ❌ Не отвергает H0 | **Неопределённость** ⚠️ |
                
                **Дополнительные тесты:**
                - ✅ **Phillips-Perron (PP)** — устойчив к гетероскедастичности, альтернатива ADF
                - ✅ **Zivot-Andrews (ZA)** — учитывает структурные разрывы (кризисы, реформы)
                
                **Как выбрать порядок дифференцирования:**
                1. Запустите ADF + KPSS на исходном ряде
                2. Если нестационарен → примените d=1, проверьте снова
                3. Если всё ещё нестационарен → попробуйте d=2 или fractional d=0.5
                4. Если есть сезонность → добавьте D=1 с периодом s
                5. **Остановитесь**, когда ряд станет стационарным (не переусложняйте!)
                
                **⚠️ Over-differencing (переусложнение):**
                - Слишком много дифференцирований делает ряд нестационарным в другом смысле
                - Увеличивает дисперсию прогноза
                - **Признак**: ACF первого лага < -0.5 после дифференцирования
                
                **Обратимость:**
                - Все методы дифференцирования **обратимы** через кумулятивную сумму (cumsum)
                - Параметры (d, D, s, метод) сохраняются в `session_state`
                - Прогноз в дифференцированной шкале → обратное преобразование → исходная шкала
                
                **⚠️ Почитать:**
                - Dickey-Fuller test: https://en.wikipedia.org/wiki/Augmented_Dickey%E2%80%93Fuller_test
                - Fractional differencing: López de Prado, "Advances in Financial Machine Learning" (2018), Ch. 5
                - Stationarity in TS: https://otexts.com/fpp3/stationarity.html
                """)
            
            # ────────────────────────────────────────────────
            # 🎮 ПЕСОЧНИЦА: ОБЕСПЕЧЕНИЕ СТАЦИОНАРНОСТИ
            # ────────────────────────────────────────────────
            
            # Инициализация session_state
            if "show_stationarity_preview" not in st.session_state:
                st.session_state.show_stationarity_preview = False
            if "stationarity_transform_params" not in st.session_state:
                st.session_state.stationarity_transform_params = {}
            
            # ── ЛЕВАЯ КОЛОНКА: ПАНЕЛЬ УПРАВЛЕНИЯ ───────────
            c1, c2, c3 = st.columns([27, 52, 21])
            
            with c1:
                st.markdown("###### Панель управления")
                
                # Выбор числовой колонки
                target_col = st.selectbox(
                    "Исследуемый признак:",
                    options=num_cols,
                    index=num_cols.index(target_col_stat) if target_col_stat in num_cols else 0,
                    key="stationarity_target_col"
                )
                
                series = df_stat_ts[target_col].dropna().astype(float)
                
                # Определение частоты для сезонного периода
                inferred_freq = pd.infer_freq(series.index)
                default_s = 12
                if inferred_freq:
                    if 'D' in inferred_freq:
                        default_s = 7
                    elif 'W' in inferred_freq:
                        default_s = 52
                    elif 'M' in inferred_freq:
                        default_s = 12
                    elif 'Q' in inferred_freq:
                        default_s = 4
                    elif 'Y' in inferred_freq:
                        default_s = 1
                
                # Выбор метода
                stationarity_method = st.radio(
                    "Метод дифференцирования:",
                    ["Первое различие (d=1)",
                    "Сезонное различие (D=1, период s)",
                    "Второе различие (d=2)",
                    "Логарифмическое различие (темпы роста)",
                    "Дробное дифференцирование (fractional, d∈(0,1))",
                    "Комбинированное (d + сезонное D)"],
                    index=0,
                    key="stationarity_method",
                    label_visibility="collapsed"
                )
                
                # Описание метода
                method_descriptions = {
                    "Первое различие (d=1)": 
                        "Базовый метод: Δy_t = y_t - y_{t-1}. Удаляет линейный тренд. Стандарт для ARIMA.",
                    "Сезонное различие (D=1, период s)": 
                        f"Δ_s y_t = y_t - y_{{t-s}}. Удаляет сезонность. Период s={default_s} (автоопределено).",
                    "Второе различие (d=2)": 
                        "Δ²y_t = Δ(Δy_t). Для квадратичного тренда. Редко используется, риск over-differencing.",
                    "Логарифмическое различие (темпы роста)": 
                        "Δln(y_t) = ln(y_t) - ln(y_{t-1}). Интерпретируется как % изменение. Только для y > 0.",
                    "Дробное дифференцирование (fractional, d∈(0,1))": 
                        "(1-L)^d, где d∈(0,1). Сохраняет долгосрочную память ряда (López de Prado, 2018).",
                    "Комбинированное (d + сезонное D)": 
                        f"Δ^d Δ_s^D y_t. Для SARIMA: сначала сезонное, потом первое различие. s={default_s}."
                }
                
                st.markdown(
                    f'<div style="background: #f0f9ff; border-left: 3px solid #0284c7; padding: 8px 12px; '
                    f'margin: 8px 0; border-radius: 4px;">'
                    f'<span style="color: #0369a1; font-size: 13px;">'
                    f'💡 <strong>{stationarity_method.split(" ")[0]}:</strong> {method_descriptions[stationarity_method]}'
                    f'</span></div>',
                    unsafe_allow_html=True
                )
                
                # Параметры метода
                st.divider()
                st.markdown("**Параметры:**")
                
                # Проверка на отрицательные значения (для log)
                has_negative = (series <= 0).any()
                if "Логарифмическое" in stationarity_method and has_negative:
                    st.error(" Логарифмическое различие требует положительных значений. Выберите другой метод.")
                
                if "Первое" in stationarity_method:
                    d_order = st.slider(
                        "Порядок d:",
                        min_value=1,
                        max_value=2,
                        value=1,
                        step=1,
                        key="stationarity_d",
                        help="d=1 для линейного тренда, d=2 для квадратичного (редко)"
                    )
                    param_value = d_order
                
                elif "Сезонное" in stationarity_method:
                    s_period = st.number_input(
                        "Сезонный период s:",
                        min_value=2,
                        max_value=365,
                        value=default_s,
                        step=1,
                        key="stationarity_s",
                        help=f"Период сезонности (автоопределено: {default_s})"
                    )
                    param_value = s_period
                
                elif "Второе" in stationarity_method:
                    param_value = 2
                    st.info("️ Второе различие = применение первого различия дважды")
                
                elif "Логарифмическое" in stationarity_method:
                    param_value = "log"
                    st.info("ℹ️ Применяется ln(y_t) - ln(y_{t-1})")
                
                elif "Дробное" in stationarity_method:
                    frac_d = st.slider(
                        "Дробный порядок d:",
                        min_value=0.1,
                        max_value=0.9,
                        value=0.5,
                        step=0.1,
                        key="stationarity_frac_d",
                        help="d∈(0,1). Меньше d = больше памяти сохраняется. Обычно 0.3-0.7."
                    )
                    param_value = frac_d
                    
                    # Подсказка по d
                    frac_hints = {
                        0.3: " Слабое дифференцирование, сохраняется ~70% памяти",
                        0.5: " Умеренное дифференцирование, сохраняется ~50% памяти",
                        0.7: " Сильное дифференцирование, сохраняется ~30% памяти"
                    }
                    st.info(f"💡 {frac_hints.get(frac_d, '')}")
                
                elif "Комбинированное" in stationarity_method:
                    c_d, c_s = st.columns(2)
                    with c_d:
                        d_order = st.slider(
                            "Порядок d (первое различие):",
                            min_value=1,
                            max_value=2,
                            value=1,
                            step=1,
                            key="stationarity_combo_d"
                        )
                    with c_s:
                        s_period = st.number_input(
                            "Сезонный период s:",
                            min_value=2,
                            max_value=365,
                            value=default_s,
                            step=1,
                            key="stationarity_combo_s"
                        )
                    param_value = (d_order, s_period)
                
                st.divider()
                
                # Кнопки действий
                if st.button("▶ Применить дифференцирование", type="primary", use_container_width=True, key="btn_apply_stationarity"):
                    if "Логарифмическое" in stationarity_method and has_negative:
                        st.error("❌ Нельзя применить: есть неположительные значения")
                    else:
                        st.session_state.show_stationarity_preview = True
                        st.rerun()
                
                if st.button("↶ Сбросить", use_container_width=True, key="btn_reset_stationarity"):
                    st.session_state.show_stationarity_preview = False
                    st.rerun()
            
            # ── ЦЕНТРАЛЬНАЯ КОЛОНКА: ВИЗУАЛИЗАЦИЯ ──────────
            with c2:
                st.markdown("###### Визуализация: До / После")
                
                if target_col in df_stat_ts.columns:
                    original_series = df_stat_ts[target_col].dropna().astype(float)
                    
                    if st.session_state.show_stationarity_preview:
                        try:
                            # Применяем дифференцирование
                            if "Первое" in stationarity_method:
                                differenced = apply_differencing(original_series, 'first', d=param_value)
                                method_name = f"First Diff (d={param_value})"
                            
                            elif "Сезонное" in stationarity_method:
                                differenced = apply_differencing(original_series, 'seasonal', s=param_value)
                                method_name = f"Seasonal Diff (s={param_value})"
                            
                            elif "Второе" in stationarity_method:
                                differenced = apply_differencing(original_series, 'second')
                                method_name = "Second Diff (d=2)"
                            
                            elif "Логарифмическое" in stationarity_method:
                                differenced = apply_differencing(original_series, 'log')
                                method_name = "Log Diff"
                            
                            elif "Дробное" in stationarity_method:
                                differenced = apply_differencing(original_series, 'fractional', frac_d=param_value)
                                method_name = f"Fractional Diff (d={param_value})"
                            
                            elif "Комбинированное" in stationarity_method:
                                d_order, s_period = param_value
                                differenced = apply_differencing(original_series, 'combined', d=d_order, s=s_period)
                                method_name = f"Combined (d={d_order}, s={s_period})"
                            
                            # ── ГРАФИК СРАВНЕНИЯ (2 ряда) ───────
                            fig = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    f" Исходный ряд: {target_col}",
                                    f" После дифференцирования ({method_name})"
                                ),
                                vertical_spacing=0.12
                            )
                            
                            # Исходный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=original_series.index, y=original_series.values,
                                    mode='lines',
                                    name='Исходные данные',
                                    line=dict(color='#048A81', width=2),
                                    showlegend=False
                                ),
                                row=1, col=1
                            )
                            
                            # Дифференцированный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=differenced.index, y=differenced.values,
                                    mode='lines',
                                    name='После дифференцирования',
                                    line=dict(color='#DC2626', width=2),
                                    showlegend=False
                                ),
                                row=2, col=1
                            )
                            
                            fig.update_layout(
                                height=600,
                                margin=dict(l=50, r=20, t=80, b=40),
                                hovermode='x unified',
                            )
                            
                            fig.update_xaxes(title_text="Дата", row=2, col=1)
                            fig.update_yaxes(title_text="Значение", row=1, col=1)
                            fig.update_yaxes(title_text="Разность", row=2, col=1)
                            
                            fig.update_annotations(font=dict(size=13, color="#1e293b"), yshift=10)
                            
                            st.plotly_chart(fig, use_container_width=True, key="stationarity_main_chart")
                            
                            # ── ACF ДО/ПОСЛЕ ────────────────────
                            st.markdown("###### Автокорреляция (ACF): До / После")
                            st.caption("После дифференцирования ACF должна быстро затухать (все лаги в синей зоне)")
                            
                            from statsmodels.tsa.stattools import acf
                            
                            max_lag = min(40, len(original_series) // 4)
                            acf_before = acf(original_series.dropna(), nlags=max_lag)
                            acf_after = acf(differenced.dropna(), nlags=max_lag)
                            
                            fig_acf = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    "ACF исходного ряда",
                                    f"ACF после дифференцирования ({method_name})"
                                ),
                                vertical_spacing=0.15
                            )
                            
                            # ACF до
                            fig_acf.add_trace(
                                go.Bar(x=list(range(len(acf_before))), y=acf_before,
                                    marker_color='#048A81', name='ACF До', showlegend=False),
                                row=1, col=1
                            )
                            conf_int = 1.96 / np.sqrt(len(original_series))
                            fig_acf.add_hline(y=conf_int, line_dash="dash", line_color="red", row=1, col=1)
                            fig_acf.add_hline(y=-conf_int, line_dash="dash", line_color="red", row=1, col=1)
                            
                            # ACF после
                            fig_acf.add_trace(
                                go.Bar(x=list(range(len(acf_after))), y=acf_after,
                                    marker_color='#DC2626', name='ACF После', showlegend=False),
                                row=2, col=1
                            )
                            conf_int_after = 1.96 / np.sqrt(len(differenced))
                            fig_acf.add_hline(y=conf_int_after, line_dash="dash", line_color="red", row=2, col=1)
                            fig_acf.add_hline(y=-conf_int_after, line_dash="dash", line_color="red", row=2, col=1)
                            
                            fig_acf.update_layout(height=500, margin=dict(l=50, r=20, t=60, b=40))
                            st.plotly_chart(fig_acf, use_container_width=True, key="stationarity_acf_chart")
                            
                            # ── ROLLING MEAN/STD ДО/ПОСЛЕ ───────
                            st.markdown("###### Скользящие mean и std (тест на постоянство)")
                            st.caption("После дифференцирования rolling mean должен быть около 0, rolling std — постоянным")
                            
                            window = min(30, len(original_series) // 5)
                            rolling_mean_before = original_series.rolling(window=window).mean()
                            rolling_std_before = original_series.rolling(window=window).std()
                            rolling_mean_after = differenced.rolling(window=window).mean()
                            rolling_std_after = differenced.rolling(window=window).std()
                            
                            fig_rolling = make_subplots(
                                rows=2, cols=2,
                                subplot_titles=(
                                    "Rolling Mean (До)", "Rolling Mean (После)",
                                    "Rolling Std (До)", "Rolling Std (После)"
                                ),
                                vertical_spacing=0.12,
                                horizontal_spacing=0.1
                            )
                            
                            fig_rolling.add_trace(
                                go.Scatter(x=rolling_mean_before.index, y=rolling_mean_before.values,
                                        mode='lines', line=dict(color='#048A81', width=2), showlegend=False),
                                row=1, col=1
                            )
                            fig_rolling.add_trace(
                                go.Scatter(x=rolling_mean_after.index, y=rolling_mean_after.values,
                                        mode='lines', line=dict(color='#DC2626', width=2), showlegend=False),
                                row=1, col=2
                            )
                            fig_rolling.add_trace(
                                go.Scatter(x=rolling_std_before.index, y=rolling_std_before.values,
                                        mode='lines', line=dict(color='#048A81', width=2), showlegend=False),
                                row=2, col=1
                            )
                            fig_rolling.add_trace(
                                go.Scatter(x=rolling_std_after.index, y=rolling_std_after.values,
                                        mode='lines', line=dict(color='#DC2626', width=2), showlegend=False),
                                row=2, col=2
                            )
                            
                            fig_rolling.update_layout(height=500, margin=dict(l=50, r=20, t=60, b=40))
                            st.plotly_chart(fig_rolling, use_container_width=True, key="stationarity_rolling_chart")
                            
                            # ── ЗАПУСК ТЕСТОВ НА ДИФФЕРЕНЦИРОВАННОМ РЯДЕ ──
                            tests_after = run_stationarity_tests(differenced)
                            
                            # Сохранение результатов
                            st.session_state.stationarity_transform_result = {
                                'method': method_name,
                                'param': param_value,
                                'original_series': original_series,
                                'differenced_series': differenced,
                                'target_col': target_col,
                                'tests_before': tests_raw,
                                'tests_after': tests_after
                            }
                            
                            st.divider()
                            
                            # ─ СРАВНЕНИЕ ТЕСТОВ ДО/ПОСЛЕ ───────
                            st.markdown("###### Сравнение тестов стационарности: До / После")
                            
                            c_test1, c_test2 = st.columns(2)
                            
                            with c_test1:
                                st.markdown("**ДО дифференцирования:**")
                                if 'adf' in tests_raw:
                                    adf_p_before = tests_raw['adf']['pvalue']
                                    adf_status_before = "✅ Стационарен" if adf_p_before < 0.05 else "❌ Нестационарен"
                                    st.metric("ADF p-value", f"{adf_p_before:.4f}", 
                                            delta=adf_status_before)
                                if 'kpss' in tests_raw and 'pvalue_level' in tests_raw['kpss']:
                                    kpss_p_before = tests_raw['kpss']['pvalue_level']
                                    kpss_status_before = "✅ Стационарен" if kpss_p_before > 0.05 else "❌ Нестационарен"
                                    st.metric("KPSS p-value (level)", f"{kpss_p_before:.4f}",
                                            delta=kpss_status_before)
                            
                            with c_test2:
                                st.markdown("**ПОСЛЕ дифференцирования:**")
                                if 'adf' in tests_after:
                                    adf_p_after = tests_after['adf']['pvalue']
                                    adf_status_after = "✅ Стационарен" if adf_p_after < 0.05 else "❌ Нестационарен"
                                    delta_adf = adf_p_after - tests_raw.get('adf', {}).get('pvalue', 0)
                                    st.metric("ADF p-value", f"{adf_p_after:.4f}",
                                            delta=f"{delta_adf:+.4f}")
                                if 'kpss' in tests_after and 'pvalue_level' in tests_after['kpss']:
                                    kpss_p_after = tests_after['kpss']['pvalue_level']
                                    kpss_status_after = "✅ Стационарен" if kpss_p_after > 0.05 else "❌ Нестационарен"
                                    delta_kpss = kpss_p_after - tests_raw.get('kpss', {}).get('pvalue_level', 0)
                                    st.metric("KPSS p-value (level)", f"{kpss_p_after:.4f}",
                                            delta=f"{delta_kpss:+.4f}")
                            
                            # Консенсус после
                            consensus_after = tests_after.get('consensus', 'unknown')
                            consensus_colors = {
                                'stationary': '#16a34a',
                                'trend-stationary': '#d97706',
                                'non-stationary': '#dc2626',
                                'inconclusive': '#6b7280'
                            }
                            consensus_texts = {
                                'stationary': '✅ Стационарен',
                                'trend-stationary': '⚠️ Тренд-стационарен',
                                'non-stationary': '❌ Нестационарен',
                                'inconclusive': '⚠️ Неопределённость'
                            }
                            color_after = consensus_colors.get(consensus_after, '#6b7280')
                            text_after = consensus_texts.get(consensus_after, 'Неизвестно')
                            
                            st.markdown(f"<div style='background: {color_after}20; border-left: 4px solid {color_after}; "
                                    f"padding: 12px; border-radius: 6px; margin: 15px 0;'>"
                                    f"<strong>Консенсус после дифференцирования:</strong> "
                                    f"<span style='color: {color_after}; font-weight: 600;'>{text_after}</span><br>"
                                    f"<small>{tests_after.get('recommendation', '')}</small>"
                                    f"</div>", unsafe_allow_html=True)
                            
                            # Проверка на over-differencing
                            if len(acf_after) > 1 and acf_after[1] < -0.5:
                                st.warning("⚠️ **Признак over-differencing:** ACF первого лага < -0.5. "
                                        "Ряд переусложнён. Попробуйте меньший порядок d.")
                            
                            # Кнопки подтверждения
                            c_ok_stat, c_cancel_stat = st.columns(2)
                            with c_ok_stat:
                                if st.button("✅ Применить к данным", type="primary", use_container_width=True, key="btn_confirm_stationarity"):
                                    # Применяем дифференцирование к основному df
                                    df_final_stat = st.session_state.df.copy()
                                    
                                    # Сохраняем оригинальный столбец
                                    orig_col_name = f"{target_col}_original"
                                    if orig_col_name not in df_final_stat.columns:
                                        df_final_stat[orig_col_name] = df_final_stat[target_col]
                                    
                                    # Применяем дифференцирование
                                    if "Первое" in stationarity_method:
                                        df_final_stat[target_col] = df_final_stat[target_col].astype(float).diff(param_value)
                                    elif "Сезонное" in stationarity_method:
                                        df_final_stat[target_col] = df_final_stat[target_col].astype(float).diff(param_value)
                                    elif "Второе" in stationarity_method:
                                        df_final_stat[target_col] = df_final_stat[target_col].astype(float).diff(2)
                                    elif "Логарифмическое" in stationarity_method:
                                        df_final_stat[target_col] = np.log(df_final_stat[target_col].astype(float)).diff()
                                    elif "Дробное" in stationarity_method:
                                        # Для дробного — применяем к основному df
                                        series_main = df_final_stat[target_col].astype(float).dropna()
                                        diff_frac = apply_differencing(series_main, 'fractional', frac_d=param_value)
                                        df_final_stat = df_final_stat.loc[diff_frac.index]
                                        df_final_stat[target_col] = diff_frac.values
                                    elif "Комбинированное" in stationarity_method:
                                        d_order, s_period = param_value
                                        df_final_stat[target_col] = df_final_stat[target_col].astype(float).diff(s_period).diff(d_order)
                                    
                                    # Удаляем NaN после дифференцирования
                                    df_final_stat = df_final_stat.dropna(subset=[target_col])
                                    
                                    # Сохраняем параметры
                                    st.session_state.stationarity_transform_params = {
                                        'method': method_name,
                                        'param': param_value,
                                        'column': target_col,
                                        'original_col_name': orig_col_name,
                                        'tests_before': tests_raw,
                                        'tests_after': tests_after
                                    }
                                    
                                    # Синхронизация
                                    st.session_state.df = df_final_stat.copy()
                                    st.session_state.validation_ready = False
                                    st.session_state.show_stationarity_preview = False
                                    
                                    # Удаляем рабочие копии
                                    work_dfs = [
                                        "df_missing_work", "df_pattern_work", "df_range_work",
                                        "df_outlier_work", "df_inclusion_work", "df_referential_work",
                                        "df_text_work", "df_consistency_work", "df_uniqueness_work",
                                        "df_regularity_work", "df_regular_work", "df_variance_work",
                                        "df_smooth_work", "df_stationarity_work"
                                    ]
                                    for work_df_name in work_dfs:
                                        if work_df_name in st.session_state:
                                            del st.session_state[work_df_name]
                                    
                                    if "val_results" in st.session_state:
                                        del st.session_state.val_results
                                    
                                    st.success(f"✅ Дифференцирование **{method_name}** применено!")
                                    st.info(f"💡 Оригинальные значения сохранены в колонке `{orig_col_name}`. "
                                        f"Удалено {len(df_final_stat) - len(df_final_stat.dropna())} строк с NaN. "
                                        f"Перезапустите валидацию для обновления статистик.")
                                    st.rerun()
                            
                            with c_cancel_stat:
                                if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_stationarity"):
                                    st.session_state.show_stationarity_preview = False
                                    st.rerun()
                        
                        except Exception as e:
                            st.error(f"❌ Ошибка дифференцирования: {e}")
                            import traceback
                            with st.expander("🔍 Stack trace"):
                                st.code(traceback.format_exc(), language="python")
                    
                    else:
                        # Показываем только исходный ряд
                        fig = px.line(
                            x=original_series.index,
                            y=original_series.values,
                            labels={'x': 'Дата', 'y': target_col},
                        )
                        fig.update_layout(height=400, margin=dict(l=40, r=20, t=40, b=40))
                        st.plotly_chart(fig, use_container_width=True, key="stationarity_main_chart")
                        
                        st.info("💡 Выберите метод дифференцирования и нажмите **'Применить дифференцирование'** для просмотра результата.")
                else:
                    st.warning("⚠️ Выбранная колонка не найдена в данных.")
            
            # ─ ПРАВАЯ КОЛОНКА: МЕТРИКИ КАЧЕСТВА ───────────
            with c3:
                st.markdown("###### Метрики качества")
                
                with st.container(border=True):
                    st.markdown("**До дифференцирования:**")
                    
                    if 'adf' in tests_raw:
                        adf_p = tests_raw['adf']['pvalue']
                        adf_status = "❌ Нестацион." if adf_p >= 0.05 else "✅ Стацион."
                        st.metric("ADF p-value", f"{adf_p:.4f}", delta=adf_status)
                    
                    if 'kpss' in tests_raw and 'pvalue_level' in tests_raw['kpss']:
                        kpss_p = tests_raw['kpss']['pvalue_level']
                        kpss_status = "✅ Стацион." if kpss_p > 0.05 else "❌ Нестацион."
                        st.metric("KPSS p-value", f"{kpss_p:.4f}", delta=kpss_status)
                    
                    if 'pp' in tests_raw:
                        pp_p = tests_raw['pp']['pvalue']
                        pp_status = "❌ Нестацион." if pp_p >= 0.05 else "✅ Стацион."
                        st.metric("PP p-value", f"{pp_p:.4f}", delta=pp_status)
                    
                    st.divider()
                    st.markdown("**Статистики ряда:**")
                    st.metric("Длина ряда", f"{len(series):,}".replace(",", " "))
                    st.metric("Среднее", f"{series.mean():.2f}")
                    st.metric("Стд. отклонение", f"{series.std():.2f}")
                
                # Метрики ПОСЛЕ дифференцирования
                if st.session_state.show_stationarity_preview and 'stationarity_transform_result' in st.session_state:
                    tests_after = st.session_state.stationarity_transform_result['tests_after']
                    differenced_series = st.session_state.stationarity_transform_result['differenced_series']
                    
                    with st.container(border=True):
                        st.markdown("**После дифференцирования:**")
                        
                        if 'adf' in tests_after:
                            adf_p_after = tests_after['adf']['pvalue']
                            adf_p_before = tests_raw.get('adf', {}).get('pvalue', 0)
                            delta_adf = adf_p_after - adf_p_before
                            adf_status = "✅ Стацион." if adf_p_after < 0.05 else "❌ Нестацион."
                            st.metric("ADF p-value", f"{adf_p_after:.4f}",
                                    delta=f"{delta_adf:+.4f}")
                        
                        if 'kpss' in tests_after and 'pvalue_level' in tests_after['kpss']:
                            kpss_p_after = tests_after['kpss']['pvalue_level']
                            kpss_p_before = tests_raw.get('kpss', {}).get('pvalue_level', 0)
                            delta_kpss = kpss_p_after - kpss_p_before
                            kpss_status = "✅ Стацион." if kpss_p_after > 0.05 else "❌ Нестацион."
                            st.metric("KPSS p-value", f"{kpss_p_after:.4f}",
                                    delta=f"{delta_kpss:+.4f}")
                        
                        if 'pp' in tests_after:
                            pp_p_after = tests_after['pp']['pvalue']
                            pp_p_before = tests_raw.get('pp', {}).get('pvalue', 0)
                            delta_pp = pp_p_after - pp_p_before
                            pp_status = "✅ Стацион." if pp_p_after < 0.05 else " Нестацион."
                            st.metric("PP p-value", f"{pp_p_after:.4f}",
                                    delta=f"{delta_pp:+.4f}")
                        
                        st.divider()
                        st.markdown("**Статистики ряда:**")
                        st.metric("Длина ряда", f"{len(differenced_series):,}".replace(",", " "))
                        st.metric("Среднее", f"{differenced_series.mean():.4f}")
                        st.metric("Стд. отклонение", f"{differenced_series.std():.4f}")
                    
                    st.divider()
                    
                    # Рекомендации
                    consensus_after = tests_after.get('consensus', 'unknown')
                    if consensus_after == 'stationary':
                        st.success("✅ **Ряд стал стационарным!** Можно применять ARIMA/VAR.")
                    elif consensus_after == 'trend-stationary':
                        st.warning("⚠️ **Тренд-стационарен.** Достаточно детренда (удалить линейный тренд).")
                    elif consensus_after == 'non-stationary':
                        st.error("❌ **Ряд всё ещё нестационарен.** Попробуйте другой метод или больший порядок.")
                    else:
                        st.info("⚠️ **Неопределённость.** Визуальный анализ + пробное дифференцирование.")
                    
                    # Информация о параметрах
                    st.info(f"**Параметры:** {st.session_state.stationarity_transform_result['method']}")
            
            # ── ИНФОРМАЦИЯ О СОХРАНЁННЫХ ТРАНСФОРМАЦИЯХ ────
            if st.session_state.stationarity_transform_params:
                st.divider()
                with st.expander("💾 Сохранённые параметры дифференцирования", expanded=False):
                    st.json(st.session_state.stationarity_transform_params)
                    st.caption("Эти параметры будут использованы для **обратного преобразования** прогноза "
                            "в исходную шкалу (через кумулятивную сумму).")
        
        else:
            st.warning("⚠️ В датасете нет числовых колонок для анализа.")
    else:
        st.warning("⚠️ Не обнаружены колонки с датами или числовыми данными. Убедитесь, что активирован режим временных рядов.")


    # ═══════════════════════════════════════════════════════
    # 🔹 8. FEATURE ENGINEERING + СПЕКТРАЛЬНЫЙ АНАЛИЗ
    # ═══════════════════════════════════════════════════════
    st.divider()
    st.markdown("### Feature Engineering + Спектральный анализ")
    st.caption("Фичинг - создание новых признаков: лаги, скользящие статистики, спектральные характеристики. "
            "Ознакомительный анализ частотной структуры ряда после преобразований.")

    # ── ЦЕЛИ И РЕЗУЛЬТАТЫ ────────────────────────────────
    with st.expander("Цели субмодуля ⁞ Feature Engineering", expanded=False):
        st.markdown("""
        **Feature Engineering** — процесс создания новых признаков из существующих данных для улучшения качества моделей.
        Фичинг позволяет моделям ML «увидеть» временные закономерности, которые они не распознали бы сами (например, циклические эффекты, длинные лаги).
                    
        **Что создаём:**
        1. **Лаги (Lags)** — предыдущие значения ряда (t-1, t-2, ...) для ARIMA, LSTM
        2. **Скользящие статистики** — mean, std, min, max за окно (тренды, волатильность)
        3. **Спектральные признаки** — доминирующие частоты, спектральная энергия, энтропия
        4. **Временные признаки** — час, день недели, месяц, квартал, год, weekend, holiday
        5. **Разности (Differences)** — для стационарности (t - t-1)
        6. **Процентные изменения** — для анализа динамики
        7. **EWMA** — экспоненциально взвешенное скользящее среднее
        8. **Отношения к скользящему среднему** — для выявления отклонений
        
        **Спектральный анализ (ознакомительный):**
        - **FFT (Fast Fourier Transform)** — выявление доминирующих периодов
        - **Periodogram** — спектральная плотность мощности
        - **Wavelet (CWT)** — временно-частотная локализация
        - **ACF/PACF** — автокорреляционная и частная автокорреляционная функции
        
        **Зачем это нужно:**
        - ML-модели (LSTM, XGBoost, Random Forest) требуют числовых признаков
        - Спектральный анализ помогает понять структуру ряда ДО моделирования
        - Лаги и скользящие окна захватывают временные зависимости
        - Спектральные признаки улучшают прогноз для сложных паттернов
        
        **Результат:** Расширенный датасет с новыми колонками, готовый для ML-моделей.
        </div>
        """, unsafe_allow_html=True)

    # ── ПРОВЕРКА ГОТОВНОСТИ ─────────────────────────────
    if st.session_state.primary_date_col and st.session_state.col_types.get("num"):
        date_col = st.session_state.primary_date_col
        num_cols = st.session_state.col_types.get("num", [])
        
        if num_cols:
            # Выбор целевой колонки
            target_col = st.selectbox(
                " Исследуемый признак:",
                options=num_cols,
                index=0,
                key="fe_target_col",
                help="Выберите числовую колонку для создания признаков"
            )
            
            # Инициализация session_state
            if "df_fe_work" not in st.session_state:
                st.session_state.df_fe_work = st.session_state.df.copy()
            if "fe_features_created" not in st.session_state:
                st.session_state.fe_features_created = []
            
            df_work = st.session_state.df_fe_work
            
            # ────────────────────────────────────────────────
            # 🔹 ЧАСТЬ 1: ОЗНАКОМИТЕЛЬНЫЙ СПЕКТРАЛЬНЫЙ АНАЛИЗ
            # ────────────────────────────────────────────────
            st.markdown("##### Ознакомительный спектральный анализ")
            st.caption("Анализ частотной структуры ряда перед созданием признаков")
            
            # Подготовка временного ряда
            try:
                df_ts = df_work.copy()
                df_ts[date_col] = pd.to_datetime(df_ts[date_col])
                df_ts = df_ts.sort_values(date_col)
                df_ts = df_ts.set_index(date_col)[[target_col]].dropna()
                
                if len(df_ts) < 30:
                    st.warning(f"️ Недостаточно данных для спектрального анализа: {len(df_ts)} точек (минимум 30)")
                else:
                    # Ряд для анализа
                    series = df_ts[target_col].astype(float)
                    
                    # ── БЫСТРЫЙ СПЕКТРАЛЬНЫЙ АНАЛИЗ ────────
                    c_spec1, c_spec2, c_spec3, c_spec4 = st.columns(4)
                    
                    with c_spec1:
                        # FFT
                        from scipy.fft import fft, fftfreq
                        from scipy.signal import find_peaks
                        
                        n = len(series)
                        y = series.values - series.mean()
                        yf = fft(y)
                        xf = fftfreq(n, 1)[:n//2]
                        amplitude = 2.0/n * np.abs(yf[0:n//2])
                        
                        # Доминирующие частоты
                        peaks, _ = find_peaks(amplitude, height=np.mean(amplitude) + np.std(amplitude))
                        dominant_periods = [1/xf[p] for p in peaks if xf[p] > 0 and xf[p] < 0.5]
                        
                        #  Форматируем периоды с разделителями тысяч
                        if dominant_periods:
                            periods_formatted = ', '.join([f'{p:,.1f}'.replace(",", " ") for p in dominant_periods[:3]])
                        else:
                            periods_formatted = "Не обнаружены"
                        
                        st.metric(" Доминирующие периоды (FFT)", 
                                f"{len(dominant_periods)}",
                                delta=periods_formatted)

                    with c_spec2:
                        # Периодограмма
                        from scipy.signal import periodogram
                        
                        freq_per, pxx_per = periodogram(series.values, fs=1.0, window='hann')
                        spectral_energy = np.sum(pxx_per)
                        
                        #  Форматируем спектральную энергию с разделителями тысяч
                        spectral_energy_formatted = f"{spectral_energy:,.2f}".replace(",", " ")
                        
                        st.metric(" Спектральная энергия", 
                                spectral_energy_formatted,
                                delta="Мощный сигнал" if spectral_energy > 100 else "Слабый сигнал")
                    
                    with c_spec3:
                        # Спектральная энтропия
                        spectrum = np.abs(yf)**2
                        spectrum_norm = spectrum / np.sum(spectrum)
                        spectral_entropy = -np.sum(spectrum_norm * np.log(spectrum_norm + 1e-10))
                        
                        st.metric(" Спектральная энтропия", 
                                f"{spectral_entropy:.3f}",
                                delta="Сложный" if spectral_entropy > 3 else "Простой")
                    
                    with c_spec4:
                        # Соотношение низких/высоких частот
                        mid_freq = len(xf)//4
                        low_energy = np.sum(amplitude[:mid_freq]**2)
                        high_energy = np.sum(amplitude[mid_freq:]**2)
                        ratio = low_energy / (high_energy + 1e-10)
                        
                        st.metric(" Low/High частоты", 
                                f"{ratio:.2f}",
                                delta="Низкочастотный" if ratio > 2 else "Высокочастотный")
                    
                    # ── ВИЗУАЛИЗАЦИЯ ─────────────────────
                    st.divider()

                    tab_fft, tab_per, tab_wave, tab_acf = st.tabs([
                        "FFT спектр", 
                        "Периодограмма", 
                        "Wavelet", 
                        "ACF/PACF"
                    ])

                    with tab_fft:
                        fig_fft = px.line(
                            x=xf[:n//2], 
                            y=amplitude,
                            title="FFT: Амплитудный спектр",
                            labels={'x': 'Частота', 'y': 'Амплитуда'}
                        )
                        # Отметить пики
                        for p in peaks[:5]:
                            if xf[p] > 0:
                                fig_fft.add_vline(x=xf[p], line_dash="dash", line_color="red")
                                fig_fft.add_annotation(
                                    x=xf[p], 
                                    y=amplitude[p],
                                    text=f"Period={1/xf[p]:.1f}",
                                    showarrow=True,
                                    arrowhead=2,
                                    ax=0,
                                    ay=-40
                                )
                        # ИЗМЕНЕНО: height=600 (вместо 400)
                        fig_fft.update_layout(height=600, showlegend=False)
                        st.plotly_chart(fig_fft, use_container_width=True, key="fft_spectrum_chart")

                    with tab_per:
                        fig_per = px.line(
                            x=freq_per, 
                            y=pxx_per,
                            title="Периодограмма (спектральная плотность мощности)",
                            labels={'x': 'Частота', 'y': 'Мощность'}
                        )
                        # 🔧 ИЗМЕНЕНО: height=600 (вместо 400)
                        fig_per.update_layout(height=600, showlegend=False)
                        st.plotly_chart(fig_per, use_container_width=True, key="periodogram_chart")

                    with tab_wave:
                        try:
                            import pywt
                            
                            widths = np.arange(1, min(128, len(series)//4))
                            cwtmatr, _ = pywt.cwt(series.values - series.mean(), widths, 'morl', sampling_period=1)
                            
                            fig_wave = px.imshow(
                                np.abs(cwtmatr),
                                title="Wavelet Scalogram (CWT)",
                                labels={'x': 'Время', 'y': 'Масштаб (период)'},
                                color_continuous_scale='Viridis',
                                aspect='auto'  # Автоматическое соотношение сторон
                            )

                            fig_wave.update_layout(
                                height=600,  # ✅ Уже 600
                                yaxis=dict(
                                    scaleanchor='x',
                                    scaleratio=0.3,  # Соотношение Y к X (увеличивает Y)
                                    title='Масштаб (период)',
                                    tickmode='linear',
                                    tick0=0,
                                    dtick=20  # Шаг меток по Y
                                ),
                                xaxis=dict(
                                    title='Время',
                                    tickmode='linear',
                                    tick0=0,
                                    dtick=1000  # Шаг меток по X (1k, 2k, 3k...)
                                )
                            )

                            st.plotly_chart(fig_wave, use_container_width=True, key="wavelet_chart")
                        except ImportError:
                            st.warning("⚠️ Установите PyWavelets: `pip install PyWavelets`")

                    with tab_acf:
                        from statsmodels.tsa.stattools import acf, pacf
                        
                        max_lag = min(60, len(series) // 4)
                        acf_values = acf(series, nlags=max_lag)
                        pacf_values = pacf(series, nlags=max_lag)
                        
                        fig_acf = make_subplots(
                            rows=2, cols=1,
                            subplot_titles=("Автокорреляционная функция (ACF)", 
                                        "Частная автокорреляционная функция (PACF)")
                        )

                        fig_acf.add_trace(
                            go.Bar(x=list(range(len(acf_values))), y=acf_values, name='ACF'),
                            row=1, col=1
                        )
                        fig_acf.add_trace(
                            go.Bar(x=list(range(len(pacf_values))), y=pacf_values, name='PACF'),
                            row=2, col=1
                        )
                        
                        # Доверительный интервал
                        conf_int = 1.96 / np.sqrt(len(series))
                        fig_acf.add_hline(y=conf_int, line_dash="dash", line_color="red")
                        fig_acf.add_hline(y=-conf_int, line_dash="dash", line_color="red")
                        fig_acf.update_layout(height=600, showlegend=False)  # ✅ Уже 600
                        st.plotly_chart(fig_acf, use_container_width=True, key="acf_pacf_chart")
                    
                    # ── АВТОПОДБОР ЛАГОВ ЧЕРЕЗ ACF/PACF ─────
                    st.divider()
                    st.markdown("###### Автоподбор лагов через ACF/PACF")
                    
                    if st.button(" Автоматически определить значимые лаги", key="btn_auto_lags"):
                        from statsmodels.tsa.stattools import acf, pacf
                        
                        max_lag = min(60, len(series) // 4)
                        acf_values = acf(series, nlags=max_lag)
                        pacf_values = pacf(series, nlags=max_lag)
                        conf_int = 1.96 / np.sqrt(len(series))
                        
                        # Находим значимые лаги
                        significant_acf = np.where(np.abs(acf_values[1:]) > conf_int)[0] + 1
                        significant_pacf = np.where(np.abs(pacf_values[1:]) > conf_int)[0] + 1
                        
                        # Рекомендуемые лаги (первые 5 значимых)
                        recommended_lags = sorted(set(significant_acf.tolist() + significant_pacf.tolist()))[:10]
                        
                        st.session_state.recommended_lags = recommended_lags
                        
                        st.success(f"✅ Обнаружено {len(recommended_lags)} значимых лагов: {recommended_lags}")
                        st.info("💡 Эти лаги будут использованы при создании признаков")
                    
                    # Сохранение результатов спектрального анализа
                    st.session_state.spectral_analysis = {
                        'dominant_periods': dominant_periods,
                        'spectral_energy': spectral_energy,
                        'spectral_entropy': spectral_entropy,
                        'low_high_ratio': ratio
                    }
            
            except Exception as e:
                st.error(f"❌ Ошибка спектрального анализа: {e}")
            
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            
            # ────────────────────────────────────────────────
            # 🔹 ЧАСТЬ 2: СОЗДАНИЕ ПРИЗНАКОВ
            # ────────────────────────────────────────────────
            st.markdown("##### Создание признаков (Feature Engineering)")
            
            # ── ПАРАМЕТРЫ ─────────────────────────────────
            st.markdown("######  Временные признаки")
            create_time = st.checkbox("Создать временные признаки", value=True, key="fe_create_time")
            if create_time:
                c_t1, c_t2 = st.columns(2)
                with c_t1:
                    create_weekend = st.checkbox("Is_weekend (выходной)", value=True, key="fe_weekend")
                    create_holiday = st.checkbox("Is_holiday (праздник)", value=False, key="fe_holiday")
                with c_t2:
                    create_dayofyear = st.checkbox("День года (1-365)", value=True, key="fe_dayofyear")
                    create_quarter = st.checkbox("Квартал", value=True, key="fe_quarter")
            
            st.markdown("###### Лаговые признаки (Lag features)")
            create_lags = st.checkbox("Создать лаги", value=True, key="fe_create_lags")
            if create_lags:
                # Определяем доступные опции
                lag_options = [1, 2, 3, 5, 7, 10, 14, 21, 30, 60, 90]

                # Получаем рекомендованные лаги и фильтруем только те, что есть в опциях
                recommended_lags = st.session_state.get('recommended_lags', [])
                default_lags = [lag for lag in recommended_lags if lag in lag_options]

                # Если после фильтрации ничего не осталось — используем стандартный набор
                if not default_lags:
                    default_lags = [1, 7, 14, 30]

                lag_periods = st.multiselect(
                    "Периоды лагов:",
                    options=lag_options,
                    default=default_lags,
                    key="lag_periods"
                )
            
            st.markdown("###### Скользящие статистики (Rolling)")
            create_rolling = st.checkbox("Скользящие статистики", value=True, key="fe_create_rolling")
            if create_rolling:
                c_r1, c_r2 = st.columns(2)
                with c_r1:
                    rolling_windows = st.multiselect(
                        "Окна (периоды):",
                        options=[3, 5, 7, 10, 14, 21, 30, 60, 90],
                        default=[7, 14, 30],
                        key="rolling_windows"
                    )
                    rolling_stats = st.multiselect(
                        "Статистики:",
                        options=['mean', 'std', 'min', 'max'],
                        default=['mean', 'std'],
                        key="rolling_stats"
                    )
                with c_r2:
                    create_ewma = st.checkbox("EWMA (экспоненциальное скользящее)", value=True, key="fe_ewma")
                    if create_ewma:
                        ewma_spans = st.multiselect(
                            "EWMA периоды (span):",
                            options=[7, 14, 30],
                            default=[7, 14],
                            key="ewma_spans"
                        )
            
            st.markdown("###### Производные признаки")
            c_d1, c_d2 = st.columns(2)
            with c_d1:
                create_diff = st.checkbox("Разности (differencing)", value=True, key="fe_create_diff")
                if create_diff:
                    diff_periods = st.multiselect(
                        "Периоды разностей:",
                        options=[1, 7, 14, 30],
                        default=[1, 7],
                        key="diff_periods"
                    )
            with c_d2:
                create_pct = st.checkbox("Процентные изменения (pct_change)", value=True, key="fe_create_pct")
                if create_pct:
                    pct_periods = st.multiselect(
                        "Периоды pct_change:",
                        options=[1, 7, 14, 30],
                        default=[1, 7],
                        key="pct_periods"
                    )
            
            create_ratio = st.checkbox("Отношения к скользящему среднему", value=True, key="fe_create_ratio")
            if create_ratio:
                ratio_windows = st.multiselect(
                    "Окна для отношений:",
                    options=[7, 14, 30],
                    default=[7, 14],
                    key="ratio_windows"
                )
            
            # ── КНОПКА ПРИМЕНЕНИЯ ────────────────────────
            if st.button(" Создать признаки", type="primary", use_container_width=True, key="btn_create_features"):
                try:
                    df_fe = df_work.copy()
                    df_fe[date_col] = pd.to_datetime(df_fe[date_col])
                    df_fe = df_fe.sort_values(date_col)
                    features_created = []
                    
                    # 1. ВРЕМЕННЫЕ ПРИЗНАКИ
                    if create_time:
                        if pd.api.types.is_datetime64_any_dtype(df_fe[date_col]):
                            df_fe['year'] = df_fe[date_col].dt.year
                            df_fe['month'] = df_fe[date_col].dt.month
                            df_fe['day'] = df_fe[date_col].dt.day
                            df_fe['dayofweek'] = df_fe[date_col].dt.dayofweek
                            if create_quarter:
                                df_fe['quarter'] = df_fe[date_col].dt.quarter
                            if create_dayofyear:
                                df_fe['dayofyear'] = df_fe[date_col].dt.dayofyear
                            if create_weekend:
                                df_fe['is_weekend'] = (df_fe[date_col].dt.dayofweek >= 5).astype(int)
                            if create_holiday:
                                # Упрощённая логика праздников (можно расширить)
                                df_fe['is_holiday'] = 0  # Заглушка, требует библиотеки holidays
                            features_created.extend(['year', 'month', 'day', 'dayofweek', 'quarter', 'dayofyear', 'is_weekend', 'is_holiday'])
                            st.success("✅ Созданы временные признаки")
                    
                    # 2. ЛАГИ
                    if create_lags and lag_periods:
                        for lag in lag_periods:
                            lag_col = f"{target_col}_lag_{lag}"
                            df_fe[lag_col] = df_fe[target_col].shift(lag)
                            features_created.append(lag_col)
                        st.success(f"✅ Создано лагов: {len(lag_periods)}")
                    
                    # 3. СКОЛЬЗЯЩИЕ СТАТИСТИКИ
                    if create_rolling and rolling_windows:
                        for window in rolling_windows:
                            for stat in rolling_stats:
                                roll_col = f"{target_col}_roll_{stat}_{window}"
                                if stat == 'mean':
                                    df_fe[roll_col] = df_fe[target_col].rolling(window=window).mean()
                                elif stat == 'std':
                                    df_fe[roll_col] = df_fe[target_col].rolling(window=window).std()
                                elif stat == 'min':
                                    df_fe[roll_col] = df_fe[target_col].rolling(window=window).min()
                                elif stat == 'max':
                                    df_fe[roll_col] = df_fe[target_col].rolling(window=window).max()
                                features_created.append(roll_col)
                        st.success(f"✅ Создано скользящих признаков: {len(rolling_windows) * len(rolling_stats)}")
                    
                    # 4. EWMA
                    if create_ewma and ewma_spans:
                        for span in ewma_spans:
                            ewma_col = f"{target_col}_ewma_{span}"
                            df_fe[ewma_col] = df_fe[target_col].ewm(span=span, adjust=False).mean()
                            features_created.append(ewma_col)
                        st.success(f"✅ Создано EWMA признаков: {len(ewma_spans)}")
                    
                    # 5. РАЗНОСТИ
                    if create_diff and diff_periods:
                        for period in diff_periods:
                            diff_col = f"{target_col}_diff_{period}"
                            df_fe[diff_col] = df_fe[target_col].diff(periods=period)
                            features_created.append(diff_col)
                        st.success(f"✅ Создано разностей: {len(diff_periods)}")
                    
                    # 6. ПРОЦЕНТНЫЕ ИЗМЕНЕНИЯ
                    if create_pct and pct_periods:
                        for period in pct_periods:
                            pct_col = f"{target_col}_pct_{period}"
                            df_fe[pct_col] = df_fe[target_col].pct_change(periods=period)
                            features_created.append(pct_col)
                        st.success(f"✅ Создано pct_change признаков: {len(pct_periods)}")
                    
                    # 7. ОТНОШЕНИЯ К СКОЛЬЗЯЩЕМУ СРЕДНЕМУ
                    if create_ratio and ratio_windows:
                        for window in ratio_windows:
                            ratio_col = f"{target_col}_ratio_ma_{window}"
                            rolling_mean = df_fe[target_col].rolling(window=window).mean()
                            df_fe[ratio_col] = df_fe[target_col] / rolling_mean
                            features_created.append(ratio_col)
                        st.success(f"✅ Создано отношений к MA: {len(ratio_windows)}")
                    
                    # 8. СПЕКТРАЛЬНЫЕ ПРИЗНАКИ (если есть анализ)
                    if hasattr(st.session_state, 'spectral_analysis'):
                        spec = st.session_state.spectral_analysis
                        df_fe['spectral_energy'] = spec.get('spectral_energy', 0)
                        df_fe['spectral_entropy'] = spec.get('spectral_entropy', 0)
                        df_fe['low_high_freq_ratio'] = spec.get('low_high_ratio', 0)
                        df_fe['n_dominant_periods'] = len(spec.get('dominant_periods', []))
                        features_created.extend(['spectral_energy', 'spectral_entropy', 'low_high_freq_ratio', 'n_dominant_periods'])
                        st.success("✅ Добавлены спектральные признаки")
                    
                    # Сохранение
                    st.session_state.df_fe_work = df_fe
                    st.session_state.fe_features_created = features_created
                    st.session_state.df = df_fe  # Синхронизация
                    
                    st.divider()
                    st.success(f"🎉 Всего создано признаков: **{len(features_created)}**")
                    st.info("💡 Новые колонки добавлены в датасет. Перезапустите валидацию для обновления статистик.")
                    
                    # Показать созданные признаки
                    with st.expander(" Список созданных признаков", expanded=True):
                        st.write(features_created)
                    
                    # Кнопка экспорта
                    st.download_button(
                        label=" Скачать датасет с признаками (CSV)",
                        data=df_fe.to_csv(index=False, encoding="utf-8-sig"),
                        file_name=f"features_{target_col}.csv",
                        mime="text/csv",
                        key="btn_export_features"
                    )
                    
                except Exception as e:
                    st.error(f"❌ Ошибка создания признаков: {e}")
                    import traceback
                    st.code(traceback.format_exc(), language="python")
            
            # ── ПРЕДПРОСМОТР ────────────────────────────
            if st.session_state.fe_features_created:
                st.divider()
                st.markdown("###### Предпросмотр датасета с признаками")
                
                # Показать только новые колонки
                new_cols = [col for col in df_work.columns if col in st.session_state.fe_features_created]
                if new_cols:
                    st.dataframe(
                        df_work[new_cols].head(20),
                        use_container_width=True,
                        height=300
                    )
                    st.caption(f"Показаны первые 20 строк из {len(new_cols)} новых признаков")
        
        else:
            st.warning("⚠️ Нет числовых колонок для создания признаков")
    else:
        st.warning("⚠️ Не найдены колонки с датами или числовыми данными")


    # ══════════════════════════════════════════════════════
    # 🔹 9. МАСШТАБИРОВАНИЕ (SCALING / NORMALIZATION)
    # ═══════════════════════════════════════════════════════
    st.divider()
    st.markdown("###  Масштабирование признаков (Scaling)")
    st.caption("Приведение числовых признаков к единому диапазону или распределению. Критично для ML-моделей "
            "(нейросети, SVM, k-NN, PCA), чувствительных к масштабу признаков. Для ARIMA/Prophet — опционально.")

    # ── ДИАГНОСТИКА ТЕКУЩЕГО СОСТОЯНИЯ ──────────────────
    if st.session_state.primary_date_col and st.session_state.col_types.get("num"):
        date_col = st.session_state.primary_date_col
        num_cols = st.session_state.col_types.get("num", [])
        
        if num_cols:
            df_scale = st.session_state.df.copy()
            df_scale[date_col] = pd.to_datetime(df_scale[date_col])
            df_scale = df_scale.sort_values(date_col)
            df_scale_ts = df_scale.set_index(date_col)
            
            # ── ФУНКЦИЯ МЕТРИК МАСШТАБИРОВАНИЯ ─────────────
            def calculate_scaling_metrics(original: pd.Series, scaled: pd.Series) -> dict:
                """Расчёт метрик качества масштабирования"""
                metrics = {}
                
                # 1. Диапазон
                metrics['range_orig'] = original.max() - original.min()
                metrics['range_scaled'] = scaled.max() - scaled.min()
                
                # 2. Среднее и стандартное отклонение
                metrics['mean_orig'] = original.mean()
                metrics['mean_scaled'] = scaled.mean()
                metrics['std_orig'] = original.std()
                metrics['std_scaled'] = scaled.std()
                
                # 3. Выбросы (по правилу 3σ)
                def count_outliers(series):
                    mean, std = series.mean(), series.std()
                    return ((series < mean - 3*std) | (series > mean + 3*std)).sum()
                
                metrics['outliers_orig'] = count_outliers(original)
                metrics['outliers_scaled'] = count_outliers(scaled)
                
                # 4. Асимметрия (skewness)
                metrics['skew_orig'] = original.skew()
                metrics['skew_scaled'] = scaled.skew()
                
                # 5. Эксцесс (kurtosis)
                metrics['kurt_orig'] = original.kurtosis()
                metrics['kurt_scaled'] = scaled.kurtosis()
                
                # 6. Коэффициент вариации
                metrics['cv_orig'] = original.std() / (abs(original.mean()) + 1e-10)
                metrics['cv_scaled'] = scaled.std() / (abs(scaled.mean()) + 1e-10)
                
                return metrics
            
            # ── МЕТРИКИ ТЕКУЩЕГО СОСТОЯНИЯ ─────────────────
            c_diag1, c_diag2, c_diag3, c_diag4 = st.columns(4)
            
            target_col_scale = st.session_state.get('ts_props_v10_target_col', num_cols[0])
            series_raw = df_scale_ts[target_col_scale].dropna().astype(float)
            
            # Базовая диагностика масштаба
            range_val = series_raw.max() - series_raw.min()
            cv_val = series_raw.std() / (abs(series_raw.mean()) + 1e-10)
            
            with c_diag1:
                st.markdown("**Анализ ряда**")
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: #1e293b;'>`{target_col_scale}`</div>", 
                        unsafe_allow_html=True)
            
            with c_diag2:
                st.markdown("**Диапазон значений**")
                range_formatted = f"{range_val:,.2f}".replace(",", " ")
                range_color = "#dc2626" if range_val > 1000 else "#d97706" if range_val > 100 else "#16a34a"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {range_color};'>"
                        f"{range_formatted}</div>", 
                        unsafe_allow_html=True)
            
            with c_diag3:
                st.markdown("**Коэф. вариации (CV)**")
                cv_color = "#dc2626" if cv_val > 1 else "#d97706" if cv_val > 0.5 else "#16a34a"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {cv_color};'>"
                        f"{cv_val:.3f}</div>", unsafe_allow_html=True)
            
            with c_diag4:
                st.markdown("**Рекомендация**")
                if cv_val > 1:
                    rec_text, rec_color = "Нужно масштабирование", "#dc2626"
                elif cv_val > 0.5:
                    rec_text, rec_color = "Желательно", "#d97706"
                else:
                    rec_text, rec_color = "Не критично", "#16a34a"
                st.markdown(f"<div style='font-size: 14px; font-weight: 600; color: {rec_color};'>"
                        f"{rec_text}</div>", unsafe_allow_html=True)
            
            # Отступ
            st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
            
            # ── ТЕХНИЧЕСКАЯ СПРАВКА ─────────────────────────
            with st.expander(" Цели субмодуля: Масштабирование признаков", expanded=False):
                st.markdown("""
                **Зачем нужно масштабирование:**
                - **ML-модели** — нейросети, SVM, k-NN, PCA требуют единый масштаб признаков
                - **Сходимость оптимизации** — градиентный сход быстрее на нормализованных данных
                - **Сравнение коэффициентов** — в линейной регрессии позволяет сравнивать важность признаков
                - **Distance-based методы** — евклидово расстояние чувствительно к масштабу
                
                **Когда НЕ нужно масштабирование:**
                -  **Деревья решений** (Random Forest, XGBoost, LightGBM) — инвариантны к масштабу
                - ❌ **ARIMA/Prophet** — работают с исходными значениями (масштабирование опционально)
                - ❌ **Интерпретируемость** — если нужны коэффициенты в исходных единицах
                
                **Методы масштабирования:**
                
                | Метод | Формула | Диапазон | Когда использовать |
                |-------|---------|----------|-------------------|
                | **Min-Max** | (x - min) / (max - min) | [0, 1] | Нейросети, изображения |
                | **Standardization** | (x - μ) / σ | ~[-3, 3] | ARIMA, PCA, линейная регрессия |
                | **Robust Scaling** | (x - median) / IQR | ~[-3, 3] | Данные с выбросами |
                | **🆕 MaxAbs** | x / max(\|x\|) | [-1, 1] | Разреженные данные, знак важен |
                | **🆕 Normalization (L2)** | x / \|\|x\|\|₂ | единичная сфера | Cosine similarity, SVM |
                | **🆕 Quantile Transform** | F¹(Uniform/Normal) | [0, 1] или N(0,1) | Сильно скошенные распределения |
                | **🆕 Log1p** | log(1 + x) | [0, ∞) | Count data с правым хвостом |
                
                **⚠️ Data Leakage (утечка данных):**
                - 🔴 **НЕЛЬЗЯ** fit на всём датасете, затем split на train/test
                - ✅ **Правильно:** fit на train → transform на train и test
                - ✅ **В production:** сохранить параметры (mean, std, min, max) для новых данных
                - ✅ **В CISStat:** параметры сохраняются в `session_state` для обратного преобразования прогноза
                
                **Выбор метода по модели:**
                - **Нейросети (LSTM, MLP):** Min-Max или Standardization
                - **SVM, k-NN:** Standardization или Robust Scaling
                - **PCA:** Standardization (обязательно!)
                - **Деревья (RF, XGBoost):** не требуется
                - **ARIMA:** не требуется (работает с исходными значениями)
                
                **Обратимость:**
                - 🔁 Все методы **обратимы** через `inverse_transform`
                -  Параметры сохраняются в `session_state.scaling_params`
                - 🔁 Прогноз в масштабированной шкале → обратное преобразование → исходная шкала
                
                **⚠️ Почитать:**
                - Feature scaling: https://en.wikipedia.org/wiki/Feature_scaling
                - Data leakage: https://en.wikipedia.org/wiki/Leakage_(machine_learning)
                - Scikit-learn preprocessing: https://scikit-learn.org/stable/modules/preprocessing.html
                """)
            
            # ── ПРЕДУПРЕЖДЕНИЕ О DATA LEAKAGE ─────────────
            st.warning("""
            ⚠️ **Важно: Data Leakage!**
            
            Масштабирование должно быть **fit на обучающей выборке**, а **transform на тестовой**.
            В CISStat параметры сохраняются автоматически для корректного обратного преобразования прогноза.
            
            Если вы используете данные для ML — разделите на train/test **ДО** масштабирования!
            """)
            
            # ────────────────────────────────────────────────
            # 🎮 ПЕСОЧНИЦА: МАСШТАБИРОВАНИЕ
            # ────────────────────────────────────────────────
            
            # Инициализация session_state
            if "show_scaling_preview" not in st.session_state:
                st.session_state.show_scaling_preview = False
            if "scaling_transform_params" not in st.session_state:
                st.session_state.scaling_transform_params = {}
            
            # ─ ЛЕВАЯ КОЛОНКА: ПАНЕЛЬ УПРАВЛЕНИЯ ───────────
            c1, c2, c3 = st.columns([27, 52, 21])
            
            with c1:
                st.markdown("###### Панель управления")
                
                # Выбор числовых колонок для масштабирования
                st.markdown("Признаки для масштабирования:")
                selected_cols = st.multiselect(
                    "Выберите колонки:",
                    options=num_cols,
                    default=[target_col_scale] if target_col_scale in num_cols else num_cols[:1],
                    key="scaling_selected_cols",
                    help="Можно выбрать несколько числовых колонок"
                )
                
                if not selected_cols:
                    st.warning("⚠️ Выберите хотя бы одну колонку")
                
                # Выбор метода
                scaling_method = st.radio(
                    "Метод масштабирования:",
                    ["Min-Max Scaling [0, 1]",
                    "Standardization (Z-score)",
                    "Robust Scaling (устойчивый к выбросам)",
                    "MaxAbs Scaling [-1, 1]",
                    "Normalization (L2 норма)",
                    "Quantile Transform (Uniform)",
                    "Quantile Transform (Normal)",
                    "Log1p Scaling"],
                    index=0,
                    key="scaling_method",
                    label_visibility="collapsed"
                )
                
                # Описание метода
                method_descriptions = {
                    "Min-Max Scaling [0, 1]": 
                        "Линейное преобразование в диапазон [0, 1]. Чувствителен к выбросам. Стандарт для нейросетей.",
                    "Standardization (Z-score)": 
                        "Центрирование (μ=0) и нормализация дисперсии (σ=1). Стандарт для PCA, линейной регрессии.",
                    "Robust Scaling (устойчивый к выбросам)": 
                        "Использует медиану и IQR вместо mean/std. Устойчив к выбросам. Для зашумлённых данных.",
                    "MaxAbs Scaling [-1, 1]": 
                        "Деление на max(|x|). Сохраняет знак, не центрирует. Для разреженных данных.",
                    "Normalization (L2 норма)": 
                        "Нормализация по евклидовой норме. Все векторы имеют длину 1. Для cosine similarity.",
                    "Quantile Transform (Uniform)": 
                        "Преобразование к равномерному распределению U[0,1]. Мощный метод для скошенных данных.",
                    "Quantile Transform (Normal)": 
                        "Преобразование к нормальному распределению N(0,1). Для методов, требующих нормальность.",
                    "Log1p Scaling": 
                        "log(1 + x). Сжимает правый хвост, безопасен для нулей. Для count data."
                }
                
                st.markdown(
                    f'<div style="background: #f0f9ff; border-left: 3px solid #0284c7; padding: 8px 12px; '
                    f'margin: 8px 0; border-radius: 4px;">'
                    f'<span style="color: #0369a1; font-size: 13px;">'
                    f' <strong>{scaling_method.split(" ")[0]}:</strong> {method_descriptions[scaling_method]}'
                    f'</span></div>',
                    unsafe_allow_html=True
                )
                
                # Параметры метода
                st.divider()
                st.markdown("**Параметры:**")
                
                if "Min-Max" in scaling_method:
                    feature_range_min = st.number_input("Минимум диапазона:", value=0.0, step=0.1, key="minmax_min")
                    feature_range_max = st.number_input("Максимум диапазона:", value=1.0, step=0.1, key="minmax_max")
                    param_value = (feature_range_min, feature_range_max)
                
                elif "Quantile" in scaling_method:
                    n_quantiles = st.slider(
                        "Число квантилей:",
                        min_value=10,
                        max_value=1000,
                        value=1000,
                        step=10,
                        key="quantile_n",
                        help="Больше = точнее, но медленнее. Обычно 1000."
                    )
                    param_value = n_quantiles
                
                elif "Log1p" in scaling_method:
                    param_value = "log1p"
                    st.info("ℹ️ Применяется log(1 + x) ко всем значениям")
                
                else:
                    param_value = None
                
                # Выбор колонок для обратного преобразования
                st.divider()
                st.markdown("**Обратное преобразование:**")
                inverse_transform = st.checkbox(
                    "Сохранить параметры для inverse_transform",
                    value=True,
                    key="scaling_inverse",
                    help="Позволит восстановить исходный масштаб прогноза"
                )
                
                st.divider()
                
                # Кнопки действий
                if st.button("▶ Применить масштабирование", type="primary", use_container_width=True, key="btn_apply_scaling"):
                    if not selected_cols:
                        st.error("❌ Выберите хотя бы одну колонку")
                    else:
                        st.session_state.show_scaling_preview = True
                        st.rerun()
                
                if st.button("↶ Сбросить", use_container_width=True, key="btn_reset_scaling"):
                    st.session_state.show_scaling_preview = False
                    st.rerun()
            
            # ─ ЦЕНТРАЛЬНАЯ КОЛОНКА: ВИЗУАЛИЗАЦИЯ ──────────
            with c2:
                st.markdown("###### Визуализация: До / После")
                
                if selected_cols:
                    target_col = selected_cols[0]  # Для визуализации берём первую колонку
                    original_series = df_scale_ts[target_col].dropna().astype(float)
                    
                    if st.session_state.show_scaling_preview:
                        try:
                            from sklearn.preprocessing import (
                                MinMaxScaler, StandardScaler, RobustScaler, 
                                MaxAbsScaler, Normalizer, QuantileTransformer
                            )
                            
                            # Применяем масштабирование
                            values = original_series.values.reshape(-1, 1)
                            
                            if "Min-Max" in scaling_method:
                                scaler = MinMaxScaler(feature_range=param_value)
                                scaled_values = scaler.fit_transform(values)
                                method_name = "Min-Max"
                                scaler_params = {
                                    'min': float(scaler.data_min_[0]),
                                    'max': float(scaler.data_max_[0]),
                                    'feature_range': param_value
                                }
                            
                            elif "Standardization" in scaling_method:
                                scaler = StandardScaler()
                                scaled_values = scaler.fit_transform(values)
                                method_name = "Standardization"
                                scaler_params = {
                                    'mean': float(scaler.mean_[0]),
                                    'std': float(scaler.scale_[0])
                                }
                            
                            elif "Robust" in scaling_method:
                                scaler = RobustScaler()
                                scaled_values = scaler.fit_transform(values)
                                method_name = "Robust"
                                scaler_params = {
                                    'median': float(scaler.center_[0]),
                                    'iqr': float(scaler.scale_[0])
                                }
                            
                            elif "MaxAbs" in scaling_method:
                                scaler = MaxAbsScaler()
                                scaled_values = scaler.fit_transform(values)
                                method_name = "MaxAbs"
                                scaler_params = {
                                    'max_abs': float(scaler.max_abs_[0])
                                }
                            
                            elif "Normalization" in scaling_method:
                                scaler = Normalizer(norm='l2')
                                scaled_values = scaler.fit_transform(values)
                                method_name = "L2-Normalization"
                                scaler_params = {'norm': 'l2'}
                            
                            elif "Quantile" in scaling_method and "Uniform" in scaling_method:
                                scaler = QuantileTransformer(n_quantiles=param_value, output_distribution='uniform')
                                scaled_values = scaler.fit_transform(values)
                                method_name = "Quantile-Uniform"
                                scaler_params = {'n_quantiles': param_value, 'distribution': 'uniform'}
                            
                            elif "Quantile" in scaling_method and "Normal" in scaling_method:
                                scaler = QuantileTransformer(n_quantiles=param_value, output_distribution='normal')
                                scaled_values = scaler.fit_transform(values)
                                method_name = "Quantile-Normal"
                                scaler_params = {'n_quantiles': param_value, 'distribution': 'normal'}
                            
                            elif "Log1p" in scaling_method:
                                scaled_values = np.log1p(values)
                                method_name = "Log1p"
                                scaler_params = {'method': 'log1p'}
                            
                            scaled_series = pd.Series(scaled_values.flatten(), index=original_series.index)
                            
                            # Сохраняем scaler для обратного преобразования
                            if inverse_transform:
                                st.session_state.scaling_scaler = scaler if 'scaler' in locals() else None
                            
                            # ── ГРАФИК СРАВНЕНИЯ (2 ряда) ──────
                            fig = make_subplots(
                                rows=2, cols=1,
                                subplot_titles=(
                                    f" Исходный ряд: {target_col}",
                                    f" После масштабирования ({method_name})"
                                ),
                                vertical_spacing=0.12
                            )
                            
                            # Исходный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=original_series.index, y=original_series.values,
                                    mode='lines',
                                    name='Исходные данные',
                                    line=dict(color='#048A81', width=2),
                                    showlegend=False
                                ),
                                row=1, col=1
                            )
                            
                            # Масштабированный ряд
                            fig.add_trace(
                                go.Scatter(
                                    x=scaled_series.index, y=scaled_series.values,
                                    mode='lines',
                                    name='После масштабирования',
                                    line=dict(color='#DC2626', width=2),
                                    showlegend=False
                                ),
                                row=2, col=1
                            )
                            
                            fig.update_layout(
                                height=600,
                                margin=dict(l=50, r=20, t=80, b=40),
                                hovermode='x unified',
                            )
                            
                            fig.update_xaxes(title_text="Дата", row=2, col=1)
                            fig.update_yaxes(title_text="Значение", row=1, col=1)
                            fig.update_yaxes(title_text="Масштабированное значение", row=2, col=1)
                            
                            fig.update_annotations(font=dict(size=13, color="#6b7280"), yshift=10)
                            
                            st.plotly_chart(fig, use_container_width=True, key="scaling_main_chart")
                            
                            # ─ ГИСТОГРАММЫ СРАВНЕНИЯ ───────────
                            st.markdown("######  Распределение значений")
                            
                            fig_hist = make_subplots(
                                rows=1, cols=2,
                                subplot_titles=("Исходное распределение", "После масштабирования")
                            )
                            
                            fig_hist.add_trace(
                                go.Histogram(x=original_series.values, nbinsx=40,
                                            marker_color='#048A81', name='До', showlegend=False),
                                row=1, col=1
                            )
                            
                            fig_hist.add_trace(
                                go.Histogram(x=scaled_series.values, nbinsx=40,
                                            marker_color='#DC2626', name='После', showlegend=False),
                                row=1, col=2
                            )
                            
                            fig_hist.update_layout(height=300, margin=dict(l=40, r=20, t=50, b=40), barmode='overlay')
                            st.plotly_chart(fig_hist, use_container_width=True, key="scaling_hist_chart")
                            
                            # ─ BOXPLOT СРАВНЕНИЕ ───────────────
                            st.markdown("###### Boxplot: До / После")
                            
                            fig_box = go.Figure()
                            
                            fig_box.add_trace(go.Box(
                                y=original_series.values,
                                name='До',
                                marker_color='#048A81',
                                boxpoints='outliers'
                            ))
                            
                            fig_box.add_trace(go.Box(
                                y=scaled_series.values,
                                name='После',
                                marker_color='#DC2626',
                                boxpoints='outliers'
                            ))
                            
                            fig_box.update_layout(
                                height=300,
                                title="Сравнение распределений (Boxplot)",
                                yaxis_title="Значение",
                                margin=dict(l=40, r=20, t=50, b=40)
                            )
                            
                            st.plotly_chart(fig_box, use_container_width=True, key="scaling_box_chart")
                            
                            # ── РАСЧЁТ МЕТРИК КАЧЕСТВА ──────────
                            metrics = calculate_scaling_metrics(original_series, scaled_series)
                            
                            # Сохранение результатов
                            st.session_state.scaling_transform_result = {
                                'method': method_name,
                                'param': param_value,
                                'original_series': original_series,
                                'scaled_series': scaled_series,
                                'target_col': target_col,
                                'metrics': metrics,
                                'scaler_params': scaler_params
                            }
                            
                            st.divider()
                            
                            # Оценка качества масштабирования
                            if "Min-Max" in method_name:
                                expected_range = param_value if param_value else (0, 1)
                                actual_min, actual_max = scaled_series.min(), scaled_series.max()
                                if abs(actual_min - expected_range[0]) < 0.01 and abs(actual_max - expected_range[1]) < 0.01:
                                    st.success(f"✅ **Масштабирование выполнено корректно!** Диапазон: [{actual_min:.3f}, {actual_max:.3f}]")
                                else:
                                    st.warning(f"⚠️ **Диапазон не соответствует ожидаемому.** Ожидалось: {expected_range}, получено: [{actual_min:.3f}, {actual_max:.3f}]")
                            
                            elif "Standardization" in method_name:
                                if abs(scaled_series.mean()) < 0.01 and abs(scaled_series.std() - 1.0) < 0.01:
                                    st.success(f"✅ **Стандартизация выполнена корректно!** μ={scaled_series.mean():.4f}, σ={scaled_series.std():.4f}")
                                else:
                                    st.warning(f"⚠️ **Параметры не соответствуют ожидаемым.** μ={scaled_series.mean():.4f}, σ={scaled_series.std():.4f}")
                            
                            else:
                                st.success(f"✅ **Масштабирование {method_name} выполнено!**")
                            
                            # Кнопки подтверждения
                            c_ok_scale, c_cancel_scale = st.columns(2)
                            with c_ok_scale:
                                if st.button("✅ Применить к данным", type="primary", use_container_width=True, key="btn_confirm_scaling"):
                                    # Применяем масштабирование к основному df
                                    df_final_scale = st.session_state.df.copy()
                                    
                                    # Сохраняем оригинальные колонки
                                    orig_col_names = []
                                    for col in selected_cols:
                                        orig_col_name = f"{col}_original"
                                        if orig_col_name not in df_final_scale.columns:
                                            df_final_scale[orig_col_name] = df_final_scale[col]
                                        orig_col_names.append(orig_col_name)
                                    
                                    # Применяем масштабирование ко всем выбранным колонкам
                                    scalers_dict = {}
                                    for col in selected_cols:
                                        col_values = df_final_scale[col].astype(float).values.reshape(-1, 1)
                                        
                                        if "Min-Max" in scaling_method:
                                            scaler = MinMaxScaler(feature_range=param_value)
                                            df_final_scale[col] = scaler.fit_transform(col_values).flatten()
                                            scalers_dict[col] = {'type': 'MinMax', 'params': scaler.get_params()}
                                        
                                        elif "Standardization" in scaling_method:
                                            scaler = StandardScaler()
                                            df_final_scale[col] = scaler.fit_transform(col_values).flatten()
                                            scalers_dict[col] = {'type': 'Standard', 'params': scaler.get_params()}
                                        
                                        elif "Robust" in scaling_method:
                                            scaler = RobustScaler()
                                            df_final_scale[col] = scaler.fit_transform(col_values).flatten()
                                            scalers_dict[col] = {'type': 'Robust', 'params': scaler.get_params()}
                                        
                                        elif "MaxAbs" in scaling_method:
                                            scaler = MaxAbsScaler()
                                            df_final_scale[col] = scaler.fit_transform(col_values).flatten()
                                            scalers_dict[col] = {'type': 'MaxAbs', 'params': scaler.get_params()}
                                        
                                        elif "Normalization" in scaling_method:
                                            scaler = Normalizer(norm='l2')
                                            df_final_scale[col] = scaler.fit_transform(col_values).flatten()
                                            scalers_dict[col] = {'type': 'Normalizer', 'params': {'norm': 'l2'}}
                                        
                                        elif "Quantile" in scaling_method:
                                            dist = 'uniform' if 'Uniform' in scaling_method else 'normal'
                                            scaler = QuantileTransformer(n_quantiles=param_value, output_distribution=dist)
                                            df_final_scale[col] = scaler.fit_transform(col_values).flatten()
                                            scalers_dict[col] = {'type': 'Quantile', 'params': scaler.get_params()}
                                        
                                        elif "Log1p" in scaling_method:
                                            df_final_scale[col] = np.log1p(df_final_scale[col].astype(float))
                                            scalers_dict[col] = {'type': 'Log1p', 'params': {}}
                                    
                                    # Сохраняем параметры
                                    st.session_state.scaling_transform_params = {
                                        'method': method_name,
                                        'param': param_value,
                                        'columns': selected_cols,
                                        'original_col_names': orig_col_names,
                                        'scalers': scalers_dict,
                                        'metrics': metrics
                                    }
                                    
                                    # Синхронизация
                                    st.session_state.df = df_final_scale.copy()
                                    st.session_state.validation_ready = False
                                    st.session_state.show_scaling_preview = False
                                    
                                    # Удаляем рабочие копии
                                    work_dfs = [
                                        "df_missing_work", "df_pattern_work", "df_range_work",
                                        "df_outlier_work", "df_inclusion_work", "df_referential_work",
                                        "df_text_work", "df_consistency_work", "df_uniqueness_work",
                                        "df_regularity_work", "df_regular_work", "df_variance_work",
                                        "df_smooth_work", "df_stationarity_work", "df_scaling_work"
                                    ]
                                    for work_df_name in work_dfs:
                                        if work_df_name in st.session_state:
                                            del st.session_state[work_df_name]
                                    
                                    if "val_results" in st.session_state:
                                        del st.session_state.val_results
                                    
                                    st.success(f"✅ Масштабирование **{method_name}** применено к {len(selected_cols)} колонкам!")
                                    st.info(f"💡 Оригинальные значения сохранены в колонках `{orig_col_names}`. "
                                        f"Перезапустите валидацию для обновления статистик.")
                                    st.rerun()
                            
                            with c_cancel_scale:
                                if st.button("❌ Отмена", use_container_width=True, key="btn_cancel_scaling"):
                                    st.session_state.show_scaling_preview = False
                                    st.rerun()
                        
                        except Exception as e:
                            st.error(f"❌ Ошибка масштабирования: {e}")
                            import traceback
                            with st.expander("🔍 Stack trace"):
                                st.code(traceback.format_exc(), language="python")
                    
                    else:
                        # Показываем только исходный ряд
                        fig = px.line(
                            x=original_series.index,
                            y=original_series.values,
                            labels={'x': 'Дата', 'y': target_col},
                        )
                        fig.update_layout(height=400, margin=dict(l=40, r=20, t=40, b=40))
                        st.plotly_chart(fig, use_container_width=True, key="scaling_main_chart")
                        
                        st.info("💡 Выберите метод масштабирования и нажмите **'Применить масштабирование'** для просмотра результата.")
                else:
                    st.warning("⚠️ Выберите хотя бы одну колонку для масштабирования")
            
            # ── ПРАВАЯ КОЛОНКА: МЕТРИКИ КАЧЕСТВА ───────────
            with c3:
                st.markdown("######  Метрики качества")
                
                # ИНИЦИАЛИЗАЦИЯ: metrics_orig всегда определён
                metrics_orig = None
                
                # 📊 ВЫЧИСЛЯЕМ БАЗОВЫЕ МЕТРИКИ ДО МАСШТАБИРОВАНИЯ (всегда доступны)
                if selected_cols and target_col in df_scale_ts.columns:
                    series_for_metrics = df_scale_ts[target_col].dropna().astype(float)
                    
                    if len(series_for_metrics) > 0:
                        # Расчёт метрик "До"
                        metrics_orig = {
                            'range_orig': series_for_metrics.max() - series_for_metrics.min(),
                            'mean_orig': series_for_metrics.mean(),
                            'std_orig': series_for_metrics.std(),
                            'outliers_orig': int(((series_for_metrics < series_for_metrics.mean() - 3*series_for_metrics.std()) | 
                                                (series_for_metrics > series_for_metrics.mean() + 3*series_for_metrics.std())).sum()),
                            'skew_orig': series_for_metrics.skew(),
                            'kurt_orig': series_for_metrics.kurtosis()
                        }
                        
                        with st.container(border=True):
                            st.markdown("**До масштабирования:**")
                            st.metric("Диапазон", f"{metrics_orig['range_orig']:,.2f}".replace(",", " "))
                            st.metric("Среднее", f"{metrics_orig['mean_orig']:.2f}")
                            st.metric("Стд. отклонение", f"{metrics_orig['std_orig']:.2f}")
                            st.metric("Выбросы (3σ)", f"{metrics_orig['outliers_orig']}")
                            st.metric("Асимметрия", f"{metrics_orig['skew_orig']:.3f}")
                    else:
                        st.info("ℹ️ Нет данных для расчёта метрик")
                else:
                    st.info("ℹ️ Выберите колонку для отображения метрик")
                
                # 🔧 Метрики ПОСЛЕ масштабирования (только если превью активно И metrics_orig определён)
                if (st.session_state.show_scaling_preview and 
                    'scaling_transform_result' in st.session_state and 
                    metrics_orig is not None):
                    
                    metrics = st.session_state.scaling_transform_result['metrics']
                    scaled_series = st.session_state.scaling_transform_result['scaled_series']
                    
                    with st.container(border=True):
                        st.markdown("**После масштабирования:**")
                        st.metric("Диапазон", f"{metrics['range_scaled']:,.2f}".replace(",", " "),
                                delta=f"{metrics['range_scaled'] - metrics_orig['range_orig']:+,.2f}".replace(",", " "))
                        st.metric("Среднее", f"{metrics['mean_scaled']:.4f}",
                                delta=f"{metrics['mean_scaled'] - metrics_orig['mean_orig']:+.4f}")
                        st.metric("Стд. отклонение", f"{metrics['std_scaled']:.4f}",
                                delta=f"{metrics['std_scaled'] - metrics_orig['std_orig']:+.4f}")
                        st.metric("Выбросы (3σ)", f"{metrics['outliers_scaled']}",
                                delta=f"{metrics['outliers_scaled'] - metrics_orig['outliers_orig']:+d}")
                        st.metric("Асимметрия", f"{metrics['skew_scaled']:.3f}",
                                delta=f"{metrics['skew_scaled'] - metrics_orig['skew_orig']:+.3f}")
                    
                    st.divider()
                    
                    # Рекомендации
                    method_name = st.session_state.scaling_transform_result['method']
                    if "Min-Max" in method_name:
                        if metrics['range_scaled'] <= 1.01:
                            st.success("✅ **Диапазон [0, 1] достигнут.** Готово для нейросетей.")
                        else:
                            st.warning("⚠️ **Диапазон не [0, 1].** Проверьте параметры.")
                    
                    elif "Standardization" in method_name:
                        if abs(metrics['mean_scaled']) < 0.01 and abs(metrics['std_scaled'] - 1.0) < 0.01:
                            st.success("✅ **μ=0, σ=1 достигнуты.** Готово для PCA/ARIMA.")
                        else:
                            st.warning("⚠️ **Параметры не μ=0, σ=1.** Проверьте данные.")
                    
                    elif "Robust" in method_name:
                        if metrics['outliers_scaled'] < metrics['outliers_orig']:
                            st.success("✅ **Выбросы уменьшены.** Robust Scaling работает корректно.")
                        else:
                            st.info("ℹ️ **Выбросы не изменились.** Это нормально для Robust Scaling.")
                    
                    else:
                        st.success(f"✅ **Масштабирование {method_name} выполнено.**")
                    
                    # Информация о параметрах
                    st.info(f" **Метод:** {method_name}")

            # ── ИНФОРМАЦИЯ О СОХРАНЁННЫХ ПАРАМЕТРАХ ────────
            if st.session_state.scaling_transform_params:
                st.divider()
                with st.expander("💾 Сохранённые параметры масштабирования", expanded=False):
                    st.json(st.session_state.scaling_transform_params)
                    st.caption("Эти параметры будут использованы для **обратного преобразования** прогноза "
                            "в исходную шкалу (через inverse_transform).")



# ═══════════════════════════════════════════════════════════
#  ВКЛАДКА 4: IH-АНАЛИЗ (Information-Entropy Analysis)
# ═══════════════════════════════════════════════════════════
with tab_exploratory:
    st.markdown("""
    <div style="padding-left: 20px; margin: 20px 0; text-align: right;">
        <p style="margin: 0 0 10px 0; color: #1e293b; line-height: 1.6; font-size: 18px; font-weight: 400;">
            "Результат любого серьёзного исследования — появление двух новых вопросов там, где был всего лишь один".
        </p>
        <p style="margin: 0; color: #64748B; font-style: italic; font-size: 16px; line-height: 1.5;">
            — Торстейн Веблен, американский экономист, социолог, публицист и футуролог,<br> один из основоположников институционализма в экономической теории
        </p>
    </div>
    """, unsafe_allow_html=True)


    st.markdown("###  IH-анализ: Информационно-энтропийное исследование")
    st.caption("Оценка информативности признаков через теорию информации Шеннона. Метрика R(Y|X) = I(X;Y) / H(Y) ∈ [0;1]")

    with st.expander("Справка по методу", expanded=False):
        st.markdown("""
        **Что измеряет IH-анализ:**
        - **Энтропия H(X)** — мера неопределённости признака (бит)
        - **Взаимная информация I(X;Y)** — сколько знаний о Y даёт знание X
        - **Нормированная связь R(Y|X)** — доля неопределённости цели, объясняемая признаком

        **Параметры:**
        - `Sharpness` (0.1–1.0) — «резкость» дискретизации непрерывных признаков
        - `Target` — целевая переменная для оценки предсказательной силы
        - `Top-K` — количество признаков для детального анализа

        **Преимущества:**
        - Работает с числовыми, категориальными данными и пропусками
        - Выявляет нелинейные и немонотонные зависимости
        - Оценивает синергию комбинаций признаков
        - Интерпретируемая метрика: «признак объясняет R×100% неопределённости цели»

        **⚠️ Ограничения для временных рядов:**
        - Автокорреляция может завышать оценки → используйте лаговые признаки
        - Требует достаточного объёма данных (>200 наблюдений для надёжности)

        **⚠️ Почитать о методе:** https://habr.com/ru/articles/1040980/
        """)

    # ───────────────────────────────────────────────────────────
    # НАСТРОЙКИ IH-АНАЛИЗА
    # ───────────────────────────────────────────────────────────
    if not df_filtered.empty and ct_f.get("num"):
        # Выбор целевой переменной
        target_options = ct_f["num"] + ([st.session_state.primary_date_col] if st.session_state.primary_date_col else [])
        ih_target = st.selectbox(
            "ЦЕЛЕВАЯ ПЕРЕМЕННАЯ (Y)",
            options=target_options,
            index=0 if target_options else None,
            key="ih_target_select"
        )

        # Выбор признаков для анализа
        feature_options = [c for c in ct_f["num"] + ct_f.get("cat", []) if c != ih_target]
        ih_features = st.multiselect(
            "ПРИЗНАКИ ДЛЯ АНАЛИЗА (X)",
            options=feature_options,
            default=feature_options[:5] if len(feature_options) >= 5 else feature_options,
            key="ih_features_select",
            help="Оставьте пустым для анализа всех доступных признаков"
        )
        if not ih_features:
            ih_features = feature_options

        # Параметры дискретизации
        c1, c2, c3 = st.columns(3)
        with c1:
            sharpness = st.slider(
                " Sharpness (резкость дискретизации)",
                min_value=0.1, max_value=1.0, value=0.25, step=0.05,
                key="ih_sharpness",
                help="Меньшее значение → больше интервалов (тоньше дискретизация)"
            )
        with c2:
            top_k = st.number_input(
                " Показать топ признаков",
                min_value=3, max_value=20, value=10, step=1,
                key="ih_top_k"
            )
        with c3:
            min_samples = st.number_input(
                " Мин. наблюдений на интервал",
                min_value=5, max_value=100, value=20, step=5,
                key="ih_min_samples",
                help="Контроль качества дискретизации"
            )

        # Кнопка запуска анализа
        if st.button("Запустить IH-анализ", type="primary", use_container_width=True, key="btn_run_ih"):
            with st.spinner("🔢 Вычисление энтропийных метрик..."):
                progress_bar = st.progress(0)

                # ───────────────────────────────────────────────
                # 🔬 ЯДРО IH-АНАЛИЗА: ФУНКЦИИ
                # ───────────────────────────────────────────────
                def discretize_feature(series: pd.Series, sharpness: float, min_samples: int) -> pd.Series:
                    """
                    Адаптивная дискретизация с параметром sharpness.
                    Меньший sharpness → больше интервалов.
                    """
                    if pd.api.types.is_categorical_dtype(series) or series.nunique() <= 10:
                        return series.astype(str)

                    # Удаление пропусков для дискретизации
                    clean = series.dropna()
                    if len(clean) < min_samples * 2:
                        return series.astype(str)  # fallback

                    # Оценка оптимального числа бинов через sharpness
                    n_bins = max(2, min(50, int(1 / sharpness)))

                    # Квантильная дискретизация (устойчива к выбросам)
                    try:
                        bins = pd.qcut(clean, q=n_bins, duplicates='drop', labels=False)
                        # Восстановление индекса с сохранением пропусков
                        result = pd.Series(index=series.index, dtype=object)
                        result[clean.index] = bins.astype(str)
                        result[series.isna()] = '_MISSING_'
                        return result
                    except Exception:
                        # Fallback на равномерную дискретизацию
                        bins = pd.cut(clean, bins=n_bins, labels=False)
                        result = pd.Series(index=series.index, dtype=object)
                        result[clean.index] = bins.astype(str)
                        result[series.isna()] = '_MISSING_'
                        return result

                def shannon_entropy(probabilities: np.ndarray, base: float = 2) -> float:
                    """Вычисление энтропии Шеннона."""
                    probabilities = probabilities[probabilities > 0]  # убрать нули
                    return -np.sum(probabilities * np.log(probabilities) / np.log(base))

                def mutual_information(x_disc: pd.Series, y_disc: pd.Series, base: float = 2) -> float:
                    """
                    Оценка взаимной информации через совместное распределение.
                    """
                    # Совместная таблица частот
                    joint = pd.crosstab(x_disc, y_disc)
                    joint_prob = joint.values / joint.values.sum()

                    # Маргинальные распределения
                    px = joint_prob.sum(axis=1)
                    py = joint_prob.sum(axis=0)

                    # MI = Σ p(x,y) * log(p(x,y) / (p(x)*p(y)))
                    mi = 0.0
                    for i in range(joint_prob.shape[0]):
                        for j in range(joint_prob.shape[1]):
                            if joint_prob[i, j] > 0 and px[i] > 0 and py[j] > 0:
                                mi += joint_prob[i, j] * np.log(joint_prob[i, j] / (px[i] * py[j]))
                    return mi / np.log(base)  # конвертация в нужное основание

                def compute_r_metric(x: pd.Series, y: pd.Series, sharpness: float, min_samples: int) -> dict:
                    """
                    Вычисление нормированной меры связи R(Y|X) = I(X;Y) / H(Y).
                    Возвращает словарь с метриками.
                    """
                    # Дискретизация
                    x_disc = discretize_feature(x, sharpness, min_samples)
                    y_disc = discretize_feature(y, sharpness, min_samples)

                    # Проверка константных признаков
                    if x_disc.nunique() <= 1:
                        return {"R": 0.0, "MI": 0.0, "H_X": 0.0, "H_Y": 0.0, 
                                "n_bins_X": 1, "n_bins_Y": y_disc.nunique(),
                                "error": "Признак X константен"}
                    
                    # Энтропии
                    _, counts_y = np.unique(y_disc, return_counts=True)
                    py = counts_y / counts_y.sum()
                    h_y = shannon_entropy(py)

                    if h_y < 1e-10:  # целевая переменная константа
                        return {"R": 0.0, "MI": 0.0, "H_X": 0.0, "H_Y": 0.0, "error": "H(Y) ≈ 0"}

                    # Взаимная информация
                    mi = mutual_information(x_disc, y_disc)

                    # Нормированная метрика
                    r_value = min(1.0, mi / h_y)  # защита от численных ошибок

                    # Энтропия признака
                    _, counts_x = np.unique(x_disc, return_counts=True)
                    px = counts_x / counts_x.sum()
                    h_x = shannon_entropy(px)

                    return {
                        "R": r_value,
                        "MI": mi,
                        "H_X": h_x,
                        "H_Y": h_y,
                        "n_bins_X": x_disc.nunique(),
                        "n_bins_Y": y_disc.nunique()
                    }

                # ───────────────────────────────────────────────
                # 📊 РАСЧЁТЫ
                # ───────────────────────────────────────────────
                results = []
                y_series = df_filtered[ih_target].copy()

                for i, feat in enumerate(ih_features):
                    progress_bar.progress(min(0.3 + 0.6 * (i + 1) / len(ih_features), 0.95))
                    x_series = df_filtered[feat].copy()

                    try:
                        metrics = compute_r_metric(x_series, y_series, sharpness, min_samples)
                        results.append({
                            "feature": feat,
                            "R": metrics["R"],
                            "MI": metrics["MI"],
                            "H_X": metrics["H_X"],
                            "H_Y": metrics["H_Y"],
                            "n_bins": metrics["n_bins_X"],
                            "dtype": str(df_filtered[feat].dtype)
                        })
                    except Exception as e:
                        results.append({
                            "feature": feat,
                            "R": 0.0,
                            "MI": 0.0,
                            "H_X": 0.0,
                            "H_Y": 0.0,
                            "n_bins": 0,
                            "dtype": str(df_filtered[feat].dtype),
                            "error": str(e)
                        })

                df_ih = pd.DataFrame(results).sort_values("R", ascending=False).reset_index(drop=True)
                progress_bar.progress(1.0)

                # ───────────────────────────────────────────────
                # 📈 ВИЗУАЛИЗАЦИЯ 1: Рейтинг признаков по R
                # ───────────────────────────────────────────────
                st.markdown("#### Рейтинг признаков по информативности (R-метрика)")

                top_df = df_ih.head(top_k).copy()
                top_df["R_pct"] = (top_df["R"] * 100).round(1)

                fig_rank = px.bar(
                    top_df,
                    x="R",
                    y="feature",
                    orientation="h",
                    color="R",
                    color_continuous_scale="Viridis",
                    labels={"R": "R(Y|X) — доля объяснённой неопределённости", "feature": "Признак"},
                    title=f"Топ-{top_k} признаков по предсказательной силе для '{ih_target}'",
                    height=min(40 + top_k * 35, 500)
                )
                fig_rank.update_layout(
                    coloraxis_showscale=False,
                    xaxis_title="R(Y|X) ∈ [0; 1]",
                    yaxis_title="",
                    hovermode="y unified"
                )
                fig_rank.update_traces(
                    texttemplate="%{x:.3f} (%{customdata}%)",
                    textposition="outside",
                    customdata=top_df["R_pct"],
                    hovertemplate="<b>%{y}</b><br>R = %{x:.3f}<br>Объясняет: %{customdata}% неопределённости<extra></extra>"
                )
                st.plotly_chart(fig_rank, use_container_width=True)

                # Интерпретация
                st.markdown("""
                <div style='background: #f0f9ff; border-left: 4px solid #0284c7; padding: 12px; border-radius: 0 6px 6px 0; margin: 10px 0;'>
                <strong>📖 Как читать:</strong> Значение R = 0.35 означает, что знание признака уменьшает
                неопределённость целевой переменной на <strong>35%</strong>. Чем ближе к 1 — тем сильнее связь.
                </div>
                """, unsafe_allow_html=True)

                # ───────────────────────────────────────────────
                # 📈 ВИЗУАЛИЗАЦИЯ 2: Тепловая карта энтропий
                # ───────────────────────────────────────────────
                st.markdown("######  Тепловая карта: Энтропия признаков и взаимная информация")

                heatmap_data = top_df[["feature", "H_X", "MI", "R"]].copy()
                heatmap_data = heatmap_data.set_index("feature")

                fig_heat = px.imshow(
                    heatmap_data.T,
                    labels=dict(x="Признак", y="Метрика", color="Значение"),
                    x=heatmap_data.index,
                    y=["H(X) — энтропия", "I(X;Y) — MI", "R(Y|X) — норм. связь"],
                    color_continuous_scale="RdYlGn",
                    aspect="auto",
                    title="Сравнение энтропийных метрик для топ-признаков"
                )
                fig_heat.update_layout(height=300)
                fig_heat.update_traces(
                    hovertemplate="<b>%{y}</b> для %{x}: <b>%{z:.3f}</b><extra></extra>",
                    text=heatmap_data.T.round(3),
                    texttemplate="%{text}",
                    textfont={"size": 10}
                )
                st.plotly_chart(fig_heat, use_container_width=True)

                # ───────────────────────────────────────────────
                # 📈 ВИЗУАЛИЗАЦИЯ 3: Синергия пар признаков
                # ───────────────────────────────────────────────
                st.markdown("######  Анализ синергии пар признаков")
                synergy_results = []

                if len(top_df) >= 2:
                    synergy_results = []
                    selected_features = top_df["feature"].head(min(6, len(top_df))).tolist()

                    for i in range(len(selected_features)):
                        for j in range(i + 1, len(selected_features)):
                            f1, f2 = selected_features[i], selected_features[j]
                            try:
                                # Индивидуальные R
                                r1 = df_ih[df_ih["feature"] == f1]["R"].values[0]
                                r2 = df_ih[df_ih["feature"] == f2]["R"].values[0]

                                # Совместный анализ: создаём комбинированный признак
                                x1_disc = discretize_feature(df_filtered[f1], sharpness, min_samples)
                                x2_disc = discretize_feature(df_filtered[f2], sharpness, min_samples)
                                x_combined = x1_disc.astype(str) + "||" + x2_disc.astype(str)

                                r_combined = compute_r_metric(x_combined, y_series, sharpness, min_samples)["R"]

                                # Синергия = R(комбо) - (R1 + R2)
                                synergy = r_combined - (r1 + r2)

                                synergy_results.append({
                                    "pair": f"{f1} + {f2}",
                                    "R1": r1,
                                    "R2": r2,
                                    "R_combined": r_combined,
                                    "synergy": synergy,
                                    "synergy_pct": synergy * 100
                                })
                            except:
                                st.warning(f"⚠️ Не удалось рассчитать синергию для {f1} + {f2}: {e}")
                                continue

                    if synergy_results:
                        df_syn = pd.DataFrame(synergy_results).sort_values("synergy", ascending=False)

                        # Визуализация синергии
                        fig_syn = px.bar(
                            df_syn.head(10),
                            x="pair",
                            y="synergy",
                            color="synergy",
                            color_continuous_scale="RdYlGn_r" if df_syn["synergy"].max() < 0 else "RdYlGn",
                            labels={"synergy": "Синергия ΔR", "pair": "Пара признаков"},
                            title="Синергия признаков: ΔR = R(комбо) − [R₁ + R₂]",
                            height=350
                        )
                        fig_syn.update_layout(
                            xaxis_tickangle=-45,
                            coloraxis_showscale=False,
                            hovermode="x unified"
                        )
                        fig_syn.update_traces(
                            texttemplate="%{y:+.3f}",
                            textposition="outside",
                            hovertemplate="<b>%{x}</b><br>Синергия: %{y:+.3f}<br>R комбо: %{customdata:.3f}<extra></extra>",
                            customdata=df_syn["R_combined"]
                        )
                        st.plotly_chart(fig_syn, use_container_width=True)

                        # Интерпретация синергии
                        st.markdown("""
                        <div style='background: #fefce8; border-left: 4px solid #ca8a04; padding: 12px; border-radius: 0 6px 6px 0; margin: 10px 0;'>
                        <strong>📖 Синергия:</strong> Положительное ΔR означает, что признаки вместе дают
                        <em>больше информации</em>, чем по отдельности (взаимодействие).
                        Отрицательное ΔR — признаки дублируют информацию (избыточность).
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.info("ℹ️ Недостаточно данных для анализа синергии")
                else:
                    st.info("ℹ️ Для анализа синергии выберите минимум 2 признака")

                # ───────────────────────────────────────────────
                # 📈 ВИЗУАЛИЗАЦИЯ 4: Детали по выбранному признаку
                # ───────────────────────────────────────────────
                st.markdown("######  Детальный разбор признака")

                if not df_ih.empty:
                    selected_feat = st.selectbox(
                        "Выберите признак для детального анализа:",
                        options=df_ih["feature"].tolist(),
                        index=0,
                        key="ih_detail_select"
                    )

                    feat_row = df_ih[df_ih["feature"] == selected_feat].iloc[0]

                    # Метрики в карточках
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("📊 R(Y|X)", f"{feat_row['R']:.3f}", f"{feat_row['R']*100:.1f}% объясняет")
                    c2.metric("💡 I(X;Y) MI", f"{feat_row['MI']:.3f} бит", "взаимная информация")
                    c3.metric("🎲 H(X)", f"{feat_row['H_X']:.3f} бит", "энтропия признака")
                    c4.metric("📦 Бины", feat_row["n_bins"], f"дискретизация")

                    # График распределения целевой переменной по бинам признака
                    x_disc = discretize_feature(df_filtered[selected_feat], sharpness, min_samples)
                    y_disc = discretize_feature(df_filtered[ih_target], sharpness, min_samples)

                    # Групповая статистика
                    cross_tab = pd.crosstab(x_disc, y_disc, normalize='index') * 100

                    if not cross_tab.empty:
                        fig_dist = px.imshow(
                            cross_tab.T,
                            labels=dict(x=f"Бины '{selected_feat}'", y=f"Бины '{ih_target}'", color="% в ячейке"),
                            color_continuous_scale="Blues",
                            aspect="auto",
                            title=f"Распределение '{ih_target}' по интервалам '{selected_feat}'"
                        )
                        fig_dist.update_layout(height=350)
                        fig_dist.update_traces(
                            hovertemplate="%{y} при %{x}: <b>%{z:.1f}%</b><extra></extra>",
                            text=cross_tab.T.round(1),
                            texttemplate="%{text:.0f}%",
                            textfont={"size": 9}
                        )
                        st.plotly_chart(fig_dist, use_container_width=True)

                # ───────────────────────────────────────────────
                # 📋 ТАБЛИЦА РЕЗУЛЬТАТОВ + ЭКСПОРТ
                # ───────────────────────────────────────────────
                st.divider()
                st.markdown("######  Полная таблица результатов")

                display_cols = ["feature", "R", "MI", "H_X", "n_bins", "dtype"]
                if "error" in df_ih.columns:
                    display_cols.append("error")

                st.dataframe(
                    df_ih[display_cols].style.format({
                        "R": "{:.3f}",
                        "MI": "{:.3f}",
                        "H_X": "{:.3f}",
                        "H_Y": "{:.3f}"
                    }).background_gradient(subset=["R"], cmap="viridis"),
                    use_container_width=True,
                    height=300
                )

                # Экспорт
                csv_ih = df_ih.to_csv(index=False, encoding="utf-8-sig")
                st.download_button(
                    label="📥 Скачать результаты IH-анализа (CSV)",
                    data=csv_ih,
                    file_name=f"IH_analysis_{ih_target}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",
                    key="btn_download_ih"
                )

                # Сохранение в session_state для дальнейшего использования
                st.session_state.ih_results = df_ih
                st.session_state.ih_target = ih_target
                st.success("✅ IH-анализ завершён! Результаты сохранены.")

                # ───────────────────────────────────────────────
                # 💡 РЕКОМЕНДАЦИИ НА ОСНОВЕ РЕЗУЛЬТАТОВ
                # ───────────────────────────────────────────────
                st.markdown("###### 💡 Автоматические рекомендации")

                recommendations = []

                # Высокая информативность
                high_r = df_ih[df_ih["R"] >= 0.5]
                if not high_r.empty:
                    rec_list = ", ".join([f"`{r}`" for r in high_r["feature"].head(3)])
                    recommendations.append(f"✅ **Сильные предикторы**: {rec_list} (R ≥ 0.5) → используйте как основные признаки в моделях")

                # Низкая информативность
                low_r = df_ih[df_ih["R"] < 0.1]
                if not low_r.empty and len(low_r) > len(df_ih) * 0.3:
                    recommendations.append("⚠️ **Много слабых признаков**: рассмотрите отбор признаков или агрегацию")

                # Высокая энтропия + низкий R (шум?)
                noisy = df_ih[(df_ih["H_X"] > df_ih["H_X"].quantile(0.75)) & (df_ih["R"] < 0.15)]
                if not noisy.empty:
                    rec_list = ", ".join([f"`{r}`" for r in noisy["feature"].head(3)])
                    recommendations.append(f"⚡ **Высокая энтропия, низкая связь**: {rec_list} → возможен шум, проверьте качество данных")

                # Синергия (если есть)
                if "synergy_results" in locals() and synergy_results:
                    best_syn = max(synergy_results, key=lambda x: x["synergy"])
                    if best_syn["synergy"] > 0.1:
                        recommendations.append(f"🤝 **Синергия**: пара `{best_syn['pair']}` даёт +{best_syn['synergy']*100:.1f}% информации вместе → создайте комбинированный признак")

                # Вывод рекомендаций
                if recommendations:
                    for i, rec in enumerate(recommendations, 1):
                        if "✅" in rec:
                            st.success(rec)
                        elif "⚠️" in rec or "⚡" in rec:
                            st.warning(rec)
                        else:
                            st.info(rec)
                else:
                    st.info("ℹ️ Явных паттернов не обнаружено — начните с признаков с наибольшим R")

                # Методологическое пояснение
                with st.expander("Методология расчёта", expanded=False):
                    st.markdown(f"""
                    **Параметры текущего запуска:**
                    - Целевая переменная: `{ih_target}`
                    - Sharpness: `{sharpness}` → ~{int(1/sharpness)} интервалов для непрерывных признаков
                    - Мин. наблюдений на бин: `{min_samples}`

                    **Формулы:**
                    1. Энтропия: `H(X) = -Σ p(x)·log₂(p(x))`
                    2. Взаимная информация: `I(X;Y) = Σ p(x,y)·log₂(p(x,y)/(p(x)·p(y)))`
                    3. Нормированная связь: `R(Y|X) = I(X;Y) / H(Y) ∈ [0; 1]`

                    **Дискретизация:**
                    - Категориальные: без изменений
                    - Числовые: квантильное разбиение на `n_bins ≈ 1/sharpness` интервалов
                    - Пропуски: кодируются как отдельный категориальный уровень `_MISSING_`

                    **Интерпретация для временных рядов:**
                    - Используйте лаговые версии признаков: `Xₜ₋ₖ → Yₜ`
                    - Проверяйте устойчивость метрик на разных временных отрезках
                    - Автокорреляция может завышать оценки → сравнивайте с перестановочным тестом
                    """)

    else:
        st.warning("⚠️ Для IH-анализа необходимы числовые данные. Загрузите датасет с метриками.")




# ────────────────────────────────────────────────────────────
#  ВКЛАДКА 5: МОДЕЛИРОВАНИЕ
# ────────────────────────────────────────────────────────────
with tab_modeling:
    st.markdown("""
    <div style="padding-left: 20px; margin: 20px 0; text-align: right;">
        <p style="margin: 0 0 10px 0; color: #1e293b; line-height: 1.6; font-size: 18px; font-weight: 400;">
            "Все модели неверны, но некоторые полезны".
        </p>
        <p style="margin: 0; color: #64748B; font-style: italic; font-size: 16px; line-height: 1.5;">
            — Джордж Бокс, британский статистик, внёсший вклад в такие области, как<br>
            контроль качества, планирование эксперимента, анализ временных рядов и байесовский вывод.
        </p>
    </div>
    """, unsafe_allow_html=True)