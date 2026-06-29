# tests/unit/test_export.py
import pytest
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from app.core.export import export_validation_passport_csv, export_validation_passport_excel


@pytest.fixture
def sample_passport_data():
    """Фикстура с тестовыми данными паспорта валидации."""
    df = pd.DataFrame({
        "Вид проверки": ["Стационарность", "Нормальность", "Пропуски"],
        "Метрика": ["ADF p-value", "Jarque-Bera", "% NaN"],
        "Значение": [0.03, 0.12, 2.5],
        "Статус": ["✅ Стационарен", "⚠️ Отклонение", "✅ Норма"]
    })
    metadata = {
        "document_title": "Паспорт валидации",
        "dataset_name": "test_dataset.csv",
        "platform_tagline": "CISStat TS Analysis",
        "verification": "Верифицировано",
        "generated_at": "2026-06-22 15:00:00",
        "n_rows": 1000,
        "n_cols": 10
    }
    recommendations = {
        "primary_recommendation": "ARIMA(1,1,1)",
        "available": ["ARIMA", "Exponential Smoothing"],
        "limited": ["Prophet"],
        "unavailable": ["LSTM"],
        "explanation": "DQ Score: 85.0%, Уровень качества: high"
    }
    return df, metadata, 85.0, recommendations


def test_export_csv_contains_headers(sample_passport_data):
    """Проверяем, что CSV содержит заголовки-комментарии и данные."""
    df, metadata, dq_score, _ = sample_passport_data
    csv_bytes = export_validation_passport_csv(df, metadata, dq_score)
    csv_text = csv_bytes.decode("utf-8-sig")
    
    # Проверяем наличие комментариев
    assert "# Паспорт валидации" in csv_text
    assert "# Датасет: test_dataset.csv" in csv_text
    assert "# DQ Score: 85.0%" in csv_text
    
    # Проверяем наличие данных
    assert "Стационарность" in csv_text
    assert "ADF p-value" in csv_text


def test_export_excel_has_two_sheets(sample_passport_data):
    """Проверяем, что Excel имеет два листа."""
    df, metadata, dq_score, recommendations = sample_passport_data
    excel_buffer = export_validation_passport_excel(df, metadata, dq_score, recommendations)
    
    # Загружаем Excel из буфера
    wb = load_workbook(excel_buffer)
    
    # Проверяем наличие листов
    assert "Паспорт валидации" in wb.sheetnames
    assert "Рекомендации по моделям" in wb.sheetnames


def test_export_excel_data_integrity(sample_passport_data):
    """Проверяем, что данные в Excel совпадают с исходным DataFrame."""
    df, metadata, dq_score, recommendations = sample_passport_data
    excel_buffer = export_validation_passport_excel(df, metadata, dq_score, recommendations)
    
    wb = load_workbook(excel_buffer)
    ws = wb["Паспорт валидации"]
    
    # Проверяем, что заголовки на месте (строка 5)
    assert ws.cell(row=5, column=1).value == "Вид проверки"
    assert ws.cell(row=5, column=2).value == "Метрика"
    
    # Проверяем, что данные на месте (строка 6+)
    assert ws.cell(row=6, column=1).value == "Стационарность"
    assert ws.cell(row=7, column=1).value == "Нормальность"


def test_export_excel_recommendations_structure(sample_passport_data):
    """Проверяем структуру листа рекомендаций."""
    df, metadata, dq_score, recommendations = sample_passport_data
    excel_buffer = export_validation_passport_excel(df, metadata, dq_score, recommendations)
    
    wb = load_workbook(excel_buffer)
    ws2 = wb["Рекомендации по моделям"]
    
    # Проверяем заголовок
    assert ws2['A1'].value == "РЕКОМЕНДАЦИИ ПО ВЫБОРУ МОДЕЛЕЙ"
    
    # Проверяем первичную рекомендацию
    assert "ARIMA(1,1,1)" in ws2['A3'].value
    
    # Проверяем, что модели записаны (строка 6+)
    assert ws2.cell(row=6, column=2).value == "ARIMA"
    assert ws2.cell(row=7, column=2).value == "Exponential Smoothing"


# tests/unit/test_export.py (дополнить существующий файл)
"""
Characterization-тесты для export_passport_to_excel.
"""
import pytest
import io
from openpyxl import load_workbook


class TestExportPassportToExcel:
    """Тесты для функции экспорта паспорта в Excel."""

    def test_export_creates_valid_excel(self):
        """Должна создавать валидный Excel-файл."""
        from app.core.export import export_passport_to_excel
        
        tech_info = {"Признак": "test_col", "Дата": "2026-01-26"}
        dist_stats = {"Среднее": 10.5, "Стд. отклонение": 2.3}
        ts_passport = [
            {"property": "Стационарность", "method": "ADF Test", "result": "✅ Стационарен"},
            {"property": "Тренд", "method": "R²", "result": "⚠️ Слабый"}
        ]
        recommendations = [
            ("ARIMA(p,d,q)", "Стационарен + есть автокорреляция", "Классический выбор"),
            ("Prophet", "Высокочастотные данные (D)", "Учет праздников")
        ]
        
        result = export_passport_to_excel(
            tech_info=tech_info,
            dist_stats=dist_stats,
            ts_passport=ts_passport,
            recommendations=recommendations,
            report_col="test_col"
        )
        
        assert isinstance(result, io.BytesIO)
        assert result.tell() == 0  # Указатель в начале
        
        # Проверяем, что файл читается
        wb = load_workbook(result)
        assert len(wb.sheetnames) == 2
        assert "1_Паспорт свойств" in wb.sheetnames
        assert "2_Рекомендации" in wb.sheetnames

    def test_export_contains_tech_info(self):
        """Должна записывать техническую информацию."""
        from app.core.export import export_passport_to_excel
        
        tech_info = {"Признак": "test_col", "Дата": "2026-01-26"}
        dist_stats = {}
        ts_passport = []
        recommendations = []
        
        result = export_passport_to_excel(
            tech_info=tech_info,
            dist_stats=dist_stats,
            ts_passport=ts_passport,
            recommendations=recommendations,
            report_col="test_col"
        )
        
        wb = load_workbook(result)
        ws = wb["1_Паспорт свойств"]
        
        # Ищем "Техническая информация" в ячейках
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Техническая информация" in str(cell.value):
                    found = True
                    break
        assert found

    def test_export_contains_passport_data(self):
        """Должна записывать данные паспорта."""
        from app.core.export import export_passport_to_excel
        
        tech_info = {}
        dist_stats = {}
        ts_passport = [
            {"property": "Стационарность", "method": "ADF Test", "result": "✅ Стационарен"}
        ]
        recommendations = []
        
        result = export_passport_to_excel(
            tech_info=tech_info,
            dist_stats=dist_stats,
            ts_passport=ts_passport,
            recommendations=recommendations,
            report_col="test_col"
        )
        
        wb = load_workbook(result)
        ws = wb["1_Паспорт свойств"]
        
        # Ищем "Стационарность" в ячейках
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Стационарность" in str(cell.value):
                    found = True
                    break
        assert found

    def test_export_contains_recommendations(self):
        """Должна записывать рекомендации на второй лист."""
        from app.core.export import export_passport_to_excel
        
        tech_info = {}
        dist_stats = {}
        ts_passport = []
        recommendations = [
            ("ARIMA(p,d,q)", "Стационарен", "Классический выбор")
        ]
        
        result = export_passport_to_excel(
            tech_info=tech_info,
            dist_stats=dist_stats,
            ts_passport=ts_passport,
            recommendations=recommendations,
            report_col="test_col"
        )
        
        wb = load_workbook(result)
        ws = wb["2_Рекомендации"]
        
        # Ищем "ARIMA" в ячейках
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "ARIMA" in str(cell.value):
                    found = True
                    break
        assert found

    def test_export_handles_empty_data(self):
        """Должна обрабатывать пустые данные."""
        from app.core.export import export_passport_to_excel
        
        result = export_passport_to_excel(
            tech_info={},
            dist_stats={},
            ts_passport=[],
            recommendations=[],
            report_col="test_col"
        )
        
        assert isinstance(result, io.BytesIO)
        wb = load_workbook(result)
        assert len(wb.sheetnames) == 2