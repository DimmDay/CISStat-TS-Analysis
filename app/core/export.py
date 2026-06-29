# app/core/export.py
"""
Единая логика экспорта результатов анализа.
⚠️ ВАЖНО: Этот модуль НЕ импортирует streamlit.
Все функции принимают явные аргументы и возвращают bytes/io.BytesIO.
"""
import io
from typing import Dict, List, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_validation_passport_csv(
    df_passport: pd.DataFrame,
    metadata: Dict[str, Any],
    dq_score: float
) -> bytes:
    """
    Генерирует CSV-файл с паспортом валидации.
    
    Args:
        df_passport: DataFrame с результатами валидации
        metadata: Словарь метаданных (document_title, dataset_name, platform_tagline, etc.)
        dq_score: Composite Data Quality Score (0-100)
    
    Returns:
        bytes: CSV-файл в кодировке UTF-8 с BOM
    """
    csv_comments = [
        f"# {metadata.get('document_title', 'Паспорт валидации')}",
        f"# Датасет: {metadata.get('dataset_name', 'Неизвестно')}",
        f"# {metadata.get('platform_tagline', 'CISStat TS Analysis')}",
        f"# {metadata.get('verification', 'Верифицировано')}",
        f"# Дата генерации: {metadata.get('generated_at', 'N/A')}",
        f"# DQ Score: {dq_score:.1f}%",
        ""
    ]
    csv_data = df_passport.to_csv(index=False, encoding="utf-8-sig")
    full_csv = "\n".join(csv_comments) + csv_data
    return full_csv.encode("utf-8-sig")


def export_validation_passport_excel(
    df_passport: pd.DataFrame,
    metadata: Dict[str, Any],
    dq_score: float,
    recommendations: Dict[str, Any]
) -> io.BytesIO:
    """
    Генерирует Excel-файл с паспортом валидации и рекомендациями по моделям.
    
    Args:
        df_passport: DataFrame с результатами валидации
        metadata: Словарь метаданных
        dq_score: Composite Data Quality Score (0-100)
        recommendations: Словарь с рекомендациями (primary_recommendation, available, limited, unavailable, explanation)
    
    Returns:
        io.BytesIO: Excel-файл в памяти
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Паспорт валидации"

    # ── ЛИСТ 1: ПАСПОРТ ВАЛИДАЦИИ ──
    # Шапка
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = metadata.get('document_title', 'Паспорт валидации')
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A2:I2')
    ds_cell = ws['A2']
    ds_cell.value = f"Датасет: {metadata.get('dataset_name', 'Неизвестно')}"
    ds_cell.font = Font(bold=True, size=12)
    ds_cell.alignment = Alignment(horizontal="left")

    ws.merge_cells('A3:I3')
    info_cell = ws['A3']
    info_cell.value = f"Строк: {metadata.get('n_rows', 'N/A')} | Колонок: {metadata.get('n_cols', 'N/A')} | Дата: {metadata.get('generated_at', 'N/A')}"
    info_cell.font = Font(size=10, color="666666")
    ws.row_dimensions[4].height = 8

    # Заголовки таблицы
    headers = list(df_passport.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="048A81", end_color="048A81", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style='thin'))

    # Данные с условным форматированием
    for row_idx, row in enumerate(df_passport.to_dict('records'), 6):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header == "Статус":
                if "✅" in str(cell.value):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "❌" in str(cell.value):
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif "⚠️" in str(cell.value):
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Подпись
    last_data_row = 5 + len(df_passport)
    ws.row_dimensions[last_data_row + 1].height = 8

    ws.merge_cells(f'A{last_data_row + 2}:I{last_data_row + 2}')
    sign_cell_1 = ws.cell(row=last_data_row + 2, column=1)
    sign_cell_1.value = metadata.get('platform_tagline', 'CISStat TS Analysis') + "."
    sign_cell_1.font = Font(bold=True, size=11, color="1D3557")
    sign_cell_1.alignment = Alignment(horizontal="left")

    ws.merge_cells(f'A{last_data_row + 3}:I{last_data_row + 3}')
    sign_cell_2 = ws.cell(row=last_data_row + 3, column=1)
    sign_cell_2.value = metadata.get('verification', 'Верифицировано СтатКомитетом СНГ') + "."
    sign_cell_2.font = Font(bold=True, size=11, color="1D3557")
    sign_cell_2.alignment = Alignment(horizontal="left")

    ws.merge_cells(f'A{last_data_row + 4}:I{last_data_row + 4}')
    date_cell = ws.cell(row=last_data_row + 4, column=1)
    date_cell.value = f"Дата генерации: {metadata.get('generated_at', 'N/A')}"
    date_cell.font = Font(italic=True, size=10, color="666666")

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    # ── ЛИСТ 2: РЕКОМЕНДАЦИИ ПО МОДЕЛЯМ ──
    ws2 = wb.create_sheet(title="Рекомендации по моделям")

    ws2.merge_cells('A1:D1')
    title2 = ws2['A1']
    title2.value = "РЕКОМЕНДАЦИИ ПО ВЫБОРУ МОДЕЛЕЙ"
    title2.font = Font(bold=True, size=14, color="FFFFFF")
    title2.fill = PatternFill(start_color="048A81", end_color="048A81", fill_type="solid")
    title2.alignment = Alignment(horizontal="center", vertical="center")

    ws2.merge_cells('A2:D2')
    info2 = ws2['A2']
    info2.value = f"Датасет: {metadata.get('dataset_name', 'Неизвестно')} | DQ Score: {dq_score:.1f}%"
    info2.font = Font(size=11)
    info2.alignment = Alignment(horizontal="left")

    ws2.merge_cells('A3:D3')
    primary_cell = ws2['A3']
    primary_cell.value = f" Первичная рекомендация: {recommendations.get('primary_recommendation', 'Нет данных')}"
    primary_cell.font = Font(bold=True, size=12, color="1D3557")
    primary_cell.fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
    primary_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[3].height = 25

    headers2 = ["Категория", "Модель", "Статус применимости", "Комментарий"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=5, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 6
    for m in recommendations.get("available", []):
        ws2.cell(row=row, column=1, value="Доступно")
        ws2.cell(row=row, column=2, value=m)
        ws2.cell(row=row, column=3, value="✅ Рекомендуется")
        ws2.cell(row=row, column=4, value="Можно применять без ограничений")
        for c in range(1, 5):
            ws2.cell(row=row, column=c).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

    for m in recommendations.get("limited", []):
        ws2.cell(row=row, column=1, value="Ограничено")
        ws2.cell(row=row, column=2, value=m)
        ws2.cell(row=row, column=3, value="️ С осторожностью")
        ws2.cell(row=row, column=4, value="Требует предобработки / валидации")
        for c in range(1, 5):
            ws2.cell(row=row, column=c).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1

    for m in recommendations.get("unavailable", []):
        ws2.cell(row=row, column=1, value="Недоступно")
        ws2.cell(row=row, column=2, value=m)
        ws2.cell(row=row, column=3, value="❌ Не применимо")
        ws2.cell(row=row, column=4, value="Недостаточно данных / низкое качество")
        for c in range(1, 5):
            ws2.cell(row=row, column=c).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1

    row += 1
    ws2.merge_cells(f'A{row}:D{row}')
    expl_cell = ws2.cell(row=row, column=1, value="Обоснование:")
    expl_cell.font = Font(bold=True, size=11, color="1D3557")
    row += 1
    ws2.merge_cells(f'A{row}:D{row}')
    expl_text = ws2.cell(row=row, column=1, value=recommendations.get("explanation", ""))
    expl_text.font = Font(size=10, italic=True)
    expl_text.alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[row].height = 50

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# app/core/export.py
"""
Модуль для экспорта данных в Excel.
Извлечено из app.py (пункт B.10 EXTRACTION_PLAN.md).
"""
import io
from datetime import datetime as dt_now
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Any


def export_passport_to_excel(
    tech_info: Dict[str, Any],
    dist_stats: Dict[str, Any],
    ts_passport: List[Dict[str, str]],
    recommendations: List[tuple],
    report_col: str
) -> io.BytesIO:
    """
    Экспортирует паспорт свойств временного ряда в Excel-файл.
    
    Args:
        tech_info: Техническая информация (словарь ключ-значение)
        dist_stats: Статистики распределения (словарь ключ-значение)
        ts_passport: Список словарей с данными паспорта
                    [{"property": "...", "method": "...", "result": "..."}]
        recommendations: Список кортежей с рекомендациями
                        [(model, condition, justification), ...]
        report_col: Название анализируемого признака
        
    Returns:
        BytesIO с Excel-файлом
        
    Examples:
        >>> tech_info = {"Признак": "test_col"}
        >>> dist_stats = {"Среднее": 10.5}
        >>> ts_passport = [{"property": "Стационарность", "method": "ADF", "result": "✅"}]
        >>> recommendations = [("ARIMA", "Стационарен", "Классика")]
        >>> result = export_passport_to_excel(tech_info, dist_stats, ts_passport, recommendations, "test_col")
        >>> isinstance(result, io.BytesIO)
        True
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "1_Паспорт свойств"
    
    # Стили
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_font = Font(bold=True, size=16)
    footer_font = Font(bold=True, color="0369A1")
    footer_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Вспомогательная функция для записи таблицы
    def write_table(data_dict: Dict[str, Any], start_row: int, title: str) -> int:
        ws.merge_cells(f"A{start_row}:B{start_row}")
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        start_row += 1
        ws.cell(row=start_row, column=1, value="Параметр").font = Font(bold=True)
        ws.cell(row=start_row, column=2, value="Значение").font = Font(bold=True)
        start_row += 1
        for k, v in data_dict.items():
            c1 = ws.cell(row=start_row, column=1, value=k)
            c2 = ws.cell(row=start_row, column=2, value=v)
            c1.border = thin_border
            c2.border = thin_border
            start_row += 1
        return start_row + 1
    
    # Заголовок
    row = 1
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value=f"Предварительный отчет о свойствах признака: {report_col}")
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    row += 2
    ws.cell(row=row, column=1, value=f"Исследуемый параметр: {report_col}").font = header_font
    ws.cell(row=row, column=3, value=f"Дата: {dt_now.now().strftime('%d.%m.%Y %H:%M')}").font = header_font
    row += 2
    
    # Записываем данные
    row = write_table(tech_info, row, "Техническая информация")
    row = write_table(dist_stats, row, "Статистики распределения")
    
    # Итоговый паспорт свойств
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws.cell(row=row, column=1, value="Итоговый паспорт свойств")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    ws.cell(row=row, column=1, value="Свойство").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Метод").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Результат").font = Font(bold=True)
    row += 1
    
    for item in ts_passport:
        c1 = ws.cell(row=row, column=1, value=item.get("property", ""))
        c2 = ws.cell(row=row, column=2, value=item.get("method", ""))
        c3 = ws.cell(row=row, column=3, value=item.get("result", ""))
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        
        result = item.get("result", "")
        if "✅" in result:
            c3.fill = green_fill
        elif "⚠️" in result:
            c3.fill = yellow_fill
        elif "❌" in result:
            c3.fill = red_fill
        row += 1
    
    # Второй лист: Рекомендации
    ws_rec = wb.create_sheet("2_Рекомендации")
    row_rec = 1
    ws_rec.merge_cells(f"A{row_rec}:C{row_rec}")
    cell = ws_rec.cell(row=row_rec, column=1, value="Рекомендуемые модели и обоснование")
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    row_rec += 2
    ws_rec.cell(row=row_rec, column=1, value="Модель").font = header_font
    ws_rec.cell(row=row_rec, column=2, value="Условие применения").font = header_font
    ws_rec.cell(row=row_rec, column=3, value="Обоснование").font = header_font
    row_rec += 1
    
    for model, condition, justification in recommendations:
        ws_rec.cell(row=row_rec, column=1, value=model).font = Font(bold=True)
        ws_rec.cell(row=row_rec, column=2, value=condition)
        ws_rec.cell(row=row_rec, column=3, value=justification)
        for col in range(1, 4):
            ws_rec.cell(row=row_rec, column=col).border = thin_border
        row_rec += 1
    
    # Методологическое пояснение
    row_rec += 1
    ws_rec.merge_cells(f"A{row_rec}:C{row_rec}")
    cell = ws_rec.cell(row=row_rec, column=1, value="🔄 Методология выбора моделей")
    cell.font = Font(bold=True, size=11, color="0369A1")
    cell.alignment = Alignment(horizontal='left')
    row_rec += 1
    ws_rec.cell(row=row_rec, column=1, value="• Статистические тесты (ADF, Ljung-Box) → выбор класса моделей")
    ws_rec.cell(row=row_rec+1, column=1, value="• Спектральный анализ (ACF, FFT) → параметры сезонности и признаки")
    ws_rec.cell(row=row_rec+2, column=1, value="• Корреляционный анализ → отбор признаков, борьба с мультиколлинеарностью")
    ws_rec.cell(row=row_rec+3, column=1, value="• Порядок действий: 1) Преобразования → 2) Признаки → 3) Подбор параметров → 4) Валидация")
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws_rec.column_dimensions['A'].width = 30
    ws_rec.column_dimensions['B'].width = 35
    ws_rec.column_dimensions['C'].width = 60
    
    # Футер
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws.cell(row=row, column=1, value=" Исследовано платформой CISStat TS Analytics | ✅ Верифицировано СтатКомитетом СНГ")
    cell.font = footer_font
    cell.fill = footer_fill
    cell.alignment = Alignment(horizontal='center')
    
    # Сохраняем в BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output